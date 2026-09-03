import gc
import hashlib
import time
from copy import copy, deepcopy
from functools import partial
from concurrent.futures import ProcessPoolExecutor, as_completed
import numpy as np
import pandas as pd
from math import isclose, isfinite, sqrt
from sklearn.preprocessing import StandardScaler
from copulas.multivariate import GaussianMultivariate
import networkx as nx
import matplotlib.cm as cm
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pyomo.opt as po
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from centralized_coordination import combine_networks
from network_data import NetworkData
from load import Load
from shared_energy_storage import SharedEnergyStorage
from planning_parameters import PlanningParameters
from shared_energy_storage_data import SharedEnergyStorageData
from model_construction_helpers import *
from helper_functions import *


# ======================================================================================================================
#   Class SHARED RESOURCES PLANNING
# ======================================================================================================================
class SharedResourcesPlanning:

    def __init__(self, data_dir, filename):
        self.name = filename.replace('.json', '')
        self.data_dir = data_dir
        self.filename = filename
        self.market_data_file = str()
        self.num_market_scenarios = int()
        self.plot_market_data = bool()
        self.results_dir = os.path.join(data_dir, 'Results')
        self.diagrams_dir = os.path.join(data_dir, 'Diagrams')
        self.logs_dir = os.path.join(self.results_dir, 'Logs')
        self.params_file = str()
        self.years = dict()
        self.days = dict()
        self.num_instants = int()
        self.random_seed = None
        self.scenario_metadata = dict()
        self.discount_factor = float()
        self.cost_energy_p = dict()
        self.cost_flex = dict()
        self.prob_market_scenarios = dict()
        self.parallel_execution = bool()
        self.distribution_networks = dict()
        self.transmission_network = NetworkData()
        self.shared_ess_data = SharedEnergyStorageData()
        self.active_distribution_network_nodes = list()
        self.params = PlanningParameters()

    def run_planning_problem(self, debug_flag=False):
        print('[INFO] Running PLANNING PROBLEM...')
        _run_planning_problem(self, debug_flag=debug_flag)

    def run_operational_planning(self, type='distributed', candidate_solution=dict(), num_steps=8,
                                 print_results=False, filename=str(), debug_flag=False,
                                 initial_state=None, return_state=False):

        if type == 'distributed':
            print('[INFO] Running OPERATIONAL PLANNING (DISTRIBUTED)...')
            if not candidate_solution:
                candidate_solution = self.get_initial_candidate_solution()
            convergence, results, models, sensitivities, primal_evolution, execution_time, state = _run_operational_planning(self, candidate_solution, initial_state=initial_state, debug_flag=debug_flag)
            if print_results and not state.get('initialization_failed', False):
                if not filename:
                    filename = f'{self.name}_distributed'
                self.write_operational_planning_results_to_excel(
                    models,
                    results,
                    filename=filename,
                    primal_evolution=primal_evolution,
                    admm_diagnostics=state.get('admm_diagnostics', []),
                    solver_recovery_diagnostics=state.get('solver_recovery_diagnostics', []),
                    execution_time=execution_time,
                )
            elif print_results:
                print('[WARNING] Operational results were not written because initialization failed.')
            output = convergence, results, models, sensitivities, primal_evolution
            if return_state:
                return (*output, state)
            return output

        elif type == 'hierarchical':
            print('[INFO] Running OPERATIONAL PLANNING (HIERARCHICAL)...')
            results, models, execution_time = _run_operational_planning_hierarchical(self, num_steps=num_steps, debug_flag=debug_flag, print_pq_map=print_results)
            if print_results:
                if not filename:
                    filename = f'{self.name}_hierarchical'
                self.write_operational_planning_results_hierarchical_to_excel(models, results, filename=filename, execution_time=execution_time)
            return results, models

        elif type == 'centralized':
            print('[INFO] Running OPERATIONAL PLANNING (CENTRALIZED)...')
            centralized_network, results, model = _run_operational_planning_centralized(self, debug_flag=debug_flag)
            if print_results:
                processed_results = centralized_network.process_results(model, results)
                filename = f'{self.name}_operational_planning_results_centralized'
                centralized_network.write_optimization_results_to_excel(processed_results, filename=filename)
            return results, model

        elif type == 'uncoordinated':
            results, models = self.run_without_coordination(print_results=print_results)
            return results, models

        else:
            print('[ERROR] Unrecognized COORDINATED OPERATIONAL PLANNING TYPE!...')
            exit(ERROR_SPECIFICATION_FILE)

    def run_without_coordination(self, print_results=False):

        print('[INFO] Running PLANNING PROBLEM WITHOUT COORDINATION...')

        # print('[INFO] \t - Transmission Network. Power factor control switched off!')
        # print('[INFO] \t - Transmission Network. Flexible load control switched off!')
        # self.transmission_network.params.fl_reg = False
        # for year in self.transmission_network.years:
        #     for day in self.transmission_network.days:
        #         for generator in self.transmission_network.network[year][day].generators:
        #             if generator.is_curtaillable():
        #                 generator.power_factor_control = False

        # print('[INFO] \t - Distribution Networks. Power factor control switched off!')
        # print('[INFO] \t - Distribution Networks. Flexible load control switched off!')
        # for node_id in self.distribution_networks:
        #     distribution_network = self.distribution_networks[node_id]
        #     distribution_network.params.fl_reg = False
        #     for year in distribution_network.years:
        #         for day in distribution_network.days:
        #             for generator in distribution_network.network[year][day].generators:
        #                 generator.power_factor_control = False

        results, models, execution_time = _run_operational_planning_without_coordination(self)
        if print_results:
            self.write_operational_planning_results_without_coordination_to_excel(models, results, execution_time=execution_time)

        # print('[INFO] \t - Transmission Network. Power factor control switched back on!')
        # print('[INFO] \t - Transmission Network. Flexible load control switched back on!')
        # self.transmission_network.params.fl_reg = True
        # for year in self.transmission_network.years:
        #     for day in self.transmission_network.days:
        #         for generator in self.transmission_network.network[year][day].generators:
        #             if generator.is_curtaillable():
        #                 generator.power_factor_control = True

        # print('[INFO] \t - Distribution Networks. Power factor control switched back on!')
        # print('[INFO] \t - Distribution Networks. Flexible load control switched back on!')
        # for node_id in self.distribution_networks:
        #     distribution_network = self.distribution_networks[node_id]
        #     distribution_network.params.fl_reg = True
        #     for year in distribution_network.years:
        #         for day in distribution_network.days:
        #             for generator in distribution_network.network[year][day].generators:
        #                 generator.power_factor_control = True

        return results, models

    def combine_networks(self):
        transmission_network = self.transmission_network
        distribution_networks = self.distribution_networks
        return combine_networks(transmission_network, distribution_networks)

    def get_operational_recourse_value(self, models):
        return _get_operational_recourse_value(self, models)

    def get_operational_recourse_components(self, models):
        return _get_operational_recourse_components(self, models)

    def get_primal_value(self, tso_model, dso_models, esso_model):
        return _get_primal_value(self, tso_model, dso_models, esso_model)

    def add_benders_cut(self, model, recourse_value, sensitivities, candidate_solution):
        return _add_benders_cut(self, model, recourse_value, sensitivities, candidate_solution)

    def update_admm_consensus_variables(self, tso_model, dso_models, esso_model, consensus_vars, dual_vars, results, params, update_tn=False, update_dns=False, update_sess=False):
        self.update_interface_power_flow_variables(tso_model, dso_models, consensus_vars, dual_vars, results, params, update_tn=update_tn, update_dns=update_dns)
        self.update_shared_energy_storage_variables(tso_model, dso_models, esso_model, consensus_vars['ess'], dual_vars['ess'], results, params, update_tn=update_tn, update_dns=update_dns, update_sess=update_sess)

    def update_interface_power_flow_variables(self, tso_model, dso_models, interface_vars, dual_vars, results, params, update_tn=True, update_dns=True):
        _update_interface_power_flow_variables(self, tso_model, dso_models, interface_vars, dual_vars, results, params, update_tn=update_tn, update_dns=update_dns)

    def update_shared_energy_storage_variables(self, tso_model, dso_models, esso_model, consensus_vars, dual_vars, results, params, update_tn=True, update_dns=True, update_sess=True):
        _update_shared_energy_storage_variables(self, tso_model, dso_models, esso_model, consensus_vars, dual_vars, results, params, update_tn=update_tn, update_dns=update_dns, update_sess=update_sess)

    def read_planning_problem(self):
        _read_planning_problem(self)

    def read_market_data_from_file(self):
        _read_market_data_from_file(self)

    def read_planning_parameters_from_file(self):
        filename = os.path.join(self.data_dir, self.params_file)
        self.params.read_parameters_from_file(filename)

    def write_planning_results_to_excel(self, master_problem_model, operational_planning_models, operational_results=dict(), bound_evolution=dict(), execution_time=float()):
        filename = os.path.join(self.results_dir, self.name + '_planning_results.xlsx')
        processed_results = _process_operational_planning_results(self, operational_planning_models['tso'], operational_planning_models['dso'], operational_planning_models['esso'], operational_results)
        shared_ess_cost = self.shared_ess_data.get_investment_cost_and_rated_capacity(master_problem_model)
        shared_ess_capacity = self.shared_ess_data.get_available_capacity(operational_planning_models['esso'])
        salvage_value_results = self.shared_ess_data.get_salvage_value_results(
            operational_planning_models['esso']
        )
        _write_planning_results_to_excel(
            self,
            processed_results,
            bound_evolution=bound_evolution,
            shared_ess_cost=shared_ess_cost,
            shared_ess_capacity=shared_ess_capacity,
            salvage_value_results=salvage_value_results,
            filename=filename,
            execution_time=execution_time,
        )

    def write_operational_planning_results_to_excel(self, optimization_models, results, filename=str(),
                                                     primal_evolution=list(), admm_diagnostics=list(),
                                                     solver_recovery_diagnostics=list(),
                                                     execution_time=float()):
        if not filename:
            filename = 'operational_planning_results'
        processed_results = _process_operational_planning_results(self, optimization_models['tso'], optimization_models['dso'], optimization_models['esso'], results)
        shared_ess_capacity = self.shared_ess_data.get_available_capacity(optimization_models['esso'])
        salvage_value_results = self.shared_ess_data.get_salvage_value_results(
            optimization_models['esso']
        )
        _write_operational_planning_results_to_excel(
            self,
            processed_results,
            primal_evolution=primal_evolution,
            admm_diagnostics=admm_diagnostics,
            solver_recovery_diagnostics=solver_recovery_diagnostics,
            shared_ess_capacity=shared_ess_capacity,
            salvage_value_results=salvage_value_results,
            filename=filename,
            execution_time=execution_time,
        )

    def write_operational_planning_results_hierarchical_to_excel(self, optimization_models, results, filename=str(), execution_time=float()):
        if not filename:
            filename = 'operational_planning_results_hierarchical'
        filename = os.path.join(self.results_dir, filename + '.xlsx')
        processed_results = _process_operational_planning_results_hierarchical(self, optimization_models['tso'], optimization_models['dso'], results)
        _write_operational_planning_results_hierarchical_to_excel(self, processed_results, filename, execution_time=execution_time)

    def write_operational_planning_results_without_coordination_to_excel(self, optimization_models, results, filename=str(), execution_time=float()):
        if not filename:
            filename = 'operational_planning_results_hierarchical_no_coordination'
        filename = os.path.join(self.results_dir, self.name + '_operational_planning_results_no_coordination.xlsx')
        processed_results = _process_operational_planning_results_no_coordination(self, optimization_models['tso'], optimization_models['dso'], results)
        _write_operational_planning_results_no_coordination_to_excel(self, processed_results, filename, execution_time=execution_time)

    def plot_market_price_scenarios(self):
        years_to_plot = list(self.years)[0]
        _plot_market_price_scenarios(self, years_to_plot=[years_to_plot], save_dir=self.diagrams_dir)

    def get_initial_candidate_solution(self):
        return _get_initial_candidate_solution(self)

    def get_test_candidate_solution(self, s_inv, e_inv, node_id, investment_year):
        return _get_test_candidate_solution(self, node_id, investment_year, s_inv=s_inv, e_inv=e_inv)

    def plot_diagram(self):
        _plot_networkx_diagram(self)


# ======================================================================================================================
#  PLANNING functions
# ======================================================================================================================
def _run_planning_problem(planning_problem, debug_flag=False):

    shared_ess_data = planning_problem.shared_ess_data
    benders_parameters = planning_problem.params.benders
    positive_bootstrap_params = benders_parameters.positive_bootstrap
    sensitivity_probe_params = benders_parameters.sensitivity_probe
    finite_difference_params = benders_parameters.finite_difference
    lower_level_models = dict()
    operational_results = dict()

    # ------------------------------------------------------------------------------------------------------------------
    # 0. Initialization
    iteration = 1
    convergence = False
    from_warm_start = False
    upper_bound = float('inf')
    master_estimate_evolution = list()
    upper_bound_evolution = list()
    investment_cost_evolution = list()
    alpha_evolution = list()
    operational_recourse_evolution = list()
    gross_operational_cost_evolution = list()
    terminal_salvage_value_evolution = list()
    candidate_total_evolution = list()
    esso_violation_evolution = list()
    gap_signed_evolution = list()
    gap_abs_evolution = list()
    gap_rel_evolution = list()
    finite_difference_results = list()
    sensitivity_probe_diagnostics = list()
    admm_diagnostics = list()
    solver_recovery_diagnostics = list()
    operational_state = None
    sensitivities = None
    incumbent = None
    positive_validation_reference = None
    incumbent_update_evolution = list()
    candidate_source_evolution = list()
    operational_initialization_evolution = list()
    candidate_source = 'master_solution'
    positive_bootstrap_used = False
    positive_bootstrap_iteration = None
    operational_reference_state = None
    termination_reason = None
    print_memory_usage("Start of planning problem", debug_flag)

    start = time.time()
    master_problem_model = planning_problem.shared_ess_data.build_master_problem()
    master_result = shared_ess_data.optimize_master_problem(master_problem_model)
    if not master_result or master_result.solver.termination_condition != po.TerminationCondition.optimal:
        print("[ERROR] Benders-type master problem did not solve to optimality. Exiting planning loop.")
        print('[INFO] Planning termination reason: initial_master_solve_failure.')
        return
    candidate_solution = shared_ess_data.get_candidate_solution(master_problem_model)

    # Benders-type main cycle
    while iteration <= benders_parameters.num_max_iters and not convergence:

        print(f'=============================================== ITERATION #{iteration} ==============================================')

        _print_candidate_solution(candidate_solution)
        print_memory_usage(f"Before subproblem (iter {iteration})", debug_flag)
        print_results = False
        if iteration == 1 or debug_flag:
            print_results = True

        # 1. Subproblem
        # 1.1. Solve operational planning, with fixed investment variables,
        # 1.2. Get coupling constraints' sensitivities (subproblem)
        # 1.3. Get the operational recourse value and local sensitivities
        candidate_initial_state = (operational_reference_state if (sensitivity_probe_params.enabled and candidate_source == 'master_solution' and positive_bootstrap_used) else None)
        if candidate_initial_state is not None:
            print('[INFO] Operational candidate initialization: positive reference state.')

        candidate_initialization_source = ('positive_reference_state' if candidate_initial_state is not None else 'cold')

        operational_convergence, operational_results, lower_level_models, sensitivities, _, operational_state = planning_problem.run_operational_planning(
            candidate_solution=candidate_solution,
            print_results=print_results,
            filename=f'{planning_problem.name}_operational_planning_results_distributed_without ESS',
            initial_state=candidate_initial_state,
            return_state=True,
        )

        for diagnostic in operational_state.get('admm_diagnostics', []):
            diagnostic_with_outer_iteration = dict(diagnostic)
            diagnostic_with_outer_iteration['outer_iteration'] = iteration
            admm_diagnostics.append(diagnostic_with_outer_iteration)
        for diagnostic in operational_state.get('solver_recovery_diagnostics', []):
            diagnostic_with_outer_iteration = dict(diagnostic)
            diagnostic_with_outer_iteration['outer_iteration'] = iteration
            solver_recovery_diagnostics.append(diagnostic_with_outer_iteration)

        initialization_failed = operational_state.get('initialization_failed', False)
        investment_cost = pe.value(master_problem_model.investment_cost)
        alpha = None
        master_estimate = None
        if candidate_source == 'master_solution':
            alpha = pe.value(master_problem_model.alpha)
            master_estimate = pe.value(master_problem_model.objective)
        esso_violation = None
        if not initialization_failed:
            esso_violation = shared_ess_data.get_feasibility_violation(lower_level_models['esso'])
        candidate_is_feasible = (
            operational_convergence
            and esso_violation is not None
            and esso_violation <= BENDERS_FEASIBILITY_TOLERANCE
        )

        operational_recourse = None
        gross_operational_cost = None
        terminal_salvage_value = None
        candidate_total = None
        incumbent_updated = False
        evaluated_candidate = None
        if candidate_is_feasible:
            recourse_components = planning_problem.get_operational_recourse_components(lower_level_models)
            gross_operational_cost = recourse_components['gross_operational_cost']
            terminal_salvage_value = recourse_components['terminal_salvage_value']
            operational_recourse = recourse_components['net_operational_recourse']
            candidate_total = investment_cost + operational_recourse
            evaluated_candidate = {
                'iteration': iteration,
                'candidate_source': candidate_source,
                'candidate_solution': deepcopy(candidate_solution),
                'investment_cost': investment_cost,
                'gross_operational_cost': gross_operational_cost,
                'terminal_salvage_value': terminal_salvage_value,
                'operational_recourse': operational_recourse,
                'candidate_total': candidate_total,
                'esso_violation': esso_violation,
                'models': lower_level_models,
                'results': operational_results,
                'sensitivities': deepcopy(sensitivities),
                'state': operational_state,
            }
            if sensitivity_probe_params.enabled and candidate_source == 'positive_bootstrap':
                operational_reference_state = operational_state
            if candidate_total < upper_bound:
                upper_bound = candidate_total
                incumbent_updated = True
                incumbent = evaluated_candidate
            if (
                    finite_difference_params.enabled
                    and _select_finite_difference_investment(
                        evaluated_candidate['candidate_solution'], finite_difference_params
                    ) is not None
                    and (
                        positive_validation_reference is None
                        or candidate_total < positive_validation_reference['candidate_total']
                    )):
                positive_validation_reference = evaluated_candidate

        gap_signed = None
        gap_abs = None
        gap_rel = None
        if isfinite(upper_bound) and master_estimate is not None:
            gap_signed = upper_bound - master_estimate
            gap_abs = abs(gap_signed)
            gap_rel = gap_abs / max(abs(upper_bound), 1e-6)

        master_estimate_evolution.append(master_estimate)
        upper_bound_evolution.append(upper_bound if isfinite(upper_bound) else None)
        investment_cost_evolution.append(investment_cost)
        alpha_evolution.append(alpha)
        operational_recourse_evolution.append(operational_recourse)
        gross_operational_cost_evolution.append(gross_operational_cost)
        terminal_salvage_value_evolution.append(terminal_salvage_value)
        candidate_total_evolution.append(candidate_total)
        esso_violation_evolution.append(esso_violation)
        gap_signed_evolution.append(gap_signed)
        gap_abs_evolution.append(gap_abs)
        gap_rel_evolution.append(gap_rel)
        incumbent_update_evolution.append(incumbent_updated)
        candidate_source_evolution.append(candidate_source)
        operational_initialization_evolution.append(candidate_initialization_source)

        master_estimate_text = (
            f'{master_estimate:.2f}' if master_estimate is not None else 'N/A'
        )
        alpha_text = f'{alpha:.2f}' if alpha is not None else 'N/A'
        recourse_text = f'{operational_recourse:.2f}' if operational_recourse is not None else 'N/A'
        gross_recourse_text = (
            f'{gross_operational_cost:.2f}' if gross_operational_cost is not None else 'N/A'
        )
        salvage_text = f'{terminal_salvage_value:.2f}' if terminal_salvage_value is not None else 'N/A'
        candidate_total_text = f'{candidate_total:.2f}' if candidate_total is not None else 'N/A'
        upper_bound_text = f'{upper_bound:.2f}' if isfinite(upper_bound) else 'N/A'
        gap_text = f'{gap_signed / max(abs(upper_bound), 1e-6) * 100:.2f}%' if gap_signed is not None else 'N/A'
        esso_violation_text = f'{esso_violation:.6f}' if esso_violation is not None else 'N/A'
        print(
            f"[INFO] Iteration #{iteration} | Source = {candidate_source} | "
            f"Master = {master_estimate_text} | Alpha = {alpha_text} | "
            f"Investment = {investment_cost:.2f} | Gross recourse = {gross_recourse_text} | "
            f"Salvage = {salvage_text} | Net recourse = {recourse_text} | "
            f"Candidate = {candidate_total_text} | UB = {upper_bound_text} | Gap = {gap_text} | "
            f"ESSO violation = {esso_violation_text}"
        )

        if planning_problem.params.gc:
            gc.collect()
        print_memory_usage(f"After subproblem (iter {iteration})", debug_flag)

        if not operational_convergence:
            if initialization_failed:
                termination_reason = 'operational_initialization_failure'
                print(
                    '[WARNING] Operational initialization failed. No ADMM cycle or formal Benders '
                    'feasibility cut is available; stopping the outer loop.'
                )
            else:
                termination_reason = 'operational_admm_failure'
                print("[WARNING] ADMM did not converge. No formal Benders feasibility cut is available; stopping the outer loop.")
            break
        if esso_violation > BENDERS_FEASIBILITY_TOLERANCE:
            termination_reason = 'operational_infeasibility'
            print(
                f"[WARNING] Shared ESS feasibility violation {esso_violation:.6f} exceeds "
                f"{BENDERS_FEASIBILITY_TOLERANCE:.6f}. No formal feasibility cut is available; stopping the outer loop."
            )
            break
        if (
                master_estimate is not None
                and master_estimate > upper_bound + benders_parameters.tol_abs):
            termination_reason = 'local_model_crossed_incumbent'
            print(
                "[WARNING] The Benders-type master estimate exceeds the incumbent feasible objective. "
                "The local cuts are not global lower bounds; stopping without claiming optimality."
            )
            break
        if (
                gap_rel is not None
                and (gap_rel < benders_parameters.tol_rel
                     or gap_abs <= benders_parameters.tol_abs)):
            convergence = True
            termination_reason = 'converged'
            break
        if iteration == benders_parameters.num_max_iters:
            termination_reason = 'maximum_iterations'
            break

        if (
                sensitivity_probe_params.enabled
                and not (
                    positive_bootstrap_params.enabled
                    and not positive_bootstrap_used
                    and _is_zero_investment_candidate(candidate_solution)
                )
                and _has_missing_investment_sensitivities(
                    planning_problem, sensitivities
                )):
            sensitivities, probe_diagnostics, probe_state = (
                _complete_missing_sensitivities_with_probe(
                    planning_problem,
                    candidate_solution,
                    sensitivities,
                    sensitivity_probe_params,
                    iteration,
                    initial_state=operational_reference_state,
                    debug_flag=debug_flag,
                )
            )
            sensitivity_probe_diagnostics.extend(probe_diagnostics)
            if evaluated_candidate is not None:
                evaluated_candidate['sensitivities'] = deepcopy(sensitivities)
            if probe_state is not None:
                for diagnostic in probe_state.get('admm_diagnostics', []):
                    diagnostic_with_outer_iteration = dict(diagnostic)
                    diagnostic_with_outer_iteration['outer_iteration'] = iteration
                    diagnostic_with_outer_iteration['evaluation_type'] = 'sensitivity_probe'
                    admm_diagnostics.append(diagnostic_with_outer_iteration)
                for diagnostic in probe_state.get('solver_recovery_diagnostics', []):
                    diagnostic_with_outer_iteration = dict(diagnostic)
                    diagnostic_with_outer_iteration['outer_iteration'] = iteration
                    diagnostic_with_outer_iteration['evaluation_type'] = 'sensitivity_probe'
                    solver_recovery_diagnostics.append(diagnostic_with_outer_iteration)

        print_memory_usage(f"Before master problem solve (iter {iteration})", debug_flag)

        # 2. Solve Master problem
        # 2.1. Add a local sensitivity cut based on the evaluated recourse value
        # 2.2. Run master problem optimization
        # 2.3. Get the next common investment plan
        cut_added = planning_problem.add_benders_cut(master_problem_model, operational_recourse, sensitivities, candidate_solution)
        if not cut_added:
            if (
                    positive_bootstrap_params.enabled
                    and not positive_bootstrap_used
                    and _is_zero_investment_candidate(candidate_solution)):
                try:
                    candidate_solution = _build_positive_bootstrap_candidate(
                        planning_problem, positive_bootstrap_params
                    )
                except ValueError as exc:
                    termination_reason = 'positive_bootstrap_failure'
                    print(f'[WARNING] Positive bootstrap could not be constructed: {exc}')
                    break
                shared_ess_data.load_candidate_solution_into_master_model(
                    master_problem_model, candidate_solution
                )
                positive_bootstrap_used = True
                positive_bootstrap_iteration = iteration + 1
                candidate_source = 'positive_bootstrap'
                print(
                    '[INFO] Sensitivities at the zero-capacity plan are incomplete. '
                    f'Evaluating a positive bootstrap plan at iteration '
                    f'{positive_bootstrap_iteration} before resolving the master problem.'
                )
                iteration += 1
                continue
            termination_reason = 'sensitivity_unavailable'
            print("[WARNING] Sensitivity information is incomplete. Stopping the outer loop without adding a cut.")
            break
        master_result = shared_ess_data.optimize_master_problem(master_problem_model, from_warm_start=from_warm_start)
        if not master_result or master_result.solver.termination_condition != po.TerminationCondition.optimal:
            termination_reason = 'master_solve_failure'
            print("[WARNING] Benders-type master problem did not solve to optimality. Stopping the outer loop.")
            break

        if planning_problem.params.gc:
            gc.collect()
        print_memory_usage(f"After master problem solve (iter {iteration})", debug_flag)

        # Get new candidate solution
        candidate_solution = shared_ess_data.get_candidate_solution(master_problem_model)
        candidate_source = 'master_solution'
        print_memory_usage(f"After GC (iter {iteration})", debug_flag)

        iteration += 1
        from_warm_start = True

    if termination_reason is None:
        termination_reason = 'converged' if convergence else 'maximum_iterations'

    if convergence:
        print(f"[INFO] Benders-type procedure converged at iteration {iteration}.")
    else:
        print('[WARNING] Convergence not obtained!')
    print(f'[INFO] Planning termination reason: {termination_reason}.')

    validation_after_stop = (
        finite_difference_params.validate_after_heuristic_stop
        and termination_reason in {'local_model_crossed_incumbent', 'maximum_iterations'}
    )
    validation_reference = None
    validation_source = None
    validation_reference_is_incumbent = None
    if finite_difference_params.enabled and incumbent is not None and (
            convergence or validation_after_stop):
        validation_reference, validation_source, validation_reference_is_incumbent = (
            _select_finite_difference_validation_reference(
                incumbent, positive_validation_reference, finite_difference_params
            )
        )
        if validation_reference is None:
            print(
                '[WARNING] Finite-difference validation skipped: no matching positive '
                'investment was found in any feasible evaluated candidate.'
            )
        else:
            if not validation_reference_is_incumbent:
                print(
                    '[INFO] The incumbent has no matching positive investment. Finite-difference '
                    f'validation will use the best positive feasible evaluation from outer iteration '
                    f'{validation_reference["iteration"]}; the incumbent and upper bound are unchanged.'
                )
            finite_difference_results = _validate_local_sensitivities_with_finite_differences(
                planning_problem,
                validation_reference['candidate_solution'],
                validation_reference['operational_recourse'],
                validation_reference['sensitivities'],
                validation_reference['models'],
                validation_reference['state'],
                finite_difference_params,
                baseline_outer_iteration=validation_reference['iteration'],
                termination_reason=termination_reason,
                validation_source=validation_source,
                validation_reference_is_incumbent=validation_reference_is_incumbent,
            )

    # Write results
    end = time.time()
    total_execution_time = end - start
    print('[INFO] Execution time: {:.2f} s'.format(total_execution_time))
    bound_evolution = {
        'master_estimate': master_estimate_evolution,
        'lower_bound': master_estimate_evolution,
        'upper_bound': upper_bound_evolution,
        'investment_cost': investment_cost_evolution,
        'alpha': alpha_evolution,
        'gross_operational_cost': gross_operational_cost_evolution,
        'terminal_salvage_value': terminal_salvage_value_evolution,
        'operational_recourse': operational_recourse_evolution,
        'candidate_total': candidate_total_evolution,
        'esso_violation': esso_violation_evolution,
        'gap_signed': gap_signed_evolution,
        'gap_abs': gap_abs_evolution,
        'gap_rel': gap_rel_evolution,
        'incumbent_updated': incumbent_update_evolution,
        'candidate_source': candidate_source_evolution,
        'operational_initialization': operational_initialization_evolution,
        'termination_reason': termination_reason,
        'convergence': convergence,
        'outer_iterations': iteration,
        'incumbent_iteration': incumbent['iteration'] if incumbent is not None else None,
        'incumbent_objective': incumbent['candidate_total'] if incumbent is not None else None,
        'positive_bootstrap_used': positive_bootstrap_used,
        'positive_bootstrap_iteration': positive_bootstrap_iteration,
        'positive_bootstrap_budget_fraction': (
            positive_bootstrap_params.budget_fraction if positive_bootstrap_used else None
        ),
        'sensitivity_probe_enabled': sensitivity_probe_params.enabled,
        'sensitivity_probe_budget_fraction': sensitivity_probe_params.budget_fraction,
        'sensitivity_probe_energy_to_power_ratio': sensitivity_probe_params.energy_to_power_ratio,
        'validation_reference_source': validation_source,
        'validation_reference_iteration': (
            validation_reference['iteration'] if validation_reference is not None else None
        ),
        'validation_reference_is_incumbent': validation_reference_is_incumbent,
        'finite_difference': finite_difference_results,
        'sensitivity_probe_diagnostics': sensitivity_probe_diagnostics,
        'admm_diagnostics': admm_diagnostics,
        'solver_recovery_diagnostics': solver_recovery_diagnostics,
    }
    if incumbent is not None:
        _restore_candidate_data(planning_problem, incumbent['candidate_solution'])
        shared_ess_data.load_candidate_solution_into_master_model(
            master_problem_model, incumbent['candidate_solution']
        )
        planning_problem.write_planning_results_to_excel(
            master_problem_model,
            incumbent['models'],
            incumbent['results'],
            bound_evolution,
            execution_time=total_execution_time,
        )
    elif operational_state and operational_state.get('initialization_failed', False):
        print('[WARNING] Planning results were not written because the final operational initialization failed.')
    else:
        print('[WARNING] Planning results were not written because no feasible incumbent is available.')


def _get_operational_recourse_value(planning_problem, models):
    return _get_operational_recourse_components(planning_problem, models)['net_operational_recourse']


def _get_operational_recourse_components(planning_problem, models):
    # Operational recourse is based on the local base SMOPF objectives.
    # It excludes scenario-deviation regularization and ADMM augmentation terms, but may include artificial penalty terms and is therefore not necessarily a pure economic operating cost.
    gross_operational_cost = planning_problem.transmission_network.get_primal_value(models['tso'])
    for node_id, distribution_network in planning_problem.distribution_networks.items():
        gross_operational_cost += distribution_network.get_primal_value(models['dso'][node_id])
    terminal_salvage_value = planning_problem.shared_ess_data.get_salvage_value(models['esso'])
    return {
        'gross_operational_cost': gross_operational_cost,
        'terminal_salvage_value': terminal_salvage_value,
        'net_operational_recourse': gross_operational_cost - terminal_salvage_value,
    }


def _get_operational_recourse_block_components(planning_problem, models):
    """
    Decompose net operational recourse into the exact weighted TSO/DSO year-day contributions used by NetworkData.get_primal_value().
    Keys:
        ('TSO', None, year, day)
        ('DSO', node_id, year, day)
        ('SALVAGE', None, None, None)
    The salvage contribution is stored with a negative sign so that:
        sum(blocks.values()) == net_operational_recourse
    """

    blocks = dict()

    # --------------------------------------------------------------------------------------------------------------
    # TSO blocks
    transmission_network = planning_problem.transmission_network

    for year in transmission_network.years:
        for day in transmission_network.days:
            local_value = transmission_network.network[year][day].get_primal_value(models['tso'][year][day], transmission_network.params)
            weight = _get_admm_block_weight(transmission_network, year, day)
            blocks[('TSO', None, year, day)] = float(weight * local_value)

    # --------------------------------------------------------------------------------------------------------------
    # DSO blocks
    for node_id, distribution_network in planning_problem.distribution_networks.items():
        for year in distribution_network.years:
            for day in distribution_network.days:
                local_value = distribution_network.network[year][day].get_primal_value(models['dso'][node_id][year][day], distribution_network.params)
                weight = _get_admm_block_weight(distribution_network, year, day)
                blocks[('DSO', node_id, year, day)] = float(weight * local_value)

    # --------------------------------------------------------------------------------------------------------------
    # Terminal salvage
    # Net recourse = gross operating cost - terminal salvage.
    terminal_salvage_value = planning_problem.shared_ess_data.get_salvage_value(models['esso'])
    blocks[('SALVAGE', None, None, None)] = -float(terminal_salvage_value)

    return blocks


def _print_recourse_jump_diagnostics(
        current_blocks,
        previous_blocks,
        cycle,
        current_recourse,
        previous_recourse,
        objective_tolerance,
        top_n=10,
        tso_proximal_movements=None,
        current_objective_component_blocks=None,
        previous_objective_component_blocks=None,
        current_slack_component_blocks = None,
        previous_slack_component_blocks = None,
        current_tso_voltage_slack_state=None,
        previous_tso_voltage_slack_state=None
):
    """
    Print the blocks responsible for a failure of the recourse-stationarity criterion.
    Block changes are ranked by absolute magnitude, while the signed change is retained so that increases and decreases can be distinguished.
    """

    if current_blocks is None or previous_blocks is None:
        return

    if current_recourse is None or previous_recourse is None:
        return

    if objective_tolerance is None:
        return

    signed_recourse_delta = current_recourse - previous_recourse
    objective_change_abs = abs(signed_recourse_delta)

    # Only print the detailed decomposition when recourse stationarity fails.
    if objective_change_abs <= objective_tolerance:
        return

    changes = list()

    all_keys = set(current_blocks) | set(previous_blocks)

    for key in all_keys:

        previous_value = previous_blocks.get(key, 0.0)
        current_value = current_blocks.get(key, 0.0)

        delta = current_value - previous_value

        agent, node_id, year, day = key

        changes.append({
            'agent': agent,
            'node_id': node_id,
            'year': year,
            'day': day,
            'previous': previous_value,
            'current': current_value,
            'delta': delta,
            'abs_delta': abs(delta),
        })

    changes.sort(key=lambda entry: entry['abs_delta'], reverse=True)

    # --------------------------------------------------------------------------------------------------------------
    # Reconciliation
    previous_block_total = sum(previous_blocks.values())
    current_block_total = sum(current_blocks.values())

    signed_block_delta = current_block_total - previous_block_total
    delta_mismatch = signed_block_delta - signed_recourse_delta

    print(
        f'[RECOURSE JUMP] cycle={cycle} | '
        f'abs_change={objective_change_abs:.6e} | '
        f'tol={objective_tolerance:.6e} | '
        f'signed_change={signed_recourse_delta:+.6e}'
    )

    print(
        f'[RECOURSE JUMP] Reconciliation | '
        f'previous={previous_block_total:.6e} | '
        f'current={current_block_total:.6e} | '
        f'block_delta={signed_block_delta:+.6e} | '
        f'mismatch={delta_mismatch:+.6e}'
    )

    # --------------------------------------------------------------------------------------------------------------
    # Aggregate changes by agent
    aggregate_changes = dict()

    for entry in changes:

        if entry['agent'] == 'DSO':
            aggregate_key = f'DSO node={entry["node_id"]}'
        else:
            aggregate_key = entry['agent']

        if aggregate_key not in aggregate_changes:
            aggregate_changes[aggregate_key] = 0.0

        aggregate_changes[aggregate_key] += entry['delta']

    print('[RECOURSE JUMP] Aggregate signed changes:')

    for aggregate_key, delta in sorted(aggregate_changes.items(), key=lambda item: abs(item[1]), reverse=True):
        print(f'  {aggregate_key} | delta={delta:+.6e}')

    # --------------------------------------------------------------------------------------------------------------
    # Largest individual year-day changes
    print(f'[RECOURSE JUMP] Largest {min(top_n, len(changes))} block changes:')
    for entry in changes[:top_n]:
        print(
            f'  {entry["agent"]} '
            f'node={entry["node_id"]} '
            f'year={entry["year"]} '
            f'day={entry["day"]} | '
            f'previous={entry["previous"]:.6e} | '
            f'current={entry["current"]:.6e} | '
            f'delta={entry["delta"]:+.6e} | '
            f'abs_delta={entry["abs_delta"]:.6e}'
        )

        if (entry['agent'] == 'TSO' and tso_proximal_movements is not None):
            block_key = (entry['year'], entry['day'])
            proximal_block = tso_proximal_movements.get(block_key)
            if (proximal_block is not None and proximal_block.get('successful', False)):
                v_data = proximal_block['v']
                pf_data = proximal_block['pf']
                ess_data = proximal_block['ess']
                v_move = (v_data['normalized_movement'] if v_data is not None else 0.0)
                pf_move = (pf_data['normalized_movement'] if pf_data is not None else 0.0)
                ess_move = (ess_data['normalized_movement'] if ess_data is not None else 0.0)
                print(
                    '    [RECOURSE JUMP][TSO PROX] '
                    f'year={entry["year"]} '
                    f'day={entry["day"]} | '
                    f'V max={v_move:.6f} | '
                    f'PF max={pf_move:.6f} | '
                    f'ESS max={ess_move:.6f}'
                )

        # ----------------------------------------------------------------------------------------------------------
        # Objective-component decomposition for this exact recourse block.
        if (current_objective_component_blocks is not None and previous_objective_component_blocks is not None and entry['agent'] in ('TSO', 'DSO')):
            component_key = (entry['agent'], entry['node_id'], entry['year'], entry['day'])
            current_components = current_objective_component_blocks.get(component_key)
            previous_components = previous_objective_component_blocks.get(component_key)
            if (current_components is not None and previous_components is not None):
                component_names = (
                    'generation_cost',
                    'flexibility_cost',
                    'load_curtailment_cost',
                    'res_curtailment_penalty',
                    'ess_usage_penalty',
                    'slack_penalties',
                    'ess_complementarity_penalties',
                )
                component_deltas = {name: current_components[name] - previous_components[name] for name in component_names}
                classified_delta = sum(component_deltas.values())
                component_mismatch = classified_delta - entry['delta']
                print(
                    '    [RECOURSE COMPONENTS] '
                    f'gen={component_deltas["generation_cost"]:+.6e} | '
                    f'flex={component_deltas["flexibility_cost"]:+.6e} | '
                    f'load_curt={component_deltas["load_curtailment_cost"]:+.6e} | '
                    f'res_curt={component_deltas["res_curtailment_penalty"]:+.6e} | '
                    f'ess_usage={component_deltas["ess_usage_penalty"]:+.6e} | '
                    f'slacks={component_deltas["slack_penalties"]:+.6e} | '
                    f'ess_comp={component_deltas["ess_complementarity_penalties"]:+.6e}'
                )
                if (current_slack_component_blocks is not None and previous_slack_component_blocks is not None and entry['agent'] in ('TSO', 'DSO')):
                    slack_key = (entry['agent'], entry['node_id'], entry['year'], entry['day'])
                    current_slacks = current_slack_component_blocks.get(slack_key)
                    previous_slacks = previous_slack_component_blocks.get(slack_key)
                    if current_slacks is not None and previous_slacks is not None:

                        slack_names = (
                            'voltage',
                            'node_balance_p',
                            'node_balance_q',
                            'branch_flow_ij',
                            'branch_flow_ji',
                            'flex_day_balance_p',
                            'flex_day_balance_q',
                        )

                        slack_deltas = {name: current_slacks[name] - previous_slacks[name] for name in slack_names}
                        classified_slack_delta = sum(slack_deltas.values())
                        objective_slack_delta = (current_components['slack_penalties'] - previous_components['slack_penalties'])
                        slack_mismatch = (classified_slack_delta - objective_slack_delta)

                        print(
                            '    [SLACK COMPONENTS] '
                            f'voltage={slack_deltas["voltage"]:+.6e} | '
                            f'node_P={slack_deltas["node_balance_p"]:+.6e} | '
                            f'node_Q={slack_deltas["node_balance_q"]:+.6e} | '
                            f'branch_ij={slack_deltas["branch_flow_ij"]:+.6e} | '
                            f'branch_ji={slack_deltas["branch_flow_ji"]:+.6e} | '
                            f'flex_P={slack_deltas["flex_day_balance_p"]:+.6e} | '
                            f'flex_Q={slack_deltas["flex_day_balance_q"]:+.6e}'
                        )

                        voltage_delta = slack_deltas['voltage']
                        if (entry['agent'] == 'TSO' and current_tso_voltage_slack_state is not None and previous_tso_voltage_slack_state is not None and abs(voltage_delta) > 0.1 * objective_tolerance):
                            _print_tso_voltage_slack_transitions(
                                year=entry['year'],
                                day=entry['day'],
                                current_state=current_tso_voltage_slack_state,
                                previous_state=previous_tso_voltage_slack_state,
                                expected_voltage_penalty_delta=voltage_delta,
                                top_n=5,
                            )

                        print(
                            '    [SLACK COMPONENT CHECK] '
                            f'classified_delta={classified_slack_delta:+.6e} | '
                            f'objective_slack_delta={objective_slack_delta:+.6e} | '
                            f'mismatch={slack_mismatch:+.6e}'
                        )
                print(
                    '    [RECOURSE COMPONENT CHECK] '
                    f'classified_delta={classified_delta:+.6e} | '
                    f'block_delta={entry["delta"]:+.6e} | '
                    f'mismatch={component_mismatch:+.6e}'
                )


def _get_operational_sensitivities(planning_problem, models):
    available_sensitivities = {'s': dict(), 'e': dict()}
    for year in planning_problem.years:
        available_sensitivities['s'][year] = {
            node_id: 0.00 for node_id in planning_problem.active_distribution_network_nodes
        }
        available_sensitivities['e'][year] = {
            node_id: 0.00 for node_id in planning_problem.active_distribution_network_nodes
        }

    local_sensitivities = [
        planning_problem.transmission_network.get_sensitivities(models['tso'])
    ]
    for node_id, distribution_network in planning_problem.distribution_networks.items():
        local_sensitivities.append(distribution_network.get_sensitivities(models['dso'][node_id]))

    for local_values in local_sensitivities:
        for capacity_type in ('s', 'e'):
            for year, node_values in local_values[capacity_type].items():
                for node_id, value in node_values.items():
                    if value is None or available_sensitivities[capacity_type][year][node_id] is None:
                        available_sensitivities[capacity_type][year][node_id] = None
                    else:
                        available_sensitivities[capacity_type][year][node_id] += value

    investment_sensitivities = planning_problem.shared_ess_data.map_available_capacity_sensitivities_to_investments(
        models['esso'], available_sensitivities
    )
    salvage_sensitivities = planning_problem.shared_ess_data.get_salvage_value_sensitivities(
        models['esso']
    )
    for capacity_type in ('s', 'e'):
        for year in planning_problem.years:
            for node_id in planning_problem.active_distribution_network_nodes:
                if investment_sensitivities[capacity_type][year][node_id] is not None:
                    investment_sensitivities[capacity_type][year][node_id] += (
                        salvage_sensitivities[capacity_type][year][node_id]
                    )
    return investment_sensitivities


def _validate_local_sensitivities_with_finite_differences(planning_problem, candidate_solution,
                                                          baseline_recourse, sensitivities, baseline_models,
                                                          baseline_state, params,
                                                          baseline_outer_iteration=None,
                                                          termination_reason=None,
                                                          validation_source='incumbent',
                                                          validation_reference_is_incumbent=True):
    selected = _select_finite_difference_investment(candidate_solution, params)
    if selected is None:
        print('[WARNING] Finite-difference validation skipped: no matching positive investment was found.')
        return []
    if baseline_state is None:
        print('[WARNING] Finite-difference validation skipped: the converged operational state is unavailable.')
        return []
    if sensitivities is None:
        print('[WARNING] Finite-difference validation skipped: reference sensitivities are unavailable.')
        return []

    node_id, year = selected
    original_sensitivity_s = sensitivities['s'][year][node_id]
    original_sensitivity_e = sensitivities['e'][year][node_id]
    if original_sensitivity_s is None or original_sensitivity_e is None:
        print('[WARNING] Finite-difference validation skipped: the selected sensitivity is unavailable.')
        return []

    base_s = candidate_solution['investment'][node_id][year]['s']
    base_e = candidate_solution['investment'][node_id][year]['e']
    ratio = (
        planning_problem.shared_ess_data.params.min_energy_to_power_ratio
        if isclose(base_s, 0.00, abs_tol=SMALL_TOLERANCE)
        else base_e / base_s
    )
    baseline_soh_margin = _get_investment_soh_margin(
        planning_problem, baseline_models['esso'], node_id, year
    )
    direction_specs = _get_finite_difference_directions(params.directions, ratio, base_s, base_e)
    display_direction = next(
        (direction for direction in direction_specs if direction['name'] == 'fixed_ratio'),
        direction_specs[0] if direction_specs else None,
    )
    validation_results = []

    print(
        '[INFO] Running finite-difference validation of the selected local sensitivity '
        f'reference (source={validation_source}, outer iteration={baseline_outer_iteration})...'
    )
    print(
        f'[INFO] Selected investment: node {node_id}, year {year}, '
        f'S = {base_s:.6f} MVA, E = {base_e:.6f} MVAh, E/S = {ratio:.6f}.'
    )

    try:
        reference = _refine_finite_difference_operational_point(
            planning_problem,
            candidate_solution,
            baseline_state,
            params,
            node_id,
            year,
            initial_sensitivities=sensitivities,
            require_sensitivity_stability=True,
        )

        for attempt in reference['attempts']:
            current_sensitivity_s = attempt['sensitivity_s']
            current_sensitivity_e = attempt['sensitivity_e']
            cumulative_drift = (
                attempt['recourse'] - baseline_recourse
                if attempt['recourse'] is not None else None
            )
            original_drift_s = _relative_value_drift(
                current_sensitivity_s, original_sensitivity_s
            )
            original_drift_e = _relative_value_drift(
                current_sensitivity_e, original_sensitivity_e
            )
            original_drift = _maximum_available_value(original_drift_s, original_drift_e)
            active_set_changed = _soh_active_state_changed(
                baseline_soh_margin, attempt['soh_margin'], params.soh_active_tolerance
            )
            original_cut_reproducible = (
                attempt['stabilized']
                and cumulative_drift is not None
                and abs(cumulative_drift) <= attempt['stationarity_tolerance']
                and original_drift is not None
                and original_drift <= params.slope_consistency_tolerance
                and not active_set_changed
            )
            original_display_slope = _directional_slope(
                display_direction, original_sensitivity_s, original_sensitivity_e
            )
            current_display_slope = _directional_slope(
                display_direction, current_sensitivity_s, current_sensitivity_e
            )
            validation_results.append({
                'run_type': 'reference_refinement',
                'direction': 'replay',
                'status': 'passed' if attempt['stabilized'] else 'inconclusive',
                'reason': '; '.join(attempt['reasons']),
                'refinement': attempt['refinement'],
                'max_refinements': params.max_replay_refinements,
                'reference_stabilized': attempt['stabilized'],
                'original_cut_reproducible': original_cut_reproducible,
                'validation_source': validation_source,
                'validation_reference_is_incumbent': validation_reference_is_incumbent,
                'baseline_outer_iteration': baseline_outer_iteration,
                'termination_reason': termination_reason,
                'node_id': node_id,
                'year': year,
                'base_s': base_s,
                'base_e': base_e,
                'energy_to_power_ratio': ratio,
                'step_fraction': 0.00,
                'step_size': 0.00,
                'delta_s': 0.00,
                'delta_e': 0.00,
                'sensitivity_s': original_sensitivity_s,
                'sensitivity_e': original_sensitivity_e,
                'replay_sensitivity_s': current_sensitivity_s,
                'replay_sensitivity_e': current_sensitivity_e,
                'original_analytic_slope': original_display_slope,
                'analytic_slope': current_display_slope,
                'replay_analytic_slope': current_display_slope,
                'baseline_recourse': baseline_recourse,
                'reference_recourse': attempt['recourse'],
                'replay_drift': cumulative_drift,
                'stationarity_drift': attempt['stationarity_drift'],
                'stationarity_tolerance': attempt['stationarity_tolerance'],
                'replay_tolerance': attempt['stationarity_tolerance'],
                'sensitivity_relative_drift': attempt['sensitivity_relative_drift'],
                'sensitivity_relative_drift_s': attempt['sensitivity_relative_drift_s'],
                'sensitivity_relative_drift_e': attempt['sensitivity_relative_drift_e'],
                'original_sensitivity_relative_drift': original_drift,
                'original_sensitivity_relative_drift_s': original_drift_s,
                'original_sensitivity_relative_drift_e': original_drift_e,
                'operational_convergence': attempt['operational_convergence'],
                'esso_violation': attempt['esso_violation'],
                'baseline_soh_margin': baseline_soh_margin,
                'reference_soh_margin': attempt['soh_margin'],
                'active_set_changed': active_set_changed,
                'passed': attempt['stabilized'],
            })

            cycle_drift_text = _format_optional_float(attempt['stationarity_drift'])
            cumulative_drift_text = _format_optional_float(cumulative_drift)
            sensitivity_drift_text = _format_optional_percent(
                attempt['sensitivity_relative_drift']
            )
            print(
                f'[INFO] Finite-difference reference refinement {attempt["refinement"]}/'
                f'{params.max_replay_refinements} | Cycle recourse drift = {cycle_drift_text} | '
                f'Cumulative baseline drift = {cumulative_drift_text} | '
                f'Partial sensitivity drift = {sensitivity_drift_text} | '
                f'Status = {"passed" if attempt["stabilized"] else "inconclusive"}'
            )

        if not reference['stabilized']:
            print(
                '[WARNING] Finite-difference perturbations skipped: the selected reference '
                'did not stabilize within the configured refinement limit.'
            )
            return validation_results

        reference_sensitivity_s = reference['sensitivities']['s'][year][node_id]
        reference_sensitivity_e = reference['sensitivities']['e'][year][node_id]
        original_drift_s = _relative_value_drift(
            reference_sensitivity_s, original_sensitivity_s
        )
        original_drift_e = _relative_value_drift(
            reference_sensitivity_e, original_sensitivity_e
        )
        original_drift = _maximum_available_value(original_drift_s, original_drift_e)
        original_active_set_changed = _soh_active_state_changed(
            baseline_soh_margin, reference['soh_margin'], params.soh_active_tolerance
        )
        original_cut_reproducible = (
            abs(reference['recourse'] - baseline_recourse) <= reference['stationarity_tolerance']
            and original_drift is not None
            and original_drift <= params.slope_consistency_tolerance
            and not original_active_set_changed
        )
        if not original_cut_reproducible:
            print(
                '[WARNING] The polished reference differs materially from the selected baseline '
                'recourse, sensitivities, or minimum-SoH active set. Perturbations will validate '
                'the polished local derivative and report the original sensitivity-point drift separately.'
            )

        for direction in direction_specs:
            previous_observed_slope = None
            original_analytic_slope = _directional_slope(
                direction, original_sensitivity_s, original_sensitivity_e
            )
            analytic_slope = _directional_slope(
                direction, reference_sensitivity_s, reference_sensitivity_e
            )

            for step_fraction in params.relative_step_sizes:
                if step_fraction <= 0.00:
                    print(f'[WARNING] Ignoring non-positive relative finite-difference step {step_fraction}.')
                    continue

                step_size = step_fraction * direction['scale']
                delta_s = direction['s'] * step_size
                delta_e = direction['e'] * step_size
                perturbed_candidate = deepcopy(candidate_solution)
                perturbed_candidate['investment'][node_id][year]['s'] += delta_s
                perturbed_candidate['investment'][node_id][year]['e'] += delta_e
                _rebuild_candidate_total_capacities(planning_problem, perturbed_candidate)
                first_stage_feasible, first_stage_reason = _check_candidate_first_stage_feasibility(
                    planning_problem, perturbed_candidate
                )

                endpoint = _refine_finite_difference_operational_point(
                    planning_problem,
                    perturbed_candidate,
                    reference['state'],
                    params,
                    node_id,
                    year,
                    require_sensitivity_stability=False,
                )
                predicted_change = analytic_slope * step_size
                perturbed_recourse = endpoint['recourse']
                observed_change = None
                absolute_error = None
                observed_slope = None
                absolute_slope_error = None
                relative_error = None
                same_sign = None
                signal_to_noise_ratio = None
                slope_consistency_error = None
                active_set_changed = _soh_active_state_changed(
                    reference['soh_margin'], endpoint['soh_margin'], params.soh_active_tolerance
                )
                reasons = list(endpoint['reasons'])
                status = 'inconclusive'

                if endpoint['stabilized'] and perturbed_recourse is not None:
                    observed_change = perturbed_recourse - reference['recourse']
                    absolute_error = abs(observed_change - predicted_change)
                    observed_slope = observed_change / step_size
                    absolute_slope_error = abs(observed_slope - analytic_slope)
                    relative_error = absolute_slope_error / max(
                        abs(observed_slope), abs(analytic_slope), 1.00
                    )
                    same_sign = (
                        analytic_slope * observed_slope >= 0.00
                        or (
                            isclose(analytic_slope, 0.00, abs_tol=1.00)
                            and isclose(observed_slope, 0.00, abs_tol=1.00)
                        )
                    )
                    noise_floor = max(
                        reference['stationarity_drift'] or 0.00,
                        endpoint['stationarity_drift'] or 0.00,
                        params.replay_absolute_tolerance,
                    )
                    signal_to_noise_ratio = abs(observed_change) / noise_floor
                    if previous_observed_slope is not None:
                        slope_consistency_error = abs(observed_slope - previous_observed_slope) / max(
                            abs(observed_slope), abs(previous_observed_slope), 1.00
                        )

                    if active_set_changed:
                        reasons.append('minimum-SoH activity changed')
                    if signal_to_noise_ratio < params.minimum_signal_to_noise_ratio:
                        reasons.append('finite-difference signal is below the noise threshold')

                    if reasons:
                        status = 'inconclusive'
                    elif not same_sign:
                        status = 'failed'
                        reasons.append('analytic and observed slopes have different signs')
                    elif relative_error > params.relative_error_tolerance:
                        status = 'failed'
                        reasons.append('relative slope error exceeds tolerance')
                    elif (
                            slope_consistency_error is not None
                            and slope_consistency_error > params.slope_consistency_tolerance):
                        status = 'failed'
                        reasons.append('finite-difference slopes are not consistent across step sizes')
                    else:
                        status = 'passed'

                    previous_observed_slope = observed_slope

                validation_results.append({
                    'run_type': 'perturbation',
                    'direction': direction['name'],
                    'direction_s': direction['s'],
                    'direction_e': direction['e'],
                    'status': status,
                    'reason': '; '.join(reasons),
                    'refinement': endpoint['refinement_count'],
                    'max_refinements': params.max_replay_refinements,
                    'reference_stabilized': reference['stabilized'],
                    'endpoint_stabilized': endpoint['stabilized'],
                    'original_cut_reproducible': original_cut_reproducible,
                    'validation_source': validation_source,
                    'validation_reference_is_incumbent': validation_reference_is_incumbent,
                    'baseline_outer_iteration': baseline_outer_iteration,
                    'termination_reason': termination_reason,
                    'node_id': node_id,
                    'year': year,
                    'base_s': base_s,
                    'base_e': base_e,
                    'energy_to_power_ratio': ratio,
                    'step_fraction': step_fraction,
                    'step_size': step_size,
                    'delta_s': delta_s,
                    'delta_e': delta_e,
                    'first_stage_feasible': first_stage_feasible,
                    'first_stage_reason': first_stage_reason,
                    'sensitivity_s': original_sensitivity_s,
                    'sensitivity_e': original_sensitivity_e,
                    'replay_sensitivity_s': reference_sensitivity_s,
                    'replay_sensitivity_e': reference_sensitivity_e,
                    'original_analytic_slope': original_analytic_slope,
                    'analytic_slope': analytic_slope,
                    'replay_analytic_slope': analytic_slope,
                    'predicted_change': predicted_change,
                    'baseline_recourse': baseline_recourse,
                    'reference_recourse': reference['recourse'],
                    'perturbed_recourse': perturbed_recourse,
                    'observed_change': observed_change,
                    'absolute_error': absolute_error,
                    'observed_slope': observed_slope,
                    'absolute_slope_error': absolute_slope_error,
                    'relative_error': relative_error,
                    'signal_to_noise_ratio': signal_to_noise_ratio,
                    'slope_consistency_error': slope_consistency_error,
                    'same_sign': same_sign,
                    'operational_convergence': endpoint['operational_convergence'],
                    'esso_violation': endpoint['esso_violation'],
                    'baseline_soh_margin': baseline_soh_margin,
                    'reference_soh_margin': reference['soh_margin'],
                    'perturbed_soh_margin': endpoint['soh_margin'],
                    'active_set_changed': active_set_changed,
                    'replay_drift': reference['recourse'] - baseline_recourse,
                    'stationarity_drift': endpoint['stationarity_drift'],
                    'stationarity_tolerance': endpoint['stationarity_tolerance'],
                    'replay_tolerance': reference['stationarity_tolerance'],
                    'sensitivity_relative_drift': reference['sensitivity_relative_drift'],
                    'sensitivity_relative_drift_s': reference['sensitivity_relative_drift_s'],
                    'sensitivity_relative_drift_e': reference['sensitivity_relative_drift_e'],
                    'original_sensitivity_relative_drift': original_drift,
                    'original_sensitivity_relative_drift_s': original_drift_s,
                    'original_sensitivity_relative_drift_e': original_drift_e,
                    'passed': status == 'passed',
                })

                observed_text = _format_optional_float(observed_change)
                error_text = _format_optional_percent(relative_error)
                endpoint_drift_text = _format_optional_float(endpoint['stationarity_drift'])
                print(
                    f'[INFO] Finite difference {direction["name"]} | '
                    f'h = {step_size:.6f} ({step_fraction:.2%}) | '
                    f'Refinements = {endpoint["refinement_count"]} | '
                    f'Endpoint drift = {endpoint_drift_text} | '
                    f'Predicted change = {predicted_change:.6f} | '
                    f'Observed change = {observed_text} | Relative error = {error_text} | '
                    f'Status = {status}'
                )
    finally:
        _restore_candidate_data(planning_problem, candidate_solution)

    return validation_results


def _refine_finite_difference_operational_point(planning_problem, candidate_solution,
                                                initial_state, params, node_id, year,
                                                initial_sensitivities=None,
                                                require_sensitivity_stability=False):
    state = initial_state
    previous_sensitivity_s = _get_selected_sensitivity(initial_sensitivities, 's', year, node_id)
    previous_sensitivity_e = _get_selected_sensitivity(initial_sensitivities, 'e', year, node_id)
    attempts = []
    final = {
        'stabilized': False,
        'operational_convergence': False,
        'models': None,
        'sensitivities': None,
        'state': state,
        'recourse': None,
        'esso_violation': None,
        'soh_margin': None,
        'stationarity_drift': None,
        'stationarity_tolerance': None,
        'sensitivity_relative_drift': None,
        'sensitivity_relative_drift_s': None,
        'sensitivity_relative_drift_e': None,
        'reasons': [],
        'refinement_count': 0,
    }

    for refinement in range(1, params.max_replay_refinements + 1):
        state_for_run = dict(state)
        state_for_run['consecutive_converged_cycles'] = 0
        convergence, _, models, current_sensitivities, _, current_state = (
            planning_problem.run_operational_planning(
                candidate_solution=candidate_solution,
                print_results=False,
                initial_state=state_for_run,
                return_state=True,
            )
        )
        esso_violation = _get_esso_violation_if_available(planning_problem, models)
        recourse = None
        soh_margin = None
        current_sensitivity_s = None
        current_sensitivity_e = None
        stationarity_drift = _get_last_admm_objective_change(current_state)
        stationarity_tolerance = None
        sensitivity_drift_s = None
        sensitivity_drift_e = None
        sensitivity_drift = None
        reasons = []

        if not convergence:
            reasons.append('ADMM refinement did not converge')
        else:
            recourse = planning_problem.get_operational_recourse_value(models)
            stationarity_tolerance = max(
                params.replay_absolute_tolerance,
                params.replay_relative_tolerance * max(abs(recourse), 1.00),
            )
            soh_margin = _get_investment_soh_margin(
                planning_problem, models['esso'], node_id, year
            )
            current_sensitivity_s = _get_selected_sensitivity(
                current_sensitivities, 's', year, node_id
            )
            current_sensitivity_e = _get_selected_sensitivity(
                current_sensitivities, 'e', year, node_id
            )
            sensitivity_drift_s = _relative_value_drift(
                current_sensitivity_s, previous_sensitivity_s
            )
            sensitivity_drift_e = _relative_value_drift(
                current_sensitivity_e, previous_sensitivity_e
            )
            sensitivity_drift = _maximum_available_value(
                sensitivity_drift_s, sensitivity_drift_e
            )

            if stationarity_drift is None:
                reasons.append('validation recourse stationarity is unavailable')
            elif stationarity_drift > stationarity_tolerance:
                reasons.append('validation recourse stationarity exceeds tolerance')
            if require_sensitivity_stability:
                if sensitivity_drift is None:
                    reasons.append('validation sensitivity stability is unavailable')
                elif sensitivity_drift > params.slope_consistency_tolerance:
                    reasons.append('partial sensitivities have not stabilized')

        if esso_violation is None:
            reasons.append('ESSO violation is unavailable')
        elif esso_violation > BENDERS_FEASIBILITY_TOLERANCE:
            reasons.append('ESSO violation exceeds tolerance')

        stabilized = not reasons
        attempt = {
            'refinement': refinement,
            'stabilized': stabilized,
            'reasons': reasons,
            'operational_convergence': convergence,
            'models': models,
            'sensitivities': current_sensitivities,
            'state': current_state,
            'recourse': recourse,
            'esso_violation': esso_violation,
            'soh_margin': soh_margin,
            'stationarity_drift': stationarity_drift,
            'stationarity_tolerance': stationarity_tolerance,
            'sensitivity_s': current_sensitivity_s,
            'sensitivity_e': current_sensitivity_e,
            'sensitivity_relative_drift': sensitivity_drift,
            'sensitivity_relative_drift_s': sensitivity_drift_s,
            'sensitivity_relative_drift_e': sensitivity_drift_e,
        }
        attempts.append(attempt)
        final.update(attempt)
        final['attempts'] = attempts
        final['refinement_count'] = refinement

        if stabilized:
            break
        if not convergence or esso_violation is None or esso_violation > BENDERS_FEASIBILITY_TOLERANCE:
            break

        state = current_state
        previous_sensitivity_s = current_sensitivity_s
        previous_sensitivity_e = current_sensitivity_e

    return final


def _get_last_admm_objective_change(state):
    if not isinstance(state, dict):
        return None
    for diagnostic in reversed(state.get('admm_diagnostics', [])):
        objective_change = diagnostic.get('objective_change_abs')
        if objective_change is not None:
            return abs(objective_change)
    return None


def _get_selected_sensitivity(sensitivities, capacity_type, year, node_id):
    if sensitivities is None:
        return None
    return sensitivities.get(capacity_type, {}).get(year, {}).get(node_id)


def _relative_value_drift(current, previous):
    if current is None or previous is None:
        return None
    return abs(current - previous) / max(abs(current), abs(previous), 1.00)


def _maximum_available_value(*values):
    available_values = [value for value in values if value is not None]
    return max(available_values) if available_values else None


def _directional_slope(direction, sensitivity_s, sensitivity_e):
    if direction is None or sensitivity_s is None or sensitivity_e is None:
        return None
    return direction['s'] * sensitivity_s + direction['e'] * sensitivity_e


def _format_optional_float(value):
    return f'{value:.6f}' if value is not None else 'N/A'


def _format_optional_percent(value):
    return f'{value * 100:.2f}%' if value is not None else 'N/A'


def _get_finite_difference_directions(direction_names, ratio, base_s, base_e):
    direction_definitions = {
        'power_only': {'s': 1.00, 'e': 0.00, 'scale': max(abs(base_s), 1.00)},
        'energy_only': {'s': 0.00, 'e': 1.00, 'scale': max(abs(base_e), 1.00)},
        'fixed_ratio': {'s': 1.00, 'e': ratio, 'scale': max(abs(base_s), 1.00)},
    }
    return [
        {'name': name, **direction_definitions[name]}
        for name in direction_names
    ]


def _get_esso_violation_if_available(planning_problem, models):
    if not isinstance(models, dict) or not models.get('esso'):
        return None
    return planning_problem.shared_ess_data.get_feasibility_violation(models['esso'])


def _check_candidate_first_stage_feasibility(planning_problem, candidate_solution):
    shared_ess_data = planning_problem.shared_ess_data
    params = shared_ess_data.params
    years = list(shared_ess_data.years)
    tolerance = 1e-8
    reasons = []
    expected_investment_cost = 0.00

    for node_id, yearly_investments in candidate_solution['investment'].items():
        for year, investment in yearly_investments.items():
            s_value = investment['s']
            e_value = investment['e']
            if s_value < -tolerance or e_value < -tolerance:
                reasons.append(f'negative investment at node {node_id}, year {year}')
            if e_value + tolerance < params.min_energy_to_power_ratio * s_value:
                reasons.append(f'minimum E/S ratio violated at node {node_id}, year {year}')
            if e_value > params.max_energy_to_power_ratio * s_value + tolerance:
                reasons.append(f'maximum E/S ratio violated at node {node_id}, year {year}')

            year_discount = 1.00 / (
                (1.00 + shared_ess_data.discount_factor) ** (int(year) - int(years[0]))
            )
            for scenario, probability in enumerate(shared_ess_data.prob_market_scenarios):
                expected_investment_cost += year_discount * probability * (
                    shared_ess_data.cost_investment['power'][scenario][year] * s_value
                    + shared_ess_data.cost_investment['energy'][scenario][year] * e_value
                )

        for year, total_capacity in candidate_solution['total_capacity'][node_id].items():
            if total_capacity['e'] > params.max_capacity + tolerance:
                reasons.append(f'maximum energy capacity violated at node {node_id}, year {year}')

    if expected_investment_cost > params.budget + tolerance:
        reasons.append('investment budget violated')

    return not reasons, '; '.join(reasons)


def _build_positive_bootstrap_candidate(planning_problem, params):
    shared_ess_data = planning_problem.shared_ess_data
    shared_ess_params = shared_ess_data.params
    ratio = params.energy_to_power_ratio
    if ratio is None:
        ratio = shared_ess_params.min_energy_to_power_ratio
    if not (
            shared_ess_params.min_energy_to_power_ratio
            <= ratio
            <= shared_ess_params.max_energy_to_power_ratio):
        raise ValueError(
            f'configured E/S ratio {ratio:.6f} is outside '
            f'[{shared_ess_params.min_energy_to_power_ratio:.6f}, '
            f'{shared_ess_params.max_energy_to_power_ratio:.6f}]'
        )

    candidate_solution = planning_problem.get_initial_candidate_solution()
    years = list(shared_ess_data.years)
    discounted_unit_cost_total = 0.00
    for node_id in shared_ess_data.active_distribution_network_nodes:
        for year in years:
            annualization = 1.00 / (
                (1.00 + shared_ess_data.discount_factor) ** (int(year) - int(years[0]))
            )
            expected_unit_cost = 0.00
            for scenario, probability in enumerate(shared_ess_data.prob_market_scenarios):
                expected_unit_cost += probability * (
                    shared_ess_data.cost_investment['power'][scenario][year]
                    + ratio * shared_ess_data.cost_investment['energy'][scenario][year]
                )
            discounted_unit_cost_total += annualization * expected_unit_cost

    if discounted_unit_cost_total <= 0.00:
        raise ValueError('expected discounted investment unit cost must be positive')

    target_cost = params.budget_fraction * shared_ess_params.budget
    power_investment = target_cost / discounted_unit_cost_total
    for node_id in shared_ess_data.active_distribution_network_nodes:
        for year in years:
            candidate_solution['investment'][node_id][year]['s'] = power_investment
            candidate_solution['investment'][node_id][year]['e'] = ratio * power_investment
    _rebuild_candidate_total_capacities(planning_problem, candidate_solution)

    maximum_energy_capacity = max(
        total_capacity['e']
        for node_capacities in candidate_solution['total_capacity'].values()
        for total_capacity in node_capacities.values()
    )
    if maximum_energy_capacity > shared_ess_params.max_capacity:
        scale = shared_ess_params.max_capacity / maximum_energy_capacity
        for node_id in shared_ess_data.active_distribution_network_nodes:
            for year in years:
                candidate_solution['investment'][node_id][year]['s'] *= scale
                candidate_solution['investment'][node_id][year]['e'] *= scale
        _rebuild_candidate_total_capacities(planning_problem, candidate_solution)

    minimum_power_investment = min(
        investment['s']
        for node_investments in candidate_solution['investment'].values()
        for investment in node_investments.values()
    )
    if minimum_power_investment <= SMALL_TOLERANCE:
        raise ValueError(
            f'power investment {minimum_power_investment:.6g} is too small to provide '
            f'regular positive-capacity sensitivities; increase budget_fraction'
        )

    feasible, reason = _check_candidate_first_stage_feasibility(
        planning_problem, candidate_solution
    )
    if not feasible:
        raise ValueError(f'constructed candidate is not master-feasible: {reason}')
    return candidate_solution


def _has_missing_investment_sensitivities(planning_problem, sensitivities):
    if sensitivities is None:
        return True
    for capacity_type in ('s', 'e'):
        for year in planning_problem.shared_ess_data.years:
            for node_id in planning_problem.shared_ess_data.active_distribution_network_nodes:
                if (
                        year not in sensitivities.get(capacity_type, {})
                        or node_id not in sensitivities[capacity_type][year]
                        or sensitivities[capacity_type][year][node_id] is None):
                    return True
    return False


def _get_missing_sensitivity_pairs(planning_problem, sensitivities):
    missing_pairs = []
    for node_id in planning_problem.shared_ess_data.active_distribution_network_nodes:
        for year in planning_problem.shared_ess_data.years:
            missing_types = []
            for capacity_type in ('s', 'e'):
                value = None
                if sensitivities is not None:
                    value = sensitivities.get(capacity_type, {}).get(year, {}).get(node_id)
                if value is None:
                    missing_types.append(capacity_type)
            if missing_types:
                missing_pairs.append((node_id, year, tuple(missing_types)))
    return missing_pairs


def _complete_missing_sensitivities_with_probe(
        planning_problem, candidate_solution, sensitivities, params,
        outer_iteration, initial_state=None, debug_flag=False):
    missing_pairs = _get_missing_sensitivity_pairs(planning_problem, sensitivities)
    completed_sensitivities = deepcopy(sensitivities)
    if completed_sensitivities is None:
        completed_sensitivities = {
            capacity_type: {
                year: {
                    node_id: None
                    for node_id in planning_problem.shared_ess_data.active_distribution_network_nodes
                }
                for year in planning_problem.shared_ess_data.years
            }
            for capacity_type in ('s', 'e')
        }
    diagnostics = []
    if not missing_pairs:
        return completed_sensitivities, diagnostics, None

    unsupported_pairs = []
    for node_id, year, _ in missing_pairs:
        investment = candidate_solution['investment'][node_id][year]
        if (
                abs(investment['s']) > SHARED_ESS_ZERO_CAPACITY_TOLERANCE
                or abs(investment['e']) > SHARED_ESS_ZERO_CAPACITY_TOLERANCE):
            unsupported_pairs.append((node_id, year))

    if unsupported_pairs:
        pair_text = ', '.join(
            f'node {node_id}, year {year}' for node_id, year in unsupported_pairs
        )
        print(
            '[WARNING] Missing sensitivities were found at positive-capacity investments '
            f'({pair_text}); the zero-capacity sensitivity probe is not applicable.'
        )
        for node_id, year, missing_types in missing_pairs:
            diagnostics.append({
                'outer_iteration': outer_iteration,
                'node_id': node_id,
                'year': year,
                'missing_types': ','.join(missing_types),
                'status': 'unsupported_positive_capacity',
                'reason': 'At least one missing sensitivity belongs to a positive-capacity investment.',
            })
        return completed_sensitivities, diagnostics, None

    try:
        positive_reference = _build_positive_bootstrap_candidate(
            planning_problem, params
        )
    except ValueError as error:
        print(f'[WARNING] Sensitivity probe could not be constructed: {error}')
        for node_id, year, missing_types in missing_pairs:
            diagnostics.append({
                'outer_iteration': outer_iteration,
                'node_id': node_id,
                'year': year,
                'missing_types': ','.join(missing_types),
                'status': 'construction_failed',
                'reason': str(error),
            })
        return completed_sensitivities, diagnostics, None

    probe_candidate = deepcopy(candidate_solution)
    for node_id, year, _ in missing_pairs:
        reference_investment = positive_reference['investment'][node_id][year]
        probe_candidate['investment'][node_id][year]['s'] = reference_investment['s']
        probe_candidate['investment'][node_id][year]['e'] = reference_investment['e']
    _rebuild_candidate_total_capacities(planning_problem, probe_candidate)

    probe_master_feasible, probe_master_reason = _check_candidate_first_stage_feasibility(
        planning_problem, probe_candidate
    )
    pair_text = ', '.join(
        f'node {node_id}, year {year}' for node_id, year, _ in missing_pairs
    )
    print(
        '[INFO] Running one-sided interior sensitivity probe for zero-capacity '
        f'investments at {pair_text}.'
    )
    initialization_source = 'positive_reference_state' if initial_state is not None else 'cold'
    print(f'[INFO] Sensitivity probe initialization: {initialization_source}.')
    if not probe_master_feasible:
        print(
            '[INFO] The auxiliary sensitivity probe is outside the master feasible set '
            f'({probe_master_reason}); it will be used only to estimate local one-sided slopes.'
        )

    probe_state = None
    probe_convergence = False
    probe_models = None
    probe_sensitivities = None
    probe_esso_violation = None
    probe_recourse = None
    try:
        (
            probe_convergence,
            _,
            probe_models,
            probe_sensitivities,
            _,
            probe_state,
        ) = planning_problem.run_operational_planning(
            candidate_solution=probe_candidate,
            print_results=False,
            debug_flag=debug_flag,
            initial_state=initial_state,
            return_state=True,
        )
        if not probe_state.get('initialization_failed', False):
            probe_esso_violation = planning_problem.shared_ess_data.get_feasibility_violation(
                probe_models['esso']
            )
        if (
                probe_convergence
                and probe_esso_violation is not None
                and probe_esso_violation <= BENDERS_FEASIBILITY_TOLERANCE):
            probe_recourse = planning_problem.get_operational_recourse_value(probe_models)
    finally:
        _restore_candidate_data(planning_problem, candidate_solution)

    probe_feasible = (
        probe_convergence
        and probe_esso_violation is not None
        and probe_esso_violation <= BENDERS_FEASIBILITY_TOLERANCE
        and probe_sensitivities is not None
    )
    completed_count = 0
    for node_id, year, missing_types in missing_pairs:
        reference_investment = probe_candidate['investment'][node_id][year]
        row = {
            'outer_iteration': outer_iteration,
            'node_id': node_id,
            'year': year,
            'missing_types': ','.join(missing_types),
            'probe_power_mva': reference_investment['s'],
            'probe_energy_mvah': reference_investment['e'],
            'probe_master_feasible': probe_master_feasible,
            'probe_master_feasibility_reason': probe_master_reason,
            'initialization_source': initialization_source,
            'operational_convergence': probe_convergence,
            'esso_violation': probe_esso_violation,
            'probe_recourse': probe_recourse,
        }
        pair_completed = probe_feasible
        for capacity_type in missing_types:
            probe_value = None
            if probe_sensitivities is not None:
                probe_value = (
                    probe_sensitivities.get(capacity_type, {})
                    .get(year, {})
                    .get(node_id)
                )
            row[f'sensitivity_{capacity_type}'] = probe_value
            if probe_value is None or not probe_feasible:
                pair_completed = False
            else:
                completed_sensitivities[capacity_type][year][node_id] = probe_value
                completed_count += 1
        row['status'] = 'completed' if pair_completed else 'failed'
        row['reason'] = '' if pair_completed else 'Probe did not return every required feasible sensitivity.'
        diagnostics.append(row)

    required_count = sum(len(missing_types) for _, _, missing_types in missing_pairs)
    if completed_count == required_count:
        print(
            f'[INFO] Sensitivity probe completed all {completed_count} missing '
            'one-sided investment sensitivities.'
        )
    else:
        print(
            f'[WARNING] Sensitivity probe completed {completed_count} of '
            f'{required_count} missing investment sensitivities.'
        )
    return completed_sensitivities, diagnostics, probe_state


def _is_zero_investment_candidate(candidate_solution):
    return all(
        abs(investment[capacity_type]) <= SHARED_ESS_ZERO_CAPACITY_TOLERANCE
        for node_investments in candidate_solution['investment'].values()
        for investment in node_investments.values()
        for capacity_type in ('s', 'e')
    )


def _soh_active_state_changed(baseline_margin, candidate_margin, tolerance):
    if baseline_margin is None or candidate_margin is None:
        return None
    return (baseline_margin <= tolerance) != (candidate_margin <= tolerance)


def _select_finite_difference_validation_reference(incumbent, positive_reference, params):
    if (
            incumbent is not None
            and _select_finite_difference_investment(
                incumbent['candidate_solution'], params
            ) is not None):
        return incumbent, 'incumbent', True
    if (
            positive_reference is not None
            and _select_finite_difference_investment(
                positive_reference['candidate_solution'], params
            ) is not None):
        return positive_reference, 'best_positive_evaluated_candidate', False
    return None, None, None


def _select_finite_difference_investment(candidate_solution, params):
    investments = candidate_solution['investment']

    if params.node_id is not None and params.year is not None:
        node_id = next((value for value in investments if str(value) == str(params.node_id)), None)
        if node_id is None:
            return None
        year = next((value for value in investments[node_id] if str(value) == str(params.year)), None)
        if year is None:
            return None
        investment = investments[node_id][year]
        if not (
                investment['s'] > SMALL_TOLERANCE
                or investment['e'] > SMALL_TOLERANCE):
            return None
        return node_id, year

    selected = None
    selected_power = -float('inf')
    for node_id, yearly_investments in investments.items():
        for year, investment in yearly_investments.items():
            if investment['s'] > selected_power and (
                    investment['s'] > SMALL_TOLERANCE or investment['e'] > SMALL_TOLERANCE):
                selected = node_id, year
                selected_power = investment['s']
    return selected


def _rebuild_candidate_total_capacities(planning_problem, candidate_solution):
    shared_ess_data = planning_problem.shared_ess_data
    years = list(shared_ess_data.years)

    for node_id in candidate_solution['investment']:
        candidate_solution['total_capacity'][node_id] = {
            year: {'s': 0.00, 'e': 0.00} for year in years
        }
        shared_ess_idx = shared_ess_data.get_shared_energy_storage_idx(node_id)
        for y_inv, year_inv in enumerate(years):
            shared_energy_storage = shared_ess_data.shared_energy_storages[year_inv][shared_ess_idx]
            num_years = shared_ess_data.years[year_inv]
            tcal_norm = round(shared_energy_storage.t_cal / num_years)
            max_tcal_norm = min(y_inv + tcal_norm, len(years))
            investment = candidate_solution['investment'][node_id][year_inv]
            for y in range(y_inv, max_tcal_norm):
                year = years[y]
                candidate_solution['total_capacity'][node_id][year]['s'] += investment['s']
                candidate_solution['total_capacity'][node_id][year]['e'] += investment['e']


def _get_investment_soh_margin(planning_problem, esso_models, node_id, year_inv):
    years = list(planning_problem.shared_ess_data.years)
    y_inv = years.index(year_inv)
    model = esso_models[node_id]
    shared_ess_idx = planning_problem.shared_ess_data.get_shared_energy_storage_idx(node_id)
    shared_energy_storage = planning_problem.shared_ess_data.shared_energy_storages[year_inv][shared_ess_idx]
    margins = []
    for y in model.years:
        if not model.es_soh_per_unit_cumul[y_inv, y].fixed:
            margins.append(pe.value(model.es_soh_per_unit_cumul[y_inv, y]) - shared_energy_storage.soh_min)
    return min(margins) if margins else None


def _restore_candidate_data(planning_problem, candidate_solution):
    total_capacity = candidate_solution['total_capacity']
    planning_problem.transmission_network.update_data_with_candidate_solution(total_capacity)
    for distribution_network in planning_problem.distribution_networks.values():
        distribution_network.update_data_with_candidate_solution(total_capacity)
    planning_problem.shared_ess_data.update_data_with_candidate_solution(
        candidate_solution['investment']
    )


def _add_benders_cut(planning_problem, model, recourse_value, sensitivities, candidate_solution):
    years = [year for year in planning_problem.years]
    benders_cut = recourse_value
    for e in model.energy_storages:
        node_id = planning_problem.active_distribution_network_nodes[e]
        for y in model.years:
            year = years[y]
            sensitivity_s = sensitivities['s'][year][node_id]
            sensitivity_e = sensitivities['e'][year][node_id]
            if sensitivity_s is None or sensitivity_e is None:
                return False
            benders_cut += sensitivity_s * (model.es_s_investment[e, y] - candidate_solution['investment'][node_id][year]['s'])
            benders_cut += sensitivity_e * (model.es_e_investment[e, y] - candidate_solution['investment'][node_id][year]['e'])
    print("[INFO] Benders-type procedure. Adding local sensitivity cut...")
    model.benders_cuts.add(model.alpha >= benders_cut)
    return True


# ======================================================================================================================
#  OPERATIONAL PLANNING (DISTRIBUTED)
# ======================================================================================================================
def _run_operational_planning(planning_problem, candidate_solution, initial_state=None, debug_flag=False):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks
    shared_ess_data = planning_problem.shared_ess_data
    admm_parameters = planning_problem.params.admm
    results = {'tso': dict(), 'dso': dict(), 'esso': dict()}
    shared_ess_data.solver_recovery_diagnostics = list()

    # ------------------------------------------------------------------------------------------------------------------
    # 0. Initialization

    print('[INFO]\t - Initializing...')

    start = time.time()
    if initial_state is not None and initial_state.get('initialization_failed', False):
        print('[WARNING] Ignoring a previously failed operational state and rebuilding initialization.')
        initial_state = None
    from_warm_start = initial_state is not None
    primal_evolution = list()
    admm_diagnostics = list()
    continuing_same_candidate = (
        initial_state is not None
        and initial_state.get('candidate_solution') == candidate_solution
    )
    previous_recourse = (
        initial_state.get('last_recourse')
        if continuing_same_candidate else None
    )
    previous_recourse_blocks = (
        deepcopy(initial_state.get('last_recourse_blocks'))
        if continuing_same_candidate else None
    )
    previous_objective_component_blocks = (
        deepcopy(initial_state.get('last_objective_component_blocks'))
        if continuing_same_candidate else None
    )
    previous_slack_component_blocks = (
        deepcopy(initial_state.get('last_slack_component_blocks'))
        if continuing_same_candidate else None
    )
    previous_tso_voltage_slack_state = (
        deepcopy(initial_state.get('last_tso_voltage_slack_state'))
        if continuing_same_candidate
        else None
    )
    consecutive_converged_cycles = (
        initial_state.get('consecutive_converged_cycles', 0)
        if continuing_same_candidate else 0
    )

    if initial_state is None:
        # Create ADMM variables and obtain the initial local solutions.
        consensus_vars, dual_vars = create_admm_variables(planning_problem)
        dso_models, results['dso'] = create_distribution_networks_models(
            distribution_networks,
            consensus_vars,
            candidate_solution['total_capacity'],
            parallel_execution=planning_problem.parallel_execution,
        )
        tso_model, results['tso'] = create_transmission_network_model(
            planning_problem, consensus_vars, candidate_solution['total_capacity']
        )
        esso_model, results['esso'] = create_shared_energy_storage_model(
            shared_ess_data, consensus_vars, candidate_solution['investment']
        )

        if not _admm_local_solves_succeeded(planning_problem, results):
            print(
                '[WARNING] Operational initialization failed because at least one local problem '
                'did not solve successfully. ADMM will not be started.'
            )
            end = time.time()
            total_execution_time = end - start
            print('[INFO] \t - Execution time: {:.2f} s'.format(total_execution_time))
            optim_models = {'tso': tso_model, 'dso': dso_models, 'esso': esso_model}
            state = {
                'models': optim_models,
                'consensus_vars': deepcopy(consensus_vars),
                'dual_vars': deepcopy(dual_vars),
                'candidate_solution': deepcopy(candidate_solution),
                'last_recourse': None,
                'last_recourse_blocks': None,
                'last_objective_component_blocks': None,
                'last_slack_component_blocks': None,
                'last_tso_voltage_slack_state': None,
                'consecutive_converged_cycles': 0,
                'admm_diagnostics': admm_diagnostics,
                'solver_recovery_diagnostics': deepcopy(shared_ess_data.solver_recovery_diagnostics),
                'initialization_failed': True,
            }
            return (
                False,
                results,
                optim_models,
                None,
                primal_evolution,
                total_execution_time,
                state,
            )

        _prepare_distribution_objectives_for_admm(distribution_networks, dso_models)
        _prepare_transmission_objectives_for_admm(transmission_network, tso_model)
        objective_scale = _compute_common_admm_objective_scale(planning_problem, tso_model, dso_models)

        update_distribution_models_to_admm(planning_problem, dso_models, admm_parameters, objective_scale)
        update_transmission_model_to_admm(planning_problem, tso_model, admm_parameters, objective_scale)
        update_shared_energy_storage_model_to_admm(planning_problem, esso_model, admm_parameters)
        _initialize_shared_ess_consensus(planning_problem, consensus_vars)

        # Initialize only the TSO-DSO interface coordination here.
        # Shared-ESS dual variables must remain zero before the first consensus-ADMM cycle.
        planning_problem.update_interface_power_flow_variables(
            tso_model,
            dso_models,
            consensus_vars,
            dual_vars,
            results,
            admm_parameters,
            update_tn=True,
            update_dns=True,
        )

        # The successful initialization populated primal values and IPOPT multipliers.
        from_warm_start = True
    else:
        models = _clone_operational_models(initial_state['models'])
        tso_model = models['tso']
        dso_models = models['dso']
        esso_model = models['esso']
        consensus_vars = deepcopy(initial_state['consensus_vars'])
        dual_vars = deepcopy(initial_state['dual_vars'])
        _update_operational_models_with_candidate(planning_problem, models, candidate_solution)

    sess_available_capacities = shared_ess_data.get_updated_capacities(esso_model)

    # ------------------------------------------------------------------------------------------------------------------
    # ADMM -- Main cycle
    # ------------------------------------------------------------------------------------------------------------------
    convergence = False
    for iter in range(1, admm_parameters.num_max_iters + 1):

        print(f'[INFO] \t - ADMM Iteration {iter}')
        log_debug(f"\t - Memory before iteration {iter}", debug_flag)
        print_memory_usage(f"\t - ADMM Iteration {iter} Start", debug_flag)

        iter_start = time.time()

        # --------------------------------------------------------------------------------------------------------------
        # 1. Solve DSOs problems
        results['dso'] = update_distribution_coordination_models_and_solve(
            distribution_networks, dso_models,
            consensus_vars['vmag'], dual_vars['vmag']['dso'],
            consensus_vars['pf'], dual_vars['pf']['dso'],
            consensus_vars['ess'], dual_vars['ess']['dso'],
            admm_parameters,
            sess_available_capacities,
            from_warm_start=from_warm_start, parallel_execution=planning_problem.parallel_execution
        )

        # Update ADMM consensus variables and primal diagnostics.
        update_and_check_convergence(
            planning_problem, tso_model, dso_models, esso_model,
            consensus_vars, dual_vars, results, admm_parameters,
            primal_evolution,
            update_flags={"update_tn": False, "update_dns": True, "update_sess": False},
            debug_flag=debug_flag,
            check_convergence=False,
        )

        # --------------------------------------------------------------------------------------------------------------
        # 2. Solve TSO problem
        results['tso'] = update_transmission_coordination_model_and_solve(
            transmission_network, tso_model,
            consensus_vars['vmag'], dual_vars['vmag']['tso'],
            consensus_vars['pf'], dual_vars['pf']['tso'],
            consensus_vars['ess'], dual_vars['ess']['tso'],
            admm_parameters,
            sess_available_capacities,
            from_warm_start=from_warm_start
        )

        # Update the proximal centre only with successful TSO solutions.
        tso_proximal_movements = _update_tso_proximal_centres_after_solve(planning_problem, tso_model, results['tso'], cycle=iter)

        # Update ADMM consensus variables and primal diagnostics.
        update_and_check_convergence(
            planning_problem, tso_model, dso_models, esso_model,
            consensus_vars, dual_vars, results, admm_parameters,
            primal_evolution,
            update_flags={"update_tn": True, "update_dns": False, "update_sess": False},
            debug_flag=debug_flag,
            check_convergence=False,
        )

        # --------------------------------------------------------------------------------------------------------------
        # 3. Solve ESSO problem
        results['esso'] = update_shared_energy_storages_coordination_model_and_solve(
            planning_problem, esso_model,
            consensus_vars['ess']['z'], dual_vars['ess']['esso'],
            admm_parameters, from_warm_start=from_warm_start
        )

        # Update the final block and evaluate convergence only after a complete cycle.
        update_and_check_convergence(
            planning_problem, tso_model, dso_models, esso_model,
            consensus_vars, dual_vars, results, admm_parameters,
            primal_evolution,
            update_flags={
                "update_tn": False,
                "update_dns": False,
                "update_sess": True
            },
            debug_flag=debug_flag,
            check_convergence=False,
        )

        # Diagnostic: compare the three shared-ESS schedules after completing the DSO -> TSO -> ESSO ADMM cycle.
        if debug_flag:
            _print_shared_ess_consensus_diagnostics(planning_problem, consensus_vars)

        residual_metrics = get_admm_residual_metrics(planning_problem, tso_model, dso_models, esso_model, consensus_vars)
        _print_worst_primal_residual_diagnostics(residual_metrics, admm_parameters)
        worst_v_primal = residual_metrics.get('worst_v_primal')
        worst_pf_primal = residual_metrics.get('worst_pf_primal')
        worst_pf_dual = residual_metrics.get('worst_pf_dual')
        worst_ess_primal = residual_metrics.get('worst_ess_primal')

        if worst_pf_primal is not None and debug_flag:
            unit = 'MW' if worst_pf_primal['power_type'] == 'p' else 'MVAr'
            print(
                '[DEBUG][PF MAX PRIMAL] '
                f'node={worst_pf_primal["node_id"]}, '
                f'year={worst_pf_primal["year"]}, '
                f'day={worst_pf_primal["day"]}, '
                f'period={worst_pf_primal["period"]}, '
                f'type={worst_pf_primal["power_type"].upper()} | '
                f'TSO={worst_pf_primal["tso_value"]:.6f}, '
                f'DSO={worst_pf_primal["dso_value"]:.6f}, '
                f'diff={worst_pf_primal["absolute_difference"]:.6f} {unit}, '
                f'rating={worst_pf_primal["interface_rating"]:.6f} MVA, '
                f'normalized={worst_pf_primal["normalized_residual"]:.6f}'
            )

        if worst_pf_dual is not None and debug_flag:
            unit = 'MW' if worst_pf_dual['power_type'] == 'p' else 'MVAr'
            print(
                '[DEBUG][PF MAX DUAL] '
                f'agent={worst_pf_dual["agent"].upper()}, '
                f'node={worst_pf_dual["node_id"]}, '
                f'year={worst_pf_dual["year"]}, '
                f'day={worst_pf_dual["day"]}, '
                f'period={worst_pf_dual["period"]}, '
                f'type={worst_pf_dual["power_type"].upper()} | '
                f'current={worst_pf_dual["current_value"]:.6f}, '
                f'previous={worst_pf_dual["previous_value"]:.6f}, '
                f'change={worst_pf_dual["absolute_change"]:.6f} {unit}, '
                f'rho={worst_pf_dual["rho"]:.6f}, '
                f'rating={worst_pf_dual["interface_rating"]:.6f} MVA, '
                f'normalized={worst_pf_dual["normalized_residual"]:.6f}'
            )

        local_solves_ok = _admm_local_solves_succeeded(planning_problem, results)
        residual_convergence = check_admm_convergence(planning_problem, consensus_vars, residual_metrics, admm_parameters, debug_flag=debug_flag)
        if not local_solves_ok:
            print('[WARNING]\t\t - At least one local ADMM problem did not solve successfully.')
            residual_convergence = False

        recourse = None
        gross_operational_cost = None
        recourse_blocks = None
        objective_component_blocks = None
        slack_component_blocks = None
        tso_voltage_slack_state = None
        terminal_salvage_value = None
        objective_change_abs = None
        objective_change_rel = None
        objective_tolerance = None
        objective_convergence = False
        if local_solves_ok:

            operational_models = {
                'tso': tso_model,
                'dso': dso_models,
                'esso': esso_model,
            }

            recourse_components = _get_operational_recourse_components(planning_problem, operational_models)
            gross_operational_cost = recourse_components['gross_operational_cost']
            terminal_salvage_value = recourse_components['terminal_salvage_value']
            recourse = recourse_components['net_operational_recourse']

            # Per-agent / per-year / per-day recourse decomposition.
            recourse_blocks = _get_operational_recourse_block_components(planning_problem, operational_models)
            objective_component_blocks = _get_operational_objective_component_blocks(planning_problem, operational_models)
            slack_component_blocks = _get_operational_slack_component_blocks(planning_problem, operational_models)
            for block_key, slack_components in slack_component_blocks.items():
                slack_total = slack_components['total_slack_penalties']
                unclassified = slack_components['unclassified']
                slack_reconciliation_tolerance = max(1e-4, 1e-10 * max(abs(slack_total), 1.0))
                if abs(unclassified) > slack_reconciliation_tolerance:
                    print(
                        '[WARNING][SLACK COMPONENTS] '
                        f'Block {block_key} does not reconcile | '
                        f'classified={slack_components["classified_total"]:.6e} | '
                        f'exact={slack_total:.6e} | '
                        f'unclassified={unclassified:+.6e}'
                    )
            tso_voltage_slack_state = _get_tso_voltage_slack_state(planning_problem, tso_model)

            # ----------------------------------------------------------------------------------------------------------
            # Check that the decomposition exactly reproduces the net recourse.
            block_recourse = sum(recourse_blocks.values())
            reconciliation_tolerance = max(1e-4, 1e-10 * max(abs(recourse), 1.0))
            if abs(block_recourse - recourse) > reconciliation_tolerance:
                print(
                    '[WARNING][RECOURSE JUMP] '
                    'Block decomposition does not reconcile with net recourse | '
                    f'blocks={block_recourse:.6f} | '
                    f'recourse={recourse:.6f} | '
                    f'difference={block_recourse - recourse:+.6e}'
                )

            # ----------------------------------------------------------------------------------------------------------
            # Existing recourse-stationarity calculation.
            if previous_recourse is not None:
                objective_change_abs = abs(recourse - previous_recourse)
                recourse_scale = max(abs(recourse), abs(previous_recourse), 1.0)
                objective_change_rel = objective_change_abs / recourse_scale
                objective_tolerance = max(admm_parameters.tol['objective']['abs'], admm_parameters.tol['objective']['rel'] * recourse_scale)
                objective_convergence = (objective_change_abs <= objective_tolerance)

            if (
                    recourse_blocks is not None and previous_recourse_blocks is not None and
                    objective_component_blocks is not None and previous_objective_component_blocks is not None and
                    objective_change_abs is not None and objective_tolerance is not None
            ):
                _print_recourse_jump_diagnostics(
                    current_blocks=recourse_blocks,
                    previous_blocks=previous_recourse_blocks,
                    cycle=iter,
                    current_recourse=recourse,
                    previous_recourse=previous_recourse,
                    objective_tolerance=objective_tolerance,
                    top_n=10,
                    tso_proximal_movements=tso_proximal_movements,
                    current_objective_component_blocks=objective_component_blocks,
                    previous_objective_component_blocks=previous_objective_component_blocks,
                    current_slack_component_blocks=slack_component_blocks,
                    previous_slack_component_blocks=previous_slack_component_blocks,
                    current_tso_voltage_slack_state=tso_voltage_slack_state,
                    previous_tso_voltage_slack_state=previous_tso_voltage_slack_state
                )

        if recourse is None:
            print('[INFO]\t\t - Recourse stationarity unavailable after a failed local solve.')
        elif objective_change_abs is None:
            print('[INFO]\t\t - Recourse stationarity requires one previous successful cycle.')
        elif objective_convergence:
            print('[INFO]\t\t - Recourse stationarity ok!')
        else:
            print(f'[INFO]\t\t - Recourse stationarity failed. {objective_change_abs:.6f} > {objective_tolerance:.6f}')

        cycle_convergence = residual_convergence and objective_convergence
        if cycle_convergence:
            consecutive_converged_cycles += 1
        else:
            consecutive_converged_cycles = 0
        convergence = (consecutive_converged_cycles >= admm_parameters.minimum_consecutive_converged_cycles)

        penalty_actions, penalties_before, penalties_after = _update_admm_penalties(tso_model, dso_models, esso_model, residual_metrics, admm_parameters, allow_update=local_solves_ok)
        admm_diagnostics.append({
            'cycle': iter,
            'local_solves_ok': local_solves_ok,
            'primal_v': residual_metrics['primal']['v'],
            'primal_v_mean': residual_metrics['primal']['v_mean'],
            'primal_pf': residual_metrics['primal']['pf'],
            'primal_pf_mean': residual_metrics['primal']['pf_mean'],
            'primal_ess': residual_metrics['primal']['ess'],
            'primal_ess_mean': residual_metrics['primal']['ess_mean'],
            'primal_v_tolerance': admm_parameters.tol['consensus']['v'],
            'primal_v_mean_tolerance': admm_parameters.tol['consensus']['v_mean'],
            'primal_pf_tolerance': admm_parameters.tol['consensus']['pf'],
            'primal_pf_mean_tolerance': admm_parameters.tol['consensus']['pf_mean'],
            'primal_ess_tolerance': admm_parameters.tol['consensus']['ess'],
            'primal_ess_mean_tolerance': admm_parameters.tol['consensus']['ess_mean'],
            'dual_v': residual_metrics['dual']['v'],
            'dual_v_mean': residual_metrics['dual']['v_mean'],
            'dual_pf': residual_metrics['dual']['pf'],
            'dual_pf_mean': residual_metrics['dual']['pf_mean'],
            'dual_ess': residual_metrics['dual']['ess'],
            'dual_ess_mean': residual_metrics['dual']['ess_mean'],
            'dual_v_tolerance': admm_parameters.tol['stationarity']['v'],
            'dual_pf_tolerance': admm_parameters.tol['stationarity']['pf'],
            'dual_ess_tolerance': admm_parameters.tol['stationarity']['ess'],
            'primal_v_ratio': residual_metrics['primal']['v'] / admm_parameters.tol['consensus']['v'],
            'primal_v_mean_ratio': residual_metrics['primal']['v_mean'] / admm_parameters.tol['consensus']['v_mean'],
            'primal_pf_ratio': residual_metrics['primal']['pf'] / admm_parameters.tol['consensus']['pf'],
            'primal_pf_mean_ratio': residual_metrics['primal']['pf_mean'] / admm_parameters.tol['consensus']['pf_mean'],
            'primal_ess_ratio': residual_metrics['primal']['ess'] / admm_parameters.tol['consensus']['ess'],
            'primal_ess_mean_ratio': residual_metrics['primal']['ess_mean'] / admm_parameters.tol['consensus']['ess_mean'],
            'dual_v_ratio': residual_metrics['dual']['v'] / admm_parameters.tol['stationarity']['v'],
            'dual_v_mean_ratio': residual_metrics['dual']['v_mean'] / admm_parameters.tol['stationarity']['v'],
            'dual_pf_ratio': residual_metrics['dual']['pf'] / admm_parameters.tol['stationarity']['pf'],
            'dual_pf_mean_ratio': residual_metrics['dual']['pf_mean'] / admm_parameters.tol['stationarity']['pf'],
            'dual_ess_ratio': residual_metrics['dual']['ess'] / admm_parameters.tol['stationarity']['ess'],
            'dual_ess_mean_ratio': residual_metrics['dual']['ess_mean'] / admm_parameters.tol['stationarity']['ess'],
            'worst_v_primal_node': worst_v_primal['node_id'] if worst_v_primal is not None else None,
            'worst_v_primal_year': worst_v_primal['year'] if worst_v_primal is not None else None,
            'worst_v_primal_day': worst_v_primal['day'] if worst_v_primal is not None else None,
            'worst_v_primal_period': worst_v_primal['period'] if worst_v_primal is not None else None,
            'worst_v_primal_tso': worst_v_primal['tso_value'] if worst_v_primal is not None else None,
            'worst_v_primal_dso': worst_v_primal['dso_value'] if worst_v_primal is not None else None,
            'worst_v_primal_difference': worst_v_primal['absolute_difference'] if worst_v_primal is not None else None,
            'worst_v_primal_base': worst_v_primal['interface_v_base'] if worst_v_primal is not None else None,
            'worst_v_primal_rho_tso': worst_v_primal['rho_tso'] if worst_v_primal is not None else None,
            'worst_v_primal_rho_dso': worst_v_primal['rho_dso'] if worst_v_primal is not None else None,
            'worst_pf_primal_node': worst_pf_primal['node_id'] if worst_pf_primal is not None else None,
            'worst_pf_primal_year': worst_pf_primal['year'] if worst_pf_primal is not None else None,
            'worst_pf_primal_day': worst_pf_primal['day'] if worst_pf_primal is not None else None,
            'worst_pf_primal_period': worst_pf_primal['period'] if worst_pf_primal is not None else None,
            'worst_pf_primal_type': worst_pf_primal['power_type'] if worst_pf_primal is not None else None,
            'worst_pf_primal_tso': worst_pf_primal['tso_value'] if worst_pf_primal is not None else None,
            'worst_pf_primal_dso': worst_pf_primal['dso_value'] if worst_pf_primal is not None else None,
            'worst_pf_primal_difference': worst_pf_primal['absolute_difference'] if worst_pf_primal is not None else None,
            'worst_pf_primal_rating': worst_pf_primal['interface_rating'] if worst_pf_primal is not None else None,
            'worst_ess_primal_node': worst_ess_primal['node_id'] if worst_ess_primal is not None else None,
            'worst_ess_primal_year': worst_ess_primal['year'] if worst_ess_primal is not None else None,
            'worst_ess_primal_day': worst_ess_primal['day'] if worst_ess_primal is not None else None,
            'worst_ess_primal_period': worst_ess_primal['period'] if worst_ess_primal is not None else None,
            'worst_ess_primal_type': worst_ess_primal['power_type'] if worst_ess_primal is not None else None,
            'worst_ess_primal_agent': worst_ess_primal['agent'] if worst_ess_primal is not None else None,
            'worst_ess_primal_agent_value': worst_ess_primal['agent_value'] if worst_ess_primal is not None else None,
            'worst_ess_primal_z': worst_ess_primal['z_value'] if worst_ess_primal is not None else None,
            'worst_ess_primal_difference': worst_ess_primal['absolute_difference'] if worst_ess_primal is not None else None,
            'worst_ess_primal_rating': worst_ess_primal['normalization_rating'] if worst_ess_primal is not None else None,
            'worst_ess_primal_rho': worst_ess_primal['rho'] if worst_ess_primal is not None else None,
            'worst_pf_dual_agent': worst_pf_dual['agent'] if worst_pf_dual is not None else None,
            'worst_pf_dual_node': worst_pf_dual['node_id'] if worst_pf_dual is not None else None,
            'worst_pf_dual_year': worst_pf_dual['year'] if worst_pf_dual is not None else None,
            'worst_pf_dual_day': worst_pf_dual['day'] if worst_pf_dual is not None else None,
            'worst_pf_dual_period': worst_pf_dual['period'] if worst_pf_dual is not None else None,
            'worst_pf_dual_type': worst_pf_dual['power_type'] if worst_pf_dual is not None else None,
            'worst_pf_dual_change': worst_pf_dual['absolute_change'] if worst_pf_dual is not None else None,
            'gross_operational_cost': gross_operational_cost,
            'terminal_salvage_value': terminal_salvage_value,
            'recourse': recourse,
            'objective_change_abs': objective_change_abs,
            'objective_change_rel': objective_change_rel,
            'objective_tolerance': objective_tolerance,
            'objective_absolute_tolerance': admm_parameters.tol['objective']['abs'],
            'objective_relative_tolerance': admm_parameters.tol['objective']['rel'],
            'residual_convergence': residual_convergence,
            'objective_convergence': objective_convergence,
            'cycle_convergence': cycle_convergence,
            'consecutive_converged_cycles': consecutive_converged_cycles,
            'required_consecutive_cycles': admm_parameters.minimum_consecutive_converged_cycles,
            'rho_v_before': penalties_before['v'],
            'rho_pf_before': penalties_before['pf'],
            'rho_ess_before': penalties_before['ess'],
            'rho_v_after': penalties_after['v'],
            'rho_pf_after': penalties_after['pf'],
            'rho_ess_after': penalties_after['ess'],
            'rho_v_action': penalty_actions['v'],
            'rho_pf_action': penalty_actions['pf'],
            'rho_ess_action': penalty_actions['ess'],
        })

        objective_change_text = (f'{objective_change_abs:.6f}' if objective_change_abs is not None else 'N/A')
        objective_tolerance_text = (f'{objective_tolerance:.6f}' if objective_tolerance is not None else 'N/A')
        print(
            f'[INFO]\t\t - ADMM cycle {iter} | '
            f'Primal '
            f'(V max/mean | PF max/mean | ESS max/mean) = '
            f'{residual_metrics["primal"]["v"]:.6f}/'
            f'{residual_metrics["primal"]["v_mean"]:.6f} | '
            f'{residual_metrics["primal"]["pf"]:.6f}/'
            f'{residual_metrics["primal"]["pf_mean"]:.6f} | '
            f'{residual_metrics["primal"]["ess"]:.6f}/'
            f'{residual_metrics["primal"]["ess_mean"]:.6f} | '
            f'Dual '
            f'(V max/mean | PF max/mean | ESS max/mean) = '
            f'{residual_metrics["dual"]["v"]:.6f}/'
            f'{residual_metrics["dual"]["v_mean"]:.6f} | '
            f'{residual_metrics["dual"]["pf"]:.6f}/'
            f'{residual_metrics["dual"]["pf_mean"]:.6f} | '
            f'{residual_metrics["dual"]["ess"]:.6f}/'
            f'{residual_metrics["dual"]["ess_mean"]:.6f} | '
            f'Recourse change = {objective_change_text} '
            f'(tol. {objective_tolerance_text}) | '
            f'Stable cycles = {consecutive_converged_cycles}/'
            f'{admm_parameters.minimum_consecutive_converged_cycles} | '
            f'Penalty actions (V/PF/ESS) = '
            f'{penalty_actions["v"]}/{penalty_actions["pf"]}/{penalty_actions["ess"]}'
        )

        previous_recourse = recourse
        previous_recourse_blocks = recourse_blocks
        previous_objective_component_blocks = objective_component_blocks
        previous_slack_component_blocks = slack_component_blocks
        previous_tso_voltage_slack_state = tso_voltage_slack_state

        sess_available_capacities = shared_ess_data.get_updated_capacities(esso_model)
        if debug_flag:
            print("[DEBUG] available_capacities:")
            for node_id in sess_available_capacities:
                print(f"\t{node_id}:")
                for year in sess_available_capacities[node_id]:
                    print(f"\t{year}: {sess_available_capacities[node_id][year]}")

        # --------------------------------------------------------------------------------------------------------------

        iter_end = time.time()
        print(f"[INFO] \t - Iteration {iter}: {iter_end - iter_start:.2f} s")

        if convergence:
            print(f"[INFO] \t - ADMM converged in {iter} iteration(s).")
            break

        if planning_problem.params.gc:
            gc.collect()
        from_warm_start = True
        log_debug(f"\t - Memory after iteration {iter}", debug_flag)
        print_memory_usage(f"\t - ADMM Iteration {iter} End", debug_flag)

    if not convergence:
        print(f'[WARNING] \t - ADMM did NOT converge in {admm_parameters.num_max_iters} iterations!')

    end = time.time()
    total_execution_time = end - start
    print('[INFO] \t - Execution time: {:.2f} s'.format(total_execution_time))

    optim_models = {'tso': tso_model, 'dso': dso_models, 'esso': esso_model}
    sensitivities = None
    if convergence:
        sensitivities = _get_operational_sensitivities(planning_problem, optim_models)

    state = {
        'models': optim_models,
        'consensus_vars': deepcopy(consensus_vars),
        'dual_vars': deepcopy(dual_vars),
        'candidate_solution': deepcopy(candidate_solution),
        'last_recourse': previous_recourse,
        'last_recourse_blocks': deepcopy(previous_recourse_blocks),
        'last_objective_component_blocks': deepcopy(previous_objective_component_blocks),
        'last_slack_component_blocks': deepcopy(previous_slack_component_blocks),
        'last_tso_voltage_slack_state': deepcopy(previous_tso_voltage_slack_state),
        'consecutive_converged_cycles': consecutive_converged_cycles,
        'admm_diagnostics': admm_diagnostics,
        'solver_recovery_diagnostics': deepcopy(shared_ess_data.solver_recovery_diagnostics),
        'initialization_failed': False,
    }

    return convergence, results, optim_models, sensitivities, primal_evolution, total_execution_time, state


def _clone_operational_models(models):
    if isinstance(models, dict):
        return {key: _clone_operational_models(value) for key, value in models.items()}
    return models.clone()


def _update_operational_models_with_candidate(planning_problem, models, candidate_solution):
    total_capacity = candidate_solution['total_capacity']
    investment = candidate_solution['investment']

    transmission_network = planning_problem.transmission_network
    transmission_network.update_data_with_candidate_solution(total_capacity)
    transmission_network.update_model_with_candidate_solution(models['tso'], total_capacity)

    for node_id, distribution_network in planning_problem.distribution_networks.items():
        distribution_network.update_data_with_candidate_solution(total_capacity)
        distribution_network.update_model_with_candidate_solution(models['dso'][node_id], total_capacity)

    planning_problem.shared_ess_data.update_data_with_candidate_solution(investment)
    planning_problem.shared_ess_data.update_model_with_candidate_solution(models['esso'], investment)


def update_and_check_convergence(planning_problem, tso_model, dso_models, esso_model,
                                 consensus_vars, dual_vars, results, admm_parameters,
                                 primal_evolution, update_flags, debug_flag=False,
                                 check_convergence=True):

    planning_problem.update_admm_consensus_variables(
        tso_model, dso_models, esso_model,
        consensus_vars, dual_vars, results, admm_parameters,
        **update_flags
    )
    primal_value = planning_problem.get_primal_value(tso_model, dso_models, esso_model)
    primal_evolution.append(primal_value)

    if not check_convergence:
        return False

    residual_metrics = get_admm_residual_metrics(
        planning_problem,
        tso_model,
        dso_models,
        esso_model,
        consensus_vars,
    )

    return check_admm_convergence(
        planning_problem,
        consensus_vars,
        residual_metrics,
        admm_parameters,
        debug_flag=debug_flag,
    )


def print_debug_info(planning_problem, consensus_vars, print_vmag=False, print_pf=False, print_ess=False):
    for node_id in planning_problem.active_distribution_network_nodes:
        for year in planning_problem.years:
            if any([print_vmag, print_pf, print_ess]):
                print(f"\tYear {year}")
            for day in planning_problem.days:
                if any([print_vmag, print_pf, print_ess]):
                    print(f"\t\tDay {day}")
                if print_vmag:
                    print(f"\t\tNode {node_id}, {year}, {day}, PF, TSO,  V  {[vmag for vmag in consensus_vars['vmag']['tso']['current'][node_id][year][day]]}")
                    print(f"\t\tNode {node_id}, {year}, {day}, PF, DSO,  V  {[vmag for vmag in consensus_vars['vmag']['dso']['current'][node_id][year][day]]}")
                # if print_pf:
                #     print(f"\t\tNode {node_id}, {year}, {day}, PF, TSO,  P {consensus_vars['pf']['tso']['current'][node_id][year][day]['p']}")
                #     print(f"\t\tNode {node_id}, {year}, {day}, PF, DSO,  P {consensus_vars['pf']['dso']['current'][node_id][year][day]['p']}")
                #     print(f"\t\tNode {node_id}, {year}, {day}, PF, TSO,  Q {consensus_vars['pf']['tso']['current'][node_id][year][day]['q']}")
                #     print(f"\t\tNode {node_id}, {year}, {day}, PF, DSO,  Q {consensus_vars['pf']['dso']['current'][node_id][year][day]['q']}")
                # if print_ess:
                #     print(f"\t\tNode {node_id}, {year}, {day}, ESS, TSO,  P {consensus_vars['ess']['tso']['current'][node_id][year][day]['p']}")
                #     print(f"\t\tNode {node_id}, {year}, {day}, ESS, DSO,  P {consensus_vars['ess']['dso']['current'][node_id][year][day]['p']}")
                #     print(f"\t\tNode {node_id}, {year}, {day}, ESS, TSO,  Q {consensus_vars['ess']['tso']['current'][node_id][year][day]['q']}")
                #     print(f"\t\tNode {node_id}, {year}, {day}, ESS, DSO,  Q {consensus_vars['ess']['dso']['current'][node_id][year][day]['q']}")


def _scenario_probability(network, market_scenario, operation_scenario):
    return network.prob_market_scenarios[market_scenario] * network.prob_operation_scenarios[operation_scenario]


def _get_admm_block_weight(network_data, year, day):
    years = list(network_data.years)
    annualization = 1.0 / ((1.0 + network_data.discount_factor) ** (int(year) - int(years[0])))
    return float(network_data.years[year]) * float(network_data.days[day]) * annualization


def _compute_common_admm_objective_scale(planning_problem, tso_model, dso_models):

    values = []
    entries = []

    # TSO
    transmission_network = planning_problem.transmission_network
    for year in transmission_network.years:
        for day in transmission_network.days:

            weight = _get_admm_block_weight(transmission_network, year, day)
            raw_value = pe.value(tso_model[year][day].objective.expr)
            weighted_value = abs(weight * raw_value)
            scenario_penalty = pe.value(tso_model[year][day].scenario_deviation_penalty) if hasattr(tso_model[year][day], 'scenario_deviation_penalty') else 0.0
            raw_base = raw_value - scenario_penalty
            if isfinite(weighted_value) and weighted_value > SMALL_TOLERANCE:
                values.append(weighted_value)

            entries.append({
                'agent': 'TSO',
                'node': None,
                'year': year,
                'day': day,
                'weight': weight,
                'raw_base': raw_base,
                'scenario_penalty': scenario_penalty,
                'raw_total': raw_value,
                'weighted_base': weight * raw_base,
                'weighted_scenario_penalty': weight * scenario_penalty,
                'weighted_total': weight * raw_value,
            })

    # DSOs
    for node_id, distribution_network in planning_problem.distribution_networks.items():
        for year in distribution_network.years:
            for day in distribution_network.days:

                weight = _get_admm_block_weight(distribution_network, year, day)
                raw_value = pe.value(dso_models[node_id][year][day].objective.expr)
                weighted_value = abs(weight * raw_value)
                scenario_penalty = pe.value(dso_models[node_id][year][day].scenario_deviation_penalty) if hasattr(dso_models[node_id][year][day], 'scenario_deviation_penalty') else 0.0
                raw_base = raw_value - scenario_penalty
                if isfinite(weighted_value) and weighted_value > SMALL_TOLERANCE:
                    values.append(weighted_value)

                entries.append({
                    'agent': 'DSO',
                    'node': node_id,
                    'year': year,
                    'day': day,
                    'weight': weight,
                    'raw_base': raw_base,
                    'scenario_penalty': scenario_penalty,
                    'raw_total': raw_value,
                    'weighted_base': weight * raw_base,
                    'weighted_scenario_penalty': weight * scenario_penalty,
                    'weighted_total': weight * raw_value,
                })

    if not values:
        raise ValueError('Cannot compute common ADMM objective scale: no valid weighted TSO/DSO objective values were found.')

    objective_scale = float(max(values))
    if (not isfinite(objective_scale) or objective_scale <= SMALL_TOLERANCE):
        raise ValueError(f'Invalid common ADMM objective scale: {objective_scale}')

    print(
        '[ADMM OF SCALE] '
        f'n={len(values)} | '
        f'min={min(values):.6e} | '
        f'median={np.median(values):.6e} | '
        f'max={max(values):.6e} | '
        f'selected={objective_scale:.6e}'
    )

    entries_sorted = sorted(entries, key=lambda x: abs(x['weighted_total']), reverse=True)
    print('[ADMM OF SCALE] Largest weighted objectives:')
    for entry in entries_sorted[:10]:
        print(
            f'  {entry["agent"]} '
            f'node={entry["node"]} '
            f'year={entry["year"]} '
            f'day={entry["day"]} | '
            f'base={entry["weighted_base"]:.6e} | '
            f'scenario={entry["weighted_scenario_penalty"]:.6e} | '
            f'total={entry["weighted_total"]:.6e}'
        )

    return objective_scale


def _add_tso_scenario_deviation_penalty(model, network, include_voltage=True):

    voltage_deviation = 0.0
    interface_power_deviation = 0.0
    shared_ess_deviation = 0.0

    for s_m in model.scenarios_market:
        for s_o in model.scenarios_operation:
            probability = _scenario_probability(network, s_m, s_o)
            for p in model.periods:
                for dn in model.active_distribution_networks:
                    if include_voltage:
                        voltage_deviation += probability * (model.vmag_adn[dn, s_m, s_o, p] - model.expected_interface_vmag[dn, p]) ** 2
                    interface_power_deviation += probability * network.baseMVA * ((model.pc_adn[dn, s_m, s_o, p] - model.expected_interface_pf_p[dn, p]) ** 2 + (model.qc_adn[dn, s_m, s_o, p] - model.expected_interface_pf_q[dn, p]) ** 2)
                for e in model.shared_energy_storages:
                    shared_ess_deviation += probability * network.baseMVA * ((model.shared_es_pnet[e, s_m, s_o, p] - model.expected_shared_ess_p[e, p]) ** 2 + (model.shared_es_qnet[e, s_m, s_o, p] - model.expected_shared_ess_q[e, p]) ** 2)

    model.scenario_deviation_weight = pe.Param(initialize=PENALTY_SCENARIO_DEVIATION)
    model.scenario_deviation_voltage = pe.Expression(expr=voltage_deviation)
    model.scenario_deviation_interface_power = pe.Expression(expr=interface_power_deviation)
    model.scenario_deviation_shared_ess = pe.Expression(expr=shared_ess_deviation)

    model.scenario_deviation_penalty = pe.Expression(expr=model.scenario_deviation_weight * (model.scenario_deviation_voltage + model.scenario_deviation_interface_power) + PENALTY_SHARED_ESS_SCENARIO_DEVIATION * model.scenario_deviation_shared_ess)
    model.objective.expr = copy(model.objective.expr) + model.scenario_deviation_penalty


def _add_dso_scenario_deviation_penalty(model, network, include_voltage=True):

    voltage_deviation = 0.0
    interface_power_deviation = 0.0
    shared_ess_deviation = 0.0
    ref_node_id = network.get_reference_node_id()
    shared_ess_idx = network.get_shared_energy_storage_idx(ref_node_id)

    for s_m in model.scenarios_market:
        for s_o in model.scenarios_operation:
            probability = _scenario_probability(network, s_m, s_o)
            for p in model.periods:
                if include_voltage:
                    voltage_deviation += probability * (model.vmag_adn[s_m, s_o, p] - model.expected_interface_vmag[p]) ** 2
                interface_power_deviation += probability * network.baseMVA * ((model.pg_adn[s_m, s_o, p] - model.expected_interface_pf_p[p]) ** 2 + (model.qg_adn[s_m, s_o, p] - model.expected_interface_pf_q[p]) ** 2)
                shared_ess_deviation += probability * network.baseMVA * ((model.shared_es_pnet[shared_ess_idx, s_m, s_o, p] - model.expected_shared_ess_p[p]) ** 2 + (model.shared_es_qnet[shared_ess_idx, s_m, s_o, p] - model.expected_shared_ess_q[p]) ** 2)

    model.scenario_deviation_weight = pe.Param(initialize=PENALTY_SCENARIO_DEVIATION)
    model.scenario_deviation_voltage = pe.Expression(expr=voltage_deviation)
    model.scenario_deviation_interface_power = pe.Expression(expr=interface_power_deviation)
    model.scenario_deviation_shared_ess = pe.Expression(expr=shared_ess_deviation)

    model.scenario_deviation_penalty = pe.Expression(expr=model.scenario_deviation_weight * (model.scenario_deviation_voltage + model.scenario_deviation_interface_power) + PENALTY_SHARED_ESS_SCENARIO_DEVIATION * model.scenario_deviation_shared_ess)
    model.objective.expr = copy(model.objective.expr) + model.scenario_deviation_penalty


def _add_tso_scenario_tracking_penalty(model, network, active_distribution_network_nodes, interface_vmag, interface_pf):

    voltage_tracking = 0.0
    interface_power_tracking = 0.0
    for dn in model.active_distribution_networks:
        node_id = active_distribution_network_nodes[dn]
        for s_m in model.scenarios_market:
            for s_o in model.scenarios_operation:
                probability = _scenario_probability(network, s_m, s_o)
                for p in model.periods:
                    vmag_req = interface_vmag[node_id][p]
                    p_req = interface_pf[node_id]['p'][p] / network.baseMVA
                    q_req = interface_pf[node_id]['q'][p] / network.baseMVA
                    voltage_tracking += probability * (
                        model.vmag_adn[dn, s_m, s_o, p] - vmag_req
                    ) ** 2
                    interface_power_tracking += probability * network.baseMVA * (
                        (model.pc_adn[dn, s_m, s_o, p] - p_req) ** 2
                        + (model.qc_adn[dn, s_m, s_o, p] - q_req) ** 2
                    )

    model.scenario_tracking_weight = pe.Param(
        initialize=PENALTY_SCENARIO_DEVIATION * 1e6
    )
    model.scenario_tracking_voltage = pe.Expression(expr=voltage_tracking)
    model.scenario_tracking_interface_power = pe.Expression(
        expr=interface_power_tracking
    )
    model.scenario_tracking_penalty = pe.Expression(
        expr=model.scenario_tracking_weight * (
            model.scenario_tracking_voltage
            + model.scenario_tracking_interface_power
        )
    )
    model.objective.expr = copy(model.objective.expr) + model.scenario_tracking_penalty


def create_transmission_network_model(planning_problem, consensus_vars, candidate_solution):

    print(f'[INFO] \t\t - Transmission Network...')

    # Build model, fix candidate solution
    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks
    transmission_network.update_data_with_candidate_solution(candidate_solution)
    tso_model = transmission_network.build_model()
    transmission_network.update_model_with_candidate_solution(tso_model, candidate_solution)

    # Update model with expected interface values
    for year in transmission_network.years:
        for day in transmission_network.days:

            s_base = transmission_network.network[year][day].baseMVA
            tso_model[year][day].active_distribution_networks = range(len(transmission_network.active_distribution_network_nodes))

            # Free Vmag, Pc, Qc at the interface nodes, fix base Pc and Qc profiles
            for dn in tso_model[year][day].active_distribution_networks:

                adn_node_id = transmission_network.active_distribution_network_nodes[dn]
                adn_node_idx = transmission_network.network[year][day].get_node_idx(adn_node_id)
                adn_load_idx = transmission_network.network[year][day].get_adn_load_idx(adn_node_id)
                distribution_network = distribution_networks[adn_node_id]
                interface_transf_rating = distribution_network.network[year][day].get_interface_branch_rating() / s_base

                for s_m in tso_model[year][day].scenarios_market:
                    for s_o in tso_model[year][day].scenarios_operation:
                        for p in tso_model[year][day].periods:

                            # Interface voltage remains governed by the explicit squared-voltage constraints; remove its slacks.
                            if transmission_network.params.slacks.grid_operation.voltage:
                                tso_model[year][day].slack_v_sqr_down[adn_node_idx, s_m, s_o, p].setub(0.00)
                                tso_model[year][day].slack_v_sqr_up[adn_node_idx, s_m, s_o, p].setub(0.00)

                            # Fix Pc and Qc (base profiles), free pc_adn and qc_adn
                            interface_pf_p = consensus_vars['pf']['dso']['current'][adn_node_id][year][day]['p'][p] / s_base
                            interface_pf_q = consensus_vars['pf']['dso']['current'][adn_node_id][year][day]['q'][p] / s_base

                            tso_model[year][day].pc[adn_load_idx, s_m, s_o, p].setub(None)
                            tso_model[year][day].pc[adn_load_idx, s_m, s_o, p].setlb(None)
                            tso_model[year][day].qc[adn_load_idx, s_m, s_o, p].setub(None)
                            tso_model[year][day].qc[adn_load_idx, s_m, s_o, p].setlb(None)
                            fix_or_set(tso_model[year][day].pc[adn_load_idx, s_m, s_o, p], interface_pf_p)
                            fix_or_set(tso_model[year][day].qc[adn_load_idx, s_m, s_o, p], interface_pf_q)

                            tso_model[year][day].flex_p_up[adn_load_idx, s_m, s_o, p].fixed = False
                            tso_model[year][day].flex_p_down[adn_load_idx, s_m, s_o, p].fixed = False
                            tso_model[year][day].flex_q_up[adn_load_idx, s_m, s_o, p].fixed = False
                            tso_model[year][day].flex_q_down[adn_load_idx, s_m, s_o, p].fixed = False
                            tso_model[year][day].flex_p_up[adn_load_idx, s_m, s_o, p].setub(interface_transf_rating)
                            tso_model[year][day].flex_p_down[adn_load_idx, s_m, s_o, p].setub(interface_transf_rating)
                            tso_model[year][day].flex_q_up[adn_load_idx, s_m, s_o, p].setub(interface_transf_rating)
                            tso_model[year][day].flex_q_down[adn_load_idx, s_m, s_o, p].setub(interface_transf_rating)

            # Add expected interface values shared-ESS schedule
            tso_model[year][day].expected_interface_vmag = pe.Var(tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, domain=pe.NonNegativeReals, initialize=1.0)
            tso_model[year][day].expected_interface_pf_p = pe.Var(tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
            tso_model[year][day].expected_interface_pf_q = pe.Var(tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
            tso_model[year][day].expected_shared_ess_p = pe.Var(tso_model[year][day].shared_energy_storages, tso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
            tso_model[year][day].expected_shared_ess_q = pe.Var(tso_model[year][day].shared_energy_storages, tso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
            tso_model[year][day].expected_interface_vmag_def = pe.Constraint( tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, rule=partial(tn_interface_expected_vmag_rule, network=transmission_network.network[year][day]))
            tso_model[year][day].expected_interface_pf_p_def = pe.Constraint(tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, rule=partial(tn_interface_expected_pf_p_rule, network=transmission_network.network[year][day]))
            tso_model[year][day].expected_interface_pf_q_def = pe.Constraint(tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, rule=partial(tn_interface_expected_pf_q_rule, network=transmission_network.network[year][day]))
            tso_model[year][day].expected_shared_ess_p_def = pe.Constraint(tso_model[year][day].shared_energy_storages, tso_model[year][day].periods, rule=partial(tn_interface_expected_sess_p_rule, network=transmission_network.network[year][day],),)
            tso_model[year][day].expected_shared_ess_q_def = pe.Constraint(tso_model[year][day].shared_energy_storages, tso_model[year][day].periods, rule=partial(tn_interface_expected_sess_q_rule, network=transmission_network.network[year][day],),)
            for e in tso_model[year][day].shared_energy_storages:
                configure_shared_ess_operational_state(tso_model[year][day], e, pe.value(tso_model[year][day].shared_es_s_rated_fixed[e]), pe.value(tso_model[year][day].shared_es_e_rated_fixed[e]),)

            # A soft, probability-weighted penalty promotes one expected interface schedule.
            _add_tso_scenario_deviation_penalty(tso_model[year][day], transmission_network.network[year][day])

    # Run SMOPF
    results = transmission_network.optimize(tso_model)

    # Get initial interface and shared ESS values
    for year in transmission_network.years:
        for day in transmission_network.days:
            if not _solver_result_succeeded(results[year][day]):
                continue
            s_base = transmission_network.network[year][day].baseMVA
            for dn in tso_model[year][day].active_distribution_networks:
                adn_node_id = transmission_network.active_distribution_network_nodes[dn]
                v_base = transmission_network.network[year][day].get_node_base_kv(adn_node_id)
                shared_ess_idx = transmission_network.network[year][day].get_shared_energy_storage_idx(adn_node_id)
                for p in tso_model[year][day].periods:
                    interface_vmag = pe.value(tso_model[year][day].expected_interface_vmag[dn, p]) * v_base
                    interface_pf_p = pe.value(tso_model[year][day].expected_interface_pf_p[dn, p]) * s_base
                    interface_pf_q = pe.value(tso_model[year][day].expected_interface_pf_q[dn, p]) * s_base
                    p_ess = pe.value(tso_model[year][day].expected_shared_ess_p[shared_ess_idx, p]) * s_base
                    q_ess = pe.value(tso_model[year][day].expected_shared_ess_q[shared_ess_idx, p]) * s_base
                    consensus_vars['vmag']['tso']['current'][adn_node_id][year][day][p] = interface_vmag
                    consensus_vars['pf']['tso']['current'][adn_node_id][year][day]['p'][p] = interface_pf_p
                    consensus_vars['pf']['tso']['current'][adn_node_id][year][day]['q'][p] = interface_pf_q
                    consensus_vars['ess']['tso']['current'][adn_node_id][year][day]['p'][p] = p_ess
                    consensus_vars['ess']['tso']['current'][adn_node_id][year][day]['q'][p] = q_ess

    return tso_model, results


def create_distribution_networks_models(distribution_networks, consensus_vars, candidate_solution, parallel_execution=False):
    if parallel_execution:
        return create_distribution_networks_models_parallel(distribution_networks, consensus_vars, candidate_solution)
    else:
        return create_distribution_networks_models_sequential(distribution_networks, consensus_vars, candidate_solution)


def create_distribution_networks_models_sequential(distribution_networks, consensus_vars, candidate_solution):

    dso_models = dict()
    results = dict()

    for node_id in distribution_networks:

        print(f'[INFO] \t\t - Distribution Network node {node_id}...')

        distribution_network = distribution_networks[node_id]

        # Build model, fix candidate solution
        distribution_network.update_data_with_candidate_solution(candidate_solution)
        dso_model = distribution_network.build_model()
        distribution_network.update_model_with_candidate_solution(dso_model, candidate_solution)

        for year in distribution_network.years:
            for day in distribution_network.days:

                ref_node_id = distribution_network.network[year][day].get_reference_node_id()
                shared_ess_idx = distribution_network.network[year][day].get_shared_energy_storage_idx(ref_node_id)

                # Add expected interface values shared-ESS schedule
                dso_model[year][day].expected_interface_vmag = pe.Var(dso_model[year][day].periods, domain=pe.NonNegativeReals, initialize=1.00)
                dso_model[year][day].expected_interface_pf_p = pe.Var(dso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
                dso_model[year][day].expected_interface_pf_q = pe.Var(dso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
                dso_model[year][day].expected_shared_ess_p = pe.Var(dso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
                dso_model[year][day].expected_shared_ess_q = pe.Var(dso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
                dso_model[year][day].expected_interface_vmag_def = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_vmag_rule, network=distribution_network.network[year][day]))
                dso_model[year][day].expected_interface_pf_p_def = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_pf_p_rule, network=distribution_network.network[year][day]))
                dso_model[year][day].expected_interface_pf_q_def = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_pf_q_rule, network=distribution_network.network[year][day]))
                dso_model[year][day].expected_shared_ess_p_def = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_sess_p_rule, network=distribution_network.network[year][day], shared_ess_idx=shared_ess_idx,),)
                dso_model[year][day].expected_shared_ess_q_def = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_sess_q_rule, network=distribution_network.network[year][day], shared_ess_idx=shared_ess_idx,),)
                configure_shared_ess_operational_state(dso_model[year][day], shared_ess_idx, pe.value(dso_model[year][day].shared_es_s_rated_fixed[shared_ess_idx]), pe.value(dso_model[year][day].shared_es_e_rated_fixed[shared_ess_idx]),)

                # A soft, probability-weighted penalty promotes one expected interface schedule.
                _add_dso_scenario_deviation_penalty(dso_model[year][day], distribution_network.network[year][day],)

        # Run SMOPF
        results[node_id] = distribution_network.optimize(dso_model)

        # Get initial interface and shared ESS values
        for year in distribution_network.years:
            for day in distribution_network.days:
                if not _solver_result_succeeded(results[node_id][year][day]):
                    continue
                ref_node_id = distribution_network.network[year][day].get_reference_node_id()
                s_base = distribution_network.network[year][day].baseMVA
                v_base = distribution_network.network[year][day].get_node_base_kv(ref_node_id)
                for p in dso_model[year][day].periods:

                    interface_vmag = pe.value(dso_model[year][day].expected_interface_vmag[p]) * v_base
                    interface_pf_p = pe.value(dso_model[year][day].expected_interface_pf_p[p]) * s_base
                    interface_pf_q = pe.value(dso_model[year][day].expected_interface_pf_q[p]) * s_base
                    p_ess = pe.value(dso_model[year][day].expected_shared_ess_p[p]) * s_base
                    q_ess = pe.value(dso_model[year][day].expected_shared_ess_q[p]) * s_base

                    consensus_vars['vmag']['dso']['current'][node_id][year][day][p] = interface_vmag
                    consensus_vars['pf']['dso']['current'][node_id][year][day]['p'][p] = interface_pf_p
                    consensus_vars['pf']['dso']['current'][node_id][year][day]['q'][p] = interface_pf_q
                    consensus_vars['ess']['dso']['current'][node_id][year][day]['p'][p] = p_ess
                    consensus_vars['ess']['dso']['current'][node_id][year][day]['q'][p] = q_ess

        dso_models[node_id] = dso_model

    return dso_models, results


def create_distribution_networks_models_parallel(distribution_networks, consensus_vars, candidate_solution):

    results = dict()
    dso_models = dict()
    for node_id in distribution_networks:
        results[node_id] = dict()
        dso_models[node_id] = dict()

    tasks = []
    max_workers = min(os.cpu_count() // 2, len(distribution_networks))  # Note: to limit memory usage
    with ProcessPoolExecutor(max_workers=max_workers) as executor:

        for node_id in distribution_networks:
            tasks.append(executor.submit(create_distribution_network_model, node_id, distribution_networks[node_id], candidate_solution))

        for future in as_completed(tasks):

            node_id, result, model = future.result()
            results[node_id] = result
            dso_models[node_id] = model

            # Get initial interface and shared ESS values
            for year in distribution_networks[node_id].years:
                for day in distribution_networks[node_id].days:
                    if not _solver_result_succeeded(results[node_id][year][day]):
                        continue
                    ref_node_id = distribution_networks[node_id].network[year][day].get_reference_node_id()
                    s_base = distribution_networks[node_id].network[year][day].baseMVA
                    v_base = distribution_networks[node_id].network[year][day].get_node_base_kv(ref_node_id)
                    for p in dso_models[node_id][year][day].periods:
                        interface_vmag = pe.value(dso_models[node_id][year][day].expected_interface_vmag[p]) * v_base
                        interface_pf_p = pe.value(dso_models[node_id][year][day].expected_interface_pf_p[p]) * s_base
                        interface_pf_q = pe.value(dso_models[node_id][year][day].expected_interface_pf_q[p]) * s_base
                        p_ess = pe.value(dso_models[node_id][year][day].expected_shared_ess_p[p]) * s_base
                        q_ess = pe.value(dso_models[node_id][year][day].expected_shared_ess_q[p]) * s_base

                        consensus_vars['vmag']['dso']['current'][node_id][year][day][p] = interface_vmag
                        consensus_vars['pf']['dso']['current'][node_id][year][day]['p'][p] = interface_pf_p
                        consensus_vars['pf']['dso']['current'][node_id][year][day]['q'][p] = interface_pf_q
                        consensus_vars['ess']['dso']['current'][node_id][year][day]['p'][p] = p_ess
                        consensus_vars['ess']['dso']['current'][node_id][year][day]['q'][p] = q_ess

    return dso_models, results


def create_distribution_network_model(node_id, distribution_network, candidate_solution):

    # Build model, fix candidate solution
    distribution_network.update_data_with_candidate_solution(candidate_solution)
    dso_model = distribution_network.build_model()
    distribution_network.update_model_with_candidate_solution(dso_model, candidate_solution)

    # Update model with expected interface values
    for year in distribution_network.years:
        for day in distribution_network.days:

            ref_node_id = distribution_network.network[year][day].get_reference_node_id()
            shared_ess_idx = distribution_network.network[year][day].get_shared_energy_storage_idx(ref_node_id)

            # Add expected interface values and shared-ESS schedule
            dso_model[year][day].expected_interface_vmag = pe.Var(dso_model[year][day].periods, domain=pe.NonNegativeReals, initialize=1.00)
            dso_model[year][day].expected_interface_pf_p = pe.Var(dso_model[year][day].periods, domain=pe.Reals, initialize=0.00)
            dso_model[year][day].expected_interface_pf_q = pe.Var(dso_model[year][day].periods, domain=pe.Reals, initialize=0.00)
            dso_model[year][day].expected_shared_ess_p = pe.Var(dso_model[year][day].periods, domain=pe.Reals, initialize=0.00)
            dso_model[year][day].expected_shared_ess_q = pe.Var(dso_model[year][day].periods, domain=pe.Reals, initialize=0.00)
            dso_model[year][day].interface_expected_values_vmag = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_vmag_rule, network=distribution_network.network[year][day]))
            dso_model[year][day].interface_expected_values_pf_p = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_pf_p_rule, network=distribution_network.network[year][day]))
            dso_model[year][day].interface_expected_values_pf_q = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_pf_q_rule, network=distribution_network.network[year][day]))
            dso_model[year][day].expected_shared_ess_p_def = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_sess_p_rule, network=distribution_network.network[year][day], shared_ess_idx=shared_ess_idx,),)
            dso_model[year][day].expected_shared_ess_q_def = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_sess_q_rule, network=distribution_network.network[year][day], shared_ess_idx=shared_ess_idx,),)
            configure_shared_ess_operational_state(dso_model[year][day], shared_ess_idx, pe.value(dso_model[year][day].shared_es_s_rated_fixed[shared_ess_idx]), pe.value(dso_model[year][day].shared_es_e_rated_fixed[shared_ess_idx]),)

    # Add probability-weighted deviations from the expected interface schedule.
    for year in distribution_network.years:
        for day in distribution_network.days:
            _add_dso_scenario_deviation_penalty(dso_model[year][day], distribution_network.network[year][day],)

    # Run SMOPF
    res = distribution_network.optimize(dso_model)

    return node_id, res, dso_model


def create_shared_energy_storage_model(shared_ess_data, consensus_vars, candidate_solution):

    years = list(shared_ess_data.years)
    days = list(shared_ess_data.days)

    # Build model, fix candidate solution
    shared_ess_data.update_data_with_candidate_solution(candidate_solution)
    esso_model = shared_ess_data.build_subproblem()
    shared_ess_data.update_model_with_candidate_solution(esso_model, candidate_solution)

    # Fix TSO's request
    for node_id in shared_ess_data.active_distribution_network_nodes:
        for y in esso_model[node_id].years:
            year = years[y]
            for d in esso_model[node_id].days:
                day = days[d]
                for p in esso_model[node_id].periods:
                    p_req = consensus_vars['ess']['tso']['current'][node_id][year][day]['p'][p]
                    q_req = consensus_vars['ess']['tso']['current'][node_id][year][day]['q'][p]
                    fix_or_set(esso_model[node_id].es_pnet[y, d, p], p_req)
                    fix_or_set(esso_model[node_id].es_qnet[y, d, p], q_req)

    # Run optimization
    results = shared_ess_data.optimize(esso_model)

    # Get initial shared ESS values
    for node_id in shared_ess_data.active_distribution_network_nodes:
        if not _solver_result_succeeded(results[node_id]):
            continue
        for y in esso_model[node_id].years:
            year = years[y]
            for d in esso_model[node_id].days:
                day = days[d]
                for p in esso_model[node_id].periods:
                    shared_ess_p = pe.value(esso_model[node_id].es_pnet[y, d, p])
                    shared_ess_q = pe.value(esso_model[node_id].es_qnet[y, d, p])
                    consensus_vars['ess']['esso']['current'][node_id][year][day]['p'][p] = shared_ess_p
                    consensus_vars['ess']['esso']['current'][node_id][year][day]['q'][p] = shared_ess_q
                    consensus_vars['ess']['esso']['prev'][node_id][year][day]['p'][p] = shared_ess_p
                    consensus_vars['ess']['esso']['prev'][node_id][year][day]['q'][p] = shared_ess_q

    return esso_model, results


def _get_primal_value(planning_problem, tso_model, dso_models, esso_model):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks
    shared_ess_data = planning_problem.shared_ess_data

    primal_value = 0.0
    primal_value += transmission_network.get_primal_value(tso_model)
    for node_id in distribution_networks:
        primal_value += distribution_networks[node_id].get_primal_value(dso_models[node_id])
    primal_value += shared_ess_data.get_primal_value(esso_model)

    return primal_value



# ======================================================================================================================
#  OPERATIONAL PLANNING (HIERARCHICAL)
# ======================================================================================================================
def _run_operational_planning_hierarchical(planning_problem, num_steps=8, print_pq_map=False, debug_flag=False):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks
    results = {'tso': dict(), 'dso': dict()}

    start = time.time()

    # Get initial DN solutions
    dso_models = dict()
    for node_id in distribution_networks:
        distribution_network = distribution_networks[node_id]
        dso_models[node_id] = distribution_network.get_pq_map(num_steps=num_steps, print_pq_map=print_pq_map)

    pf_requested = dict()
    tso_model = transmission_network.build_model()
    for year in transmission_network.years:
        pf_requested[year] = dict()
        for day in transmission_network.days:

            pf_requested[year][day] = dict()
            for adn_node_id in transmission_network.active_distribution_network_nodes:
                pf_requested[year][day][adn_node_id] = dict()
                for p in tso_model[year][day].periods:
                    pf_requested[year][day][adn_node_id][p] = dict()

            s_base = transmission_network.network[year][day].baseMVA
            tso_model[year][day].active_distribution_networks = range(len(transmission_network.active_distribution_network_nodes))

            # TN, Fix Pc, Qc at the interface nodes, free flexibility
            for dn in tso_model[year][day].active_distribution_networks:
                adn_node_id = transmission_network.active_distribution_network_nodes[dn]
                adn_load_idx = transmission_network.network[year][day].get_adn_load_idx(adn_node_id)
                for s_m in tso_model[year][day].scenarios_market:
                    for s_o in tso_model[year][day].scenarios_operation:
                        for p in tso_model[year][day].periods:
                            init_solution = dso_models[adn_node_id][year][day][p]['initial_solution']
                            tso_model[year][day].pc[adn_load_idx, s_m, s_o, p].setub(init_solution['Pg'] / s_base + EQUALITY_TOLERANCE)
                            tso_model[year][day].pc[adn_load_idx, s_m, s_o, p].setlb(init_solution['Pg'] / s_base - EQUALITY_TOLERANCE)
                            tso_model[year][day].qc[adn_load_idx, s_m, s_o, p].setub(init_solution['Qg'] / s_base + EQUALITY_TOLERANCE)
                            tso_model[year][day].qc[adn_load_idx, s_m, s_o, p].setlb(init_solution['Qg'] / s_base - EQUALITY_TOLERANCE)
                            tso_model[year][day].flex_p_up[adn_load_idx, s_m, s_o, p].setub(None)
                            tso_model[year][day].flex_p_down[adn_load_idx, s_m, s_o, p].setub(None)
                            tso_model[year][day].flex_q_up[adn_load_idx, s_m, s_o, p].setub(None)
                            tso_model[year][day].flex_q_down[adn_load_idx, s_m, s_o, p].setub(None)
                            if transmission_network.params.l_curt:
                                tso_model[year][day].pc_curt_down[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)
                                tso_model[year][day].pc_curt_up[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)
                                tso_model[year][day].qc_curt_down[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)
                                tso_model[year][day].qc_curt_up[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)

            # TN, Add expected interface values
            tso_model[year][day].expected_interface_vmag = pe.Var(tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, domain=pe.NonNegativeReals, initialize=1.0)
            tso_model[year][day].expected_interface_pf_p = pe.Var(tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
            tso_model[year][day].expected_interface_pf_q = pe.Var(tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
            tso_model[year][day].expected_interface_vmag_def = pe.Constraint( tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, rule=partial(tn_interface_expected_vmag_rule, network=transmission_network.network[year][day]))
            tso_model[year][day].expected_interface_pf_p_def = pe.Constraint(tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, rule=partial(tn_interface_expected_pf_p_rule, network=transmission_network.network[year][day]))
            tso_model[year][day].expected_interface_pf_q_def = pe.Constraint(tso_model[year][day].active_distribution_networks, tso_model[year][day].periods, rule=partial(tn_interface_expected_pf_q_rule, network=transmission_network.network[year][day]))

            # TN, Add ADNs' PQ maps constraints, Fix expected vmag
            tso_model[year][day].pq_maps = pe.ConstraintList()
            for dn in tso_model[year][day].active_distribution_networks:
                adn_node_id = transmission_network.active_distribution_network_nodes[dn]
                for p in tso_model[year][day].periods:
                    adn_pq_map = dso_models[adn_node_id][year][day][p]
                    # initial_solution = adn_pq_map['initial_solution']
                    # tso_model[year][day].pq_maps.add(tso_model[year][day].expected_interface_vmag[dn, p] <= initial_solution['Vg'] + EQUALITY_TOLERANCE)
                    # tso_model[year][day].pq_maps.add(tso_model[year][day].expected_interface_vmag[dn, p] >= initial_solution['Vg'] - EQUALITY_TOLERANCE)
                    for ineq in adn_pq_map['inequalities']:
                        a = ineq['Pg']
                        b = ineq['Qg']
                        c = ineq['c'] / s_base
                        tso_model[year][day].pq_maps.add(a * tso_model[year][day].expected_interface_pf_p[dn, p] + b * tso_model[year][day].expected_interface_pf_q[dn, p] <= c)

            # Promote the expected interface schedule represented by the PQ maps.
            _add_tso_scenario_deviation_penalty(tso_model[year][day], transmission_network.network[year][day], include_voltage=False,)

    # Optimize TN, Get resulting interface PFs
    print(f'[INFO] - Running OPF on {transmission_network.name} with hierarchical constraints...')
    results['tso'] = transmission_network.optimize(tso_model)
    for year in transmission_network.years:
        for day in transmission_network.days:
            s_base = transmission_network.network[year][day].baseMVA
            for dn in tso_model[year][day].active_distribution_networks:
                adn_node_id = transmission_network.active_distribution_network_nodes[dn]
                for p in tso_model[year][day].periods:
                    vmag = pe.value(tso_model[year][day].expected_interface_vmag[dn, p])
                    pc = pe.value(tso_model[year][day].expected_interface_pf_p[dn, p]) * s_base
                    qc = pe.value(tso_model[year][day].expected_interface_pf_q[dn, p]) * s_base
                    pf_requested[year][day][adn_node_id][p] = {'Pg': pc, 'Qg': qc, 'Vg': vmag}
                    # if print_pq_map:
                    #     init_oper_point = dso_models[adn_node_id][year][day][p]
                    #     final_oper_point = pf_requested[year][day][adn_node_id]
                    #     distribution_network = operational_planning.distribution_networks[adn_node_id]
                    #     distribution_network.pq_map_comparison(t=t, num_steps_max=num_steps, initial_solution=init_oper_point, final_solution=final_oper_point)

    # Run OPF on DNs, considering established power flow (settlement)
    dso_models = dict()
    print('[INFO] - Running settlement on Distribution Networks...')
    for node_id in distribution_networks:

        distribution_network = distribution_networks[node_id]
        print(f'[INFO]\t - Network {distribution_network.name}...')

        dso_model = distribution_network.build_model()
        distribution_network.update_of_to_settlement(dso_model)

        for year in planning_problem.years:
            for day in planning_problem.days:
                for p in dso_model[year][day].periods:
                    dso_model[year][day].interface_vmag_req[p].fix(pf_requested[year][day][node_id][p]['Vg'])
                    dso_model[year][day].interface_pf_p_req[p].fix(pf_requested[year][day][node_id][p]['Pg'] / distribution_network.network[year][day].baseMVA)
                    dso_model[year][day].interface_pf_q_req[p].fix(pf_requested[year][day][node_id][p]['Qg'] / distribution_network.network[year][day].baseMVA)

        results['dso'][node_id] = distribution_network.optimize(dso_model)
        dso_models[node_id] = dso_model

    end = time.time()
    total_execution_time = end - start
    print('[INFO] \t - Execution time: {:.2f} s'.format(total_execution_time))

    optim_models = {'tso': tso_model, 'dso': dso_models}

    return results, optim_models, total_execution_time


# ======================================================================================================================
#  OPERATIONAL PLANNING (CENTRALIZED)
# ======================================================================================================================
def _run_operational_planning_centralized(planning_problem, debug_flag=False):

    # Combined networks
    centralized_network = planning_problem.combine_networks()
    for year in centralized_network.years:
        for day in centralized_network.days:
            if centralized_network.params.print_to_screen:
                centralized_network.network[year][day].print_network_to_screen()
            if centralized_network.params.plot_diagram:
                centralized_network.network[year][day].plot_diagram()

    # Run SMOPF
    centralized_model = centralized_network.build_model()

    print(f'[INFO] - Running SMOPF, Network {centralized_network.name}...')
    results = centralized_network.optimize(centralized_model, print_header=False)

    return centralized_network, results, centralized_model


# ======================================================================================================================
#  ADMM functions
# ======================================================================================================================
def create_admm_variables(planning_problem):

    num_instants = planning_problem.num_instants
    use_ess_previous_iter = any(planning_problem.params.admm.previous_iter['ess'].values())

    consensus_variables = {
        'vmag': {
            'tso': {'current': dict(), 'prev': dict()},
            'dso': {'current': dict(), 'prev': dict()}
        },
        'pf': {
            'tso': {'current': dict(), 'prev': dict()},
            'dso': {'current': dict(), 'prev': dict()}
        },
        'ess': {
            'tso': {'current': dict(), 'prev': dict()},
            'dso': {'current': dict(), 'prev': dict()},
            'esso': {'current': dict(), 'prev': dict()},
            'z': {'current': dict(), 'prev': dict()}
        }
    }

    dual_variables = {
        'vmag': {'tso': {'current': dict()}, 'dso': {'current': dict()}},
        'pf': {'tso': {'current': dict()}, 'dso': {'current': dict()}},
        'ess': {'tso': {'current': dict()}, 'dso': {'current': dict()}, 'esso': {'current': dict()}}
    }

    if use_ess_previous_iter:
        dual_variables['ess']['tso']['prev'] = dict()
        dual_variables['ess']['dso']['prev'] = dict()

    for dn in range(len(planning_problem.active_distribution_network_nodes)):

        node_id = planning_problem.active_distribution_network_nodes[dn]

        consensus_variables['vmag']['tso']['current'][node_id] = dict()
        consensus_variables['vmag']['dso']['current'][node_id] = dict()
        consensus_variables['pf']['tso']['current'][node_id] = dict()
        consensus_variables['pf']['dso']['current'][node_id] = dict()
        consensus_variables['ess']['tso']['current'][node_id] = dict()
        consensus_variables['ess']['dso']['current'][node_id] = dict()
        consensus_variables['ess']['esso']['current'][node_id] = dict()
        consensus_variables['ess']['z']['current'][node_id] = dict()

        consensus_variables['vmag']['tso']['prev'][node_id] = dict()
        consensus_variables['vmag']['dso']['prev'][node_id] = dict()
        consensus_variables['pf']['tso']['prev'][node_id] = dict()
        consensus_variables['pf']['dso']['prev'][node_id] = dict()
        consensus_variables['ess']['tso']['prev'][node_id] = dict()
        consensus_variables['ess']['dso']['prev'][node_id] = dict()
        consensus_variables['ess']['esso']['prev'][node_id] = dict()
        consensus_variables['ess']['z']['prev'][node_id] = dict()

        dual_variables['vmag']['tso']['current'][node_id] = dict()
        dual_variables['vmag']['dso']['current'][node_id] = dict()
        dual_variables['pf']['tso']['current'][node_id] = dict()
        dual_variables['pf']['dso']['current'][node_id] = dict()
        dual_variables['ess']['tso']['current'][node_id] = dict()
        dual_variables['ess']['dso']['current'][node_id] = dict()
        dual_variables['ess']['esso']['current'][node_id] = dict()

        if use_ess_previous_iter:
            dual_variables['ess']['tso']['prev'][node_id] = dict()
            dual_variables['ess']['dso']['prev'][node_id] = dict()

        for year in planning_problem.years:

            consensus_variables['vmag']['tso']['current'][node_id][year] = dict()
            consensus_variables['vmag']['dso']['current'][node_id][year] = dict()
            consensus_variables['pf']['tso']['current'][node_id][year] = dict()
            consensus_variables['pf']['dso']['current'][node_id][year] = dict()
            consensus_variables['ess']['tso']['current'][node_id][year] = dict()
            consensus_variables['ess']['dso']['current'][node_id][year] = dict()
            consensus_variables['ess']['esso']['current'][node_id][year] = dict()
            consensus_variables['ess']['z']['current'][node_id][year] = dict()

            consensus_variables['vmag']['tso']['prev'][node_id][year] = dict()
            consensus_variables['vmag']['dso']['prev'][node_id][year] = dict()
            consensus_variables['pf']['tso']['prev'][node_id][year] = dict()
            consensus_variables['pf']['dso']['prev'][node_id][year] = dict()
            consensus_variables['ess']['tso']['prev'][node_id][year] = dict()
            consensus_variables['ess']['dso']['prev'][node_id][year] = dict()
            consensus_variables['ess']['esso']['prev'][node_id][year] = dict()
            consensus_variables['ess']['z']['prev'][node_id][year] = dict()

            dual_variables['vmag']['tso']['current'][node_id][year] = dict()
            dual_variables['vmag']['dso']['current'][node_id][year] = dict()
            dual_variables['pf']['tso']['current'][node_id][year] = dict()
            dual_variables['pf']['dso']['current'][node_id][year] = dict()
            dual_variables['ess']['tso']['current'][node_id][year] = dict()
            dual_variables['ess']['dso']['current'][node_id][year] = dict()
            dual_variables['ess']['esso']['current'][node_id][year] = dict()

            if use_ess_previous_iter:
                dual_variables['ess']['tso']['prev'][node_id][year] = dict()
                dual_variables['ess']['dso']['prev'][node_id][year] = dict()

            for day in planning_problem.days:

                node_base_kv = planning_problem.transmission_network.network[year][day].get_node_base_kv(node_id)

                consensus_variables['vmag']['tso']['current'][node_id][year][day] = [node_base_kv] * num_instants
                consensus_variables['vmag']['dso']['current'][node_id][year][day] = [node_base_kv] * num_instants
                consensus_variables['pf']['tso']['current'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['pf']['dso']['current'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['ess']['tso']['current'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['ess']['dso']['current'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['ess']['esso']['current'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['ess']['z']['current'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}

                consensus_variables['vmag']['tso']['prev'][node_id][year][day] = [node_base_kv] * num_instants
                consensus_variables['vmag']['dso']['prev'][node_id][year][day] = [node_base_kv] * num_instants
                consensus_variables['pf']['tso']['prev'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['pf']['dso']['prev'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['ess']['tso']['prev'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['ess']['dso']['prev'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['ess']['esso']['prev'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                consensus_variables['ess']['z']['prev'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}

                dual_variables['vmag']['tso']['current'][node_id][year][day] = [0.0] * planning_problem.num_instants
                dual_variables['vmag']['dso']['current'][node_id][year][day] = [0.0] * planning_problem.num_instants
                dual_variables['pf']['tso']['current'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                dual_variables['pf']['dso']['current'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                dual_variables['ess']['tso']['current'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                dual_variables['ess']['dso']['current'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                dual_variables['ess']['esso']['current'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}

                if use_ess_previous_iter:
                    dual_variables['ess']['tso']['prev'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}
                    dual_variables['ess']['dso']['prev'][node_id][year][day] = {'p': [0.0] * num_instants, 'q': [0.0] * num_instants}

    return consensus_variables, dual_variables


def _initialize_shared_ess_consensus(planning_problem, consensus_vars):

    for node_id in planning_problem.active_distribution_network_nodes:
        for year in planning_problem.years:
            for day in planning_problem.days:
                for power_type in ('p', 'q'):
                    for p in range(planning_problem.num_instants):

                        tso_value = consensus_vars['ess']['tso']['current'][node_id][year][day][power_type][p]
                        dso_value = consensus_vars['ess']['dso']['current'][node_id][year][day][power_type][p]
                        esso_value = consensus_vars['ess']['esso']['current'][node_id][year][day][power_type][p]
                        z_value = (tso_value + dso_value + esso_value) / 3.0

                        consensus_vars['ess']['z']['current'][node_id][year][day][power_type][p] = z_value
                        consensus_vars['ess']['z']['prev'][node_id][year][day][power_type][p] = z_value


def _shared_ess_admm_normalization_mva(rating_mva, floor_mva):
    return max(abs(rating_mva), floor_mva)


def _shared_ess_admm_normalization_pu(rating_pu, s_base, floor_mva):
    if s_base <= 0.00:
        raise ValueError('Network base power must be positive for ADMM normalization.')
    rating_mva = abs(rating_pu) * s_base
    return _shared_ess_admm_normalization_mva(rating_mva, floor_mva) / s_base


def _prepare_transmission_objectives_for_admm(transmission_network, model):
    for year in transmission_network.years:
        for day in transmission_network.days:
            model[year][day].penalty_ess_usage.set_value(0.00)
            model[year][day].penalty_gen_curtailment.set_value(0.00)
            if transmission_network.params.obj_type == OBJ_MIN_COST:
                model[year][day].cost_load_curtailment.set_value(COST_CONSUMPTION_CURTAILMENT)
            elif transmission_network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
                model[year][day].penalty_load_curtailment.set_value(PENALTY_LOAD_CURTAILMENT)
                model[year][day].penalty_flex_usage.set_value(0.00)


def update_transmission_model_to_admm(planning_problem, model, params, objective_scale):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks

    proximal_cfg = params.proximal_regularization
    use_tso_proximal = (proximal_cfg['enabled'] and proximal_cfg['tso']['enabled'])
    tso_proximal_cfg = proximal_cfg['tso']

    for year in transmission_network.years:
        for day in transmission_network.days:

            s_base = transmission_network.network[year][day].baseMVA

            if use_tso_proximal:
                model[year][day].prox_gamma_v = pe.Param(initialize=tso_proximal_cfg['gamma']['v'])
                model[year][day].prox_gamma_pf = pe.Param(initialize=tso_proximal_cfg['gamma']['pf'])
                model[year][day].prox_gamma_ess = pe.Param(initialize=tso_proximal_cfg['gamma']['ess'])

                # Previous successful TSO iterate: interface voltage
                model[year][day].prox_v_prev = pe.Param(model[year][day].active_distribution_networks, model[year][day].periods, mutable=True, domain=pe.Reals, initialize=1.0)

                # Previous successful TSO iterate: interface P/Q
                model[year][day].prox_pf_p_prev = pe.Param(model[year][day].active_distribution_networks, model[year][day].periods, mutable=True, domain=pe.Reals, initialize=0.0)
                model[year][day].prox_pf_q_prev = pe.Param(model[year][day].active_distribution_networks, model[year][day].periods, mutable=True, domain=pe.Reals, initialize=0.0)

                # Previous successful TSO iterate: shared-ESS P/Q
                model[year][day].prox_ess_p_prev = pe.Param(model[year][day].shared_energy_storages, model[year][day].periods, mutable=True, domain=pe.Reals, initialize=0.0)
                model[year][day].prox_ess_q_prev = pe.Param(model[year][day].shared_energy_storages, model[year][day].periods, mutable=True, domain=pe.Reals, initialize=0.0)

                for dn in model[year][day].active_distribution_networks:
                    for p in model[year][day].periods:
                        model[year][day].prox_v_prev[dn, p].set_value(pe.value(model[year][day].expected_interface_vmag[dn, p]))
                        model[year][day].prox_pf_p_prev[dn, p].set_value(pe.value(model[year][day].expected_interface_pf_p[dn, p]))
                        model[year][day].prox_pf_q_prev[dn, p].set_value(pe.value(model[year][day].expected_interface_pf_q[dn, p]))

                for e in model[year][day].shared_energy_storages:
                    for p in model[year][day].periods:
                        model[year][day].prox_ess_p_prev[e, p].set_value(pe.value(model[year][day].expected_shared_ess_p[e, p]))
                        model[year][day].prox_ess_q_prev[e, p].set_value(pe.value(model[year][day].expected_shared_ess_q[e, p]))

            # Add ADMM variables
            model[year][day].rho_v = pe.Param(mutable=True, domain=pe.NonNegativeReals, initialize=params.rho['v'][transmission_network.name])
            model[year][day].vmag_req = pe.Param(model[year][day].active_distribution_networks, model[year][day].periods, mutable=True, domain=pe.NonNegativeReals)    # Square of voltage magnitude
            model[year][day].dual_vmag_req = pe.Param(model[year][day].active_distribution_networks, model[year][day].periods, mutable=True, domain=pe.Reals)          # Dual variable - voltage magnitude requested

            model[year][day].rho_pf = pe.Param(mutable=True, domain=pe.NonNegativeReals, initialize=params.rho['pf'][transmission_network.name])
            model[year][day].p_pf_req = pe.Param(model[year][day].active_distribution_networks, model[year][day].periods, mutable=True, domain=pe.Reals)                # Active power - requested by distribution networks
            model[year][day].q_pf_req = pe.Param(model[year][day].active_distribution_networks, model[year][day].periods, mutable=True, domain=pe.Reals)                # Reactive power - requested by distribution networks
            model[year][day].dual_pf_p_req = pe.Param(model[year][day].active_distribution_networks, model[year][day].periods, mutable=True, domain=pe.Reals)           # Dual variable - active power requested
            model[year][day].dual_pf_q_req = pe.Param(model[year][day].active_distribution_networks, model[year][day].periods, mutable=True, domain=pe.Reals)           # Dual variable - reactive power requested

            model[year][day].rho_ess = pe.Param(mutable=True, domain=pe.NonNegativeReals, initialize=params.rho['ess'][transmission_network.name])
            model[year][day].p_ess_req = pe.Param(model[year][day].shared_energy_storages, model[year][day].periods, mutable=True, domain=pe.Reals)                     # SharedESS - Active power requested (DSO)
            model[year][day].q_ess_req = pe.Param(model[year][day].shared_energy_storages, model[year][day].periods, mutable=True, domain=pe.Reals)                     # SharedESS - Reactive power requested (DSO)
            model[year][day].dual_ess_p_req = pe.Param(model[year][day].shared_energy_storages, model[year][day].periods, mutable=True, domain=pe.Reals)                # Dual variable - SharedESS active power
            model[year][day].dual_ess_q_req = pe.Param(model[year][day].shared_energy_storages, model[year][day].periods, mutable=True, domain=pe.Reals)                # Dual variable - SharedESS reactive power
            if params.previous_iter['ess']['tso']:
                model[year][day].rho_ess_prev = pe.Param(mutable=True, domain=pe.NonNegativeReals, initialize=params.rho_previous_iter['ess'][transmission_network.name])
                model[year][day].p_ess_prev = pe.Param(model[year][day].shared_energy_storages, model[year][day].periods, mutable=True, domain=pe.Reals)                # SharedESS - previous iteration active power
                model[year][day].q_ess_prev = pe.Param(model[year][day].shared_energy_storages, model[year][day].periods, mutable=True, domain=pe.Reals)                # SharedESS - previous iteration reactive power
                model[year][day].dual_ess_p_prev = pe.Param(model[year][day].shared_energy_storages, model[year][day].periods, mutable=True, domain=pe.Reals)           # Dual variable - previous iteration shared ESS active power
                model[year][day].dual_ess_q_prev = pe.Param(model[year][day].shared_energy_storages, model[year][day].periods, mutable=True, domain=pe.Reals)           # Dual variable - previous iteration shared ESS reactive power


            # Objective function - augmented Lagrangian
            block_weight = _get_admm_block_weight(transmission_network, year, day)
            effective_scale = objective_scale / block_weight

            model[year][day].admm_common_objective_scale = pe.Param(initialize=objective_scale)
            model[year][day].admm_block_weight = pe.Param(initialize=block_weight)
            model[year][day].admm_objective_scale = pe.Param(initialize=effective_scale)
            obj = copy(model[year][day].objective.expr) / effective_scale

            for dn in model[year][day].active_distribution_networks:

                adn_node_id = transmission_network.active_distribution_network_nodes[dn]
                distribution_network = distribution_networks[adn_node_id]
                interface_transf_rating = distribution_network.network[year][day].get_interface_branch_rating() / s_base

                for p in model[year][day].periods:

                    constraint_v_req = (model[year][day].expected_interface_vmag[dn, p] - model[year][day].vmag_req[dn, p])
                    obj += model[year][day].dual_vmag_req[dn, p] * constraint_v_req
                    obj += (model[year][day].rho_v / 2) * (constraint_v_req ** 2)

                    constraint_p_req = (model[year][day].expected_interface_pf_p[dn, p] - model[year][day].p_pf_req[dn, p]) / interface_transf_rating
                    constraint_q_req = (model[year][day].expected_interface_pf_q[dn, p] - model[year][day].q_pf_req[dn, p]) / interface_transf_rating
                    obj += model[year][day].dual_pf_p_req[dn, p] * constraint_p_req
                    obj += model[year][day].dual_pf_q_req[dn, p] * constraint_q_req
                    obj += (model[year][day].rho_pf / 2) * (constraint_p_req ** 2)
                    obj += (model[year][day].rho_pf / 2) * (constraint_q_req ** 2)

                    if use_tso_proximal:

                        # ----------------------------------------------------------------------
                        # Pure proximal regularization: interface voltage
                        proximal_v = model[year][day].expected_interface_vmag[dn, p] - model[year][day].prox_v_prev[dn, p]
                        obj += (model[year][day].prox_gamma_v / 2) * proximal_v ** 2

                        # ----------------------------------------------------------------------
                        # Pure proximal regularization: interface active/reactive power
                        proximal_pf_p = (model[year][day].expected_interface_pf_p[dn, p] - model[year][day].prox_pf_p_prev[dn, p]) / interface_transf_rating
                        proximal_pf_q = (model[year][day].expected_interface_pf_q[dn, p] - model[year][day].prox_pf_q_prev[dn, p]) / interface_transf_rating
                        obj += (model[year][day].prox_gamma_pf / 2) * proximal_pf_p ** 2
                        obj += (model[year][day].prox_gamma_pf / 2) * proximal_pf_q ** 2

            for e in model[year][day].shared_energy_storages:

                shared_ess_rating = _shared_ess_admm_normalization_pu(transmission_network.network[year][day].shared_energy_storages[e].s, s_base, params.shared_ess_normalization_floor_mva)

                for p in model[year][day].periods:

                    constraint_ess_p_req = (model[year][day].expected_shared_ess_p[e, p] - model[year][day].p_ess_req[e, p]) / (2 * shared_ess_rating)
                    constraint_ess_q_req = (model[year][day].expected_shared_ess_q[e, p] - model[year][day].q_ess_req[e, p]) / (2 * shared_ess_rating)
                    obj += (model[year][day].dual_ess_p_req[e, p]) * constraint_ess_p_req
                    obj += (model[year][day].dual_ess_q_req[e, p]) * constraint_ess_q_req
                    obj += (model[year][day].rho_ess / 2) * constraint_ess_p_req ** 2
                    obj += (model[year][day].rho_ess / 2) * constraint_ess_q_req ** 2

                    if params.previous_iter['ess']['tso']:
                        constraint_ess_p_prev = (model[year][day].expected_shared_ess_p[e, p] - model[year][day].p_ess_prev[e, p]) / (2 * shared_ess_rating)
                        constraint_ess_q_prev = (model[year][day].expected_shared_ess_q[e, p] - model[year][day].q_ess_prev[e, p]) / (2 * shared_ess_rating)
                        obj += (model[year][day].rho_ess_prev / 2) * constraint_ess_p_prev ** 2
                        obj += (model[year][day].rho_ess_prev / 2) * constraint_ess_q_prev ** 2

                    if use_tso_proximal:
                        proximal_ess_p = (model[year][day].expected_shared_ess_p[e, p] - model[year][day].prox_ess_p_prev[e, p]) / (2 * shared_ess_rating)
                        proximal_ess_q = (model[year][day].expected_shared_ess_q[e, p] - model[year][day].prox_ess_q_prev[e, p]) / (2 * shared_ess_rating)
                        obj += (model[year][day].prox_gamma_ess / 2) * proximal_ess_p ** 2
                        obj += (model[year][day].prox_gamma_ess / 2) * proximal_ess_q ** 2

            # Add ADMM OF, deactivate original OF
            model[year][day].objective.deactivate()
            model[year][day].admm_objective = pe.Objective(sense=pe.minimize, expr=obj)


def _prepare_distribution_objectives_for_admm(distribution_networks, models):
    for node_id in distribution_networks:
        dso_model = models[node_id]
        distribution_network = distribution_networks[node_id]
        for year in distribution_network.years:
            for day in distribution_network.days:
                dso_model[year][day].penalty_ess_usage.set_value(0.00)
                # dso_model[year][day].penalty_gen_curtailment.set_value(0.00)
                if distribution_network.params.obj_type == OBJ_MIN_COST:
                    dso_model[year][day].cost_load_curtailment.set_value(COST_CONSUMPTION_CURTAILMENT)
                elif distribution_network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
                    dso_model[year][day].penalty_load_curtailment.set_value(PENALTY_LOAD_CURTAILMENT)
                    dso_model[year][day].penalty_flex_usage.set_value(0.00)


def update_distribution_models_to_admm(planning_problem, models, params, objective_scale):

    distribution_networks = planning_problem.distribution_networks

    for node_id in distribution_networks:

        dso_model = models[node_id]
        distribution_network = distribution_networks[node_id]

        for year in distribution_network.years:
            for day in distribution_network.days:

                s_base = distribution_network.network[year][day].baseMVA
                ref_node_id = distribution_network.network[year][day].get_reference_node_id()
                ref_node_idx = distribution_network.network[year][day].get_node_idx(ref_node_id)
                ref_node = distribution_network.network[year][day].nodes[ref_node_idx]
                voltage_upper = voltage_numerical_upper_bound(ref_node)

                # Free the interface magnitude while retaining the reference angle.
                for s_m in dso_model[year][day].scenarios_market:
                    for s_o in dso_model[year][day].scenarios_operation:
                        for p in dso_model[year][day].periods:
                            dso_model[year][day].e[ref_node_idx, s_m, s_o, p].fixed = False
                            dso_model[year][day].e[ref_node_idx, s_m, s_o, p].setub(voltage_upper)
                            dso_model[year][day].e[ref_node_idx, s_m, s_o, p].setlb(0.00)
                            dso_model[year][day].f[ref_node_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)
                            dso_model[year][day].f[ref_node_idx, s_m, s_o, p].setlb(-EQUALITY_TOLERANCE)
                            if distribution_network.params.slacks.grid_operation.voltage:
                                dso_model[year][day].slack_v_sqr_down[ref_node_idx, s_m, s_o, p].setub(0.00)
                                dso_model[year][day].slack_v_sqr_up[ref_node_idx, s_m, s_o, p].setub(0.00)

                # Add ADMM variables
                dso_model[year][day].rho_v = pe.Param(mutable=True, domain=pe.NonNegativeReals, initialize=params.rho['v'][distribution_network.network[year][day].name])
                dso_model[year][day].vmag_req = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.NonNegativeReals)       # Voltage magnitude - requested by TSO
                dso_model[year][day].dual_vmag_req = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)             # Dual variable - voltage magnitude

                dso_model[year][day].rho_pf = pe.Param(mutable=True, domain=pe.NonNegativeReals, initialize=params.rho['pf'][distribution_network.network[year][day].name])
                dso_model[year][day].p_pf_req = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)                   # Active power - requested by TSO
                dso_model[year][day].q_pf_req = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)                   # Reactive power - requested by TSO
                dso_model[year][day].dual_pf_p_req = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)              # Dual variable - active power
                dso_model[year][day].dual_pf_q_req = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)              # Dual variable - reactive power

                dso_model[year][day].rho_ess = pe.Param(mutable=True, domain=pe.NonNegativeReals, initialize=params.rho['ess'][distribution_network.network[year][day].name])
                dso_model[year][day].p_ess_req = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)                  # SharedESS - active power requested (TSO)
                dso_model[year][day].q_ess_req = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)                  # SharedESS - reactive power requested (TSO)
                dso_model[year][day].dual_ess_p_req = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)             # Dual variable - SharedESS active power
                dso_model[year][day].dual_ess_q_req = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)             # Dual variable - SharedESS reactive power
                if params.previous_iter['ess']['dso']:
                    dso_model[year][day].rho_ess_prev = pe.Param(mutable=True, domain=pe.NonNegativeReals, initialize=params.rho_previous_iter['ess'][distribution_network.network[year][day].name])
                    dso_model[year][day].p_ess_prev = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)             # SharedESS - previous iteration active power
                    dso_model[year][day].q_ess_prev = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)             # SharedESS - previous iteration reactive power
                    dso_model[year][day].dual_ess_p_prev = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)        # Dual variable - SharedESS previous iteration active power
                    dso_model[year][day].dual_ess_q_prev = pe.Param(dso_model[year][day].periods, mutable=True, domain=pe.Reals)        # Dual variable - SharedESS previous iteration reactive power

                # Objective function - augmented Lagrangian
                block_weight = _get_admm_block_weight(distribution_network, year, day)
                effective_scale = objective_scale / block_weight

                dso_model[year][day].admm_common_objective_scale = pe.Param(initialize=objective_scale)
                dso_model[year][day].admm_block_weight = pe.Param(initialize=block_weight)
                dso_model[year][day].admm_objective_scale = pe.Param(initialize=effective_scale)
                obj = copy(dso_model[year][day].objective.expr) / effective_scale

                shared_ess_idx = distribution_network.network[year][day].get_shared_energy_storage_idx(ref_node_id)
                shared_ess_rating = _shared_ess_admm_normalization_pu(
                    distribution_network.network[year][day].shared_energy_storages[shared_ess_idx].s,
                    s_base,
                    params.shared_ess_normalization_floor_mva,
                )

                interface_transf_rating = distribution_network.network[year][day].get_interface_branch_rating() / s_base

                # Augmented Lagrangian -- Interface power flow (residual balancing)
                for p in dso_model[year][day].periods:

                    # Voltage magnitude
                    constraint_vmag_req = (dso_model[year][day].expected_interface_vmag[p] - dso_model[year][day].vmag_req[p])
                    obj += (dso_model[year][day].dual_vmag_req[p]) * constraint_vmag_req
                    obj += (dso_model[year][day].rho_v / 2) * (constraint_vmag_req ** 2)

                    # Interface power flow
                    constraint_p_req = (dso_model[year][day].expected_interface_pf_p[p] - dso_model[year][day].p_pf_req[p]) / interface_transf_rating
                    constraint_q_req = (dso_model[year][day].expected_interface_pf_q[p] - dso_model[year][day].q_pf_req[p]) / interface_transf_rating
                    obj += (dso_model[year][day].dual_pf_p_req[p]) * constraint_p_req
                    obj += (dso_model[year][day].dual_pf_q_req[p]) * constraint_q_req
                    obj += (dso_model[year][day].rho_pf / 2) * (constraint_p_req ** 2)
                    obj += (dso_model[year][day].rho_pf / 2) * (constraint_q_req ** 2)

                    # SharedESS
                    constraint_ess_p_req = (dso_model[year][day].expected_shared_ess_p[p] - dso_model[year][day].p_ess_req[p]) / (2 * shared_ess_rating)
                    constraint_ess_q_req = (dso_model[year][day].expected_shared_ess_q[p] - dso_model[year][day].q_ess_req[p]) / (2 * shared_ess_rating)
                    obj += (dso_model[year][day].dual_ess_p_req[p]) * constraint_ess_p_req
                    obj += (dso_model[year][day].dual_ess_q_req[p]) * constraint_ess_q_req
                    obj += (dso_model[year][day].rho_ess / 2) * constraint_ess_p_req ** 2
                    obj += (dso_model[year][day].rho_ess / 2) * constraint_ess_q_req ** 2
                    if params.previous_iter['ess']['dso']:
                        constraint_ess_p_prev = (dso_model[year][day].expected_shared_ess_p[p] - dso_model[year][day].p_ess_prev[p]) / (2 * shared_ess_rating)
                        constraint_ess_q_prev = (dso_model[year][day].expected_shared_ess_q[p] - dso_model[year][day].q_ess_prev[p]) / (2 * shared_ess_rating)
                        obj += (dso_model[year][day].dual_ess_p_prev[p]) * constraint_ess_p_prev
                        obj += (dso_model[year][day].dual_ess_q_prev[p]) * constraint_ess_q_prev
                        obj += (dso_model[year][day].rho_ess_prev / 2) * constraint_ess_p_prev ** 2
                        obj += (dso_model[year][day].rho_ess_prev / 2) * constraint_ess_q_prev ** 2

                # Add ADMM OF, deactivate original OF
                dso_model[year][day].objective.deactivate()
                dso_model[year][day].admm_objective = pe.Objective(sense=pe.minimize, expr=obj)


def update_shared_energy_storage_model_to_admm(planning_problem, models, params):

    shared_ess_data = planning_problem.shared_ess_data
    years = list(shared_ess_data.years)

    for node_id in shared_ess_data.active_distribution_network_nodes:

        shared_ess_idx = shared_ess_data.get_shared_energy_storage_idx(node_id)

        # Add ADMM variables
        models[node_id].rho = pe.Param(mutable=True, domain=pe.NonNegativeReals, initialize=params.rho['ess']['esso'])

        # Free Pnet, Qnet
        for y in models[node_id].years:
            for d in models[node_id].days:
                for p in models[node_id].periods:
                    models[node_id].es_pnet[y, d, p].fixed = False
                    models[node_id].es_qnet[y, d, p].fixed = False

        # Active and Reactive power requested by TSO and DSOs
        models[node_id].p_req = pe.Param(models[node_id].years, models[node_id].days, models[node_id].periods, mutable=True, domain=pe.Reals)
        models[node_id].q_req = pe.Param(models[node_id].years, models[node_id].days, models[node_id].periods, mutable=True, domain=pe.Reals)
        models[node_id].dual_p_req = pe.Param(models[node_id].years, models[node_id].days, models[node_id].periods, mutable=True, domain=pe.Reals)
        models[node_id].dual_q_req = pe.Param(models[node_id].years, models[node_id].days, models[node_id].periods, mutable=True, domain=pe.Reals)

        # Objective function - augmented Lagrangian
        obj = copy(models[node_id].objective.expr)
        for y in models[node_id].years:
            year = years[y]
            shared_ess_rating = _shared_ess_admm_normalization_mva(
                shared_ess_data.shared_energy_storages[year][shared_ess_idx].s,
                params.shared_ess_normalization_floor_mva,
            )
            for d in models[node_id].days:
                for p in models[node_id].periods:
                    constraint_p_req = (models[node_id].es_pnet[y, d, p] - models[node_id].p_req[y, d, p]) / (2 * shared_ess_rating)
                    constraint_q_req = (models[node_id].es_qnet[y, d, p] - models[node_id].q_req[y, d, p]) / (2 * shared_ess_rating)
                    obj += models[node_id].dual_p_req[y, d, p] * constraint_p_req
                    obj += models[node_id].dual_q_req[y, d, p] * constraint_q_req
                    obj += (models[node_id].rho / 2) * constraint_p_req ** 2
                    obj += (models[node_id].rho / 2) * constraint_q_req ** 2

        # Add ADMM OF, deactivate original OF
        models[node_id].admm_objective = pe.Objective(sense=pe.minimize, expr=obj)
        models[node_id].objective.deactivate()

    return models


def _update_tso_proximal_centres_after_solve(planning_problem, model, results, cycle=None):

    params = planning_problem.params.admm
    proximal_cfg = params.proximal_regularization
    block_movements = dict()

    use_tso_proximal = (proximal_cfg['enabled'] and proximal_cfg['tso']['enabled'])

    if not use_tso_proximal:
        return {}

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks

    successful_blocks = 0
    failed_blocks = 0

    worst_v = None
    worst_pf = None
    worst_ess = None

    # ------------------------------------------------------------------------------------------------------------------
    # Update each year-day independently.
    # A failed TSO block keeps its previous proximal centre.
    for year in transmission_network.years:
        for day in transmission_network.days:

            block_key = (year, day)

            block_movements[block_key] = {
                'successful': False,
                'v': None,
                'pf': None,
                'ess': None,
            }

            if not _solver_result_succeeded(results[year][day]):
                failed_blocks += 1
                continue

            block_movements[block_key]['successful'] = True
            successful_blocks += 1

            local_model = model[year][day]
            network = transmission_network.network[year][day]
            s_base = network.baseMVA

            # ----------------------------------------------------------------------------------------------------------
            # Interface voltage and interface active/reactive power
            for dn in local_model.active_distribution_networks:

                node_id = transmission_network.active_distribution_network_nodes[dn]
                distribution_network = distribution_networks[node_id]
                interface_transf_rating = (distribution_network.network[year][day].get_interface_branch_rating() / s_base)
                v_base = network.get_node_base_kv(node_id)

                for p in local_model.periods:

                    # --------------------------------------------------------------------------------------------------
                    # Voltage
                    current_v = pe.value(local_model.expected_interface_vmag[dn, p])
                    previous_v = pe.value(local_model.prox_v_prev[dn, p])
                    movement_v = abs(current_v - previous_v)
                    movement_entry = {
                        'normalized_movement': movement_v,
                        'physical_movement': movement_v * v_base,
                        'node_id': node_id,
                        'year': year,
                        'day': day,
                        'period': p,
                    }

                    block_v = block_movements[block_key]['v']
                    if (block_v is None or movement_v > block_v['normalized_movement']):
                        block_movements[block_key]['v'] = movement_entry
                    if (worst_v is None or movement_v > worst_v['normalized_movement']):
                        worst_v = movement_entry

                    # --------------------------------------------------------------------------------------------------
                    # Interface P
                    current_pf_p = pe.value(local_model.expected_interface_pf_p[dn, p])
                    previous_pf_p = pe.value(local_model.prox_pf_p_prev[dn, p])
                    movement_pf_p = (abs(current_pf_p - previous_pf_p) / interface_transf_rating)
                    movement_entry = {
                        'normalized_movement': movement_pf_p,
                        'physical_movement': abs(current_pf_p - previous_pf_p) * s_base,
                        'power_type': 'P',
                        'node_id': node_id,
                        'year': year,
                        'day': day,
                        'period': p,
                    }

                    block_pf = block_movements[block_key]['pf']
                    if (block_pf is None or movement_pf_p > block_pf['normalized_movement']):
                        block_movements[block_key]['pf'] = movement_entry
                    if (worst_pf is None or movement_pf_p > worst_pf['normalized_movement']):
                        worst_pf = movement_entry

                    # --------------------------------------------------------------------------------------------------
                    # Interface Q
                    current_pf_q = pe.value(local_model.expected_interface_pf_q[dn, p])
                    previous_pf_q = pe.value(local_model.prox_pf_q_prev[dn, p])
                    movement_pf_q = (abs(current_pf_q - previous_pf_q) / interface_transf_rating)
                    movement_entry = {
                        'normalized_movement': movement_pf_q,
                        'physical_movement': abs(current_pf_q - previous_pf_q) * s_base,
                        'power_type': 'Q',
                        'node_id': node_id,
                        'year': year,
                        'day': day,
                        'period': p,
                    }

                    block_pf = block_movements[block_key]['pf']
                    if (block_pf is None or movement_pf_q > block_pf['normalized_movement']):
                        block_movements[block_key]['pf'] = movement_entry
                    if (worst_pf is None or movement_pf_q > worst_pf['normalized_movement']):
                        worst_pf = movement_entry

                    # --------------------------------------------------------------------------------------------------
                    # The current successful solution becomes the centre for the next TSO solve.
                    local_model.prox_v_prev[dn, p].set_value(current_v)
                    local_model.prox_pf_p_prev[dn, p].set_value(current_pf_p)
                    local_model.prox_pf_q_prev[dn, p].set_value(current_pf_q)

            # ----------------------------------------------------------------------------------------------------------
            # Shared ESS P/Q
            ess_node_by_idx = dict()

            for dn in local_model.active_distribution_networks:
                node_id = transmission_network.active_distribution_network_nodes[dn]
                shared_ess_idx = network.get_shared_energy_storage_idx(node_id)
                ess_node_by_idx[shared_ess_idx] = node_id

            for e in local_model.shared_energy_storages:

                shared_ess_rating = _shared_ess_admm_normalization_pu(network.shared_energy_storages[e].s, s_base, params.shared_ess_normalization_floor_mva)
                normalization = 2 * shared_ess_rating
                node_id = ess_node_by_idx.get(e)

                for p in local_model.periods:

                    # --------------------------------------------------------------------------------------------------
                    # Shared ESS P
                    current_ess_p = pe.value(local_model.expected_shared_ess_p[e, p])
                    previous_ess_p = pe.value(local_model.prox_ess_p_prev[e, p])
                    movement_ess_p = (abs(current_ess_p - previous_ess_p) / normalization)
                    movement_entry = {
                        'normalized_movement': movement_ess_p,
                        'physical_movement': abs(current_ess_p - previous_ess_p) * s_base,
                        'power_type': 'P',
                        'node_id': node_id,
                        'year': year,
                        'day': day,
                        'period': p,
                    }

                    block_ess = block_movements[block_key]['ess']
                    if (block_ess is None or movement_ess_p > block_ess['normalized_movement']):
                        block_movements[block_key]['ess'] = movement_entry
                    if (worst_ess is None or movement_ess_p > worst_ess['normalized_movement']):
                        worst_ess = movement_entry

                    # --------------------------------------------------------------------------------------------------
                    # Shared ESS Q
                    current_ess_q = pe.value(local_model.expected_shared_ess_q[e, p])
                    previous_ess_q = pe.value(local_model.prox_ess_q_prev[e, p])
                    movement_ess_q = (abs(current_ess_q - previous_ess_q) / normalization)
                    movement_entry = {
                        'normalized_movement': movement_ess_q,
                        'physical_movement': abs(current_ess_q - previous_ess_q) * s_base,
                        'power_type': 'Q',
                        'node_id': node_id,
                        'year': year,
                        'day': day,
                        'period': p,
                    }

                    block_ess = block_movements[block_key]['ess']
                    if (block_ess is None or movement_ess_q > block_ess['normalized_movement']):
                        block_movements[block_key]['ess'] = movement_entry
                    if (worst_ess is None or movement_ess_q > worst_ess['normalized_movement']):
                        worst_ess = movement_entry

                    # --------------------------------------------------------------------------------------------------
                    # Successful solution becomes next proximal centre.
                    local_model.prox_ess_p_prev[e, p].set_value(current_ess_p)
                    local_model.prox_ess_q_prev[e, p].set_value(current_ess_q)

    # ------------------------------------------------------------------------------------------------------------------
    # Diagnostics
    cycle_text = cycle if cycle is not None else 'N/A'
    v_max = (worst_v['normalized_movement'] if worst_v is not None else 0.0)
    pf_max = (worst_pf['normalized_movement'] if worst_pf is not None else 0.0)
    ess_max = (worst_ess['normalized_movement'] if worst_ess is not None else 0.0)

    print(
        f'[TSO PROX] cycle={cycle_text} | '
        f'updated_blocks={successful_blocks} | '
        f'held_failed_blocks={failed_blocks} | '
        f'V max={v_max:.6f} | '
        f'PF max={pf_max:.6f} | '
        f'ESS max={ess_max:.6f}'
    )

    if worst_v is not None:
        print(
            '[TSO PROX][V MAX] '
            f'node={worst_v["node_id"]}, '
            f'year={worst_v["year"]}, '
            f'day={worst_v["day"]}, '
            f'period={worst_v["period"]} | '
            f'dV={worst_v["physical_movement"]:.6f} kV | '
            f'normalized={worst_v["normalized_movement"]:.6f}'
        )

    if worst_pf is not None:
        unit = 'MW' if worst_pf['power_type'] == 'P' else 'MVAr'
        print(
            '[TSO PROX][PF MAX] '
            f'node={worst_pf["node_id"]}, '
            f'year={worst_pf["year"]}, '
            f'day={worst_pf["day"]}, '
            f'period={worst_pf["period"]}, '
            f'type={worst_pf["power_type"]} | '
            f'delta={worst_pf["physical_movement"]:.6f} {unit} | '
            f'normalized={worst_pf["normalized_movement"]:.6f}'
        )

    if worst_ess is not None:
        unit = 'MW' if worst_ess['power_type'] == 'P' else 'MVAr'
        print(
            '[TSO PROX][ESS MAX] '
            f'node={worst_ess["node_id"]}, '
            f'year={worst_ess["year"]}, '
            f'day={worst_ess["day"]}, '
            f'period={worst_ess["period"]}, '
            f'type={worst_ess["power_type"]} | '
            f'delta={worst_ess["physical_movement"]:.6f} {unit} | '
            f'normalized={worst_ess["normalized_movement"]:.6f}'
        )

    return block_movements


def _get_local_slack_penalty_components(model, network, params):

    base = network.baseMVA

    components = {
        'voltage': 0.0,
        'node_balance_p': 0.0,
        'node_balance_q': 0.0,
        'branch_flow_ij': 0.0,
        'branch_flow_ji': 0.0,
        'flex_day_balance_p': 0.0,
        'flex_day_balance_q': 0.0,
    }

    for s_m in model.scenarios_market:
        for s_o in model.scenarios_operation:

            probability = (network.prob_market_scenarios[s_m] * network.prob_operation_scenarios[s_o])

            # ----------------------------------------------------------------------------------------------------------
            # Voltage and node-balance slacks
            for i in model.nodes:
                for p in model.periods:
                    if params.slacks.grid_operation.voltage:
                        components['voltage'] += (probability * PENALTY_VOLTAGE_SQUARED * (pe.value(model.slack_v_sqr_down[i, s_m, s_o, p]) + pe.value(model.slack_v_sqr_up[i, s_m, s_o, p])))
                    if params.slacks.node_balance.active_power:
                        components['node_balance_p'] += (probability * base * PENALTY_NODE_BALANCE * (pe.value(model.slack_node_balance_p_up[i, s_m, s_o, p])+ pe.value(model.slack_node_balance_p_down[i, s_m, s_o, p])))
                    if params.slacks.node_balance.reactive_power:
                        components['node_balance_q'] += (probability * base * PENALTY_NODE_BALANCE * (pe.value(model.slack_node_balance_q_up[i, s_m, s_o, p]) + pe.value(model.slack_node_balance_q_down[i, s_m, s_o, p])))

            # ----------------------------------------------------------------------------------------------------------
            # Branch-flow slacks: i -> j
            if params.slacks.grid_operation.branch_flow:
                for b in model.branches:
                    for p in model.periods:
                        components['branch_flow_ij'] += (probability * base * PENALTY_CURRENT * pe.value(model.slack_flow_ij_sqr[b, s_m, s_o, p]))

                # ------------------------------------------------------------------------------------------------------
                # Branch-flow slacks: j -> i
                for b in model.apparent_power_limited_branches:
                    for p in model.periods:
                        components['branch_flow_ji'] += (probability * base * PENALTY_CURRENT * pe.value(model.slack_flow_ji_sqr[b, s_m, s_o, p]))

            # ----------------------------------------------------------------------------------------------------------
            # Flexibility day-balance slacks
            if params.fl_reg and params.slacks.flexibility.day_balance:
                for c in model.loads:

                    if not network.loads[c].fl_reg:
                        continue

                    components['flex_day_balance_p'] += (probability * base * PENALTY_FLEXIBILITY * (pe.value(model.slack_flex_p_balance_up[c, s_m, s_o]) + pe.value(model.slack_flex_p_balance_down[c, s_m, s_o])))
                    components['flex_day_balance_q'] += (probability * base * PENALTY_FLEXIBILITY * (pe.value(model.slack_flex_q_balance_up[c, s_m, s_o]) + pe.value(model.slack_flex_q_balance_down[c, s_m, s_o])))

    components['classified_total'] = sum(
        components[name]
        for name in (
            'voltage',
            'node_balance_p',
            'node_balance_q',
            'branch_flow_ij',
            'branch_flow_ji',
            'flex_day_balance_p',
            'flex_day_balance_q',
        )
    )

    exact_total = float(pe.value(model.total_slack_penalties))

    components['total_slack_penalties'] = exact_total
    components['unclassified'] = (exact_total - components['classified_total'])

    return components



def _get_local_objective_components(model, params):

    if params.obj_type != OBJ_MIN_COST:
        raise ValueError('Objective-component diagnostic currently implemented for OBJ_MIN_COST only.')

    components = {
        'generation_cost': float(pe.value(model.total_gen_cost)),
        'flexibility_cost': float(pe.value(model.total_flex_cost)),
        'load_curtailment_cost': float(pe.value(model.total_load_curt_cost)),
        'res_curtailment_penalty': float(pe.value(model.total_gen_curt_penalty)),
        'ess_usage_penalty': float(pe.value(model.total_ess_utilization_cost_penalty)),
        'slack_penalties': float(pe.value(model.total_slack_penalties)),
        'ess_complementarity_penalties': float(pe.value(model.total_ess_complementarity_penalties))
    }

    components['economic_market_cost'] = (
        components['generation_cost']
        + components['flexibility_cost']
    )

    components['classified_total'] = (
        components['generation_cost']
        + components['flexibility_cost']
        + components['load_curtailment_cost']
        + components['res_curtailment_penalty']
        + components['ess_usage_penalty']
        + components['slack_penalties']
        + components['ess_complementarity_penalties']
    )

    # Useful consistency check because the original objective may also
    # contain scenario-deviation regularization.
    objective_value = float(pe.value(model.objective.expr))

    scenario_deviation = (
        float(pe.value(model.scenario_deviation_penalty))
        if hasattr(model, 'scenario_deviation_penalty')
        else 0.0
    )

    components['scenario_deviation_penalty'] = scenario_deviation
    components['objective_value'] = objective_value

    components['unclassified'] = (
        objective_value
        - components['classified_total']
        - scenario_deviation
    )

    return components


def _get_operational_objective_component_blocks(planning_problem, models):

    blocks = {}

    transmission_network = planning_problem.transmission_network

    for year in transmission_network.years:
        for day in transmission_network.days:
            local_model = models['tso'][year][day]
            components = _get_local_objective_components(local_model, transmission_network.params)
            weight = _get_admm_block_weight(transmission_network, year, day)
            blocks[('TSO', None, year, day)] = {name: weight * value for name, value in components.items()}

    for node_id, distribution_network in planning_problem.distribution_networks.items():
        for year in distribution_network.years:
            for day in distribution_network.days:
                local_model = models['dso'][node_id][year][day]
                components = _get_local_objective_components(local_model, distribution_network.params)
                weight = _get_admm_block_weight(distribution_network, year, day)
                blocks[('DSO', node_id, year, day)] = {name: weight * value for name, value in components.items()}

    return blocks


def _get_operational_slack_component_blocks(planning_problem, models):

    blocks = {}

    # --------------------------------------------------------------------------------------------------------------
    # TSO
    transmission_network = planning_problem.transmission_network

    for year in transmission_network.years:
        for day in transmission_network.days:
            network = transmission_network.network[year][day]
            local_model = models['tso'][year][day]
            components = _get_local_slack_penalty_components(local_model, network, transmission_network.params)
            weight = _get_admm_block_weight(transmission_network, year, day)
            blocks[('TSO', None, year, day)] = {
                name: weight * value
                for name, value in components.items()
            }

    # --------------------------------------------------------------------------------------------------------------
    # DSOs
    for node_id, distribution_network in (planning_problem.distribution_networks.items()):
        for year in distribution_network.years:
            for day in distribution_network.days:
                network = distribution_network.network[year][day]
                local_model = models['dso'][node_id][year][day]
                components = _get_local_slack_penalty_components(local_model, network, distribution_network.params)
                weight = _get_admm_block_weight(distribution_network, year, day)
                blocks[('DSO', node_id, year, day)] = {
                    name: weight * value
                    for name, value in components.items()
                }

    return blocks


def _get_tso_voltage_slack_state(planning_problem, tso_models):

    transmission_network = planning_problem.transmission_network
    blocks = {}

    if not transmission_network.params.slacks.grid_operation.voltage:
        return blocks

    for year in transmission_network.years:
        for day in transmission_network.days:

            model = tso_models[year][day]
            network = transmission_network.network[year][day]
            weight = _get_admm_block_weight(transmission_network, year, day)

            block = {}

            for s_m in model.scenarios_market:
                for s_o in model.scenarios_operation:

                    probability = (network.prob_market_scenarios[s_m] * network.prob_operation_scenarios[s_o])

                    for i in model.nodes:

                        node = network.nodes[i]
                        node_id = node.bus_i
                        v_base_kv = network.get_node_base_kv(node_id)

                        for p in model.periods:

                            slack_down_var = model.slack_v_sqr_down[i, s_m, s_o, p]
                            slack_up_var = model.slack_v_sqr_up[i, s_m, s_o, p]
                            slack_down = float(pe.value(slack_down_var))
                            slack_up = float(pe.value(slack_up_var))
                            e = float(pe.value(model.e[i, s_m, s_o, p]))
                            f = float(pe.value(model.f[i, s_m, s_o, p]))
                            vmag_sqr = e ** 2 + f ** 2
                            vmag = sqrt(max(vmag_sqr, 0.0))
                            diagnostics = voltage_slack_diagnostics(node.v_min, node.v_max, vmag_sqr, slack_down, slack_up)

                            common = {
                                'node_idx': i,
                                'node_id': node_id,
                                'market_scenario': s_m,
                                'operation_scenario': s_o,
                                'period': p,
                                'vmag': vmag,
                                'vmag_kv': vmag * v_base_kv,
                                'v_min': node.v_min,
                                'v_max': node.v_max,
                                'v_base_kv': v_base_kv,
                                'probability': probability,
                                'weight': weight,
                            }

                            # ------------------------------------------------------------------
                            # Lower-voltage relaxation
                            ub_down = slack_down_var.ub
                            ub_down = (float(ub_down) if ub_down is not None else None)
                            block[(i, s_m, s_o, p, 'down')] = {
                                **common,
                                'direction': 'down',
                                'slack_sqr': slack_down,
                                'slack_ub': ub_down,
                                'ub_fraction': (slack_down / ub_down if (ub_down is not None and ub_down > SMALL_TOLERANCE) else 0.0),
                                'physical_relaxation': diagnostics['physical_down'],
                                'realized_violation': diagnostics['violation_down'],
                                'realized_violation_kv': diagnostics['violation_down'] * v_base_kv,
                                'weighted_penalty': weight * probability * PENALTY_VOLTAGE_SQUARED * slack_down,
                            }

                            # ------------------------------------------------------------------
                            # Upper-voltage relaxation
                            ub_up = slack_up_var.ub
                            ub_up = (float(ub_up) if ub_up is not None else None)
                            block[(i, s_m, s_o, p, 'up')] = {
                                **common,
                                'direction': 'up',
                                'slack_sqr': slack_up,
                                'slack_ub': ub_up,
                                'ub_fraction': (slack_up / ub_up if (ub_up is not None and ub_up > SMALL_TOLERANCE) else 0.0),
                                'physical_relaxation': diagnostics['physical_up'],
                                'realized_violation': diagnostics['violation_up'],
                                'realized_violation_kv': diagnostics['violation_up'] * v_base_kv,
                                'weighted_penalty': weight * probability * PENALTY_VOLTAGE_SQUARED * slack_up,
                            }

            blocks[(year, day)] = block

    return blocks


def update_transmission_coordination_model_and_solve(transmission_network, model, vmag_req, dual_vmag, pf_req, dual_pf, ess_req, dual_ess, params, sess_estimated_capacities, from_warm_start=False):

    print('[INFO] \t\t - Updating transmission network...')

    for year in transmission_network.years:
        for day in transmission_network.days:

            s_base = transmission_network.network[year][day].baseMVA

            for dn in model[year][day].active_distribution_networks:

                node_id = transmission_network.active_distribution_network_nodes[dn]
                v_base = transmission_network.network[year][day].get_node_base_kv(node_id)
                shared_ess_idx = transmission_network.network[year][day].get_shared_energy_storage_idx(node_id)
                sess_estimated_capacity = sess_estimated_capacities[node_id]

                # Update estimated rated power and energy capacity
                model[year][day].shared_es_s_rated_fixed[shared_ess_idx].set_value(sess_estimated_capacity[year]['s_available'] / s_base)
                model[year][day].shared_es_e_rated_fixed[shared_ess_idx].set_value(sess_estimated_capacity[year]['e_available'] / s_base)
                configure_shared_ess_operational_state(
                    model[year][day],
                    shared_ess_idx,
                    pe.value(model[year][day].shared_es_s_rated_fixed[shared_ess_idx]),
                    pe.value(model[year][day].shared_es_e_rated_fixed[shared_ess_idx]),
                )

                # Update VOLTAGE and POWER FLOW variables at connection point
                for p in model[year][day].periods:
                    model[year][day].dual_vmag_req[dn, p].set_value(dual_vmag['current'][node_id][year][day][p] / v_base)
                    model[year][day].vmag_req[dn, p].set_value(vmag_req['dso']['current'][node_id][year][day][p] / v_base)
                    model[year][day].dual_pf_p_req[dn, p].set_value(dual_pf['current'][node_id][year][day]['p'][p] / s_base)
                    model[year][day].dual_pf_q_req[dn, p].set_value(dual_pf['current'][node_id][year][day]['q'][p] / s_base)
                    model[year][day].p_pf_req[dn, p].set_value(pf_req['dso']['current'][node_id][year][day]['p'][p] / s_base)
                    model[year][day].q_pf_req[dn, p].set_value(pf_req['dso']['current'][node_id][year][day]['q'][p] / s_base)

                # Update shared ESS capacity and power requests
                shared_ess_idx = transmission_network.network[year][day].get_shared_energy_storage_idx(node_id)
                for p in model[year][day].periods:
                    model[year][day].dual_ess_p_req[shared_ess_idx, p].set_value(dual_ess['current'][node_id][year][day]['p'][p])
                    model[year][day].dual_ess_q_req[shared_ess_idx, p].set_value(dual_ess['current'][node_id][year][day]['q'][p])
                    model[year][day].p_ess_req[shared_ess_idx, p].set_value(ess_req['z']['current'][node_id][year][day]['p'][p] / s_base)
                    model[year][day].q_ess_req[shared_ess_idx, p].set_value(ess_req['z']['current'][node_id][year][day]['q'][p] / s_base)
                    if params.previous_iter['ess']['tso']:
                        model[year][day].dual_ess_p_prev[shared_ess_idx, p].set_value(dual_ess['prev'][node_id][year][day]['p'][p] / s_base)
                        model[year][day].dual_ess_q_prev[shared_ess_idx, p].set_value(dual_ess['prev'][node_id][year][day]['q'][p] / s_base)
                        model[year][day].p_ess_prev[shared_ess_idx, p].set_value(ess_req['tso']['prev'][node_id][year][day]['p'][p] / s_base)
                        model[year][day].q_ess_prev[shared_ess_idx, p].set_value(ess_req['tso']['prev'][node_id][year][day]['q'][p] / s_base)

    # Solve!
    res = transmission_network.optimize(model, from_warm_start=from_warm_start)
    for year in transmission_network.years:
        for day in transmission_network.days:
            if not _solver_result_succeeded(res[year][day]):
                print(
                    f'[ERROR] Transmission network {model[year][day].name}, '
                    f'year={year}, day={day} did not converge: '
                    f'{solver_result_summary(res[year][day])}'
                )
                # exit(ERROR_NETWORK_OPTIMIZATION)
    return res


def update_distribution_coordination_models_and_solve(distribution_networks, models, vmag_req, dual_vmag, pf_req, dual_pf, ess_req, dual_ess, params, sess_estimated_capacities, from_warm_start=False, parallel_execution=False):
    if parallel_execution:
        return update_distribution_coordination_models_and_solve_parallel(distribution_networks, models, vmag_req, dual_vmag, pf_req, dual_pf, ess_req, dual_ess, params, sess_estimated_capacities, from_warm_start=from_warm_start)
    else:
        return update_distribution_coordination_models_and_solve_sequential(distribution_networks, models, vmag_req, dual_vmag, pf_req, dual_pf, ess_req, dual_ess, params, sess_estimated_capacities, from_warm_start=from_warm_start)


def update_distribution_coordination_models_and_solve_sequential(distribution_networks, models, vmag_req, dual_vmag, pf_req, dual_pf, ess_req, dual_ess, params, sess_estimated_capacities, from_warm_start=False):

    print('[INFO] \t\t - Updating distribution networks:')
    res = dict()

    for node_id in distribution_networks:

        model = models[node_id]
        distribution_network = distribution_networks[node_id]
        sess_estimated_capacity = sess_estimated_capacities[node_id]

        for year in distribution_network.years:
            for day in distribution_network.days:

                ref_node_id = distribution_network.network[year][day].get_reference_node_id()
                v_base = distribution_network.network[year][day].get_node_base_kv(ref_node_id)
                s_base = distribution_network.network[year][day].baseMVA
                shared_ess_idx = distribution_network.network[year][day].get_shared_energy_storage_idx(ref_node_id)

                # Update estimated rated power and energy capacity
                model[year][day].shared_es_s_rated_fixed[shared_ess_idx].set_value(sess_estimated_capacity[year]['s_available'] / s_base)
                model[year][day].shared_es_e_rated_fixed[shared_ess_idx].set_value(sess_estimated_capacity[year]['e_available'] / s_base)
                configure_shared_ess_operational_state(model[year][day], shared_ess_idx, pe.value(model[year][day].shared_es_s_rated_fixed[shared_ess_idx]), pe.value(model[year][day].shared_es_e_rated_fixed[shared_ess_idx]))

                # Update VOLTAGE and POWER FLOW variables at connection point
                for p in model[year][day].periods:
                    model[year][day].dual_vmag_req[p].set_value(dual_vmag['current'][node_id][year][day][p] / v_base)
                    model[year][day].vmag_req[p].set_value(vmag_req['tso']['current'][node_id][year][day][p] / v_base)
                    model[year][day].dual_pf_p_req[p].set_value(dual_pf['current'][node_id][year][day]['p'][p] / s_base)
                    model[year][day].dual_pf_q_req[p].set_value(dual_pf['current'][node_id][year][day]['q'][p] / s_base)
                    model[year][day].p_pf_req[p].set_value(pf_req['tso']['current'][node_id][year][day]['p'][p] / s_base)
                    model[year][day].q_pf_req[p].set_value(pf_req['tso']['current'][node_id][year][day]['q'][p] / s_base)

                # Update SHARED ENERGY STORAGE variables (if existent)
                for p in model[year][day].periods:
                    model[year][day].dual_ess_p_req[p].set_value(dual_ess['current'][node_id][year][day]['p'][p])
                    model[year][day].dual_ess_q_req[p].set_value(dual_ess['current'][node_id][year][day]['q'][p])
                    model[year][day].p_ess_req[p].set_value(ess_req['z']['current'][node_id][year][day]['p'][p] / s_base)
                    model[year][day].q_ess_req[p].set_value(ess_req['z']['current'][node_id][year][day]['q'][p] / s_base)
                    if params.previous_iter['ess']['dso']:
                        model[year][day].dual_ess_p_prev[p].set_value(dual_ess['prev'][node_id][year][day]['p'][p] / s_base)
                        model[year][day].dual_ess_q_prev[p].set_value(dual_ess['prev'][node_id][year][day]['q'][p] / s_base)
                        model[year][day].p_ess_prev[p].set_value(ess_req['dso']['prev'][node_id][year][day]['p'][p] / s_base)
                        model[year][day].q_ess_prev[p].set_value(ess_req['dso']['prev'][node_id][year][day]['q'][p] / s_base)

        # Solve!
        res[node_id] = distribution_network.optimize(model, from_warm_start=from_warm_start)
        for year in distribution_network.years:
            for day in distribution_network.days:
                if not _solver_result_succeeded(res[node_id][year][day]):
                    print(
                        f'[WARNING] Distribution network node={node_id}, '
                        f'network={model[year][day].name}, year={year}, day={day} '
                        f'did not converge: {solver_result_summary(res[node_id][year][day])}'
                    )
                    #exit(ERROR_NETWORK_OPTIMIZATION)
    return res


def update_distribution_coordination_models_and_solve_parallel(distribution_networks, models, vmag_req, dual_vmag, pf_req, dual_pf, ess_req, dual_ess, params, sess_estimated_capacities, from_warm_start=False):

    print('[INFO] \t\t - Updating distribution networks in parallel:')
    res = dict()

    tasks = []
    max_workers = os.cpu_count() // 2
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        for node_id in distribution_networks:
            sess_estimated_capacity = sess_estimated_capacities[node_id]
            tasks.append(executor.submit(update_and_solve_dso, node_id, distribution_networks[node_id], models[node_id],
                                         vmag_req, dual_vmag, pf_req, dual_pf, ess_req, dual_ess,
                                         params, sess_estimated_capacity,
                                         from_warm_start=from_warm_start))

        for future in as_completed(tasks):
            node_id, result, updated_model = future.result()
            res[node_id] = result
            models[node_id] = updated_model

    return res


def update_and_solve_dso(node_id, distribution_network, model, vmag_req, dual_vmag, pf_req, dual_pf, ess_req, dual_ess, params, sess_estimated_capacity, from_warm_start=False):

    for year in distribution_network.years:
        for day in distribution_network.days:

            ref_node_id = distribution_network.network[year][day].get_reference_node_id()
            v_base = distribution_network.network[year][day].get_node_base_kv(ref_node_id)
            s_base = distribution_network.network[year][day].baseMVA
            shared_ess_idx = distribution_network.network[year][day].get_shared_energy_storage_idx(ref_node_id)

            # Update estimated rated power and energy capacity
            model[year][day].shared_es_s_rated_fixed[shared_ess_idx].set_value(sess_estimated_capacity[year]['s_available'] / s_base)
            model[year][day].shared_es_e_rated_fixed[shared_ess_idx].set_value(sess_estimated_capacity[year]['e_available'] / s_base)
            configure_shared_ess_operational_state(model[year][day], shared_ess_idx, pe.value(model[year][day].shared_es_s_rated_fixed[shared_ess_idx]), pe.value(model[year][day].shared_es_e_rated_fixed[shared_ess_idx]))

            # Update VOLTAGE and POWER FLOW variables at connection point
            for p in model[year][day].periods:
                fix_or_set(model[year][day].dual_vmag_req[p], dual_vmag['current'][node_id][year][day][p] / v_base)
                fix_or_set(model[year][day].vmag_req[p], vmag_req['tso']['current'][node_id][year][day][p] / v_base)
                fix_or_set(model[year][day].dual_pf_p_req[p], dual_pf['current'][node_id][year][day]['p'][p] / s_base)
                fix_or_set(model[year][day].dual_pf_q_req[p], dual_pf['current'][node_id][year][day]['q'][p] / s_base)
                fix_or_set(model[year][day].p_pf_req[p], pf_req['tso']['current'][node_id][year][day]['p'][p] / s_base)
                fix_or_set(model[year][day].q_pf_req[p], pf_req['tso']['current'][node_id][year][day]['q'][p] / s_base)

            # Update SHARED ENERGY STORAGE variables (if existent)
            for p in model[year][day].periods:
                fix_or_set(model[year][day].dual_ess_p_req[p], dual_ess['current'][node_id][year][day]['p'][p])
                fix_or_set(model[year][day].dual_ess_q_req[p], dual_ess['current'][node_id][year][day]['q'][p])
                fix_or_set(model[year][day].p_ess_req[p], ess_req['z']['current'][node_id][year][day]['p'][p] / s_base)
                fix_or_set(model[year][day].q_ess_req[p], ess_req['z']['current'][node_id][year][day]['q'][p] / s_base)
                if params.previous_iter['ess']['dso']:
                    fix_or_set(model[year][day].dual_ess_p_prev[p], dual_ess['prev'][node_id][year][day]['p'][p] / s_base)
                    fix_or_set(model[year][day].dual_ess_q_prev[p], dual_ess['prev'][node_id][year][day]['q'][p] / s_base)
                    fix_or_set(model[year][day].p_ess_prev[p], ess_req['dso']['prev'][node_id][year][day]['p'][p] / s_base)
                    fix_or_set(model[year][day].q_ess_prev[p], ess_req['dso']['prev'][node_id][year][day]['q'][p] / s_base)

    # Solve
    res = distribution_network.optimize(model, from_warm_start=from_warm_start)
    for year in distribution_network.years:
        for day in distribution_network.days:
            if not _solver_result_succeeded(res[year][day]):
                print(
                    f'[WARNING] Distribution network node={node_id}, '
                    f'network={model[year][day].name}, year={year}, day={day} '
                    f'did not converge: {solver_result_summary(res[year][day])}'
                )

    return (node_id, res, model)


def update_shared_energy_storages_coordination_model_and_solve(planning_problem, models, ess_req, dual_ess, params, from_warm_start=False):

    print('[INFO] \t\t - Updating SharedESS...')
    shared_ess_data = planning_problem.shared_ess_data
    days = [day for day in planning_problem.days]
    years = [year for year in planning_problem.years]

    for node_id in planning_problem.active_distribution_network_nodes:

        for y in models[node_id].years:
            year = years[y]
            for d in models[node_id].days:
                day = days[d]
                for p in models[node_id].periods:

                    p_req = ess_req['current'][node_id][year][day]['p'][p]
                    q_req = ess_req['current'][node_id][year][day]['q'][p]
                    dual_p_req = dual_ess['current'][node_id][year][day]['p'][p]
                    dual_q_req = dual_ess['current'][node_id][year][day]['q'][p]

                    models[node_id].p_req[y, d, p].set_value(p_req)
                    models[node_id].q_req[y, d, p].set_value(q_req)
                    models[node_id].dual_p_req[y, d, p].set_value(dual_p_req)
                    models[node_id].dual_q_req[y, d, p].set_value(dual_q_req)

    # Solve!
    res = shared_ess_data.optimize(models, from_warm_start=from_warm_start)
    for node_id in planning_problem.active_distribution_network_nodes:
        if not _solver_result_succeeded(res[node_id]):
            print(
                f'[WARNING] SharedESS operational planning node={node_id} did not converge: '
                f'{solver_result_summary(res[node_id])}'
            )

    return res


def _get_expected_network_shared_ess_charge_discharge_mva(model, network, shared_ess_idx, p):
    """
    Return probability-weighted expected shared-ESS charging and discharging apparent power in physical MVA.
    Network-model shared_es_sch/shared_es_sdch variables are in p.u., so the expected values are converted using network.baseMVA.
    """
    sch = 0.0
    sdch = 0.0
    for s_m in model.scenarios_market:
        for s_o in model.scenarios_operation:
            probability = (network.prob_market_scenarios[s_m] * network.prob_operation_scenarios[s_o])
            sch += probability * pe.value(model.shared_es_sch[shared_ess_idx, s_m, s_o, p])
            sdch += probability * pe.value(model.shared_es_sdch[shared_ess_idx, s_m, s_o, p])
    return sch * network.baseMVA, sdch * network.baseMVA


def _get_esso_shared_ess_charge_discharge_mva(model, y, d, p):
    """
    Return aggregate ESSO charging and discharging apparent power in physical MVA.
    ESSO variables are defined separately for each investment cohort, so aggregate them over all cohorts.
    """
    sch = sum(pe.value(model.es_sch_per_unit[y_inv, y, d, p]) for y_inv in model.years)
    sdch = sum(pe.value(model.es_sdch_per_unit[y_inv, y, d, p]) for y_inv in model.years)
    return sch, sdch


def get_admm_residual_metrics(planning_problem, tso_model, dso_models, esso_model, consensus_vars):

    repr_years = list(planning_problem.years)
    repr_days = list(planning_problem.days)

    year_idx = {year: idx for idx, year in enumerate(repr_years)}
    day_idx = {day: idx for idx, day in enumerate(repr_days)}

    sums = {
        'primal': {'v': 0.0, 'pf': 0.0, 'ess': 0.0},
        'dual': {'v': 0.0, 'pf': 0.0, 'ess': 0.0},
    }

    counts = {
        'primal': {'v': 0, 'pf': 0, 'ess': 0},
        'dual': {'v': 0, 'pf': 0, 'ess': 0},
    }

    primal_max = {
        'v': 0.0,
        'pf': 0.0,
        'ess': 0.0,
    }

    dual_max = {
        'v': 0.0,
        'pf': 0.0,
        'ess': 0.0,
    }

    worst_v_primal = None
    worst_pf_primal = None
    worst_pf_dual = None
    worst_ess_primal = None

    for node_id in planning_problem.active_distribution_network_nodes:

        dso_model = dso_models[node_id]

        for year in planning_problem.years:
            for day in planning_problem.days:

                # ------------------------------------------------------------------
                # Network data and normalization factors
                # ------------------------------------------------------------------

                network = planning_problem.transmission_network.network[year][day]
                s_base = network.baseMVA
                shared_ess_idx = network.get_shared_energy_storage_idx(node_id)
                interface_v_base = network.get_node_base_kv(node_id)
                interface_rating = (planning_problem.distribution_networks[node_id].network[year][day].get_interface_branch_rating())
                normalization_floor = (planning_problem.params.admm.shared_ess_normalization_floor_mva)

                # TSO shared-ESS normalization, in MVA
                tso_rating = _shared_ess_admm_normalization_mva(network.shared_energy_storages[shared_ess_idx].s * s_base, normalization_floor)

                # DSO shared-ESS normalization, in MVA
                dso_network = (planning_problem.distribution_networks[node_id].network[year][day])
                dso_ref_node_id = dso_network.get_reference_node_id()
                dso_shared_ess_idx = dso_network.get_shared_energy_storage_idx(dso_ref_node_id)
                dso_rating = _shared_ess_admm_normalization_mva(dso_network.shared_energy_storages[dso_shared_ess_idx].s * dso_network.baseMVA, normalization_floor)

                # ESSO shared-ESS normalization, in MVA
                esso_shared_ess_idx = planning_problem.shared_ess_data.get_shared_energy_storage_idx(node_id)
                esso_rating = _shared_ess_admm_normalization_mva(planning_problem.shared_ess_data.shared_energy_storages[year][esso_shared_ess_idx].s, normalization_floor)
                ess_ratings = {
                    'tso': tso_rating,
                    'dso': dso_rating,
                    'esso': esso_rating,
                }

                # ------------------------------------------------------------------
                # ADMM penalties
                # ------------------------------------------------------------------
                rho_tso_v = pe.value(tso_model[year][day].rho_v)
                rho_tso_pf = pe.value(tso_model[year][day].rho_pf)
                rho_tso_ess = pe.value(tso_model[year][day].rho_ess)
                rho_dso_v = pe.value(dso_model[year][day].rho_v)
                rho_dso_pf = pe.value(dso_model[year][day].rho_pf)
                rho_dso_ess = pe.value(dso_model[year][day].rho_ess)
                rho_esso_ess = pe.value(esso_model[node_id].rho)
                ess_rhos = {
                    'tso': rho_tso_ess,
                    'dso': rho_dso_ess,
                    'esso': rho_esso_ess,
                }

                # ------------------------------------------------------------------
                # Residuals
                # ------------------------------------------------------------------
                for p in range(planning_problem.num_instants):

                    # ==============================================================
                    # Interface voltage
                    # ==============================================================
                    tso_v = consensus_vars['vmag']['tso']['current'][node_id][year][day][p]
                    dso_v = consensus_vars['vmag']['dso']['current'][node_id][year][day][p]

                    # Primal voltage residual
                    absolute_difference = abs(tso_v - dso_v)
                    normalized_primal_residual = absolute_difference / interface_v_base
                    sums['primal']['v'] += normalized_primal_residual
                    counts['primal']['v'] += 1
                    if worst_v_primal is None or normalized_primal_residual > primal_max['v']:
                        primal_max['v'] = normalized_primal_residual
                        worst_v_primal = {
                            'node_id': node_id,
                            'year': year,
                            'day': day,
                            'period': p,
                            'tso_value': tso_v,
                            'dso_value': dso_v,
                            'absolute_difference': absolute_difference,
                            'interface_v_base': interface_v_base,
                            'rho_tso': rho_tso_v,
                            'rho_dso': rho_dso_v,
                            'normalized_residual': normalized_primal_residual,
                        }


                    # Dual voltage residual
                    for agent, rho in (('tso', rho_tso_v), ('dso', rho_dso_v)):
                        current = consensus_vars['vmag'][agent]['current'][node_id][year][day][p]
                        previous = consensus_vars['vmag'][agent]['prev'][node_id][year][day][p]
                        normalized_dual_residual = rho * abs(current - previous) / interface_v_base
                        sums['dual']['v'] += normalized_dual_residual
                        counts['dual']['v'] += 1
                        dual_max['v'] = max(dual_max['v'], normalized_dual_residual)

                    # ==============================================================
                    # Active and reactive power
                    # ==============================================================
                    for power_type in ('p', 'q'):

                        # ----------------------------------------------------------
                        # TSO-DSO interface power flow
                        # ----------------------------------------------------------
                        tso_pf = consensus_vars['pf']['tso']['current'][node_id][year][day][power_type][p]
                        dso_pf = consensus_vars['pf']['dso']['current'][node_id][year][day][power_type][p]

                        # Primal PF residual
                        normalized_primal_residual = abs(tso_pf - dso_pf) / interface_rating
                        sums['primal']['pf'] += normalized_primal_residual
                        counts['primal']['pf'] += 1
                        if normalized_primal_residual > primal_max['pf']:
                            primal_max['pf'] = normalized_primal_residual
                            worst_pf_primal = {
                                'node_id': node_id,
                                'year': year,
                                'day': day,
                                'period': p,
                                'power_type': power_type,
                                'tso_value': tso_pf,
                                'dso_value': dso_pf,
                                'absolute_difference': abs(tso_pf - dso_pf),
                                'interface_rating': interface_rating,
                                'rho_tso': rho_tso_pf,
                                'rho_dso': rho_dso_pf,
                                'normalized_residual': normalized_primal_residual,
                            }

                        # Dual PF residual
                        for agent, rho in (('tso', rho_tso_pf), ('dso', rho_dso_pf)):
                            current = consensus_vars['pf'][agent]['current'][node_id][year][day][power_type][p]
                            previous = consensus_vars['pf'][agent]['prev'][node_id][year][day][power_type][p]
                            normalized_dual_residual = rho * abs(current - previous) / interface_rating
                            sums['dual']['pf'] += normalized_dual_residual
                            counts['dual']['pf'] += 1
                            if normalized_dual_residual > dual_max['pf']:
                                dual_max['pf'] = normalized_dual_residual
                                worst_pf_dual = {
                                    'agent': agent,
                                    'node_id': node_id,
                                    'year': year,
                                    'day': day,
                                    'period': p,
                                    'power_type': power_type,
                                    'current_value': current,
                                    'previous_value': previous,
                                    'absolute_change': abs(current - previous),
                                    'rho': rho,
                                    'interface_rating': interface_rating,
                                    'normalized_residual': normalized_dual_residual,
                                }

                        # ----------------------------------------------------------
                        # Shared ESS consensus
                        # ----------------------------------------------------------
                        z_current = consensus_vars['ess']['z']['current'][node_id][year][day][power_type][p]
                        z_previous = consensus_vars['ess']['z']['prev'][node_id][year][day][power_type][p]
                        z_change = abs(z_current - z_previous)

                        for agent in ('tso', 'dso', 'esso'):

                            x_current = consensus_vars['ess'][agent]['current'][node_id][year][day][power_type][p]

                            # Primal consensus residual:
                            #
                            #       |x_i - z|
                            # r_i = ----------
                            #         2 S_i
                            #
                            absolute_difference = abs(x_current - z_current)
                            normalized_primal_residual = absolute_difference / (2.0 * ess_ratings[agent])
                            sums['primal']['ess'] += normalized_primal_residual
                            counts['primal']['ess'] += 1
                            if (worst_ess_primal is None or normalized_primal_residual > primal_max['ess']):

                                primal_max['ess'] = normalized_primal_residual

                                # --------------------------------------------------------------
                                # Charging/discharging diagnostics at the same point
                                # --------------------------------------------------------------
                                y = year_idx[year]
                                d = day_idx[day]
                                tso_sch, tso_sdch = _get_expected_network_shared_ess_charge_discharge_mva(tso_model[year][day], network, shared_ess_idx, p)
                                dso_sch, dso_sdch = _get_expected_network_shared_ess_charge_discharge_mva(dso_model[year][day], dso_network, dso_shared_ess_idx, p)
                                esso_sch, esso_sdch = _get_esso_shared_ess_charge_discharge_mva(esso_model[node_id], y, d, p)
                                worst_ess_primal = {
                                    'node_id': node_id,
                                    'year': year,
                                    'day': day,
                                    'period': p,
                                    'power_type': power_type,
                                    'agent': agent,
                                    'agent_value': x_current,
                                    'z_value': z_current,
                                    'absolute_difference': absolute_difference,
                                    'normalization_rating': ess_ratings[agent],
                                    'rho': ess_rhos[agent],
                                    'rho_tso': rho_tso_ess,
                                    'rho_dso': rho_dso_ess,
                                    'rho_esso': rho_esso_ess,
                                    'normalized_residual': normalized_primal_residual,

                                    # ----------------------------------------------------------
                                    # Complementarity diagnostics
                                    # ----------------------------------------------------------
                                    'charge_discharge': {

                                        'tso': {
                                            'sch': tso_sch,
                                            'sdch': tso_sdch,
                                            'product': tso_sch * tso_sdch,
                                            'simultaneous': min(tso_sch, tso_sdch),
                                            'net': tso_sch - tso_sdch,
                                            'base_mva': network.baseMVA,
                                        },

                                        'dso': {
                                            'sch': dso_sch,
                                            'sdch': dso_sdch,
                                            'product': dso_sch * dso_sdch,
                                            'simultaneous': min(dso_sch, dso_sdch),
                                            'net': dso_sch - dso_sdch,
                                            'base_mva': dso_network.baseMVA,
                                        },

                                        'esso': {
                                            'sch': esso_sch,
                                            'sdch': esso_sdch,
                                            'product': esso_sch * esso_sdch,
                                            'simultaneous': min(esso_sch, esso_sdch),
                                            'net': esso_sch - esso_sdch,
                                            'base_mva': None,
                                        },
                                    },
                                }

                            # Dual consensus residual:
                            #
                            #             |z^k - z^(k-1)|
                            # s_i = rho_i ----------------
                            #                  2 S_i
                            #
                            normalized_dual_residual = ess_rhos[agent] * z_change / (2.0 * ess_ratings[agent])
                            sums['dual']['ess'] += normalized_dual_residual
                            counts['dual']['ess'] += 1
                            dual_max['ess'] = max(dual_max['ess'], normalized_dual_residual)

    # Mean residuals over all corresponding coordinates.
    residual_metrics = {
        residual_type: {
            group: (sums[residual_type][group] / max(counts[residual_type][group], 1))
            for group in ('v', 'pf', 'ess')
        }
        for residual_type in ('primal', 'dual')
    }

    # Preserve mean residuals for:
    #   1. aggregate convergence criteria on the primal side;
    #   2. adaptive ADMM penalty balancing.
    for group in ('v', 'pf', 'ess'):
        residual_metrics['primal'][f'{group}_mean'] = residual_metrics['primal'][group]
        residual_metrics['dual'][f'{group}_mean'] = residual_metrics['dual'][group]

    # The base group keys represent strict worst-case residuals.
    # Primal maxima are used for pointwise convergence checks.
    # Dual maxima are retained for diagnostics only.
    for group in ('v', 'pf', 'ess'):
        residual_metrics['primal'][group] = primal_max[group]
        residual_metrics['dual'][group] = dual_max[group]

    residual_metrics['worst_v_primal'] = worst_v_primal
    residual_metrics['worst_pf_primal'] = worst_pf_primal
    residual_metrics['worst_pf_dual'] = worst_pf_dual
    residual_metrics['worst_ess_primal'] = worst_ess_primal

    return residual_metrics


def check_admm_convergence(planning_problem, consensus_vars, residual_metrics, params, debug_flag=False):
    consensus_convergence = check_consensus_convergence(residual_metrics, params)
    stationary_convergence = check_stationary_convergence(residual_metrics, params)
    if not consensus_convergence and debug_flag:
        print_debug_info(
            planning_problem,
            consensus_vars,
            print_vmag=True,
            print_pf=True,
            print_ess=True,
        )
    return consensus_convergence and stationary_convergence


def _print_worst_primal_residual_diagnostics(residual_metrics, params):

    # ------------------------------------------------------------------
    # Voltage
    # ------------------------------------------------------------------
    worst_v = residual_metrics.get('worst_v_primal')
    if (worst_v is not None and residual_metrics['primal']['v'] > params.tol['consensus']['v']):
        print(
            '[DIAG][V MAX] '
            f'node={worst_v["node_id"]}, '
            f'year={worst_v["year"]}, '
            f'day={worst_v["day"]}, '
            f'period={worst_v["period"]} | '
            f'TSO={worst_v["tso_value"]:.6f} kV, '
            f'DSO={worst_v["dso_value"]:.6f} kV, '
            f'diff={worst_v["absolute_difference"]:.6f} kV, '
            f'base={worst_v["interface_v_base"]:.6f} kV, '
            f'normalized={worst_v["normalized_residual"]:.6f}, '
            f'rho(TSO/DSO)='
            f'{worst_v["rho_tso"]:.6f}/'
            f'{worst_v["rho_dso"]:.6f}'
        )

    # ------------------------------------------------------------------
    # Interface power flow
    # ------------------------------------------------------------------
    worst_pf = residual_metrics.get('worst_pf_primal')
    if (worst_pf is not None and residual_metrics['primal']['pf'] > params.tol['consensus']['pf']):
        unit = ('MW' if worst_pf['power_type'] == 'p' else 'MVAr')
        print(
            '[DIAG][PF MAX] '
            f'node={worst_pf["node_id"]}, '
            f'year={worst_pf["year"]}, '
            f'day={worst_pf["day"]}, '
            f'period={worst_pf["period"]}, '
            f'type={worst_pf["power_type"].upper()} | '
            f'TSO={worst_pf["tso_value"]:.6f} {unit}, '
            f'DSO={worst_pf["dso_value"]:.6f} {unit}, '
            f'diff={worst_pf["absolute_difference"]:.6f} {unit}, '
            f'rating={worst_pf["interface_rating"]:.6f} MVA, '
            f'normalized={worst_pf["normalized_residual"]:.6f}, '
            f'rho(TSO/DSO)='
            f'{worst_pf["rho_tso"]:.6f}/'
            f'{worst_pf["rho_dso"]:.6f}'
        )

    # ------------------------------------------------------------------
    # Shared ESS
    # ------------------------------------------------------------------
    worst_ess = residual_metrics.get('worst_ess_primal')
    if (worst_ess is not None and residual_metrics['primal']['ess'] > params.tol['consensus']['ess']):

        unit = ('MW' if worst_ess['power_type'] == 'p' else 'MVAr')
        print(
            '[DIAG][ESS MAX] '
            f'agent={worst_ess["agent"].upper()}, '
            f'node={worst_ess["node_id"]}, '
            f'year={worst_ess["year"]}, '
            f'day={worst_ess["day"]}, '
            f'period={worst_ess["period"]}, '
            f'type={worst_ess["power_type"].upper()} | '
            f'x={worst_ess["agent_value"]:.6f} {unit}, '
            f'z={worst_ess["z_value"]:.6f} {unit}, '
            f'diff={worst_ess["absolute_difference"]:.6f} {unit}, '
            f'norm_rating='
            f'{worst_ess["normalization_rating"]:.6f} MVA, '
            f'normalized='
            f'{worst_ess["normalized_residual"]:.6f}, '
            f'rho={worst_ess["rho"]:.6f}, '
            f'rho(TSO/DSO/ESSO)='
            f'{worst_ess["rho_tso"]:.6f}/'
            f'{worst_ess["rho_dso"]:.6f}/'
            f'{worst_ess["rho_esso"]:.6f}'
        )

        charge_discharge = worst_ess.get('charge_discharge')
        if charge_discharge is not None:
            for agent in ('tso', 'dso', 'esso'):
                values = charge_discharge[agent]
                base_text = ''
                if values['base_mva'] is not None:
                    base_text = f', base={values["base_mva"]:.6f} MVA'
                print(
                    f'[DIAG][ESS COMP] {agent.upper()} | '
                    f'Sch={values["sch"]:.6f} MVA, '
                    f'Sdch={values["sdch"]:.6f} MVA, '
                    f'net={values["net"]:.6f} MVA, '
                    f'product={values["product"]:.6e} MVA^2, '
                    f'simultaneous={values["simultaneous"]:.6f} MVA'
                    f'{base_text}'
                )


def _print_tso_voltage_slack_transitions(year, day, current_state, previous_state, expected_voltage_penalty_delta, top_n=5):

    current_block = current_state.get((year, day))
    previous_block = previous_state.get((year, day))

    if current_block is None or previous_block is None:
        return

    transitions = []

    all_keys = set(current_block) | set(previous_block)

    for key in all_keys:

        current = current_block.get(key)
        previous = previous_block.get(key)

        if current is None or previous is None:
            continue

        penalty_delta = (current['weighted_penalty'] - previous['weighted_penalty'])
        transitions.append({
            'key': key,
            'previous': previous,
            'current': current,
            'penalty_delta': penalty_delta,
            'abs_penalty_delta': abs(penalty_delta),
        })

    transitions.sort(key=lambda x: x['abs_penalty_delta'], reverse=True)
    individual_delta = sum(transition['penalty_delta'] for transition in transitions)

    print(
        '    [VOLTAGE SLACK CHECK] '
        f'individual_delta={individual_delta:+.6e} | '
        f'component_delta={expected_voltage_penalty_delta:+.6e} | '
        f'mismatch='
        f'{individual_delta - expected_voltage_penalty_delta:+.6e}'
    )

    for transition in transitions[:top_n]:

        previous = transition['previous']
        current = transition['current']

        print(
            '    [VOLTAGE SLACK JUMP] '
            f'node={current["node_id"]} | '
            f's_m={current["market_scenario"]} | '
            f's_o={current["operation_scenario"]} | '
            f'period={current["period"]} | '
            f'direction={current["direction"]} | '
            f'penalty_delta={transition["penalty_delta"]:+.6e}'
        )

        print(
            '        '
            f'slack_sqr: '
            f'{previous["slack_sqr"]:.6e} -> '
            f'{current["slack_sqr"]:.6e} | '
            f'ub_fraction: '
            f'{previous["ub_fraction"]:.4f} -> '
            f'{current["ub_fraction"]:.4f}'
        )

        print(
            '        '
            f'V: '
            f'{previous["vmag"]:.6f} -> '
            f'{current["vmag"]:.6f} p.u. | '
            f'limits=[{current["v_min"]:.6f}, '
            f'{current["v_max"]:.6f}]'
        )

        print(
            '        '
            f'realized_violation: '
            f'{previous["realized_violation"]:.6e} -> '
            f'{current["realized_violation"]:.6e} p.u. | '
            f'violation_kV: '
            f'{previous["realized_violation_kv"]:.6f} -> '
            f'{current["realized_violation_kv"]:.6f} kV | '
            f'physical_relaxation: '
            f'{previous["physical_relaxation"]:.6e} -> '
            f'{current["physical_relaxation"]:.6e} p.u.'
        )


def _admm_local_solves_succeeded(planning_problem, results):
    for year in planning_problem.years:
        for day in planning_problem.days:
            if not _solver_result_succeeded(results['tso'][year][day]):
                return False
            for node_id in planning_problem.active_distribution_network_nodes:
                if not _solver_result_succeeded(results['dso'][node_id][year][day]):
                    return False
    for node_id in planning_problem.active_distribution_network_nodes:
        if not _solver_result_succeeded(results['esso'][node_id]):
            return False
    return True


def _solver_result_succeeded(result):
    return solver_result_succeeded(result)


def check_consensus_convergence(residual_metrics, params):

    convergence = True

    labels = {
        'v': 'interface Vmag',
        'pf': 'interface PF',
        'ess': 'shared ESS',
    }

    for group in ('v', 'pf', 'ess'):

        # --------------------------------------------------------------
        # Worst-case pointwise residual
        # --------------------------------------------------------------
        max_residual = residual_metrics['primal'][group]
        max_tolerance = params.tol['consensus'][group]
        if not _admm_metric_within_tolerance(max_residual, max_tolerance):
            print(f'[INFO]\t\t - {labels[group]} max primal residual failed. {max_residual:.6f} > {max_tolerance:.6f}')
            convergence = False

        # --------------------------------------------------------------
        # Aggregate mean residual
        # --------------------------------------------------------------
        mean_residual = residual_metrics['primal'][f'{group}_mean']
        mean_tolerance = params.tol['consensus'][f'{group}_mean']
        if not _admm_metric_within_tolerance(mean_residual, mean_tolerance):
            print(f'[INFO]\t\t - {labels[group]} mean primal residual failed. {mean_residual:.6f} > {mean_tolerance:.6f}')
            convergence = False

    if convergence:
        print('[INFO]\t\t - Primal residuals ok!')

    return convergence


def check_stationary_convergence(residual_metrics, params):
    convergence = True
    labels = {'v': 'interface Vmag', 'pf': 'interface PF', 'ess': 'shared ESS'}
    for group in ('v', 'pf', 'ess'):
        residual = residual_metrics['dual'][f'{group}_mean']
        tolerance = params.tol['stationarity'][group]
        if not _admm_metric_within_tolerance(residual, tolerance):
            print(f'[INFO]\t\t - {labels[group]} mean dual residual failed. {residual:.6f} > {tolerance:.6f}')
            convergence = False
    if convergence:
        print('[INFO]\t\t - Dual residuals ok!')
    return convergence


def _admm_metric_within_tolerance(value, tolerance):
    return value <= tolerance


def _max_difference_with_index(values_a, values_b):
    max_diff = -1.0
    max_idx = None
    value_a = None
    value_b = None

    for idx, (a, b) in enumerate(zip(values_a, values_b)):
        diff = abs(a - b)

        if diff > max_diff:
            max_diff = diff
            max_idx = idx
            value_a = a
            value_b = b

    return max_diff, max_idx, value_a, value_b


def _print_shared_ess_consensus_diagnostics(planning_problem, consensus_vars):

    for node_id in planning_problem.active_distribution_network_nodes:
        for year in planning_problem.years:
            for day in planning_problem.days:

                print(f'[DEBUG][ESS CONSENSUS] node={node_id}, year={year}, day={day}')

                for power_type in ('p', 'q'):

                    tso_values = (consensus_vars['ess']['tso']['current'][node_id][year][day][power_type])
                    dso_values = (consensus_vars['ess']['dso']['current'][node_id][year][day][power_type])
                    esso_values = (consensus_vars['ess']['esso']['current'][node_id][year][day][power_type])
                    z_values = (consensus_vars['ess']['z']['current'][node_id][year][day][power_type])

                    tso_dso = _max_difference_with_index(tso_values, dso_values)
                    tso_esso = _max_difference_with_index(tso_values, esso_values)
                    dso_esso = _max_difference_with_index(dso_values, esso_values)
                    tso_z = _max_difference_with_index(tso_values, z_values)
                    dso_z = _max_difference_with_index(dso_values, z_values)
                    esso_z = _max_difference_with_index(esso_values, z_values)

                    unit = 'MW' if power_type == 'p' else 'MVAr'
                    print(
                        f'\t{power_type.upper()} | '
                        f'TSO-DSO={tso_dso[0]:.6f}, '
                        f'TSO-ESSO={tso_esso[0]:.6f}, '
                        f'DSO-ESSO={dso_esso[0]:.6f} {unit} | '
                        f'TSO-z={tso_z[0]:.6f}, '
                        f'DSO-z={dso_z[0]:.6f}, '
                        f'ESSO-z={esso_z[0]:.6f} {unit}'
                    )


def _get_admm_penalty_summary(tso_model, dso_models, esso_model):
    penalties = {'v': [], 'pf': [], 'ess': []}
    for year_models in tso_model.values():
        for model in year_models.values():
            penalties['v'].append(pe.value(model.rho_v))
            penalties['pf'].append(pe.value(model.rho_pf))
            penalties['ess'].append(pe.value(model.rho_ess))
    for node_models in dso_models.values():
        for year_models in node_models.values():
            for model in year_models.values():
                penalties['v'].append(pe.value(model.rho_v))
                penalties['pf'].append(pe.value(model.rho_pf))
                penalties['ess'].append(pe.value(model.rho_ess))
    for model in esso_model.values():
        penalties['ess'].append(pe.value(model.rho))
    return {
        group: sum(values) / max(len(values), 1)
        for group, values in penalties.items()
    }


def _update_admm_penalties(tso_model, dso_models, esso_model, residual_metrics, params, allow_update=True):

    before = _get_admm_penalty_summary(tso_model, dso_models, esso_model)

    actions = dict()
    factors = dict()
    update_params = params.penalty_update

    for group in ('v', 'pf', 'ess'):

        # --------------------------------------------------------------
        # Adaptive penalty balancing uses normalized residual severity.
        # For the primal residual, consider both the worst-case and mean
        # convergence criteria; retain the mean dual residual.
        # --------------------------------------------------------------
        primal_max = residual_metrics['primal'][group]
        primal_mean = residual_metrics['primal'][f'{group}_mean']
        dual_mean = residual_metrics['dual'][f'{group}_mean']

        primal_max_tol = params.tol['consensus'][group]
        primal_mean_tol = params.tol['consensus'][f'{group}_mean']
        dual_tol = params.tol['stationarity'][group]

        # Normalized residual severities
        primal_max_ratio = primal_max / primal_max_tol
        primal_mean_ratio = primal_mean / primal_mean_tol

        primal_ratio = max(primal_max_ratio, primal_mean_ratio)
        dual_ratio = dual_mean / dual_tol

        adaptation_converged = (primal_ratio <= 1.0 and dual_ratio <= 1.0)

        # --------------------------------------------------------------
        # Residual-balance thresholds
        # --------------------------------------------------------------
        increase_balance_ratio = update_params['residual_balance_ratio']
        decrease_balance_ratio = (update_params.get('residual_balance_ratio_pf_decrease', update_params['residual_balance_ratio']) if group == 'pf' else update_params['residual_balance_ratio'])

        factor = 1.0
        action = 'held'

        # --------------------------------------------------------------
        # Determine penalty update
        # --------------------------------------------------------------
        if not params.adaptive_penalty:
            action = 'fixed'
        elif not allow_update:
            action = 'held after solver failure'
        elif not adaptation_converged:
            if (primal_ratio > increase_balance_ratio * dual_ratio):
                factor = update_params['increase_factor']
                action = 'increased'
            elif (dual_ratio > decrease_balance_ratio * primal_ratio):
                factor = 1.0 / update_params['decrease_factor']
                action = 'decreased'

        print(
            f'[ADMM RHO] {group.upper()} | '
            f'primal max ratio={primal_max_ratio:.3f} | '
            f'primal mean ratio={primal_mean_ratio:.3f} | '
            f'primal selected={primal_ratio:.3f} | '
            f'dual mean ratio={dual_ratio:.3f} | '
            f'increase threshold={increase_balance_ratio:.1f} | '
            f'decrease threshold={decrease_balance_ratio:.1f} | '
            f'action={action}'
        )

        actions[group] = action
        factors[group] = factor

    # ------------------------------------------------------------------
    # Apply common group-wise scaling factors
    # ------------------------------------------------------------------
    if params.adaptive_penalty and allow_update:

        # TSO
        for year_models in tso_model.values():
            for model in year_models.values():
                _scale_admm_penalty(model.rho_v, factors['v'], update_params)
                _scale_admm_penalty(model.rho_pf, factors['pf'], update_params)
                _scale_admm_penalty(model.rho_ess, factors['ess'], update_params)
                if hasattr(model, 'rho_ess_prev'):
                    _scale_admm_penalty(model.rho_ess_prev, factors['ess'], update_params)

        # DSOs
        for node_models in dso_models.values():
            for year_models in node_models.values():
                for model in year_models.values():
                    _scale_admm_penalty(model.rho_v, factors['v'], update_params)
                    _scale_admm_penalty(model.rho_pf, factors['pf'], update_params)
                    _scale_admm_penalty(model.rho_ess, factors['ess'], update_params)
                    if hasattr(model, 'rho_ess_prev'):
                        _scale_admm_penalty(model.rho_ess_prev, factors['ess'], update_params)

        # ESSO
        for model in esso_model.values():
            _scale_admm_penalty(model.rho, factors['ess'], update_params)

    after = _get_admm_penalty_summary(tso_model, dso_models, esso_model)

    return actions, before, after


def _scale_admm_penalty(penalty, factor, params):
    value = pe.value(penalty) * factor
    penalty.set_value(min(max(value, params['min']), params['max']))


def _update_interface_power_flow_variables(planning_problem, tso_model, dso_models, interface_vars, dual_vars, results, params, update_tn=True, update_dns=True):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks

    # Transmission network - Update Vmag and PF at the TN-DN interface
    if update_tn:
        for dn in range(len(planning_problem.active_distribution_network_nodes)):
            node_id = planning_problem.active_distribution_network_nodes[dn]
            for year in planning_problem.years:
                for day in planning_problem.days:
                    v_base = transmission_network.network[year][day].get_node_base_kv(node_id)
                    s_base = transmission_network.network[year][day].baseMVA
                    if _solver_result_succeeded(results['tso'][year][day]):
                        for p in tso_model[year][day].periods:
                            interface_vars['vmag']['tso']['prev'][node_id][year][day][p] = copy(interface_vars['vmag']['tso']['current'][node_id][year][day][p])
                            interface_vars['pf']['tso']['prev'][node_id][year][day]['p'][p] = copy(interface_vars['pf']['tso']['current'][node_id][year][day]['p'][p])
                            interface_vars['pf']['tso']['prev'][node_id][year][day]['q'][p] = copy(interface_vars['pf']['tso']['current'][node_id][year][day]['q'][p])

                            vmag_req = pe.value(tso_model[year][day].expected_interface_vmag[dn, p]) * v_base
                            p_req = pe.value(tso_model[year][day].expected_interface_pf_p[dn, p]) * s_base
                            q_req = pe.value(tso_model[year][day].expected_interface_pf_q[dn, p]) * s_base
                            interface_vars['vmag']['tso']['current'][node_id][year][day][p] = vmag_req
                            interface_vars['pf']['tso']['current'][node_id][year][day]['p'][p] = p_req
                            interface_vars['pf']['tso']['current'][node_id][year][day]['q'][p] = q_req

    # Distribution Network - Update PF at the TN-DN interface
    if update_dns:
        for node_id in distribution_networks:
            distribution_network = distribution_networks[node_id]
            dso_model = dso_models[node_id]
            for year in planning_problem.years:
                for day in planning_problem.days:
                    ref_node_id = distribution_network.network[year][day].get_reference_node_id()
                    v_base = distribution_network.network[year][day].get_node_base_kv(ref_node_id)
                    s_base = distribution_network.network[year][day].baseMVA
                    if _solver_result_succeeded(results['dso'][node_id][year][day]):
                        for p in dso_model[year][day].periods:
                            interface_vars['vmag']['dso']['prev'][node_id][year][day][p] = copy(interface_vars['vmag']['dso']['current'][node_id][year][day][p])
                            interface_vars['pf']['dso']['prev'][node_id][year][day]['p'][p] = copy(interface_vars['pf']['dso']['current'][node_id][year][day]['p'][p])
                            interface_vars['pf']['dso']['prev'][node_id][year][day]['q'][p] = copy(interface_vars['pf']['dso']['current'][node_id][year][day]['q'][p])

                            vmag_req = pe.value(dso_model[year][day].expected_interface_vmag[p]) * v_base
                            p_req = pe.value(dso_model[year][day].expected_interface_pf_p[p]) * s_base
                            q_req = pe.value(dso_model[year][day].expected_interface_pf_q[p]) * s_base
                            interface_vars['vmag']['dso']['current'][node_id][year][day][p] = vmag_req
                            interface_vars['pf']['dso']['current'][node_id][year][day]['p'][p] = p_req
                            interface_vars['pf']['dso']['current'][node_id][year][day]['q'][p] = q_req

    # Update Lambdas
    for node_id in distribution_networks:

        distribution_network = distribution_networks[node_id]

        for year in planning_problem.years:
            for day in planning_problem.days:

                tso_s_base = transmission_network.network[year][day].baseMVA
                dso_s_base = distribution_network.network[year][day].baseMVA
                interface_rating = (distribution_network.network[year][day].get_interface_branch_rating())

                tso_succeeded = (_solver_result_succeeded(results['tso'][year][day]) if update_tn else False)
                dso_succeeded = (_solver_result_succeeded(results['dso'][node_id][year][day]) if update_tn or update_dns else False)

                for p in range(planning_problem.num_instants):
                    if update_tn and tso_succeeded and dso_succeeded:
                        rho_v_tso = pe.value(tso_model[year][day].rho_v)
                        rho_pf_tso = pe.value(tso_model[year][day].rho_pf)
                        error_v_req_tso = interface_vars['vmag']['tso']['current'][node_id][year][day][p] - interface_vars['vmag']['dso']['current'][node_id][year][day][p]
                        error_p_pf_req_tso = interface_vars['pf']['tso']['current'][node_id][year][day]['p'][p] - interface_vars['pf']['dso']['current'][node_id][year][day]['p'][p]
                        error_q_pf_req_tso = interface_vars['pf']['tso']['current'][node_id][year][day]['q'][p] - interface_vars['pf']['dso']['current'][node_id][year][day]['q'][p]
                        dual_vars['vmag']['tso']['current'][node_id][year][day][p] += rho_v_tso * error_v_req_tso
                        dual_vars['pf']['tso']['current'][node_id][year][day]['p'][p] += rho_pf_tso * error_p_pf_req_tso / interface_rating * tso_s_base
                        dual_vars['pf']['tso']['current'][node_id][year][day]['q'][p] += rho_pf_tso * error_q_pf_req_tso / interface_rating * tso_s_base

                    if update_tn and tso_succeeded and dso_succeeded:
                        rho_v_dso = pe.value(dso_models[node_id][year][day].rho_v)
                        rho_pf_dso = pe.value(dso_models[node_id][year][day].rho_pf)
                        error_v_req_dso = interface_vars['vmag']['dso']['current'][node_id][year][day][p] - interface_vars['vmag']['tso']['current'][node_id][year][day][p]
                        error_p_pf_req_dso = interface_vars['pf']['dso']['current'][node_id][year][day]['p'][p] - interface_vars['pf']['tso']['current'][node_id][year][day]['p'][p]
                        error_q_pf_req_dso = interface_vars['pf']['dso']['current'][node_id][year][day]['q'][p] - interface_vars['pf']['tso']['current'][node_id][year][day]['q'][p]
                        dual_vars['vmag']['dso']['current'][node_id][year][day][p] += rho_v_dso * error_v_req_dso
                        dual_vars['pf']['dso']['current'][node_id][year][day]['p'][p] += rho_pf_dso * error_p_pf_req_dso / interface_rating * dso_s_base
                        dual_vars['pf']['dso']['current'][node_id][year][day]['q'][p] += rho_pf_dso * error_q_pf_req_dso / interface_rating * dso_s_base


def _update_shared_energy_storage_variables(planning_problem, tso_model, dso_models, sess_model, shared_ess_vars, dual_vars, results, params, update_tn=True, update_dns=True, update_sess=True):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks
    shared_ess_data = planning_problem.shared_ess_data
    repr_days = [day for day in planning_problem.days]
    repr_years = [year for year in planning_problem.years]

    for node_id in planning_problem.active_distribution_network_nodes:

        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]

        # Power requested by ESSO
        if update_sess:
            for y in sess_model[node_id].years:
                year = repr_years[y]
                if _solver_result_succeeded(results['esso'][node_id]):
                    for d in sess_model[node_id].days:
                        day = repr_days[d]
                        for p in sess_model[node_id].periods:
                            shared_ess_vars['esso']['prev'][node_id][year][day]['p'][p] = copy(shared_ess_vars['esso']['current'][node_id][year][day]['p'][p])
                            shared_ess_vars['esso']['prev'][node_id][year][day]['q'][p] = copy(shared_ess_vars['esso']['current'][node_id][year][day]['q'][p])

                            p_req = pe.value(sess_model[node_id].es_pnet[y, d, p])
                            q_req = pe.value(sess_model[node_id].es_qnet[y, d, p])
                            shared_ess_vars['esso']['current'][node_id][year][day]['p'][p] = p_req
                            shared_ess_vars['esso']['current'][node_id][year][day]['q'][p] = q_req

        # Power requested by TSO
        if update_tn:
            for y in range(len(repr_years)):
                year = repr_years[y]
                for d in range(len(repr_days)):
                    day = repr_days[d]
                    if _solver_result_succeeded(results['tso'][year][day]):
                        s_base = transmission_network.network[year][day].baseMVA
                        shared_ess_idx = transmission_network.network[year][day].get_shared_energy_storage_idx(node_id)
                        for p in tso_model[year][day].periods:
                            shared_ess_vars['tso']['prev'][node_id][year][day]['p'][p] = copy(shared_ess_vars['tso']['current'][node_id][year][day]['p'][p])
                            shared_ess_vars['tso']['prev'][node_id][year][day]['q'][p] = copy(shared_ess_vars['tso']['current'][node_id][year][day]['q'][p])

                            p_req = pe.value(tso_model[year][day].expected_shared_ess_p[shared_ess_idx, p]) * s_base
                            q_req = pe.value(tso_model[year][day].expected_shared_ess_q[shared_ess_idx, p]) * s_base
                            shared_ess_vars['tso']['current'][node_id][year][day]['p'][p] = p_req
                            shared_ess_vars['tso']['current'][node_id][year][day]['q'][p] = q_req

        # Power requested by DSO
        if update_dns:
            for y in range(len(repr_years)):
                year = repr_years[y]
                for d in range(len(repr_days)):
                    day = repr_days[d]
                    if _solver_result_succeeded(results['dso'][node_id][year][day]):
                        s_base = distribution_network.network[year][day].baseMVA
                        for p in dso_model[year][day].periods:
                            shared_ess_vars['dso']['prev'][node_id][year][day]['p'][p] = copy(shared_ess_vars['dso']['current'][node_id][year][day]['p'][p])
                            shared_ess_vars['dso']['prev'][node_id][year][day]['q'][p] = copy(shared_ess_vars['dso']['current'][node_id][year][day]['q'][p])

                            p_req = pe.value(dso_model[year][day].expected_shared_ess_p[p]) * s_base
                            q_req = pe.value(dso_model[year][day].expected_shared_ess_q[p]) * s_base
                            shared_ess_vars['dso']['current'][node_id][year][day]['p'][p] = p_req
                            shared_ess_vars['dso']['current'][node_id][year][day]['q'][p] = q_req

        # --------------------------------------------------------------------------------------------------------------
        # Shared-ESS consensus ADMM
        #
        # All three local blocks are solved against the same z during the
        # complete DSO -> TSO -> ESSO cycle. Only after the ESSO solution
        # is available do we:
        #
        #   1. update z,
        #   2. update the three dual variables.
        #
        # No shared-ESS dual update is performed after the intermediate
        # DSO or TSO solves.
        # --------------------------------------------------------------------------------------------------------------
        if update_sess:

            for year in planning_problem.years:
                for day in planning_problem.days:

                    tso_succeeded = _solver_result_succeeded(
                        results['tso'][year][day]
                    )
                    dso_succeeded = _solver_result_succeeded(
                        results['dso'][node_id][year][day]
                    )
                    esso_succeeded = _solver_result_succeeded(
                        results['esso'][node_id]
                    )

                    # Do not update the consensus or multipliers from a
                    # partially failed ADMM cycle.
                    if not (
                        tso_succeeded
                        and dso_succeeded
                        and esso_succeeded
                    ):
                        continue

                    # Physical MVA normalization used by each local problem.
                    normalization_floor = params.shared_ess_normalization_floor_mva

                    # - TSO
                    tso_network = transmission_network.network[year][day]
                    tso_s_base = tso_network.baseMVA
                    tso_shared_ess_idx = tso_network.get_shared_energy_storage_idx(node_id)
                    tso_rating = _shared_ess_admm_normalization_mva(tso_network.shared_energy_storages[tso_shared_ess_idx].s * tso_s_base, normalization_floor)

                    # - DSO
                    dso_network = distribution_network.network[year][day]
                    dso_s_base = dso_network.baseMVA
                    dso_ref_node_id = dso_network.get_reference_node_id()
                    dso_shared_ess_idx = dso_network.get_shared_energy_storage_idx(dso_ref_node_id)
                    dso_rating = _shared_ess_admm_normalization_mva(dso_network.shared_energy_storages[dso_shared_ess_idx].s * dso_s_base, normalization_floor)

                    # - ESSO
                    esso_shared_ess_idx = shared_ess_data.get_shared_energy_storage_idx(node_id)
                    esso_rating = _shared_ess_admm_normalization_mva(shared_ess_data.shared_energy_storages[year][esso_shared_ess_idx].s, normalization_floor)

                    ratings = {
                        'tso': tso_rating,
                        'dso': dso_rating,
                        'esso': esso_rating
                    }

                    rhos = {
                        'tso': pe.value(tso_model[year][day].rho_ess),
                        'dso': pe.value(dso_models[node_id][year][day].rho_ess),
                        'esso': pe.value(sess_model[node_id].rho)
                    }

                    # a_i = 1 / (2 S_i)
                    normalization = {
                        agent: 1.0 / (2.0 * ratings[agent])
                        for agent in ('tso', 'dso', 'esso')
                    }

                    for p in range(planning_problem.num_instants):

                        for power_type in ('p', 'q'):

                            x = {
                                agent:
                                    shared_ess_vars[agent]['current']
                                    [node_id][year][day][power_type][p]
                                for agent in ('tso', 'dso', 'esso')
                            }

                            lambdas = {
                                agent:
                                    dual_vars[agent]['current']
                                    [node_id][year][day][power_type][p]
                                for agent in ('tso', 'dso', 'esso')
                            }

                            denominator = sum(
                                rhos[agent]
                                * normalization[agent] ** 2
                                for agent in ('tso', 'dso', 'esso')
                            )

                            if denominator <= SMALL_TOLERANCE:
                                raise ValueError('Shared-ESS consensus ADMM has a non-positive consensus denominator.')

                            numerator = sum(
                                (
                                    rhos[agent]
                                    * normalization[agent] ** 2
                                    * x[agent]
                                    + lambdas[agent]
                                    * normalization[agent]
                                )
                                for agent in ('tso', 'dso', 'esso')
                            )

                            z_new = numerator / denominator

                            # Save z^k before replacing it with z^(k+1).
                            shared_ess_vars['z']['prev'][node_id][year][day][power_type][p] = copy(shared_ess_vars['z']['current'][node_id][year][day][power_type][p])
                            shared_ess_vars['z']['current'][node_id][year][day][power_type][p] = z_new

                            # ----------------------------------------------------------
                            # lambda_i^(k+1) = lambda_i^k  + rho_i * (x_i^(k+1) - z^(k+1)) / (2 S_i)
                            # ----------------------------------------------------------
                            for agent in ('tso', 'dso', 'esso'):
                                normalized_residual = normalization[agent] * (x[agent] - z_new)
                                dual_vars[agent]['current'][node_id][year][day][power_type][p] += rhos[agent] * normalized_residual


# ======================================================================================================================
#  OPERATIONAL PLANNING WITHOUT COORDINATION functions
# ======================================================================================================================
def _run_operational_planning_without_coordination(planning_problem):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks
    results = {'tso': dict(), 'dso': dict(), 'esso': dict()}

    # SharedESS candidate solution (no shared ESS)
    candidate_solution = dict()
    for e in range(len(planning_problem.active_distribution_network_nodes)):
        node_id = planning_problem.active_distribution_network_nodes[e]
        candidate_solution[node_id] = dict()
        for year in planning_problem.years:
            candidate_solution[node_id][year] = dict()
            candidate_solution[node_id][year]['s'] = 0.00
            candidate_solution[node_id][year]['e'] = 0.00

    start = time.time()

    # Create interface PF variables
    interface_vmag, interface_pf = create_interface_power_flow_variables(planning_problem)

    # Create DSOs' Operational Planning models, run initial SMOPF
    dso_models = dict()
    for node_id in distribution_networks:

        distribution_network = distribution_networks[node_id]
        results['dso'][node_id] = dict()

        # Build model, fix candidate solution, and Run S-MPOPF model
        distribution_network.update_data_with_candidate_solution(candidate_solution)
        dso_model = distribution_network.build_model()
        distribution_network.update_model_with_candidate_solution(dso_model, candidate_solution)

        # Update model with expected interface values.
        for year in distribution_network.years:
            for day in distribution_network.days:

                # Add interface expected variables, and their definition
                dso_model[year][day].expected_interface_vmag = pe.Var(dso_model[year][day].periods, domain=pe.NonNegativeReals, initialize=1.00)
                dso_model[year][day].expected_interface_pf_p = pe.Var(dso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
                dso_model[year][day].expected_interface_pf_q = pe.Var(dso_model[year][day].periods, domain=pe.Reals, initialize=0.0)
                dso_model[year][day].expected_interface_vmag_def = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_vmag_rule, network=distribution_network.network[year][day]))
                dso_model[year][day].expected_interface_pf_p_def = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_pf_p_rule, network=distribution_network.network[year][day]))
                dso_model[year][day].expected_interface_pf_q_def = pe.Constraint(dso_model[year][day].periods, rule=partial(dn_interface_expected_pf_q_rule, network=distribution_network.network[year][day]))

                _add_dso_scenario_deviation_penalty(
                    dso_model[year][day],
                    distribution_network.network[year][day],
                )

        results['dso'][node_id] = distribution_network.optimize(dso_model)

        # Get initial interface PF values
        for year in distribution_network.years:
            for day in distribution_network.days:
                s_base = distribution_network.network[year][day].baseMVA
                for p in dso_model[year][day].periods:
                    interface_vmag[node_id][year][day][p] = pe.value(dso_model[year][day].expected_interface_vmag[p])
                    interface_pf[node_id][year][day]['p'][p] = pe.value(dso_model[year][day].expected_interface_pf_p[p]) * s_base
                    interface_pf[node_id][year][day]['q'][p] = pe.value(dso_model[year][day].expected_interface_pf_q[p]) * s_base

        dso_models[node_id] = dso_model

    # Create TSO Operational Planning model
    transmission_network.update_data_with_candidate_solution(candidate_solution)
    tso_model = transmission_network.build_model()
    transmission_network.update_model_with_candidate_solution(tso_model, candidate_solution)

    # TSO -- Add expected values
    for year in transmission_network.years:
        for day in transmission_network.days:

            tso_model[year][day].active_distribution_networks = range(len(transmission_network.active_distribution_network_nodes))

            # Free Pc, Qc at the interface nodes
            for dn in tso_model[year][day].active_distribution_networks:
                adn_node_id = transmission_network.active_distribution_network_nodes[dn]
                adn_load_idx = transmission_network.network[year][day].get_adn_load_idx(adn_node_id)
                for s_m in tso_model[year][day].scenarios_market:
                    for s_o in tso_model[year][day].scenarios_operation:
                        for p in tso_model[year][day].periods:
                            tso_model[year][day].pc[adn_load_idx, s_m, s_o, p].fixed = False
                            tso_model[year][day].pc[adn_load_idx, s_m, s_o, p].setub(None)
                            tso_model[year][day].pc[adn_load_idx, s_m, s_o, p].setlb(None)
                            tso_model[year][day].qc[adn_load_idx, s_m, s_o, p].fixed = False
                            tso_model[year][day].qc[adn_load_idx, s_m, s_o, p].setub(None)
                            tso_model[year][day].qc[adn_load_idx, s_m, s_o, p].setlb(None)
                            if transmission_network.params.fl_reg:
                                tso_model[year][day].flex_p_up[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)
                                tso_model[year][day].flex_p_down[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)
                                tso_model[year][day].flex_q_up[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)
                                tso_model[year][day].flex_q_down[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)
                            if transmission_network.params.l_curt:
                                tso_model[year][day].pc_curt_down[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)
                                tso_model[year][day].pc_curt_up[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)
                                tso_model[year][day].qc_curt_down[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)
                                tso_model[year][day].qc_curt_up[adn_load_idx, s_m, s_o, p].setub(EQUALITY_TOLERANCE)

            # Track the DSO schedules with a scenario-weighted soft penalty.
            year_day_vmag = {
                node_id: interface_vmag[node_id][year][day]
                for node_id in transmission_network.active_distribution_network_nodes
            }
            year_day_pf = {
                node_id: interface_pf[node_id][year][day]
                for node_id in transmission_network.active_distribution_network_nodes
            }
            _add_tso_scenario_tracking_penalty(
                tso_model[year][day],
                transmission_network.network[year][day],
                transmission_network.active_distribution_network_nodes,
                year_day_vmag,
                year_day_pf,
            )

    results['tso'] = transmission_network.optimize(tso_model)

    end = time.time()
    total_execution_time = end - start
    print('[INFO] \t - Execution time: {:.2f} s'.format(total_execution_time))

    models = {'tso': tso_model, 'dso': dso_models}

    return results, models, total_execution_time


def create_interface_power_flow_variables(planning_problem):
    consensus_vars, _ = create_admm_variables(planning_problem)
    return consensus_vars['vmag']['dso']['current'], consensus_vars['pf']['dso']['current']


# ======================================================================================================================
#  PLANNING PROBLEM read functions
# ======================================================================================================================
def _read_planning_problem(planning_problem):

    # Create results and diagrams folder
    os.makedirs(planning_problem.results_dir, exist_ok=True)
    os.makedirs(planning_problem.diagrams_dir, exist_ok=True)
    os.makedirs(planning_problem.logs_dir, exist_ok=True)

    # Read specification file
    filename = os.path.join(planning_problem.data_dir, planning_problem.filename)
    planning_data = convert_json_to_dict(read_json_file(filename))

    # General Parameters
    for year in planning_data['Years']:
        planning_problem.years[int(year)] = planning_data['Years'][year]
    planning_problem.days = planning_data['Days']
    planning_problem.num_instants = planning_data['NumInstants']
    random_seed = planning_data.get('RandomSeed')
    if random_seed is not None:
        try:
            random_seed = int(random_seed)
        except (TypeError, ValueError):
            print('[ERROR] RandomSeed must be an integer. Exiting...')
            exit(ERROR_SPECIFICATION_FILE)
        if random_seed < 0 or random_seed > (2 ** 32 - 1):
            print('[ERROR] RandomSeed must be between 0 and 2^32 - 1. Exiting...')
            exit(ERROR_SPECIFICATION_FILE)
    planning_problem.random_seed = random_seed
    print(f'[INFO] Scenario random seed: {random_seed if random_seed is not None else "unseeded"}')

    # MarketData
    print('[INFO] Reading MARKET DATA from file(s)...')
    planning_problem.discount_factor = planning_data['DiscountFactor']
    planning_problem.market_data_file = planning_data['MarketData']
    planning_problem.num_market_scenarios = planning_data['NumMarketScenarios']
    planning_problem.plot_market_data = planning_data['PlotMarketData']
    planning_problem.read_market_data_from_file()
    if planning_problem.plot_market_data:
        planning_problem.plot_market_price_scenarios()

    # Distribution Networks
    planning_problem.parallel_execution = planning_data['ParallelExecution']
    for distribution_network in planning_data['DistributionNetworks']:

        print('[INFO] Reading DISTRIBUTION NETWORK DATA from file(s)...')

        network_name = distribution_network['name']                             # Network filename
        operational_data_file = distribution_network['operational_data_file']   # Operational data filename
        num_oper_scenarios = distribution_network['num_operation_scenarios']    # Number of operational scenarios
        plot_oper_data = distribution_network['plot_operational_data']          # Plot operational data
        params_file = distribution_network['params_file']                       # Params filename
        connection_nodeid = distribution_network['connection_node_id']          # Connection node ID

        distribution_network = NetworkData()
        distribution_network.name = network_name
        distribution_network.is_transmission = False
        distribution_network.data_dir = planning_problem.data_dir
        distribution_network.results_dir = planning_problem.results_dir
        distribution_network.diagrams_dir = planning_problem.diagrams_dir
        distribution_network.logs_dir = planning_problem.logs_dir
        distribution_network.years = planning_problem.years
        distribution_network.days = planning_problem.days
        distribution_network.num_oper_scenarios = num_oper_scenarios
        distribution_network.random_seed = derive_random_seed(
            planning_problem.random_seed,
            'network',
            'dso',
            network_name,
            connection_nodeid,
        )
        distribution_network.plot_operational_data = plot_oper_data
        distribution_network.num_instants = planning_problem.num_instants
        distribution_network.discount_factor = planning_problem.discount_factor
        distribution_network.cost_energy_p = planning_problem.cost_energy_p
        distribution_network.cost_flex = planning_problem.cost_flex
        distribution_network.params_file = params_file
        distribution_network.read_network_parameters()
        distribution_network.operational_data_file = operational_data_file
        distribution_network.read_network_data()
        if distribution_network.plot_operational_data:
            distribution_network.plot_operational_data_scenarios()
        for year in distribution_network.years:
            for day in distribution_network.days:
                distribution_network.network[year][day].is_transmission = False
                distribution_network.network[year][day].tn_connection_nodeid = connection_nodeid
                if distribution_network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
                    distribution_network.network[year][day].prob_market_scenarios = [1.00]
                else:
                    distribution_network.network[year][day].prob_market_scenarios = planning_problem.prob_market_scenarios[year]
                    distribution_network.network[year][day].cost_energy_p = planning_problem.cost_energy_p[year][day]
                    distribution_network.network[year][day].cost_flex = planning_problem.cost_flex[year][day]
        distribution_network.tn_connection_nodeid = connection_nodeid
        planning_problem.distribution_networks[connection_nodeid] = distribution_network
    planning_problem.active_distribution_network_nodes = [node_id for node_id in planning_problem.distribution_networks]

    # Transmission Network
    print('[INFO] Reading TRANSMISSION NETWORK DATA from file(s)...')
    transmission_network = NetworkData()
    transmission_network.name = planning_data['TransmissionNetwork']['name']
    transmission_network.is_transmission = True
    transmission_network.data_dir = planning_problem.data_dir
    transmission_network.results_dir = planning_problem.results_dir
    transmission_network.diagrams_dir = planning_problem.diagrams_dir
    transmission_network.logs_dir = planning_problem.logs_dir
    transmission_network.years = planning_problem.years
    transmission_network.days = planning_problem.days
    transmission_network.num_oper_scenarios = planning_data['TransmissionNetwork']['num_operation_scenarios']
    transmission_network.random_seed = derive_random_seed(
        planning_problem.random_seed,
        'network',
        'tso',
        transmission_network.name,
    )
    transmission_network.plot_operational_data = planning_data['TransmissionNetwork']['plot_operational_data']
    transmission_network.num_instants = planning_problem.num_instants
    transmission_network.discount_factor = planning_problem.discount_factor
    transmission_network.cost_energy_p = planning_problem.cost_energy_p
    transmission_network.cost_flex = planning_problem.cost_flex
    transmission_network.params_file = planning_data['TransmissionNetwork']['params_file']
    transmission_network.read_network_parameters()
    transmission_network.operational_data_file = planning_data['TransmissionNetwork']['operational_data_file']
    transmission_network.read_network_data()
    if transmission_network.plot_operational_data:
        transmission_network.plot_operational_data_scenarios()
    for year in transmission_network.years:
        for day in transmission_network.days:
            transmission_network.network[year][day].is_transmission = True
            transmission_network.network[year][day].active_distribution_network_nodes = [node_id for node_id in planning_problem.distribution_networks]
            if transmission_network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
                transmission_network.network[year][day].prob_market_scenarios = [1.00]
            else:
                transmission_network.network[year][day].prob_market_scenarios = planning_problem.prob_market_scenarios[year]
                transmission_network.network[year][day].cost_energy_p = planning_problem.cost_energy_p[year][day]
                transmission_network.network[year][day].cost_flex = planning_problem.cost_flex[year][day]
    transmission_network.active_distribution_network_nodes = [node_id for node_id in planning_problem.distribution_networks]
    planning_problem.transmission_network = transmission_network
    planning_problem.scenario_metadata = _compute_scenario_metadata(planning_problem)
    print(
        f'[INFO] Scenario checksum: '
        f'{planning_problem.scenario_metadata["combined_scenario_checksum"]}'
    )

    # SharedESS
    print('[INFO] Reading SHARED ESS DATA from file(s)...')
    shared_ess_data = SharedEnergyStorageData()
    shared_ess_data.name = planning_problem.name
    shared_ess_data.data_dir = planning_problem.data_dir
    shared_ess_data.results_dir = planning_problem.results_dir
    shared_ess_data.years = planning_problem.years
    shared_ess_data.days = planning_problem.days
    shared_ess_data.num_instants = planning_problem.num_instants
    shared_ess_data.discount_factor = planning_problem.discount_factor
    shared_ess_data.prob_market_scenarios = planning_problem.prob_market_scenarios
    shared_ess_data.cost_energy_p = planning_problem.cost_energy_p
    shared_ess_data.params_file = planning_data['SharedEnergyStorage']['params_file']
    shared_ess_data.read_parameters_from_file()
    shared_ess_data.create_shared_energy_storages(planning_problem)
    shared_ess_data.data_file = planning_data['SharedEnergyStorage']['data_file']
    shared_ess_data.read_shared_energy_storage_data_from_file()
    shared_ess_data.active_distribution_network_nodes = [node_id for node_id in planning_problem.distribution_networks]
    planning_problem.shared_ess_data = shared_ess_data

    # Planning Parameters
    print(f'[INFO] Reading PLANNING PARAMETERS from file...')
    planning_problem.params_file = planning_data['PlanningParameters']['params_file']
    planning_problem.read_planning_parameters_from_file()

    _check_interface_nodes_base_voltage_consistency(planning_problem)

    # Add ADN nodes to Transmission Network
    _add_adn_node_to_transmission_network(planning_problem)

    # Add Shared Energy Storages to Transmission and Distribution Networks
    _add_shared_energy_storage_to_transmission_network(planning_problem)
    _add_shared_energy_storage_to_distribution_network(planning_problem)


def _update_scenario_digest(digest, label, values):
    array = np.asarray(values, dtype=np.float64).astype('<f8', copy=False)
    array = np.ascontiguousarray(array)
    digest.update(repr(label).encode('utf-8'))
    digest.update(repr(array.shape).encode('ascii'))
    digest.update(array.tobytes(order='C'))


def _compute_scenario_metadata(planning_problem):
    market_digest = hashlib.sha256()
    for year in sorted(planning_problem.years):
        for day in sorted(planning_problem.days, key=str):
            _update_scenario_digest(
                market_digest,
                ('market', 'energy', year, day),
                planning_problem.cost_energy_p[year][day],
            )
            _update_scenario_digest(
                market_digest,
                ('market', 'flexibility', year, day),
                planning_problem.cost_flex[year][day],
            )

    operational_digest = hashlib.sha256()
    network_groups = [('tso', None, planning_problem.transmission_network)]
    network_groups.extend(
        ('dso', node_id, planning_problem.distribution_networks[node_id])
        for node_id in sorted(planning_problem.distribution_networks, key=str)
    )
    for subsystem, node_id, network_data in network_groups:
        for year in sorted(network_data.years):
            for day in sorted(network_data.days, key=str):
                network = network_data.network[year][day]
                prefix = (subsystem, node_id, network.name, year, day)
                for load in sorted(network.loads, key=lambda item: str(item.load_id)):
                    load_prefix = (*prefix, 'load', load.load_id)
                    _update_scenario_digest(operational_digest, (*load_prefix, 'pd'), load.pd)
                    _update_scenario_digest(operational_digest, (*load_prefix, 'qd'), load.qd)
                    _update_scenario_digest(
                        operational_digest,
                        (*load_prefix, 'flex_p_up'),
                        load.flexibility.active_power.upward,
                    )
                    _update_scenario_digest(
                        operational_digest,
                        (*load_prefix, 'flex_p_down'),
                        load.flexibility.active_power.downward,
                    )
                for generator in sorted(network.generators, key=lambda item: str(item.gen_id)):
                    generator_prefix = (*prefix, 'generator', generator.gen_id)
                    _update_scenario_digest(
                        operational_digest, (*generator_prefix, 'pg'), generator.pg
                    )
                    _update_scenario_digest(
                        operational_digest, (*generator_prefix, 'qg'), generator.qg
                    )

    market_checksum = market_digest.hexdigest()
    operational_checksum = operational_digest.hexdigest()
    combined_digest = hashlib.sha256()
    combined_digest.update(market_checksum.encode('ascii'))
    combined_digest.update(operational_checksum.encode('ascii'))
    return {
        'random_seed': planning_problem.random_seed,
        'deterministic_scenarios': planning_problem.random_seed is not None,
        'market_scenario_checksum': market_checksum,
        'operational_scenario_checksum': operational_checksum,
        'combined_scenario_checksum': combined_digest.hexdigest(),
    }


# ======================================================================================================================
#  MARKET DATA read functions
# ======================================================================================================================
def _read_market_data_from_file(planning_problem):

    filename = os.path.join(planning_problem.data_dir, 'MarketData', planning_problem.market_data_file)

    try:
        base_profiles = _read_market_base_profiles(filename)
    except:
        print(f'[ERROR] Reading market data from file(s). Exiting...')
        exit(ERROR_SPECIFICATION_FILE)

    market_seed = derive_random_seed(planning_problem.random_seed, 'market')
    synthetic_profiles = _generate_market_price_scenarios(
        base_profiles,
        random_seed=derive_random_seed(market_seed, 'synthetic_profiles'),
    )

    # Update subsequent years
    initial_year = list(planning_problem.years)[0]
    growth_factors = base_profiles['growth_factors']
    energy_growth_factor = float(growth_factors[growth_factors['Growth factors'] == 'Energy']['Value, [%]'].iloc[0])
    flexibility_growth_factor = float(growth_factors[growth_factors['Growth factors'] == 'Flexibility']['Value, [%]'].iloc[0])

    for year in planning_problem.years:

        planning_problem.prob_market_scenarios[year] = [(1 / planning_problem.num_market_scenarios)] * planning_problem.num_market_scenarios
        planning_problem.cost_energy_p[year] = dict()
        planning_problem.cost_flex[year] = dict()

        for day in planning_problem.days:

            energy_growth_cumul = (1 + energy_growth_factor) ** (year - initial_year)
            flexibility_growth_cumul = (1 + flexibility_growth_factor) ** (year - initial_year)

            energy_selected_profiles = synthetic_profiles['energy'][day].sample(
                n=planning_problem.num_market_scenarios,
                random_state=derive_random_seed(market_seed, 'selection', 'energy', year, str(day)),
            )
            flexibility_selected_profiles = synthetic_profiles['flexibility'][day].sample(
                n=planning_problem.num_market_scenarios,
                random_state=derive_random_seed(
                    market_seed, 'selection', 'flexibility', year, str(day)
                ),
            )

            planning_problem.cost_energy_p[year][day] = np.array(energy_selected_profiles * energy_growth_cumul)      # n_scenarios x n_instants
            planning_problem.cost_flex[year][day] = np.array(flexibility_selected_profiles * flexibility_growth_cumul)


def _read_market_base_profiles(filename):

    base_cost_data = {
        'growth_factors': pd.read_excel(filename, sheet_name='Growth Factors'),
        'energy': pd.read_excel(filename, sheet_name='Energy'),
        'flexibility': pd.read_excel(filename, sheet_name='Flexibility')
    }

    return base_cost_data


def _generate_market_price_scenarios(base_profiles, n_samples=100, bandwidth=0.10,
                                     random_seed=None):

    print('[INFO] \t - Generating market scenarios...')

    energy_df = base_profiles['energy']
    flex_df = base_profiles['flexibility']

    synthetic_profiles = {
        'energy': _generate_market_price_scenarios_per_type(
            energy_df,
            n_samples=n_samples,
            bandwidth=bandwidth,
            random_seed=derive_random_seed(random_seed, 'energy'),
        ),
        'flexibility': _generate_market_price_scenarios_per_type(
            flex_df,
            n_samples=n_samples,
            bandwidth=bandwidth,
            random_seed=derive_random_seed(random_seed, 'flexibility'),
        ),
    }

    return synthetic_profiles


def _generate_market_price_scenarios_per_type(base_profiles, n_samples=100, bandwidth=0.05,
                                              random_seed=None):

    seasons = base_profiles['Season'].unique()
    synthetic_profiles = {}

    for season in seasons:

        # Filter
        price_subset = base_profiles[(base_profiles['Season'] == season)]
        if price_subset.empty:
            print(f'[ERROR] No market data provided for season {season}')
            exit(ERROR_MARKET_DATA_FILE)

        # Prepare data
        price_hours = price_subset.iloc[:, 2:].copy()

        # Normalize and fit copula
        scaler = StandardScaler()
        price_scaled = scaler.fit_transform(price_hours)

        model = GaussianMultivariate(
            distribution=CustomGaussianKDE(bandwidth=bandwidth),
            random_state=derive_random_seed(random_seed, str(season)),
        )
        model.fit(pd.DataFrame(price_scaled))

        # Sample
        samples = model.sample(n_samples)
        samples = scaler.inverse_transform(samples)

        synthetic_profiles[season] = pd.DataFrame(samples)

    return synthetic_profiles


def _plot_market_price_scenarios(planning_problem, years_to_plot, save_dir, save_format='pdf'):

    print('[INFO] \t - Plotting market scenarios...')

    hours = np.arange(planning_problem.num_instants)
    xticks = np.arange(0, planning_problem.num_instants, 4)
    xtick_labels = [f"{h:02d}:00" for h in xticks]

    for year in years_to_plot:
        for season in planning_problem.days:

            cost_energy_p = planning_problem.cost_energy_p[year][season]
            cost_flex = planning_problem.cost_flex[year][season]

            # Calculate statistics
            mean_energy_p = cost_energy_p.mean(axis=0)
            std_energy_p = cost_energy_p.std(axis=0)
            mean_flex = cost_flex.mean(axis=0)
            std_flex = cost_flex.std(axis=0)

            # Plot
            fig, ax = plt.subplots(1, 1, figsize=(11, 6), sharex=True)

            color_energy_p = cm.tab10(0)
            color_flex = cm.tab10(1)
            ax.plot(hours, mean_energy_p, label='Energy', color=color_energy_p)
            ax.fill_between(hours, mean_energy_p - std_energy_p, mean_energy_p + std_energy_p, alpha=0.2, color=color_energy_p)
            ax.plot(hours, mean_flex, label='Flexibility', color=color_flex)
            ax.fill_between(hours, mean_flex - std_flex, mean_flex + std_flex, alpha=0.2, color=color_flex)

            ax.set_xticks(xticks)
            ax.set_xticklabels(xtick_labels)
            ax.set_xlim(0, 23)
            ax.set_xticklabels(xtick_labels, fontsize=12)
            ax.set_xlabel("Hour", loc='center', fontsize=14)
            ax.set_ylabel("Market Price, [€/MW]", fontsize=16)
            ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.2f'))
            ax.grid(True)
            ax.legend(loc='lower right', fontsize=11)
            plt.tight_layout()

            filename = os.path.join(save_dir, f"{planning_problem.name}_market_prices_{year}_{season}.{save_format}")
            plt.tight_layout()
            plt.savefig(filename)
            plt.close(fig)


# ======================================================================================================================
#  RESULTS PROCESSING functions
# ======================================================================================================================
def _process_operational_planning_results(operational_planning_problem, tso_model, dso_models, esso_model, optimization_results):

    transmission_network = operational_planning_problem.transmission_network
    distribution_networks = operational_planning_problem.distribution_networks
    shared_ess_data = operational_planning_problem.shared_ess_data

    processed_results = dict()
    processed_results['tso'] = dict()
    processed_results['dso'] = dict()
    processed_results['esso'] = dict()
    processed_results['interface'] = dict()
    processed_results['scenario_dispersion'] = list()
    processed_results['summary_detail'] = dict()

    processed_results['tso'] = transmission_network.process_results(tso_model, optimization_results['tso'])
    for node_id in distribution_networks:
        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]
        processed_results['dso'][node_id] = distribution_network.process_results(dso_model, optimization_results['dso'][node_id])
    processed_results['esso'] = shared_ess_data.process_results(esso_model)
    processed_results['interface'] = _process_results_interface(operational_planning_problem, tso_model, dso_models)
    processed_results['scenario_dispersion'] = _process_scenario_dispersion_results(
        operational_planning_problem, tso_model, dso_models
    )
    processed_results['summary_detail'] = _process_results_summary_detail(operational_planning_problem, tso_model, dso_models)

    return processed_results


def _weighted_scenario_dispersion(model, network, expected_value, scenario_value, scale=1.0):
    probabilities = {
        (s_m, s_o): _scenario_probability(network, s_m, s_o)
        for s_m in model.scenarios_market
        for s_o in model.scenarios_operation
    }
    probability_sum = sum(probabilities.values())
    scenario_count = len(probabilities)
    weighted_standard_deviation = list()
    maximum_absolute_deviation = list()

    for p in model.periods:
        expected = pe.value(expected_value(p)) * scale
        weighted_square_deviation = 0.0
        max_deviation = 0.0
        for (s_m, s_o), probability in probabilities.items():
            value = pe.value(scenario_value(s_m, s_o, p)) * scale
            deviation = value - expected
            weighted_square_deviation += probability * deviation ** 2
            max_deviation = max(max_deviation, abs(deviation))

        if probability_sum > 0.0:
            weighted_square_deviation /= probability_sum
        weighted_standard_deviation.append(
            sqrt(max(weighted_square_deviation, 0.0))
        )
        maximum_absolute_deviation.append(max_deviation)

    return {
        'weighted_standard_deviation': weighted_standard_deviation,
        'maximum_absolute_deviation': maximum_absolute_deviation,
        'scenario_count': scenario_count,
        'probability_sum': probability_sum,
    }


def _add_scenario_dispersion_records(
        records, operator, node_id, year, day, quantity, dispersion):
    metrics = (
        ('Weighted Standard Deviation', dispersion['weighted_standard_deviation']),
        ('Maximum Absolute Deviation', dispersion['maximum_absolute_deviation']),
    )
    for metric, values in metrics:
        records.append({
            'operator': operator,
            'node_id': node_id,
            'year': year,
            'day': day,
            'quantity': quantity,
            'metric': metric,
            'scenario_count': dispersion['scenario_count'],
            'probability_sum': dispersion['probability_sum'],
            'values': values,
            'maximum': max(values, default=0.0),
        })


def _process_scenario_dispersion_results(planning_problem, tso_model, dso_models):
    records = list()
    transmission_network = planning_problem.transmission_network

    for year in transmission_network.years:
        for day in transmission_network.days:
            model = tso_model[year][day]
            network = transmission_network.network[year][day]
            for dn in model.active_distribution_networks:
                node_id = transmission_network.active_distribution_network_nodes[dn]
                shared_ess_idx = network.get_shared_energy_storage_idx(node_id)
                quantities = (
                    (
                        'Interface Vmag, [p.u.]',
                        lambda p, dn=dn: model.expected_interface_vmag[dn, p],
                        lambda s_m, s_o, p, dn=dn: model.vmag_adn[dn, s_m, s_o, p],
                        1.0,
                    ),
                    (
                        'Interface P, [MW]',
                        lambda p, dn=dn: model.expected_interface_pf_p[dn, p],
                        lambda s_m, s_o, p, dn=dn: model.pc_adn[dn, s_m, s_o, p],
                        network.baseMVA,
                    ),
                    (
                        'Interface Q, [MVAr]',
                        lambda p, dn=dn: model.expected_interface_pf_q[dn, p],
                        lambda s_m, s_o, p, dn=dn: model.qc_adn[dn, s_m, s_o, p],
                        network.baseMVA,
                    ),
                    (
                        'Shared ESS P, [MW]',
                        lambda p, e=shared_ess_idx: model.expected_shared_ess_p[e, p],
                        lambda s_m, s_o, p, e=shared_ess_idx: model.shared_es_pnet[e, s_m, s_o, p],
                        network.baseMVA,
                    ),
                    (
                        'Shared ESS Q, [MVAr]',
                        lambda p, e=shared_ess_idx: model.expected_shared_ess_q[e, p],
                        lambda s_m, s_o, p, e=shared_ess_idx: model.shared_es_qnet[e, s_m, s_o, p],
                        network.baseMVA,
                    ),
                )
                for quantity, expected_value, scenario_value, scale in quantities:
                    dispersion = _weighted_scenario_dispersion(
                        model, network, expected_value, scenario_value, scale=scale
                    )
                    _add_scenario_dispersion_records(
                        records, 'TSO', node_id, year, day, quantity, dispersion
                    )

    for node_id, distribution_network in planning_problem.distribution_networks.items():
        for year in distribution_network.years:
            for day in distribution_network.days:
                model = dso_models[node_id][year][day]
                network = distribution_network.network[year][day]
                ref_node_id = network.get_reference_node_id()
                shared_ess_idx = network.get_shared_energy_storage_idx(ref_node_id)
                quantities = (
                    (
                        'Interface Vmag, [p.u.]',
                        lambda p: model.expected_interface_vmag[p],
                        lambda s_m, s_o, p: model.vmag_adn[s_m, s_o, p],
                        1.0,
                    ),
                    (
                        'Interface P, [MW]',
                        lambda p: model.expected_interface_pf_p[p],
                        lambda s_m, s_o, p: model.pg_adn[s_m, s_o, p],
                        network.baseMVA,
                    ),
                    (
                        'Interface Q, [MVAr]',
                        lambda p: model.expected_interface_pf_q[p],
                        lambda s_m, s_o, p: model.qg_adn[s_m, s_o, p],
                        network.baseMVA,
                    ),
                    (
                        'Shared ESS P, [MW]',
                        lambda p: model.expected_shared_ess_p[p],
                        lambda s_m, s_o, p: model.shared_es_pnet[
                            shared_ess_idx, s_m, s_o, p
                        ],
                        network.baseMVA,
                    ),
                    (
                        'Shared ESS Q, [MVAr]',
                        lambda p: model.expected_shared_ess_q[p],
                        lambda s_m, s_o, p: model.shared_es_qnet[
                            shared_ess_idx, s_m, s_o, p
                        ],
                        network.baseMVA,
                    ),
                )
                for quantity, expected_value, scenario_value, scale in quantities:
                    dispersion = _weighted_scenario_dispersion(
                        model, network, expected_value, scenario_value, scale=scale
                    )
                    _add_scenario_dispersion_records(
                        records, 'DSO', node_id, year, day, quantity, dispersion
                    )

    return records


def _process_operational_planning_results_hierarchical(planning_problem, tso_model, dso_models, optimization_results):
    return _process_operational_planning_results_no_coordination(planning_problem, tso_model, dso_models, optimization_results)


def _process_operational_planning_results_no_coordination(planning_problem, tso_model, dso_models, optimization_results, execution_time=float()):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks

    processed_results = dict()
    processed_results['tso'] = dict()
    processed_results['dso'] = dict()
    processed_results['summary_detail'] = dict()

    processed_results['tso'] = transmission_network.process_results(tso_model, optimization_results['tso'])
    for node_id in distribution_networks:
        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]
        processed_results['dso'][node_id] = distribution_network.process_results(dso_model, optimization_results['dso'][node_id])
    processed_results['summary_detail'] = _process_results_summary_detail(planning_problem, tso_model, dso_models)

    return processed_results


def _process_results_interface(planning_problem, tso_model, dso_models):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks

    processed_results = dict()
    processed_results['tso'] = dict()
    processed_results['dso'] = dict()

    processed_results['tso'] = transmission_network.process_results_interface(tso_model)
    for node_id in distribution_networks:
        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]
        processed_results['dso'][node_id] = distribution_network.process_results_interface(dso_model)

    return processed_results


def _process_results_summary_detail(planning_problem, tso_model, dso_models):

    transmission_network = planning_problem.transmission_network
    distribution_networks = planning_problem.distribution_networks

    processed_results = dict()
    processed_results['tso'] = dict()
    processed_results['dso'] = dict()

    processed_results['tso'] = transmission_network.process_results_summary_detail(tso_model)
    for node_id in distribution_networks:
        dso_model = dso_models[node_id]
        distribution_network = distribution_networks[node_id]
        processed_results['dso'][node_id] = distribution_network.process_results_summary_detail(dso_model)

    return processed_results


# ======================================================================================================================
#  RESULTS PLANNING - write functions
# ======================================================================================================================
def _write_planning_results_to_excel(planning_problem, results, bound_evolution=dict(), shared_ess_cost=dict(), shared_ess_capacity=dict(), salvage_value_results=dict(), filename='planing_results', execution_time=float()):

    wb = Workbook()

    _write_operational_planning_main_info_to_excel(planning_problem, wb, results, execution_time=execution_time)
    _write_operational_planning_main_info_to_excel_detailed(planning_problem, wb, results['summary_detail'])
    _write_shared_ess_specifications(wb, planning_problem.shared_ess_data)
    _write_operational_planning_market_data_to_excel(planning_problem, wb)

    if bound_evolution:
        _write_bound_evolution_to_excel(wb, bound_evolution)
        _write_planning_termination_to_excel(wb, bound_evolution)
        admm_diagnostics = bound_evolution.get('admm_diagnostics', [])
        if admm_diagnostics:
            _write_admm_diagnostics_to_excel(wb, admm_diagnostics)
        solver_recovery_diagnostics = bound_evolution.get('solver_recovery_diagnostics', [])
        if solver_recovery_diagnostics:
            _write_solver_recovery_diagnostics_to_excel(wb, solver_recovery_diagnostics)
        finite_difference_results = bound_evolution.get('finite_difference', [])
        if finite_difference_results:
            _write_finite_difference_validation_to_excel(wb, finite_difference_results)
        sensitivity_probe_diagnostics = bound_evolution.get(
            'sensitivity_probe_diagnostics', []
        )
        if sensitivity_probe_diagnostics:
            _write_sensitivity_probe_diagnostics_to_excel(
                wb, sensitivity_probe_diagnostics
            )

    if shared_ess_capacity:
        write_investment = True
        if shared_ess_cost:
            write_investment = False
        planning_problem.shared_ess_data.write_ess_capacity_results_to_excel(wb, shared_ess_capacity, write_investment=write_investment)

    if shared_ess_cost:
        planning_problem.shared_ess_data.write_ess_costs_to_excel(wb, shared_ess_cost)

    if salvage_value_results:
        planning_problem.shared_ess_data.write_salvage_value_results_to_excel(
            wb, salvage_value_results
        )

    # Interface Power Flow
    _write_interface_results_to_excel(planning_problem, wb, results['interface'])
    if results.get('scenario_dispersion'):
        _write_scenario_dispersion_to_excel(
            planning_problem, wb, results['scenario_dispersion']
        )

    # Shared Energy Storages results
    _write_shared_energy_storages_results_to_excel(planning_problem, wb, results)

    #  TSO and DSOs' results
    _write_network_voltage_results_to_excel(planning_problem, wb, results)
    _write_network_consumption_results_to_excel(planning_problem, wb, results)
    _write_network_generation_results_to_excel(planning_problem, wb, results)
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'losses')
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'ratio')
    _write_network_branch_loading_results_to_excel(planning_problem, wb, results)
    _write_network_branch_power_flow_results_to_excel(planning_problem, wb, results)
    _write_network_energy_storages_results_to_excel(planning_problem, wb, results)
    _write_relaxation_slacks_results_to_excel(planning_problem, wb, results)
    planning_problem.shared_ess_data.write_relaxation_slacks_results_to_excel(wb, results['esso'])

    # Save results
    try:
        wb.save(filename)
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = f"{filename.replace('.xlsx', '')}_{current_time}.xlsx"
        print(f"[WARNING] Results saved to file {backup_filename}.xlsx")
        wb.save(backup_filename)


def _write_bound_evolution_to_excel(workbook, bound_evolution):

    sheet = workbook.create_sheet('Convergence Characteristic')

    master_estimate = bound_evolution.get('master_estimate', bound_evolution.get('lower_bound', []))
    columns = [
        ('candidate_source', 'Candidate Source', bound_evolution.get('candidate_source', []), None, 'General'),
        ('operational_initialization', 'Operational Initialization', bound_evolution.get('operational_initialization', []), None, 'General'),
        ('master_estimate', 'Master Estimate (nominal LB), [NPV Mm.u.]', master_estimate, 1e6, '0.00'),
        ('alpha', 'Alpha, [NPV Mm.u.]', bound_evolution.get('alpha', []), 1e6, '0.00'),
        ('investment_cost', 'Investment Cost, [NPV Mm.u.]', bound_evolution.get('investment_cost', []), 1e6, '0.00'),
        ('gross_operational_cost', 'Gross Operational Objective, [NPV Mm.u.]', bound_evolution.get('gross_operational_cost', []), 1e6, '0.00'),
        ('terminal_salvage_value', 'Terminal Salvage Value, [NPV Mm.u.]', bound_evolution.get('terminal_salvage_value', []), 1e6, '0.00'),
        ('operational_recourse', 'Net Operational Recourse, [NPV Mm.u.]', bound_evolution.get('operational_recourse', []), 1e6, '0.00'),
        ('candidate_total', 'Candidate Total Objective, [NPV Mm.u.]', bound_evolution.get('candidate_total', []), 1e6, '0.00'),
        ('upper_bound', 'Incumbent Upper Bound, [NPV Mm.u.]', bound_evolution.get('upper_bound', []), 1e6, '0.00'),
        ('gap_signed', 'Signed Nominal Gap (UB - Master), [NPV Mm.u.]', bound_evolution.get('gap_signed', []), 1e6, '0.00'),
        ('gap_abs', 'Absolute Nominal Gap, [NPV Mm.u.]', bound_evolution.get('gap_abs', []), 1e6, '0.00'),
        ('gap_rel', 'Relative Nominal Gap, [%]', bound_evolution.get('gap_rel', []), 0.01, '0.00'),
        ('esso_violation', 'ESSO Aggregate Feasibility Slack, [N/A]', bound_evolution.get('esso_violation', []), 1.00, '0.000000'),
        ('incumbent_updated', 'Incumbent Updated', bound_evolution.get('incumbent_updated', []), None, 'General'),
    ]
    num_lines = max((len(values) for _, _, values, _, _ in columns), default=0)

    sheet.cell(row=1, column=1).value = 'Iteration'
    for column_idx, (_, label, _, _, _) in enumerate(columns, start=2):
        sheet.cell(row=1, column=column_idx).value = label

    for iteration in range(num_lines):
        row_idx = iteration + 2
        sheet.cell(row=row_idx, column=1).value = iteration + 1
        for column_idx, (_, _, values, divisor, number_format) in enumerate(columns, start=2):
            if iteration >= len(values) or values[iteration] is None:
                continue
            value = values[iteration]
            sheet.cell(row=row_idx, column=column_idx).value = (
                value if divisor is None else value / divisor
            )
            sheet.cell(row=row_idx, column=column_idx).number_format = number_format


def _write_planning_termination_to_excel(workbook, bound_evolution):
    sheet = workbook.create_sheet('Planning Termination')
    sensitivity_probe_diagnostics = bound_evolution.get(
        'sensitivity_probe_diagnostics', []
    )
    sensitivity_probe_evaluations = len({
        diagnostic.get('outer_iteration')
        for diagnostic in sensitivity_probe_diagnostics
        if diagnostic.get('outer_iteration') is not None
    })
    values = [
        ('Termination Reason', bound_evolution.get('termination_reason')),
        ('Converged', bound_evolution.get('convergence')),
        ('Outer Iterations', bound_evolution.get('outer_iterations')),
        ('Incumbent Iteration', bound_evolution.get('incumbent_iteration')),
        ('Incumbent Objective, [NPV m.u.]', bound_evolution.get('incumbent_objective')),
        ('Positive Bootstrap Used', bound_evolution.get('positive_bootstrap_used')),
        ('Positive Bootstrap Iteration', bound_evolution.get('positive_bootstrap_iteration')),
        ('Positive Bootstrap Budget Fraction', bound_evolution.get('positive_bootstrap_budget_fraction')),
        ('Sensitivity Probe Enabled', bound_evolution.get('sensitivity_probe_enabled')),
        ('Sensitivity Probe Budget Fraction', bound_evolution.get('sensitivity_probe_budget_fraction')),
        ('Sensitivity Probe E/S Ratio', bound_evolution.get('sensitivity_probe_energy_to_power_ratio')),
        ('Sensitivity Probe Evaluations', sensitivity_probe_evaluations),
        ('Validation Reference Source', bound_evolution.get('validation_reference_source')),
        ('Validation Reference Iteration', bound_evolution.get('validation_reference_iteration')),
        ('Validation Reference Is Incumbent', bound_evolution.get('validation_reference_is_incumbent')),
    ]
    for row_idx, (label, value) in enumerate(values, start=1):
        sheet.cell(row=row_idx, column=1).value = label
        sheet.cell(row=row_idx, column=2).value = value


def _write_sensitivity_probe_diagnostics_to_excel(workbook, diagnostics):
    sheet = workbook.create_sheet('Sensitivity Probes')
    columns = [
        ('outer_iteration', 'Outer Iteration', '0'),
        ('node_id', 'ESS Node', 'General'),
        ('year', 'Investment Year', 'General'),
        ('missing_types', 'Missing Capacity Types', 'General'),
        ('status', 'Status', 'General'),
        ('reason', 'Reason', 'General'),
        ('probe_power_mva', 'Probe Power, [MVA]', '0.000000'),
        ('probe_energy_mvah', 'Probe Energy, [MVAh]', '0.000000'),
        ('probe_master_feasible', 'Probe Master Feasible', 'General'),
        ('probe_master_feasibility_reason', 'Master Feasibility Reason', 'General'),
        ('initialization_source', 'Initialization Source', 'General'),
        ('operational_convergence', 'Operational Convergence', 'General'),
        ('esso_violation', 'ESSO Feasibility Violation', '0.000000'),
        ('probe_recourse', 'Probe Net Recourse, [NPV m.u.]', '0.00'),
        ('sensitivity_s', 'One-Sided Power Sensitivity', '0.000000'),
        ('sensitivity_e', 'One-Sided Energy Sensitivity', '0.000000'),
    ]
    for column_idx, (_, label, _) in enumerate(columns, start=1):
        sheet.cell(row=1, column=column_idx).value = label
    for row_idx, diagnostic in enumerate(diagnostics, start=2):
        for column_idx, (key, _, number_format) in enumerate(columns, start=1):
            value = diagnostic.get(key)
            if value is None:
                continue
            cell = sheet.cell(row=row_idx, column=column_idx)
            cell.value = value
            cell.number_format = number_format


def _write_finite_difference_validation_to_excel(workbook, validation_results):
    sheet = workbook.create_sheet('Sensitivity Validation')
    columns = [
        ('run_type', 'Run Type', 'General'),
        ('direction', 'Direction', 'General'),
        ('status', 'Status', 'General'),
        ('reason', 'Reason', 'General'),
        ('refinement', 'Refinement Count', '0'),
        ('max_refinements', 'Maximum Refinements', '0'),
        ('reference_stabilized', 'Reference Stabilized', 'General'),
        ('endpoint_stabilized', 'Perturbed Endpoint Stabilized', 'General'),
        ('original_cut_reproducible', 'Original Sensitivity Point Reproducible', 'General'),
        ('validation_source', 'Validation Reference Source', 'General'),
        ('validation_reference_is_incumbent', 'Validation Reference Is Incumbent', 'General'),
        ('baseline_outer_iteration', 'Baseline Outer Iteration', '0'),
        ('termination_reason', 'Planning Termination Reason', 'General'),
        ('node_id', 'Node ID', '0'),
        ('year', 'Investment Year', '0'),
        ('base_s', 'Base S Investment, [MVA]', '0.000000'),
        ('base_e', 'Base E Investment, [MVAh]', '0.000000'),
        ('energy_to_power_ratio', 'E/S Ratio, [h]', '0.000000'),
        ('step_fraction', 'Relative Directional Step, [%]', '0.00%'),
        ('step_size', 'Directional Step Scalar h', '0.000000'),
        ('delta_s', 'Delta S, [MVA]', '0.000000'),
        ('delta_e', 'Delta E, [MVAh]', '0.000000'),
        ('direction_s', 'Direction S Component', '0.000000'),
        ('direction_e', 'Direction E Component', '0.000000'),
        ('first_stage_feasible', 'Perturbed Point First-Stage Feasible', 'General'),
        ('first_stage_reason', 'First-Stage Feasibility Detail', 'General'),
        ('sensitivity_s', 'Sensitivity S, [NPV m.u./MVA]', '0.000000'),
        ('sensitivity_e', 'Sensitivity E, [NPV m.u./MVAh]', '0.000000'),
        ('replay_sensitivity_s', 'Replay Sensitivity S, [NPV m.u./MVA]', '0.000000'),
        ('replay_sensitivity_e', 'Replay Sensitivity E, [NPV m.u./MVAh]', '0.000000'),
        ('original_analytic_slope', 'Baseline Directional Slope, [NPV m.u./h]', '0.000000'),
        ('analytic_slope', 'Analytic Directional Slope, [NPV m.u./h]', '0.000000'),
        ('replay_analytic_slope', 'Replay Directional Slope, [NPV m.u./h]', '0.000000'),
        ('predicted_change', 'Predicted Recourse Change, [NPV m.u.]', '0.000000'),
        ('baseline_recourse', 'Baseline Recourse, [NPV m.u.]', '0.000000'),
        ('reference_recourse', 'Replay Reference Recourse, [NPV m.u.]', '0.000000'),
        ('perturbed_recourse', 'Perturbed Recourse, [NPV m.u.]', '0.000000'),
        ('observed_change', 'Observed Recourse Change, [NPV m.u.]', '0.000000'),
        ('absolute_error', 'Absolute Recourse-Change Error, [NPV m.u.]', '0.000000'),
        ('observed_slope', 'Observed Directional Slope, [NPV m.u./h]', '0.000000'),
        ('absolute_slope_error', 'Absolute Slope Error, [NPV m.u./h]', '0.000000'),
        ('relative_error', 'Relative Slope Error, [%]', '0.00%'),
        ('signal_to_noise_ratio', 'Signal-to-Noise Ratio', '0.00'),
        ('slope_consistency_error', 'Step-to-Step Slope Difference, [%]', '0.00%'),
        ('replay_drift', 'Replay Recourse Drift, [NPV m.u.]', '0.000000'),
        ('replay_tolerance', 'Replay Drift Tolerance, [NPV m.u.]', '0.000000'),
        ('stationarity_drift', 'Endpoint Cycle Recourse Drift, [NPV m.u.]', '0.000000'),
        ('stationarity_tolerance', 'Endpoint Recourse Tolerance, [NPV m.u.]', '0.000000'),
        ('sensitivity_relative_drift', 'Replay Sensitivity Drift, [%]', '0.00%'),
        ('sensitivity_relative_drift_s', 'Replay Sensitivity S Drift, [%]', '0.00%'),
        ('sensitivity_relative_drift_e', 'Replay Sensitivity E Drift, [%]', '0.00%'),
        ('original_sensitivity_relative_drift', 'Baseline Sensitivity Drift, [%]', '0.00%'),
        ('original_sensitivity_relative_drift_s', 'Baseline Sensitivity S Drift, [%]', '0.00%'),
        ('original_sensitivity_relative_drift_e', 'Baseline Sensitivity E Drift, [%]', '0.00%'),
        ('same_sign', 'Same Sign', 'General'),
        ('operational_convergence', 'ADMM Converged', 'General'),
        ('esso_violation', 'ESSO Feasibility Slack, [N/A]', '0.000000'),
        ('baseline_soh_margin', 'Baseline Minimum SoH Margin, [p.u.]', '0.000000'),
        ('reference_soh_margin', 'Replay Minimum SoH Margin, [p.u.]', '0.000000'),
        ('perturbed_soh_margin', 'Perturbed Minimum SoH Margin, [p.u.]', '0.000000'),
        ('active_set_changed', 'Minimum SoH Activity Changed', 'General'),
        ('passed', 'Passed', 'General'),
    ]

    for column_idx, (_, label, _) in enumerate(columns, start=1):
        sheet.cell(row=1, column=column_idx).value = label

    for row_idx, result in enumerate(validation_results, start=2):
        for column_idx, (key, _, number_format) in enumerate(columns, start=1):
            value = result.get(key)
            if value is None:
                continue
            sheet.cell(row=row_idx, column=column_idx).value = value
            sheet.cell(row=row_idx, column=column_idx).number_format = number_format


def _write_admm_diagnostics_to_excel(workbook, diagnostics):
    sheet = workbook.create_sheet('ADMM Convergence')

    columns = [
        ('outer_iteration', 'Planning Iteration', '0'),
        ('cycle', 'ADMM Cycle', '0'),
        ('local_solves_ok', 'Local Solves Successful', 'General'),
        ('primal_v', 'Primal V Max Residual', '0.000000'),
        ('primal_v_tolerance', 'Primal V Max Tolerance', '0.000000'),
        ('primal_v_mean', 'Primal V Mean Residual', '0.000000'),
        ('primal_v_mean_tolerance', 'Primal V Mean Tolerance', '0.000000'),
        ('primal_pf', 'Primal PF Max Residual', '0.000000'),
        ('primal_pf_tolerance', 'Primal PF Max Tolerance', '0.000000'),
        ('primal_pf_mean', 'Primal PF Mean Residual', '0.000000'),
        ('primal_pf_mean_tolerance', 'Primal PF Mean Tolerance', '0.000000'),
        ('primal_ess', 'Primal ESS Max Residual', '0.000000'),
        ('primal_ess_tolerance', 'Primal ESS Max Tolerance', '0.000000'),
        ('primal_ess_mean', 'Primal ESS Mean Residual', '0.000000'),
        ('primal_ess_mean_tolerance', 'Primal ESS Mean Tolerance', '0.000000'),
        ('primal_v_ratio', 'Primal V / Tolerance', '0.000'),
        ('primal_v_mean_ratio', 'Primal V Mean / Tolerance', '0.000'),
        ('primal_pf_ratio', 'Primal PF / Tolerance', '0.000'),
        ('primal_pf_mean_ratio', 'Primal PF Mean / Tolerance', '0.000'),
        ('primal_ess_ratio', 'Primal ESS Max / Tolerance', '0.000'),
        ('primal_ess_mean_ratio', 'Primal ESS Mean / Tolerance', '0.000'),
        ('dual_v', 'Dual V Max Residual', '0.000000'),
        ('dual_v_mean', 'Dual V Mean Residual', '0.000000'),
        ('dual_v_tolerance', 'Dual V Tolerance', '0.000000'),
        ('dual_pf', 'Dual PF Max Residual', '0.000000'),
        ('dual_pf_mean', 'Dual PF Mean Residual', '0.000000'),
        ('dual_pf_tolerance', 'Dual PF Tolerance', '0.000000'),
        ('dual_ess', 'Dual ESS Max Residual', '0.000000'),
        ('dual_ess_mean', 'Dual ESS Mean Residual', '0.000000'),
        ('dual_ess_tolerance', 'Dual ESS Tolerance', '0.000000'),
        ('dual_v_ratio', 'Dual V Max / Tolerance', '0.000'),
        ('dual_v_mean_ratio', 'Dual V Mean / Tolerance', '0.000'),
        ('dual_pf_ratio', 'Dual PF Max / Tolerance', '0.000'),
        ('dual_pf_mean_ratio', 'Dual PF Mean / Tolerance', '0.000'),
        ('dual_ess_ratio', 'Dual ESS Max / Tolerance', '0.000'),
        ('dual_ess_mean_ratio', 'Dual ESS Mean / Tolerance', '0.000'),
        ('gross_operational_cost', 'Gross Operational Objective, [NPV m.u.]', '0.000000'),
        ('terminal_salvage_value', 'Terminal Salvage Value, [NPV m.u.]', '0.000000'),
        ('recourse', 'Net Operational Recourse, [NPV m.u.]', '0.000000'),
        ('objective_change_abs', 'Absolute Recourse Change, [NPV m.u.]', '0.000000'),
        ('objective_change_rel', 'Relative Recourse Change, [%]', '0.00%'),
        ('objective_tolerance', 'Applied Recourse Tolerance, [NPV m.u.]', '0.000000'),
        ('objective_absolute_tolerance', 'Absolute Recourse Tolerance, [NPV m.u.]', '0.000000'),
        ('objective_relative_tolerance', 'Relative Recourse Tolerance, [%]', '0.00%'),
        ('residual_convergence', 'Residuals Converged', 'General'),
        ('objective_convergence', 'Recourse Converged', 'General'),
        ('cycle_convergence', 'Cycle Converged', 'General'),
        ('consecutive_converged_cycles', 'Consecutive Converged Cycles', '0'),
        ('required_consecutive_cycles', 'Required Consecutive Cycles', '0'),
        ('rho_v_before', 'Mean Rho V Before', '0.000000'),
        ('rho_v_after', 'Mean Rho V After', '0.000000'),
        ('rho_v_action', 'Rho V Action', 'General'),
        ('rho_pf_before', 'Mean Rho PF Before', '0.000000'),
        ('rho_pf_after', 'Mean Rho PF After', '0.000000'),
        ('rho_pf_action', 'Rho PF Action', 'General'),
        ('rho_ess_before', 'Mean Rho ESS Before', '0.000000'),
        ('rho_ess_after', 'Mean Rho ESS After', '0.000000'),
        ('rho_ess_action', 'Rho ESS Action', 'General'),
        ('worst_v_primal_node', 'Worst V Node', '0'),
        ('worst_v_primal_year', 'Worst V Year', '0'),
        ('worst_v_primal_day', 'Worst V Day', 'General'),
        ('worst_v_primal_period', 'Worst V Period', '0'),
        ('worst_v_primal_difference', 'Worst V Difference, [kV]', '0.000000'),
        ('worst_pf_primal_node', 'Worst PF Node', '0'),
        ('worst_pf_primal_year', 'Worst PF Year', '0'),
        ('worst_pf_primal_day', 'Worst PF Day', 'General'),
        ('worst_pf_primal_period', 'Worst PF Period', '0'),
        ('worst_pf_primal_type', 'Worst PF Type', 'General'),
        ('worst_pf_primal_difference', 'Worst PF Difference, [MW/MVAr]', '0.000000'),
        ('worst_pf_primal_rating', 'Worst PF Interface Rating, [MVA]','0.000000'),
        ('worst_pf_primal_rho_tso', 'Worst PF Rho TSO', '0.000000'),
        ('worst_pf_primal_rho_dso', 'Worst PF Rho DSO', '0.000000'),
        ('worst_ess_primal_node', 'Worst ESS Node', '0'),
        ('worst_ess_primal_year', 'Worst ESS Year', '0'),
        ('worst_ess_primal_day', 'Worst ESS Day', 'General'),
        ('worst_ess_primal_period', 'Worst ESS Period', '0'),
        ('worst_ess_primal_type', 'Worst ESS Type', 'General'),
        ('worst_ess_primal_agent', 'Worst ESS Agent', 'General'),
        ('worst_ess_primal_difference', 'Worst ESS Difference, [MW/MVAr]', '0.000000'),
        ('worst_ess_primal_rating', 'Worst ESS Normalization Rating, [MVA]', '0.000000'),
        ('worst_ess_primal_rho', 'Worst ESS Rho', '0.000000')
    ]
    for column_idx, (_, label, _) in enumerate(columns, start=1):
        sheet.cell(row=1, column=column_idx).value = label
    for row_idx, diagnostic in enumerate(diagnostics, start=2):
        for column_idx, (key, _, number_format) in enumerate(columns, start=1):
            value = diagnostic.get(key)
            if value is None:
                continue
            sheet.cell(row=row_idx, column=column_idx).value = value
            sheet.cell(row=row_idx, column=column_idx).number_format = number_format


def _write_solver_recovery_diagnostics_to_excel(workbook, diagnostics):
    sheet = workbook.create_sheet('Solver Recovery')
    columns = [
        ('outer_iteration', 'Planning Iteration', '0'),
        ('subsystem', 'Subsystem', 'General'),
        ('node_id', 'Node ID', '0'),
        ('warm_start', 'Warm Start', 'General'),
        ('primary_result', 'Primary Result', 'General'),
        ('recovery_options', 'Recovery Options', 'General'),
        ('recovery_result', 'Recovery Result', 'General'),
        ('recovery_succeeded', 'Recovery Successful', 'General'),
        ('primary_log', 'Primary Log', 'General'),
        ('recovery_log', 'Recovery Log', 'General'),
    ]

    for column_idx, (_, label, _) in enumerate(columns, start=1):
        sheet.cell(row=1, column=column_idx).value = label
    for row_idx, diagnostic in enumerate(diagnostics, start=2):
        for column_idx, (key, _, number_format) in enumerate(columns, start=1):
            value = diagnostic.get(key)
            if value is None:
                continue
            sheet.cell(row=row_idx, column=column_idx).value = value
            sheet.cell(row=row_idx, column=column_idx).number_format = number_format


# ======================================================================================================================
#  RESULTS OPERATIONAL PLANNING - write functions
# ======================================================================================================================
def _write_operational_planning_results_to_excel(planning_problem, results, primal_evolution=list(),
                                                 admm_diagnostics=list(), shared_ess_capacity=dict(),
                                                 salvage_value_results=dict(),
                                                 solver_recovery_diagnostics=list(),
                                                 filename='operation_planning', execution_time=float()):

    wb = Workbook()

    _write_operational_planning_main_info_to_excel(planning_problem, wb, results, execution_time=execution_time)
    _write_operational_planning_main_info_to_excel_detailed(planning_problem, wb, results['summary_detail'])
    _write_shared_ess_specifications(wb, planning_problem.shared_ess_data)
    if shared_ess_capacity:
        planning_problem.shared_ess_data.write_ess_capacity_results_to_excel(wb, shared_ess_capacity)
    if salvage_value_results:
        planning_problem.shared_ess_data.write_salvage_value_results_to_excel(
            wb, salvage_value_results
        )
    _write_operational_planning_market_data_to_excel(planning_problem, wb)

    if primal_evolution:
        _write_objective_function_evolution_to_excel(wb, primal_evolution)
    if admm_diagnostics:
        _write_admm_diagnostics_to_excel(wb, admm_diagnostics)
    if solver_recovery_diagnostics:
        _write_solver_recovery_diagnostics_to_excel(wb, solver_recovery_diagnostics)

    # Interface Power Flow
    _write_interface_results_to_excel(planning_problem, wb, results['interface'])
    if results.get('scenario_dispersion'):
        _write_scenario_dispersion_to_excel(
            planning_problem, wb, results['scenario_dispersion']
        )

    # Shared Energy Storages results
    _write_shared_energy_storages_results_to_excel(planning_problem, wb, results)

    #  TSO and DSOs' results
    _write_network_voltage_results_to_excel(planning_problem, wb, results)
    _write_network_consumption_results_to_excel(planning_problem, wb, results)
    _write_network_generation_results_to_excel(planning_problem, wb, results)
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'losses')
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'ratio')
    _write_network_branch_loading_results_to_excel(planning_problem, wb, results)
    _write_network_branch_power_flow_results_to_excel(planning_problem, wb, results)
    _write_network_energy_storages_results_to_excel(planning_problem, wb, results)
    _write_relaxation_slacks_results_to_excel(planning_problem, wb, results)
    planning_problem.shared_ess_data.write_relaxation_slacks_results_to_excel(wb, results['esso'])

    # Save results
    results_filename = os.path.join(planning_problem.results_dir, f'{filename}.xlsx')
    try:
        wb.save(results_filename)
        print('[INFO] Operational Planning Results written to {}.'.format(results_filename))
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = os.path.join(planning_problem.results_dir, f"{filename.replace('.xlsx', '')}_{current_time}.xlsx")
        print(f"[WARNING] Results saved to file {backup_filename}.xlsx")
        wb.save(backup_filename)


def _write_operational_planning_results_hierarchical_to_excel(planning_problem, results, filename='operation_planning_results_hierarchical', execution_time=float()):
    _write_operational_planning_results_no_coordination_to_excel(planning_problem, results, filename=filename, execution_time=execution_time)


def _write_operational_planning_results_no_coordination_to_excel(planning_problem, results, filename='operation_planning_results_no_coordination', execution_time=float()):

    wb = Workbook()

    _write_operational_planning_main_info_to_excel(planning_problem, wb, results, execution_time=execution_time)
    _write_operational_planning_main_info_to_excel_detailed(planning_problem, wb, results['summary_detail'])
    _write_operational_planning_market_data_to_excel(planning_problem, wb)

    #  TSO and DSOs' results
    _write_network_voltage_results_to_excel(planning_problem, wb, results)
    _write_network_consumption_results_to_excel(planning_problem, wb, results)
    _write_network_generation_results_to_excel(planning_problem, wb, results)
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'losses')
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'ratio')
    _write_network_branch_loading_results_to_excel(planning_problem, wb, results)
    _write_network_branch_power_flow_results_to_excel(planning_problem, wb, results)
    _write_network_energy_storages_results_to_excel(planning_problem, wb, results)
    _write_relaxation_slacks_results_no_coordination_to_excel(planning_problem, wb, results)

    # Save results
    try:
        wb.save(filename)
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = f"{filename.replace('.xlsx', '')}_{current_time}.xlsx"
        print(f"[WARNING] Results saved to file {backup_filename}.xlsx")
        wb.save(backup_filename)


def _write_operational_planning_main_info_to_excel(planning_problem, workbook, results, execution_time=float()):

    sheet = workbook.worksheets[0]
    sheet.title = 'Main Info'

    # Write Header
    col_idx = 4
    line_idx = 1
    for year in planning_problem.years:
        for _ in planning_problem.days:
            sheet.cell(row=line_idx, column=col_idx).value = year
            col_idx += 1

    col_idx = 1
    line_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Agent'
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Node ID'
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Value'
    col_idx += 1

    for _ in planning_problem.years:
        for day in planning_problem.days:
            sheet.cell(row=line_idx, column=col_idx).value = day
            col_idx += 1

    # TSO
    line_idx = _write_operational_planning_main_info_per_operator(planning_problem.transmission_network, sheet, 'TSO', line_idx, results['tso']['results'])

    # DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id]
        line_idx = _write_operational_planning_main_info_per_operator(distribution_network, sheet, 'DSO', line_idx, dso_results, tn_node_id=tn_node_id)

    if execution_time:
        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = '-'
        sheet.cell(row=line_idx, column=2).value = '-'
        sheet.cell(row=line_idx, column=3).value = 'Execution time, [s]'
        sheet.cell(row=line_idx, column=4).value = execution_time
        sheet.cell(row=line_idx, column=4).number_format = '0.00'

    _write_run_metadata_to_excel(
        planning_problem,
        workbook,
        include_scenario_deviation=bool(results.get('scenario_dispersion')),
    )


def _write_run_metadata_to_excel(planning_problem, workbook, include_scenario_deviation=False):

    sheet = workbook.create_sheet('Run Metadata')
    metadata = planning_problem.scenario_metadata
    random_seed = metadata.get('random_seed')
    rows = [
        ('Random Seed', random_seed if random_seed is not None else 'Unseeded'),
        ('Deterministic Scenario Generation', metadata.get('deterministic_scenarios', False)),
        ('Market Scenario SHA-256', metadata.get('market_scenario_checksum')),
        ('Operational Scenario SHA-256', metadata.get('operational_scenario_checksum')),
        ('Combined Scenario SHA-256', metadata.get('combined_scenario_checksum')),
        ('Number of Market Scenarios', planning_problem.num_market_scenarios),
        ('Transmission Operation Scenarios', planning_problem.transmission_network.num_oper_scenarios,),
    ]
    if include_scenario_deviation:
        rows.extend([
            ('Distributed Coupling Basis', 'Expected interface and shared-ESS schedules',),
            ('Distributed Scenario Dispersion Treatment', 'Interface/voltage: probability-weighted quadratic deviation penalty',),
            ('Shared ESS Scenario Treatment', 'Probability-weighted expected P/Q with quadratic scenario-deviation penalty'),
            ('Scenario Deviation Penalty Coefficient', PENALTY_SCENARIO_DEVIATION),
            ('Scenario Deviation Included in Operational Recourse', False),
        ])
    rows.extend(
        (f'Distribution Node {node_id} Operation Scenarios', planning_problem.distribution_networks[node_id].num_oper_scenarios,) for node_id in sorted(planning_problem.distribution_networks, key=str)
    )

    sheet.cell(row=1, column=1).value = 'Property'
    sheet.cell(row=1, column=2).value = 'Value'
    for row_idx, (label, value) in enumerate(rows, start=2):
        sheet.cell(row=row_idx, column=1).value = label
        sheet.cell(row=row_idx, column=2).value = value


def _write_operational_planning_main_info_per_operator(network, sheet, operator_type, line_idx, results, tn_node_id='-'):

    decimal_style = '0.00'

    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1

    # - Objective
    sheet.cell(row=line_idx, column=col_idx).value = 'Base SMOPF objective value, [N/A]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['obj']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Total Load
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Load, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_load']['p']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Load, [MVArh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_load']['q']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Flexibility used
    if network.params.fl_reg:
        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = operator_type
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Flexibility procurement, [MWh]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['flex_used']['p']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = operator_type
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Flexibility procurement, [MVArh]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['flex_used']['q']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

        if network.params.obj_type == OBJ_MIN_COST:
            line_idx += 1
            col_idx = 1
            sheet.cell(row=line_idx, column=col_idx).value = operator_type
            col_idx += 1
            sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
            col_idx += 1
            sheet.cell(row=line_idx, column=col_idx).value = 'Flexibility remuneration, [€]'
            col_idx += 1
            for year in results:
                for day in results[year]:
                    sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['flex_cost']
                    sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                    col_idx += 1

    # Total Load curtailed
    if network.params.l_curt:

        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = operator_type
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Load curtailed, [MWh]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['load_curt']['p']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = operator_type
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Load curtailed, [MVArh]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['load_curt']['q']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

        if network.params.obj_type == OBJ_MIN_COST:
            line_idx += 1
            col_idx = 1
            sheet.cell(row=line_idx, column=col_idx).value = operator_type
            col_idx += 1
            sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
            col_idx += 1
            sheet.cell(row=line_idx, column=col_idx).value = 'Load curtailment cost, [€]'
            col_idx += 1
            for year in results:
                for day in results[year]:
                    sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['load_curt_cost']
                    sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                    col_idx += 1

    # Total Generation
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Generation, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_gen']['p']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Generation, [MVArh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_gen']['q']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Total Conventional Generation
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Conventional Generation, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_conventional_gen']['p']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Conventional Generation, [MVArh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_conventional_gen']['q']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    if network.params.obj_type == OBJ_MIN_COST:
        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = operator_type
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Conventional generation cost, [€]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['gen_cost']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

    # Total Renewable Generation
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Renewable generation, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_renewable_gen']['p']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Renewable generation, [MVArh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_renewable_gen']['q']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Renewable generation, [MVAh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_renewable_gen']['s']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Renewable Generation Curtailed
    if network.params.rg_curt:

        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = operator_type
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Renewable generation curtailed, [MVAh]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['gen_curt']['s']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = operator_type
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
        col_idx += 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Renewable generation curtailment penalty, [N/A]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['gen_curt_penalty']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

    # Losses
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Losses, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['losses']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Number of price (market) scenarios
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Number of market scenarios'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = len(network.network[year][day].prob_market_scenarios)
            col_idx += 1

    # Number of operation (generation and consumption) scenarios
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = operator_type
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = tn_node_id
    col_idx += 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Number of operation scenarios'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = len(network.network[year][day].prob_operation_scenarios)
            col_idx += 1

    return line_idx


def _write_operational_planning_main_info_to_excel_detailed(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Main Info, Detailed')

    # Write Header -- Year
    line_idx = 1
    sheet.cell(row=line_idx, column=1).value = 'Operator'
    sheet.cell(row=line_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=line_idx, column=3).value = 'Year'
    sheet.cell(row=line_idx, column=4).value = 'Day'
    sheet.cell(row=line_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=line_idx, column=6).value = 'Operation Scenario'
    sheet.cell(row=line_idx, column=7).value = 'Probability, [%]'
    sheet.cell(row=line_idx, column=8).value = 'Base SMOPF Objective Value, [N/A]'
    sheet.cell(row=line_idx, column=9).value = 'Load, [MWh]'
    sheet.cell(row=line_idx, column=10).value = 'Load, [MVArh]'
    sheet.cell(row=line_idx, column=11).value = 'Flexibility procurement, [MWh]'
    sheet.cell(row=line_idx, column=12).value = 'Flexibility procurement, [MVArh]'
    sheet.cell(row=line_idx, column=13).value = 'Flexibility remuneration, [€]'
    sheet.cell(row=line_idx, column=14).value = 'Generation, [MWh]'
    sheet.cell(row=line_idx, column=15).value = 'Generation, [MVArh]'
    sheet.cell(row=line_idx, column=16).value = 'Conventional Generation, [MWh]'
    sheet.cell(row=line_idx, column=17).value = 'Conventional Generation, [MVArh]'
    sheet.cell(row=line_idx, column=18).value = 'Conventional Generation Cost, [€]'
    sheet.cell(row=line_idx, column=19).value = 'Renewable Generation, [MWh]'
    sheet.cell(row=line_idx, column=20).value = 'Renewable Generation, [MVArh]'
    sheet.cell(row=line_idx, column=21).value = 'Renewable Generation, [MVAh]'
    sheet.cell(row=line_idx, column=22).value = 'Renewable Generation Curtailed, [MVAh]'
    sheet.cell(row=line_idx, column=23).value = 'Renewable Generation Curtailment Penalty, [N/A]'
    sheet.cell(row=line_idx, column=24).value = 'Losses, [MWh]'

    # TSO
    line_idx += 1
    line_idx = _write_operational_planning_main_info_per_operator_detailed(planning_problem.transmission_network, sheet, 'TSO', line_idx, results['tso'])

    # DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]
        distribution_network = planning_problem.distribution_networks[tn_node_id]
        line_idx = _write_operational_planning_main_info_per_operator_detailed(distribution_network, sheet, 'DSO', line_idx, dso_results, tn_node_id=tn_node_id)


def _write_operational_planning_main_info_per_operator_detailed(network, sheet, operator_type, line_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    percent_style = '0.00%'

    sheet.cell(row=line_idx, column=1).value = operator_type
    for year in results:
        for day in results[year]:
            for s_m in results[year][day]['scenarios']:
                for s_o in results[year][day]['scenarios'][s_m]:

                    sheet.cell(row=line_idx, column=1).value = operator_type
                    sheet.cell(row=line_idx, column=2).value = tn_node_id
                    sheet.cell(row=line_idx, column=3).value = int(year)
                    sheet.cell(row=line_idx, column=4).value = day
                    sheet.cell(row=line_idx, column=5).value = s_m
                    sheet.cell(row=line_idx, column=6).value = s_o

                    # Probability, [%]
                    sheet.cell(row=line_idx, column=7).value = results[year][day]['scenarios'][s_m][s_o]['probability']
                    sheet.cell(row=line_idx, column=7).number_format = percent_style

                    # OF
                    sheet.cell(row=line_idx, column=8).value = results[year][day]['scenarios'][s_m][s_o]['obj']
                    sheet.cell(row=line_idx, column=8).number_format = decimal_style

                    # Load
                    sheet.cell(row=line_idx, column=9).value = results[year][day]['scenarios'][s_m][s_o]['load']['p']
                    sheet.cell(row=line_idx, column=9).number_format = decimal_style
                    sheet.cell(row=line_idx, column=10).value = results[year][day]['scenarios'][s_m][s_o]['load']['q']
                    sheet.cell(row=line_idx, column=10).number_format = decimal_style

                    # Flexibility, [MWh]
                    sheet.cell(row=line_idx, column=11).value = results[year][day]['scenarios'][s_m][s_o]['flexibility']['p']
                    sheet.cell(row=line_idx, column=11).number_format = decimal_style
                    sheet.cell(row=line_idx, column=12).value = results[year][day]['scenarios'][s_m][s_o]['flexibility']['q']
                    sheet.cell(row=line_idx, column=12).number_format = decimal_style

                    # Flexibility Cost, [€]
                    sheet.cell(row=line_idx, column=13).value = results[year][day]['scenarios'][s_m][s_o]['cost_flexibility']
                    sheet.cell(row=line_idx, column=13).number_format = decimal_style

                    # Generation
                    sheet.cell(row=line_idx, column=14).value = results[year][day]['scenarios'][s_m][s_o]['generation']['p']
                    sheet.cell(row=line_idx, column=14).number_format = decimal_style
                    sheet.cell(row=line_idx, column=15).value = results[year][day]['scenarios'][s_m][s_o]['generation']['q']
                    sheet.cell(row=line_idx, column=15).number_format = decimal_style

                    # Conventional Generation
                    sheet.cell(row=line_idx, column=16).value = results[year][day]['scenarios'][s_m][s_o]['generation_conventional']['p']
                    sheet.cell(row=line_idx, column=16).number_format = decimal_style
                    sheet.cell(row=line_idx, column=17).value = results[year][day]['scenarios'][s_m][s_o]['generation_conventional']['q']
                    sheet.cell(row=line_idx, column=17).number_format = decimal_style

                    # Conventional Generation Cost
                    sheet.cell(row=line_idx, column=18).value = results[year][day]['scenarios'][s_m][s_o]['generation_conventional_cost']
                    sheet.cell(row=line_idx, column=18).number_format = decimal_style

                    # Renewable Generation
                    sheet.cell(row=line_idx, column=19).value = results[year][day]['scenarios'][s_m][s_o]['generation_renewable']['p']
                    sheet.cell(row=line_idx, column=19).number_format = decimal_style
                    sheet.cell(row=line_idx, column=20).value = results[year][day]['scenarios'][s_m][s_o]['generation_renewable']['q']
                    sheet.cell(row=line_idx, column=20).number_format = decimal_style
                    sheet.cell(row=line_idx, column=21).value = results[year][day]['scenarios'][s_m][s_o]['generation_renewable']['s']
                    sheet.cell(row=line_idx, column=21).number_format = decimal_style
                    sheet.cell(row=line_idx, column=22).value = results[year][day]['scenarios'][s_m][s_o]['generation_renewable_curtailed']['s']
                    sheet.cell(row=line_idx, column=22).number_format = decimal_style
                    sheet.cell(row=line_idx, column=23).value = results[year][day]['scenarios'][s_m][s_o]['generation_renewable_curtailment_penalty']
                    sheet.cell(row=line_idx, column=23).number_format = decimal_style

                    # Losses
                    sheet.cell(row=line_idx, column=24).value = results[year][day]['scenarios'][s_m][s_o]['losses']
                    sheet.cell(row=line_idx, column=24).number_format = decimal_style

                    line_idx += 1

    return line_idx


def _write_shared_ess_specifications(workbook, shared_ess_info):

    sheet = workbook.create_sheet('SharedESS Specifications')

    decimal_style = '0.000'

    # Write Header
    row_idx = 1
    sheet.cell(row=row_idx, column=1).value = 'Year'
    sheet.cell(row=row_idx, column=2).value = 'Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Sinst, [MVA]'
    sheet.cell(row=row_idx, column=4).value = 'Einst, [MVAh]'

    # Write SharedESS specifications
    for year in shared_ess_info.years:
        for shared_ess in shared_ess_info.shared_energy_storages[year]:
            row_idx = row_idx + 1
            sheet.cell(row=row_idx, column=1).value = year
            sheet.cell(row=row_idx, column=2).value = shared_ess.bus
            sheet.cell(row=row_idx, column=3).value = shared_ess.s
            sheet.cell(row=row_idx, column=3).number_format = decimal_style
            sheet.cell(row=row_idx, column=4).value = shared_ess.e
            sheet.cell(row=row_idx, column=4).number_format = decimal_style


def _write_operational_planning_market_data_to_excel(planning_problem, workbook):

    sheet = workbook.create_sheet('MarketData')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Year'
    sheet.cell(row=row_idx, column=2).value = 'Day'
    sheet.cell(row=row_idx, column=3).value = 'Quantity'
    sheet.cell(row=row_idx, column=4).value = 'Market Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 5).value = p
    row_idx = row_idx + 1


    for year in planning_problem.years:
        for day in planning_problem.days:

            cost_energy = planning_problem.cost_energy_p[year][day]
            cost_flexibility = planning_problem.cost_flex[year][day]

            # - Energy
            for s_m in range(planning_problem.num_market_scenarios):
                sheet.cell(row=row_idx, column=1).value = int(year)
                sheet.cell(row=row_idx, column=2).value = day
                sheet.cell(row=row_idx, column=3).value = 'Energy'
                sheet.cell(row=row_idx, column=4).value = s_m
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 5).value = cost_energy[s_m][p]
                    sheet.cell(row=row_idx, column=p + 5).number_format = decimal_style
                row_idx += 1

            # - Flexibility
            for s_m in range(planning_problem.num_market_scenarios):
                sheet.cell(row=row_idx, column=1).value = int(year)
                sheet.cell(row=row_idx, column=2).value = day
                sheet.cell(row=row_idx, column=3).value = 'Flexibility'
                sheet.cell(row=row_idx, column=4).value = s_m
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 5).value = cost_flexibility[s_m][p]
                    sheet.cell(row=row_idx, column=p + 5).number_format = decimal_style
                row_idx += 1


def _write_objective_function_evolution_to_excel(workbook, primal_evolution):

    sheet = workbook.create_sheet('Primal Evolution')

    decimal_style = '0.000000'
    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Iteration'
    sheet.cell(row=row_idx, column=2).value = 'OF value'
    row_idx = row_idx + 1
    for i in range(len(primal_evolution)):
        sheet.cell(row=row_idx, column=1).value = i
        sheet.cell(row=row_idx, column=2).value = primal_evolution[i]
        sheet.cell(row=row_idx, column=2).number_format = decimal_style
        sheet.cell(row=row_idx, column=2).value = primal_evolution[i]
        sheet.cell(row=row_idx, column=2).number_format = decimal_style
        row_idx = row_idx + 1


def _write_scenario_dispersion_to_excel(planning_problem, workbook, records):
    sheet = workbook.create_sheet('Scenario Dispersion')
    headers = [
        'Operator',
        'Node ID',
        'Year',
        'Day',
        'Quantity',
        'Metric',
        'Scenario Count',
        'Probability Sum',
    ]
    headers.extend(range(planning_problem.num_instants))
    headers.append('Maximum')

    for column_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column_idx).value = header

    for row_idx, record in enumerate(records, start=2):
        sheet.cell(row=row_idx, column=1).value = record['operator']
        sheet.cell(row=row_idx, column=2).value = record['node_id']
        sheet.cell(row=row_idx, column=3).value = record['year']
        sheet.cell(row=row_idx, column=4).value = record['day']
        sheet.cell(row=row_idx, column=5).value = record['quantity']
        sheet.cell(row=row_idx, column=6).value = record['metric']
        sheet.cell(row=row_idx, column=7).value = record['scenario_count']
        sheet.cell(row=row_idx, column=8).value = record['probability_sum']
        sheet.cell(row=row_idx, column=8).number_format = '0.000000'
        for period_idx, value in enumerate(record['values']):
            column_idx = period_idx + 9
            sheet.cell(row=row_idx, column=column_idx).value = value
            sheet.cell(row=row_idx, column=column_idx).number_format = '0.000000'
        maximum_column = planning_problem.num_instants + 9
        sheet.cell(row=row_idx, column=maximum_column).value = record['maximum']
        sheet.cell(row=row_idx, column=maximum_column).number_format = '0.000000'


def _write_interface_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Interface PF')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Operator'
    sheet.cell(row=row_idx, column=3).value = 'Year'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 8).value = p
    row_idx = row_idx + 1

    # TSO's results
    for year in results['tso']:
        for day in results['tso'][year]:
            for node_id in results['tso'][year][day]:
                expected_vmag = [0.0 for _ in range(planning_problem.num_instants)]
                expected_p = [0.0 for _ in range(planning_problem.num_instants)]
                expected_q = [0.0 for _ in range(planning_problem.num_instants)]
                for s_m in results['tso'][year][day][node_id]:
                    omega_m = planning_problem.transmission_network.network[year][day].prob_market_scenarios[s_m]
                    for s_o in results['tso'][year][day][node_id][s_m]:
                        omega_s = planning_problem.transmission_network.network[year][day].prob_operation_scenarios[s_o]

                        # Voltage magnitude
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'TSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Vmag, [p.u.]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            interface_vmag = results['tso'][year][day][node_id][s_m][s_o]['v'][p]
                            sheet.cell(row=row_idx, column=p + 8).value = interface_vmag
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_vmag[p] += interface_vmag * omega_m * omega_s
                        row_idx += 1

                        # Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'TSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            interface_p = results['tso'][year][day][node_id][s_m][s_o]['p'][p]
                            sheet.cell(row=row_idx, column=p + 8).value = interface_p
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_p[p] += interface_p * omega_m * omega_s
                        row_idx += 1

                        # Reactive Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'TSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            interface_q = results['tso'][year][day][node_id][s_m][s_o]['q'][p]
                            sheet.cell(row=row_idx, column=p + 8).value = interface_q
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_q[p] += interface_q * omega_m * omega_s
                        row_idx += 1

                # Expected Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Vmag, [p.u.]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_vmag[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx += 1

                # Expected Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_p[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx += 1

                # Expected Reactive Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_q[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx += 1

    # DSOs' results
    for node_id in results['dso']:
        for year in results['dso'][node_id]:
            for day in results['dso'][node_id][year]:
                expected_vmag = [0.0 for _ in range(planning_problem.num_instants)]
                expected_p = [0.0 for _ in range(planning_problem.num_instants)]
                expected_q = [0.0 for _ in range(planning_problem.num_instants)]
                for s_m in results['dso'][node_id][year][day]:
                    omega_m = planning_problem.distribution_networks[node_id].network[year][day].prob_market_scenarios[s_m]
                    for s_o in results['dso'][node_id][year][day][s_m]:
                        omega_s = planning_problem.distribution_networks[node_id].network[year][day].prob_operation_scenarios[s_o]

                        # Voltage magnitude
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'DSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Vmag, [p.u.]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            interface_vmag = results['dso'][node_id][year][day][s_m][s_o]['v'][p]
                            sheet.cell(row=row_idx, column=p + 8).value = interface_vmag
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_vmag[p] += interface_vmag * omega_m * omega_s
                        row_idx += 1

                        # Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'DSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            interface_p = results['dso'][node_id][year][day][s_m][s_o]['p'][p]
                            sheet.cell(row=row_idx, column=p + 8).value = interface_p
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_p[p] += interface_p * omega_m * omega_s
                        row_idx += 1

                        # Reactive Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'DSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(len(results['dso'][node_id][year][day][s_m][s_o]['q'])):
                            interface_q = results['dso'][node_id][year][day][s_m][s_o]['q'][p]
                            sheet.cell(row=row_idx, column=p + 8).value = interface_q
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_q[p] += interface_q * omega_m * omega_s
                        row_idx += 1

                # Expected Voltage magnitude
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Vmag, [p.u.]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_vmag[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx += 1

                # Expected Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_p[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx += 1

                # Expected Reactive Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_q[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx += 1


def _write_shared_energy_storages_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('SharedESS')

    row_idx = 1
    decimal_style = '0.00'
    percent_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Operator'
    sheet.cell(row=row_idx, column=3).value = 'Year'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 8).value = p

    # TSO's results
    for year in results['tso']['results']:
        for day in results['tso']['results'][year]:

            expected_p = dict()
            expected_q = dict()
            expected_s = dict()
            expected_soc = dict()
            expected_soc_percent = dict()
            for node_id in planning_problem.active_distribution_network_nodes:
                expected_p[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_q[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_s[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc_percent[node_id] = [0.0 for _ in range(planning_problem.num_instants)]

            for s_m in results['tso']['results'][year][day]['scenarios']:

                omega_m = planning_problem.transmission_network.network[year][day].prob_market_scenarios[s_m]

                for s_o in results['tso']['results'][year][day]['scenarios'][s_m]:

                    omega_s = planning_problem.transmission_network.network[year][day].prob_operation_scenarios[s_o]

                    for node_id in planning_problem.active_distribution_network_nodes:

                        # Active power
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'TSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_p = results['tso']['results'][year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['p'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = ess_p
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if ess_p != 'N/A':
                                expected_p[node_id][p] += ess_p * omega_m * omega_s
                            else:
                                expected_p[node_id][p] = ess_p

                        # Reactive power
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'TSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_q = results['tso']['results'][year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['q'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = ess_q
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if ess_q != 'N/A':
                                expected_q[node_id][p] += ess_q * omega_m * omega_s
                            else:
                                expected_q[node_id][p] = ess_q

                        # Apparent power
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'TSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_s = results['tso']['results'][year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['s'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = ess_s
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if ess_s != 'N/A':
                                expected_s[node_id][p] += ess_s * omega_m * omega_s
                            else:
                                expected_s[node_id][p] = ess_s

                        # State-of-Charge, [MVAh]
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'TSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'SoC, [MVAh]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_soc = results['tso']['results'][year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['soc'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = ess_soc
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if ess_soc != 'N/A':
                                expected_soc[node_id][p] += ess_soc * omega_m * omega_s
                            else:
                                expected_soc[node_id][p] = ess_soc

                        # State-of-Charge, [%]
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'TSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_soc_percent = results['tso']['results'][year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['soc_percent'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = ess_soc_percent
                            sheet.cell(row=row_idx, column=p + 8).number_format = percent_style
                            if ess_soc_percent != 'N/A':
                                expected_soc_percent[node_id][p] += ess_soc_percent * omega_m * omega_s
                            else:
                                expected_soc_percent[node_id][p] = ess_soc_percent

            for node_id in planning_problem.active_distribution_network_nodes:

                # Active Power, [MW]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_p[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # Reactive Power, [MVAr]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_q[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # Apparent Power, [MVA]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_s[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # State-of-Charge, [MVAh]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [MVAh]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # State-of-Charge, [%]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'TSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc_percent[node_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = percent_style

    # DSO's results
    for node_id in results['dso']:
        for year in results['dso'][node_id]['results']:
            for day in results['dso'][node_id]['results'][year]:

                distribution_network = planning_problem.distribution_networks[node_id].network[year][day]
                ref_node_id = distribution_network.get_reference_node_id()

                expected_p = [0.0 for _ in range(planning_problem.num_instants)]
                expected_q = [0.0 for _ in range(planning_problem.num_instants)]
                expected_s = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc_percent = [0.0 for _ in range(planning_problem.num_instants)]

                for s_m in results['dso'][node_id]['results'][year][day]['scenarios']:

                    omega_m = distribution_network.prob_market_scenarios[s_m]

                    for s_o in results['dso'][node_id]['results'][year][day]['scenarios'][s_m]:

                        omega_s = distribution_network.prob_operation_scenarios[s_o]

                        # Active power
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'DSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_p = results['dso'][node_id]['results'][year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['p'][ref_node_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = ess_p
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if ess_p != 'N/A':
                                expected_p[p] += ess_p * omega_m * omega_s
                            else:
                                expected_p[p] = ess_p

                        # Reactive power
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'DSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_q = results['dso'][node_id]['results'][year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['q'][ref_node_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = ess_q
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if ess_q != 'N/A':
                                expected_q[p] += ess_q * omega_m * omega_s
                            else:
                                expected_q[p] = ess_q

                        # Apparent power
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'DSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_s = results['dso'][node_id]['results'][year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['s'][ref_node_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = ess_s
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if ess_s != 'N/A':
                                expected_s[p] += ess_s * omega_m * omega_s
                            else:
                                expected_s[p] = ess_s

                        # State-of-Charge, [MVAh]
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'DSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'SoC, [MVAh]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_soc = results['dso'][node_id]['results'][year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['soc'][ref_node_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = ess_soc
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if ess_soc != 'N/A':
                                expected_soc[p] += ess_soc * omega_m * omega_s
                            else:
                                expected_soc[p] = ess_soc

                        # State-of-Charge, [%]
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = 'DSO'
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_soc_percent = results['dso'][node_id]['results'][year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['soc_percent'][ref_node_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = ess_soc_percent
                            sheet.cell(row=row_idx, column=p + 8).number_format = percent_style
                            if ess_soc_percent != 'N/A':
                                expected_soc_percent[p] += ess_soc_percent * omega_m * omega_s
                            else:
                                expected_soc_percent[p] = ess_soc_percent

                # Expected values

                # Active Power, [MW]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_p[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # Reactive Power, [MW]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_q[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # Apparent Power, [MW]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_s[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # State-of-Charge, [MVAh]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [MVAh]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # State-of-Charge, [%]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'DSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc_percent[p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = percent_style

    # ESSO's results
    for year in results['esso']['operation']['aggregated']:
        for day in results['esso']['operation']['aggregated'][year]:
            for node_id in planning_problem.active_distribution_network_nodes:

                # Active power
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'ESSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    ess_p = results['esso']['operation']['aggregated'][year][day][node_id]['p'][p]
                    sheet.cell(row=row_idx, column=p + 8).value = ess_p
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style

                # Reactive power
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = 'ESSO'
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(planning_problem.num_instants):
                    ess_q = results['esso']['operation']['aggregated'][year][day][node_id]['q'][p]
                    sheet.cell(row=row_idx, column=p + 8).value = ess_q
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style


def _write_network_voltage_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Voltage')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Network Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Year'
    sheet.cell(row=row_idx, column=5).value = 'Day'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=8).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 9).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = planning_problem.transmission_network.network
    row_idx = _write_network_voltage_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso']['results'])

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        row_idx = _write_network_voltage_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_voltage_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'

    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    for year in results:
        for day in results[year]:

            ref_node_id = network[year][day].get_reference_node_id()
            expected_vmag = dict()
            expected_vang = dict()
            for node in network[year][day].nodes:
                expected_vmag[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_vang[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['voltage']['vmag']:

                        v_min, v_max = network[year][day].get_node_voltage_limits(node_id)

                        # Voltage magnitude
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = node_id
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Vmag, [p.u.]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network[year][day].num_instants):
                            v_mag = results[year][day]['scenarios'][s_m][s_o]['voltage']['vmag'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 9).value = v_mag
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            if node_id != ref_node_id and (v_mag > v_max + SMALL_TOLERANCE or v_mag < v_min - SMALL_TOLERANCE):
                                sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                            expected_vmag[node_id][p] += v_mag * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Voltage angle
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = node_id
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Vang, [º]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network[year][day].num_instants):
                            v_ang = results[year][day]['scenarios'][s_m][s_o]['voltage']['vang'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 9).value = v_ang
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_vang[node_id][p] += v_ang * omega_m * omega_s
                        row_idx = row_idx + 1

            for node in network[year][day].nodes:

                node_id = node.bus_i
                v_min, v_max = network[year][day].get_node_voltage_limits(node_id)

                # Expected voltage magnitude
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Vmag, [p.u.]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_vmag[node_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    if node_id != ref_node_id and (expected_vmag[node_id][p] > v_max + SMALL_TOLERANCE or expected_vmag[node_id][p] < v_min - SMALL_TOLERANCE):
                        sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                row_idx = row_idx + 1

                # Expected voltage angle
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = node_id
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Vang, [º]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_vang[node_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

    return row_idx


def _write_network_consumption_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Consumption')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Load ID'
    sheet.cell(row=row_idx, column=4).value = 'Node ID'
    sheet.cell(row=row_idx, column=5).value = 'Year'
    sheet.cell(row=row_idx, column=6).value = 'Day'
    sheet.cell(row=row_idx, column=7).value = 'Quantity'
    sheet.cell(row=row_idx, column=8).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=9).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 10).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    tso_results = results['tso']['results']
    transmission_network = planning_problem.transmission_network.network
    tn_params = planning_problem.transmission_network.params
    row_idx = _write_network_consumption_results_per_operator(transmission_network, tn_params, sheet, 'TSO', row_idx, tso_results)

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        dn_params = planning_problem.distribution_networks[tn_node_id].params
        row_idx = _write_network_consumption_results_per_operator(distribution_network, dn_params, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_consumption_results_per_operator(network, params, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    for year in results:
        for day in results[year]:

            expected_pc = dict()
            expected_pc_flex = dict()
            expected_pc_curt = dict()
            expected_pnet = dict()
            expected_qc = dict()
            expected_qc_flex = dict()
            expected_qc_curt = dict()
            expected_qnet = dict()
            for load in network[year][day].loads:
                expected_pc[load.load_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pc_flex[load.load_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pc_curt[load.load_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pnet[load.load_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_qc[load.load_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_qc_flex[load.load_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_qc_curt[load.load_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_qnet[load.load_id] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for load in network[year][day].loads:

                        # - Active Power
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = load.load_id
                        sheet.cell(row=row_idx, column=4).value = load.bus
                        sheet.cell(row=row_idx, column=5).value = int(year)
                        sheet.cell(row=row_idx, column=6).value = day
                        sheet.cell(row=row_idx, column=7).value = 'Pc, [MW]'
                        sheet.cell(row=row_idx, column=8).value = s_m
                        sheet.cell(row=row_idx, column=9).value = s_o
                        for p in range(network[year][day].num_instants):
                            pc = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc'][load.load_id][p]
                            sheet.cell(row=row_idx, column=p + 10).value = pc
                            sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                            expected_pc[load.load_id][p] += pc * omega_m * omega_s
                        row_idx = row_idx + 1

                        if params.fl_reg:

                            # - Flexibility, Pc
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = load.load_id
                            sheet.cell(row=row_idx, column=4).value = load.bus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'Pc_flex, [MW]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                pc_flex = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc_flex'][load.load_id][p]
                                sheet.cell(row=row_idx, column=p + 10).value = pc_flex
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                expected_pc_flex[load.load_id][p] += pc_flex * omega_m * omega_s
                            row_idx = row_idx + 1

                        if params.l_curt:

                            # - Active power curtailment
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = load.load_id
                            sheet.cell(row=row_idx, column=4).value = load.bus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'Pc_curt, [MW]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                pc_curt = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc_curt'][load.load_id][p]
                                sheet.cell(row=row_idx, column=p + 10).value = pc_curt
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                if pc_curt >= SMALL_TOLERANCE:
                                    sheet.cell(row=row_idx, column=p + 10).fill = violation_fill
                                expected_pc_curt[load.load_id][p] += pc_curt * omega_m * omega_s
                            row_idx = row_idx + 1

                        if params.fl_reg or params.l_curt:

                            # - Active power net consumption
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = load.load_id
                            sheet.cell(row=row_idx, column=4).value = load.bus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'Pc_net, [MW]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                p_net = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc_net'][load.load_id][p]
                                sheet.cell(row=row_idx, column=p + 10).value = p_net
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                expected_pnet[load.load_id][p] += p_net * omega_m * omega_s
                            row_idx = row_idx + 1

                        # - Reactive power
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = load.load_id
                        sheet.cell(row=row_idx, column=4).value = load.bus
                        sheet.cell(row=row_idx, column=5).value = int(year)
                        sheet.cell(row=row_idx, column=6).value = day
                        sheet.cell(row=row_idx, column=7).value = 'Qc, [MVAr]'
                        sheet.cell(row=row_idx, column=8).value = s_m
                        sheet.cell(row=row_idx, column=9).value = s_o
                        for p in range(network[year][day].num_instants):
                            qc = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc'][load.load_id][p]
                            sheet.cell(row=row_idx, column=p + 10).value = qc
                            sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                            expected_qc[load.load_id][p] += qc * omega_m * omega_s
                        row_idx = row_idx + 1

                        if params.fl_reg:

                            # - Flexibility, Qc
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = load.load_id
                            sheet.cell(row=row_idx, column=4).value = load.bus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'Qc_flex, [MVAr]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                qc_flex = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc_flex'][load.load_id][p]
                                sheet.cell(row=row_idx, column=p + 10).value = qc_flex
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                expected_qc_flex[load.load_id][p] += qc_flex * omega_m * omega_s
                            row_idx = row_idx + 1

                        if params.l_curt:

                            # - Reactive power curtailment
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = load.load_id
                            sheet.cell(row=row_idx, column=4).value = load.bus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'Qc_curt, [MVAr]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                qc_curt = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc_curt'][load.load_id][p]
                                sheet.cell(row=row_idx, column=p + 10).value = qc_curt
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                if qc_curt >= SMALL_TOLERANCE:
                                    sheet.cell(row=row_idx, column=p + 10).fill = violation_fill
                                expected_qc_curt[load.load_id][p] += qc_curt * omega_m * omega_s
                            row_idx = row_idx + 1

                        if params.fl_reg or params.l_curt:

                            # - Reactive power net consumption
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = load.load_id
                            sheet.cell(row=row_idx, column=4).value = load.bus
                            sheet.cell(row=row_idx, column=5).value = int(year)
                            sheet.cell(row=row_idx, column=6).value = day
                            sheet.cell(row=row_idx, column=7).value = 'Qc_net, [MVAr]'
                            sheet.cell(row=row_idx, column=8).value = s_m
                            sheet.cell(row=row_idx, column=9).value = s_o
                            for p in range(network[year][day].num_instants):
                                q_net = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc_net'][load.load_id][p]
                                sheet.cell(row=row_idx, column=p + 10).value = q_net
                                sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                                expected_qnet[load.load_id][p] += q_net * omega_m * omega_s
                            row_idx = row_idx + 1

            for load in network[year][day].loads:

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = load.load_id
                sheet.cell(row=row_idx, column=4).value = load.bus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'Pc, [MW]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_pc[load.load_id][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                if params.fl_reg:

                    # - Flexibility, active power
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = load.load_id
                    sheet.cell(row=row_idx, column=4).value = load.bus
                    sheet.cell(row=row_idx, column=5).value = int(year)
                    sheet.cell(row=row_idx, column=6).value = day
                    sheet.cell(row=row_idx, column=7).value = 'Pc_flex, [MW]'
                    sheet.cell(row=row_idx, column=8).value = 'Expected'
                    sheet.cell(row=row_idx, column=9).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 10).value = expected_pc_flex[load.load_id][p]
                        sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                    row_idx = row_idx + 1

                if params.l_curt:

                    # - Load curtailment (active power)
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = load.load_id
                    sheet.cell(row=row_idx, column=4).value = load.bus
                    sheet.cell(row=row_idx, column=5).value = int(year)
                    sheet.cell(row=row_idx, column=6).value = day
                    sheet.cell(row=row_idx, column=7).value = 'Pc_curt, [MW]'
                    sheet.cell(row=row_idx, column=8).value = 'Expected'
                    sheet.cell(row=row_idx, column=9).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 10).value = expected_pc_curt[load.load_id][p]
                        sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                        if expected_pc_curt[load.load_id][p] >= SMALL_TOLERANCE:
                            sheet.cell(row=row_idx, column=p + 10).fill = violation_fill
                    row_idx = row_idx + 1

                if params.fl_reg or params.l_curt:

                    # - Active power net consumption
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = load.load_id
                    sheet.cell(row=row_idx, column=4).value = load.bus
                    sheet.cell(row=row_idx, column=5).value = int(year)
                    sheet.cell(row=row_idx, column=6).value = day
                    sheet.cell(row=row_idx, column=7).value = 'Pc_net, [MW]'
                    sheet.cell(row=row_idx, column=8).value = 'Expected'
                    sheet.cell(row=row_idx, column=9).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 10).value = expected_pnet[load.load_id][p]
                        sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                    row_idx = row_idx + 1

                # - Reactive power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = load.load_id
                sheet.cell(row=row_idx, column=4).value = load.bus
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'Qc, [MVAr]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_qc[load.load_id][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                if params.fl_reg:

                    # - Flexibility, reactive power
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = load.load_id
                    sheet.cell(row=row_idx, column=4).value = load.bus
                    sheet.cell(row=row_idx, column=5).value = int(year)
                    sheet.cell(row=row_idx, column=6).value = day
                    sheet.cell(row=row_idx, column=7).value = 'Qc_flex, [MVAr]'
                    sheet.cell(row=row_idx, column=8).value = 'Expected'
                    sheet.cell(row=row_idx, column=9).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 10).value = expected_qc_flex[load.load_id][p]
                        sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                    row_idx = row_idx + 1

                if params.l_curt:

                    # - Load curtailment (reactive power)
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = load.load_id
                    sheet.cell(row=row_idx, column=4).value = load.bus
                    sheet.cell(row=row_idx, column=5).value = int(year)
                    sheet.cell(row=row_idx, column=6).value = day
                    sheet.cell(row=row_idx, column=7).value = 'Qc_curt, [MVAr]'
                    sheet.cell(row=row_idx, column=8).value = 'Expected'
                    sheet.cell(row=row_idx, column=9).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 10).value = expected_qc_curt[load.load_id][p]
                        sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                        if expected_qc_curt[load.load_id][p] >= SMALL_TOLERANCE:
                            sheet.cell(row=row_idx, column=p + 10).fill = violation_fill
                    row_idx = row_idx + 1

                if params.fl_reg or params.l_curt:

                    # - Reactive power net consumption
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = load.load_id
                    sheet.cell(row=row_idx, column=4).value = load.bus
                    sheet.cell(row=row_idx, column=5).value = int(year)
                    sheet.cell(row=row_idx, column=6).value = day
                    sheet.cell(row=row_idx, column=7).value = 'Qc_net, [MVAr]'
                    sheet.cell(row=row_idx, column=8).value = 'Expected'
                    sheet.cell(row=row_idx, column=9).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 10).value = expected_qnet[load.load_id][p]
                        sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                    row_idx = row_idx + 1

    return row_idx


def _write_network_generation_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Generation')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Generator ID'
    sheet.cell(row=row_idx, column=4).value = 'Node ID'
    sheet.cell(row=row_idx, column=5).value = 'Type'
    sheet.cell(row=row_idx, column=6).value = 'Year'
    sheet.cell(row=row_idx, column=7).value = 'Day'
    sheet.cell(row=row_idx, column=8).value = 'Quantity'
    sheet.cell(row=row_idx, column=9).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=10).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 11).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = planning_problem.transmission_network.network
    tn_params = planning_problem.transmission_network.params
    row_idx = _write_network_generation_results_per_operator(transmission_network, tn_params, sheet, 'TSO', row_idx, results['tso']['results'])

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        dn_params = planning_problem.distribution_networks[tn_node_id].params
        row_idx = _write_network_generation_results_per_operator(distribution_network, dn_params, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_generation_results_per_operator(network, params, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    for year in results:
        for day in results[year]:

            expected_pg = dict()
            expected_pg_net = dict()
            expected_qg = dict()
            expected_qg_net = dict()
            expected_sg = dict()
            expected_sg_curt = dict()
            expected_sg_net = dict()
            for generator in network[year][day].generators:
                expected_pg[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_qg[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_sg[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pg_net[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_qg_net[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_sg_curt[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_sg_net[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for generator in network[year][day].generators:

                        node_id = generator.bus
                        gen_id = generator.gen_id
                        gen_type = network[year][day].get_gen_type(gen_id)

                        # Active Power
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = gen_id
                        sheet.cell(row=row_idx, column=4).value = node_id
                        sheet.cell(row=row_idx, column=5).value = gen_type
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'Pg, [MW]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            pg = results[year][day]['scenarios'][s_m][s_o]['generation']['pg'][gen_id][p]
                            sheet.cell(row=row_idx, column=p + 11).value = pg
                            sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                            expected_pg[gen_id][p] += pg * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Active Power net
                        if generator.is_curtaillable() and params.rg_curt:

                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = gen_id
                            sheet.cell(row=row_idx, column=4).value = node_id
                            sheet.cell(row=row_idx, column=5).value = gen_type
                            sheet.cell(row=row_idx, column=6).value = int(year)
                            sheet.cell(row=row_idx, column=7).value = day
                            sheet.cell(row=row_idx, column=8).value = 'Pg_net, [MW]'
                            sheet.cell(row=row_idx, column=9).value = s_m
                            sheet.cell(row=row_idx, column=10).value = s_o
                            for p in range(network[year][day].num_instants):
                                pg_net = results[year][day]['scenarios'][s_m][s_o]['generation']['pg_net'][gen_id][p]
                                sheet.cell(row=row_idx, column=p + 11).value = pg_net
                                sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                                expected_pg_net[gen_id][p] += pg_net * omega_m * omega_s
                            row_idx = row_idx + 1

                        # Reactive Power
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = gen_id
                        sheet.cell(row=row_idx, column=4).value = node_id
                        sheet.cell(row=row_idx, column=5).value = gen_type
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'Qg, [MVAr]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            qg = results[year][day]['scenarios'][s_m][s_o]['generation']['qg'][gen_id][p]
                            sheet.cell(row=row_idx, column=p + 11).value = qg
                            sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                            expected_qg[gen_id][p] += qg * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Reactive Power net
                        if generator.is_curtaillable() and params.rg_curt:

                            # Reactive Power net
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = gen_id
                            sheet.cell(row=row_idx, column=4).value = node_id
                            sheet.cell(row=row_idx, column=5).value = gen_type
                            sheet.cell(row=row_idx, column=6).value = int(year)
                            sheet.cell(row=row_idx, column=7).value = day
                            sheet.cell(row=row_idx, column=8).value = 'Qg_net, [MVAr]'
                            sheet.cell(row=row_idx, column=9).value = s_m
                            sheet.cell(row=row_idx, column=10).value = s_o
                            for p in range(network[year][day].num_instants):
                                qg_net = results[year][day]['scenarios'][s_m][s_o]['generation']['qg_net'][gen_id][p]
                                sheet.cell(row=row_idx, column=p + 11).value = qg_net
                                sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                                expected_qg_net[gen_id][p] += qg_net * omega_m * omega_s
                            row_idx = row_idx + 1

                        # Apparent Power
                        if generator.is_curtaillable() and params.rg_curt:

                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = gen_id
                            sheet.cell(row=row_idx, column=4).value = node_id
                            sheet.cell(row=row_idx, column=5).value = gen_type
                            sheet.cell(row=row_idx, column=6).value = int(year)
                            sheet.cell(row=row_idx, column=7).value = day
                            sheet.cell(row=row_idx, column=8).value = 'Sg, [MVA]'
                            sheet.cell(row=row_idx, column=9).value = s_m
                            sheet.cell(row=row_idx, column=10).value = s_o
                            for p in range(network[year][day].num_instants):
                                sg = results[year][day]['scenarios'][s_m][s_o]['generation']['sg'][gen_id][p]
                                sheet.cell(row=row_idx, column=p + 11).value = sg
                                sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                                expected_sg[gen_id][p] += sg * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Apparent Power curtailment
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = gen_id
                            sheet.cell(row=row_idx, column=4).value = node_id
                            sheet.cell(row=row_idx, column=5).value = gen_type
                            sheet.cell(row=row_idx, column=6).value = int(year)
                            sheet.cell(row=row_idx, column=7).value = day
                            sheet.cell(row=row_idx, column=8).value = 'Sg_curt, [MVA]'
                            sheet.cell(row=row_idx, column=9).value = s_m
                            sheet.cell(row=row_idx, column=10).value = s_o
                            for p in range(network[year][day].num_instants):
                                sg_curt = results[year][day]['scenarios'][s_m][s_o]['generation']['sg_curt'][gen_id][p]
                                sheet.cell(row=row_idx, column=p + 11).value = sg_curt
                                sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                                if not isclose(sg_curt, 0.00, abs_tol=VIOLATION_TOLERANCE):
                                    sheet.cell(row=row_idx, column=p + 11).fill = violation_fill
                                expected_sg_curt[gen_id][p] += sg_curt * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Apparent Power Net
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = gen_id
                            sheet.cell(row=row_idx, column=4).value = node_id
                            sheet.cell(row=row_idx, column=5).value = gen_type
                            sheet.cell(row=row_idx, column=6).value = int(year)
                            sheet.cell(row=row_idx, column=7).value = day
                            sheet.cell(row=row_idx, column=8).value = 'Sg_net, [MVA]'
                            sheet.cell(row=row_idx, column=9).value = s_m
                            sheet.cell(row=row_idx, column=10).value = s_o
                            for p in range(network[year][day].num_instants):
                                sg_net = results[year][day]['scenarios'][s_m][s_o]['generation']['sg_net'][gen_id][p]
                                sheet.cell(row=row_idx, column=p + 11).value = sg_net
                                sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                                expected_sg_net[gen_id][p] += sg_net * omega_m * omega_s
                            row_idx = row_idx + 1

            for generator in network[year][day].generators:

                node_id = generator.bus
                gen_id = generator.gen_id
                gen_type = network[year][day].get_gen_type(gen_id)

                # Active Power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = gen_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = gen_type
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'Pg, [MW]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = expected_pg[gen_id][p]
                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                row_idx = row_idx + 1

                # Active Power Net
                if generator.is_curtaillable() and params.rg_curt:

                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = gen_id
                    sheet.cell(row=row_idx, column=4).value = node_id
                    sheet.cell(row=row_idx, column=5).value = gen_type
                    sheet.cell(row=row_idx, column=6).value = int(year)
                    sheet.cell(row=row_idx, column=7).value = day
                    sheet.cell(row=row_idx, column=8).value = 'Pg_net, [MW]'
                    sheet.cell(row=row_idx, column=9).value = 'Expected'
                    sheet.cell(row=row_idx, column=10).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 11).value = expected_pg_net[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                    row_idx = row_idx + 1

                # Reactive Power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = gen_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = gen_type
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'Qg, [MVAr]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = expected_qg[gen_id][p]
                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                row_idx = row_idx + 1

                # Reactive Power net
                if generator.is_curtaillable() and params.rg_curt:

                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = gen_id
                    sheet.cell(row=row_idx, column=4).value = node_id
                    sheet.cell(row=row_idx, column=5).value = gen_type
                    sheet.cell(row=row_idx, column=6).value = int(year)
                    sheet.cell(row=row_idx, column=7).value = day
                    sheet.cell(row=row_idx, column=8).value = 'Qg_net, [MVAr]'
                    sheet.cell(row=row_idx, column=9).value = 'Expected'
                    sheet.cell(row=row_idx, column=10).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 11).value = expected_qg_net[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                    row_idx = row_idx + 1

                # Apparent Power
                if generator.is_curtaillable() and params.rg_curt:

                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = gen_id
                    sheet.cell(row=row_idx, column=4).value = node_id
                    sheet.cell(row=row_idx, column=5).value = gen_type
                    sheet.cell(row=row_idx, column=6).value = int(year)
                    sheet.cell(row=row_idx, column=7).value = day
                    sheet.cell(row=row_idx, column=8).value = 'Sg, [MVA]'
                    sheet.cell(row=row_idx, column=9).value = 'Expected'
                    sheet.cell(row=row_idx, column=10).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 11).value = expected_sg[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                    row_idx = row_idx + 1

                    # Apparent Power curtailment
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = gen_id
                    sheet.cell(row=row_idx, column=4).value = node_id
                    sheet.cell(row=row_idx, column=5).value = gen_type
                    sheet.cell(row=row_idx, column=6).value = int(year)
                    sheet.cell(row=row_idx, column=7).value = day
                    sheet.cell(row=row_idx, column=8).value = 'Sg_curt, [MVA]'
                    sheet.cell(row=row_idx, column=9).value = 'Expected'
                    sheet.cell(row=row_idx, column=10).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 11).value = expected_sg_curt[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                        if not isclose(expected_sg_curt[gen_id][p], 0.00, abs_tol=VIOLATION_TOLERANCE):
                            sheet.cell(row=row_idx, column=p + 11).fill = violation_fill
                    row_idx = row_idx + 1

                    # Apparent Power Net
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = gen_id
                    sheet.cell(row=row_idx, column=4).value = node_id
                    sheet.cell(row=row_idx, column=5).value = gen_type
                    sheet.cell(row=row_idx, column=6).value = int(year)
                    sheet.cell(row=row_idx, column=7).value = day
                    sheet.cell(row=row_idx, column=8).value = 'Sg_net, [MVA]'
                    sheet.cell(row=row_idx, column=9).value = 'Expected'
                    sheet.cell(row=row_idx, column=10).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 11).value = expected_sg_net[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                    row_idx = row_idx + 1

    return row_idx


def _write_network_branch_results_to_excel(planning_problem, workbook, results, result_type):

    sheet_name = str()
    if result_type == 'losses':
        sheet_name = 'Branch Losses'
    elif result_type == 'ratio':
        sheet_name = 'Transformer Ratio'

    sheet = workbook.create_sheet(sheet_name)

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Branch ID'
    sheet.cell(row=row_idx, column=4).value = 'From Node ID'
    sheet.cell(row=row_idx, column=5).value = 'To Node ID'
    sheet.cell(row=row_idx, column=6).value = 'Year'
    sheet.cell(row=row_idx, column=7).value = 'Day'
    sheet.cell(row=row_idx, column=8).value = 'Quantity'
    sheet.cell(row=row_idx, column=9).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=10).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 11).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = planning_problem.transmission_network.network
    row_idx = _write_network_branch_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso']['results'], result_type)

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        row_idx = _write_network_branch_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, result_type, tn_node_id=tn_node_id)


def _write_network_branch_results_per_operator(network, sheet, operator_type, row_idx, results, result_type, tn_node_id='-'):

    decimal_style = '0.00'

    aux_string = str()
    if result_type == 'losses':
        aux_string = 'P, [MW]'
    elif result_type == 'ratio':
        aux_string = 'Ratio'

    for year in results:
        for day in results[year]:

            expected_values = dict()
            for branch in network[year][day].branches:
                expected_values[branch.branch_id] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for branch in network[year][day].branches:

                        branch_id = branch.branch_id

                        if not(result_type == 'ratio' and not branch.is_transformer):

                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.branch_id
                            sheet.cell(row=row_idx, column=4).value = branch.fbus
                            sheet.cell(row=row_idx, column=5).value = branch.tbus
                            sheet.cell(row=row_idx, column=6).value = int(year)
                            sheet.cell(row=row_idx, column=7).value = day
                            sheet.cell(row=row_idx, column=8).value = aux_string
                            sheet.cell(row=row_idx, column=9).value = s_m
                            sheet.cell(row=row_idx, column=10).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = results[year][day]['scenarios'][s_m][s_o]['branches'][result_type][branch_id][p]
                                sheet.cell(row=row_idx, column=p + 11).value = value
                                sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                                expected_values[branch_id][p] += value * omega_m * omega_s
                            row_idx = row_idx + 1

            for branch in network[year][day].branches:
                branch_id = branch.branch_id
                if not (result_type == 'ratio' and not branch.is_transformer):

                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = branch.branch_id
                    sheet.cell(row=row_idx, column=4).value = branch.fbus
                    sheet.cell(row=row_idx, column=5).value = branch.tbus
                    sheet.cell(row=row_idx, column=6).value = int(year)
                    sheet.cell(row=row_idx, column=7).value = day
                    sheet.cell(row=row_idx, column=8).value = aux_string
                    sheet.cell(row=row_idx, column=9).value = 'Expected'
                    sheet.cell(row=row_idx, column=10).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 11).value = expected_values[branch_id][p]
                        sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                    row_idx = row_idx + 1

    return row_idx


def _write_network_branch_loading_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Branch Loading')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Branch ID'
    sheet.cell(row=row_idx, column=4).value = 'From Node ID'
    sheet.cell(row=row_idx, column=5).value = 'To Node ID'
    sheet.cell(row=row_idx, column=6).value = 'Year'
    sheet.cell(row=row_idx, column=7).value = 'Day'
    sheet.cell(row=row_idx, column=8).value = 'Quantity'
    sheet.cell(row=row_idx, column=9).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=10).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 11).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = planning_problem.transmission_network.network
    row_idx = _write_network_branch_loading_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso']['results'])

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        row_idx = _write_network_branch_loading_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_branch_loading_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    perc_style = '0.00%'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    for year in results:
        for day in results[year]:

            expected_values = {'flow_ij_perc': {}, 'flow_ji_perc': {}}
            for branch in network[year][day].branches:
                expected_values['flow_ij_perc'][branch.branch_id] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for branch in network[year][day].branches:
                        branch_loading = results[year][day]['scenarios'][s_m][s_o]['branches']['branch_flow']
                        directions = (
                            ('flow_ij_perc', 'Flow_ij, [%]', branch.fbus, branch.tbus),
                            ('flow_ji_perc', 'Flow_ji, [%]', branch.tbus, branch.fbus),
                        )
                        for quantity, label, fbus, tbus in directions:
                            if branch.branch_id not in branch_loading[quantity]:
                                continue
                            expected_values[quantity].setdefault(
                                branch.branch_id,
                                [0.0 for _ in range(network[year][day].num_instants)],
                            )
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = branch.branch_id
                            sheet.cell(row=row_idx, column=4).value = fbus
                            sheet.cell(row=row_idx, column=5).value = tbus
                            sheet.cell(row=row_idx, column=6).value = int(year)
                            sheet.cell(row=row_idx, column=7).value = day
                            sheet.cell(row=row_idx, column=8).value = label
                            sheet.cell(row=row_idx, column=9).value = s_m
                            sheet.cell(row=row_idx, column=10).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = branch_loading[quantity][branch.branch_id][p]
                                sheet.cell(row=row_idx, column=p + 11).value = value
                                sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                                if value > 1.00 + VIOLATION_TOLERANCE:
                                    sheet.cell(row=row_idx, column=p + 11).fill = violation_fill
                                expected_values[quantity][branch.branch_id][p] += value * omega_m * omega_s
                            row_idx = row_idx + 1

            for branch in network[year][day].branches:
                directions = (
                    ('flow_ij_perc', 'Flow_ij, [%]', branch.fbus, branch.tbus),
                    ('flow_ji_perc', 'Flow_ji, [%]', branch.tbus, branch.fbus),
                )
                for quantity, label, fbus, tbus in directions:
                    if branch.branch_id not in expected_values[quantity]:
                        continue
                    sheet.cell(row=row_idx, column=1).value = operator_type
                    sheet.cell(row=row_idx, column=2).value = tn_node_id
                    sheet.cell(row=row_idx, column=3).value = branch.branch_id
                    sheet.cell(row=row_idx, column=4).value = fbus
                    sheet.cell(row=row_idx, column=5).value = tbus
                    sheet.cell(row=row_idx, column=6).value = int(year)
                    sheet.cell(row=row_idx, column=7).value = day
                    sheet.cell(row=row_idx, column=8).value = label
                    sheet.cell(row=row_idx, column=9).value = 'Expected'
                    sheet.cell(row=row_idx, column=10).value = '-'
                    for p in range(network[year][day].num_instants):
                        value = expected_values[quantity][branch.branch_id][p]
                        sheet.cell(row=row_idx, column=p + 11).value = value
                        sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                        if value > 1.00 + VIOLATION_TOLERANCE:
                            sheet.cell(row=row_idx, column=p + 11).fill = violation_fill
                    row_idx = row_idx + 1

    return row_idx


def _write_network_branch_power_flow_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Power Flows')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Branch ID'
    sheet.cell(row=row_idx, column=4).value = 'From Node ID'
    sheet.cell(row=row_idx, column=5).value = 'To Node ID'
    sheet.cell(row=row_idx, column=6).value = 'Year'
    sheet.cell(row=row_idx, column=7).value = 'Day'
    sheet.cell(row=row_idx, column=8).value = 'Quantity'
    sheet.cell(row=row_idx, column=9).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=10).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 11).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    transmission_network = planning_problem.transmission_network.network
    row_idx = _write_network_power_flow_results_per_operator(transmission_network, sheet, 'TSO', row_idx, results['tso']['results'])

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        row_idx = _write_network_power_flow_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_power_flow_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    perc_style = '0.00%'

    for year in results:
        for day in results[year]:

            expected_values = {'pij': {}, 'pji': {}, 'qij': {}, 'qji': {}, 'sij': {}, 'sji': {}}
            for branch in network[year][day].branches:
                expected_values['pij'][branch.branch_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['pji'][branch.branch_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['qij'][branch.branch_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['qji'][branch.branch_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['sij'][branch.branch_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['sji'][branch.branch_id] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for branch in network[year][day].branches:

                        branch_id = branch.branch_id
                        rating = branch.rate
                        if rating == 0.0:
                            rating = BRANCH_UNKNOWN_RATING

                        # Pij, [MW]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.fbus
                        sheet.cell(row=row_idx, column=5).value = branch.tbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pij'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                            expected_values['pij'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Pij, [%]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.fbus
                        sheet.cell(row=row_idx, column=5).value = branch.tbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'P, [%]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pij'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                        row_idx = row_idx + 1

                        # Pji, [MW]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.tbus
                        sheet.cell(row=row_idx, column=5).value = branch.fbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pji'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                            expected_values['pji'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Pji, [%]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.tbus
                        sheet.cell(row=row_idx, column=5).value = branch.fbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'P, [%]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pji'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                        row_idx = row_idx + 1

                        # Qij, [MVAr]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.fbus
                        sheet.cell(row=row_idx, column=5).value = branch.tbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qij'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                            expected_values['qij'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Qij, [%]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.fbus
                        sheet.cell(row=row_idx, column=5).value = branch.tbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'Q, [%]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qij'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                        row_idx = row_idx + 1

                        # Qji, [MW]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.tbus
                        sheet.cell(row=row_idx, column=5).value = branch.fbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qji'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                            expected_values['qji'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Qji, [%]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.tbus
                        sheet.cell(row=row_idx, column=5).value = branch.fbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'Q, [%]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qji'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                        row_idx = row_idx + 1

                        # Sij, [MVA]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.fbus
                        sheet.cell(row=row_idx, column=5).value = branch.tbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sij'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                            expected_values['sij'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Sij, [%]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.fbus
                        sheet.cell(row=row_idx, column=5).value = branch.tbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'S, [%]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sij'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                        row_idx = row_idx + 1

                        # Sji, [MW]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.tbus
                        sheet.cell(row=row_idx, column=5).value = branch.fbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sji'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                            expected_values['sji'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Sji, [%]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = branch.branch_id
                        sheet.cell(row=row_idx, column=4).value = branch.tbus
                        sheet.cell(row=row_idx, column=5).value = branch.fbus
                        sheet.cell(row=row_idx, column=6).value = int(year)
                        sheet.cell(row=row_idx, column=7).value = day
                        sheet.cell(row=row_idx, column=8).value = 'S, [%]'
                        sheet.cell(row=row_idx, column=9).value = s_m
                        sheet.cell(row=row_idx, column=10).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sji'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 11).value = value
                            sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                        row_idx = row_idx + 1

            for branch in network[year][day].branches:

                branch_id = branch.branch_id
                rating = branch.rate
                if rating == 0.0:
                    rating = BRANCH_UNKNOWN_RATING

                # Pij, [MW]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = branch.tbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = expected_values['pij'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                row_idx = row_idx + 1

                # Pij, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = branch.tbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'P, [%]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = abs(expected_values['pij'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                row_idx = row_idx + 1

                # Pji, [MW]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = branch.fbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = expected_values['pji'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                row_idx = row_idx + 1

                # Pji, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = branch.fbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'P, [%]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = abs(expected_values['pji'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                row_idx = row_idx + 1

                # Qij, [MVAr]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = branch.tbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = expected_values['qij'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                row_idx = row_idx + 1

                # Qij, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = branch.tbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'Q, [%]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = abs(expected_values['qij'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                row_idx = row_idx + 1

                # Qji, [MVAr]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = branch.fbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = expected_values['qji'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                row_idx = row_idx + 1

                # Qji, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = branch.fbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'Q, [%]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = abs(expected_values['qji'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                row_idx = row_idx + 1

                # Sij, [MVA]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = branch.tbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = expected_values['sij'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                row_idx = row_idx + 1

                # Sij, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.fbus
                sheet.cell(row=row_idx, column=5).value = branch.tbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'S, [%]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = abs(expected_values['sij'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                row_idx = row_idx + 1

                # Sji, [MVA]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = branch.fbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = expected_values['sji'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 11).number_format = decimal_style
                row_idx = row_idx + 1

                # Sji, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = branch.branch_id
                sheet.cell(row=row_idx, column=4).value = branch.tbus
                sheet.cell(row=row_idx, column=5).value = branch.fbus
                sheet.cell(row=row_idx, column=6).value = int(year)
                sheet.cell(row=row_idx, column=7).value = day
                sheet.cell(row=row_idx, column=8).value = 'S, [%]'
                sheet.cell(row=row_idx, column=9).value = 'Expected'
                sheet.cell(row=row_idx, column=10).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 11).value = abs(expected_values['sji'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 11).number_format = perc_style
                row_idx = row_idx + 1

    return row_idx


def _write_network_energy_storages_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Energy Storage')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'ESS ID'
    sheet.cell(row=row_idx, column=4).value = 'Node ID'
    sheet.cell(row=row_idx, column=5).value = 'Year'
    sheet.cell(row=row_idx, column=6).value = 'Day'
    sheet.cell(row=row_idx, column=7).value = 'Quantity'
    sheet.cell(row=row_idx, column=8).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=9).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 10).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    tso_results = results['tso']['results']
    transmission_network = planning_problem.transmission_network.network
    if planning_problem.transmission_network.params.es_reg:
        row_idx = _write_network_energy_storages_results_per_operator(transmission_network, sheet, 'TSO', row_idx, tso_results)

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        if planning_problem.distribution_networks[tn_node_id].params.es_reg:
            row_idx = _write_network_energy_storages_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, tn_node_id=tn_node_id)


def _write_network_energy_storages_results_per_operator(network, sheet, operator_type, row_idx, results, tn_node_id='-'):

    decimal_style = '0.00'
    percent_style = '0.00%'

    for year in results:
        for day in results[year]:

            expected_p = dict()
            expected_q = dict()
            expected_s = dict()
            expected_soc = dict()
            expected_soc_percent = dict()
            for energy_storage in network[year][day].energy_storages:
                es_id = energy_storage.es_id
                expected_p[es_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_q[es_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_s[es_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_soc[es_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_soc_percent[es_id] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for energy_storage in network[year][day].energy_storages:

                        es_id = energy_storage.es_id
                        node_id = energy_storage.bus

                        # - Active Power
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = es_id
                        sheet.cell(row=row_idx, column=4).value = node_id
                        sheet.cell(row=row_idx, column=5).value = int(year)
                        sheet.cell(row=row_idx, column=6).value = day
                        sheet.cell(row=row_idx, column=7).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=8).value = s_m
                        sheet.cell(row=row_idx, column=9).value = s_o
                        for p in range(network[year][day].num_instants):
                            ess_p = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['p'][es_id][p]
                            sheet.cell(row=row_idx, column=p + 10).value = ess_p
                            sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                            expected_p[es_id][p] += ess_p * omega_m * omega_s
                        row_idx = row_idx + 1

                        # - Reactive Power
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = es_id
                        sheet.cell(row=row_idx, column=4).value = node_id
                        sheet.cell(row=row_idx, column=5).value = int(year)
                        sheet.cell(row=row_idx, column=6).value = day
                        sheet.cell(row=row_idx, column=7).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=8).value = s_m
                        sheet.cell(row=row_idx, column=9).value = s_o
                        for p in range(network[year][day].num_instants):
                            ess_q = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['q'][es_id][p]
                            sheet.cell(row=row_idx, column=p + 10).value = ess_q
                            sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                            expected_q[es_id][p] += ess_q * omega_m * omega_s
                        row_idx = row_idx + 1

                        # - Apparent Power
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = es_id
                        sheet.cell(row=row_idx, column=4).value = node_id
                        sheet.cell(row=row_idx, column=5).value = int(year)
                        sheet.cell(row=row_idx, column=6).value = day
                        sheet.cell(row=row_idx, column=7).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=8).value = s_m
                        sheet.cell(row=row_idx, column=9).value = s_o
                        for p in range(network[year][day].num_instants):
                            ess_s = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['s'][es_id][p]
                            sheet.cell(row=row_idx, column=p + 10).value = ess_s
                            sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                            expected_s[es_id][p] += ess_s * omega_m * omega_s
                        row_idx = row_idx + 1

                        # State-of-Charge, [MWh]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = es_id
                        sheet.cell(row=row_idx, column=4).value = node_id
                        sheet.cell(row=row_idx, column=5).value = int(year)
                        sheet.cell(row=row_idx, column=6).value = day
                        sheet.cell(row=row_idx, column=7).value = 'SoC, [MWh]'
                        sheet.cell(row=row_idx, column=8).value = s_m
                        sheet.cell(row=row_idx, column=9).value = s_o
                        for p in range(network[year][day].num_instants):
                            ess_soc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['soc'][es_id][p]
                            sheet.cell(row=row_idx, column=p + 10).value = ess_soc
                            sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                            if ess_soc != 'N/A':
                                expected_soc[es_id][p] += ess_soc * omega_m * omega_s
                            else:
                                expected_soc[es_id][p] = ess_soc
                        row_idx = row_idx + 1

                        # State-of-Charge, [%]
                        sheet.cell(row=row_idx, column=1).value = operator_type
                        sheet.cell(row=row_idx, column=2).value = tn_node_id
                        sheet.cell(row=row_idx, column=3).value = es_id
                        sheet.cell(row=row_idx, column=4).value = node_id
                        sheet.cell(row=row_idx, column=5).value = int(year)
                        sheet.cell(row=row_idx, column=6).value = day
                        sheet.cell(row=row_idx, column=7).value = 'SoC, [%]'
                        sheet.cell(row=row_idx, column=8).value = s_m
                        sheet.cell(row=row_idx, column=9).value = s_o
                        for p in range(network[year][day].num_instants):
                            ess_soc_percent = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['soc_percent'][es_id][p]
                            sheet.cell(row=row_idx, column=p + 10).value = ess_soc_percent
                            sheet.cell(row=row_idx, column=p + 10).number_format = percent_style
                            if ess_soc_percent != 'N/A':
                                expected_soc_percent[es_id][p] += ess_soc_percent * omega_m * omega_s
                            else:
                                expected_soc_percent[es_id][p] = ess_soc_percent
                        row_idx = row_idx + 1

            for energy_storage in network[year][day].energy_storages:

                es_id = energy_storage.es_id
                node_id = energy_storage.bus

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = es_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_p[es_id][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                # - Reactive Power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = es_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_q[es_id][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                # - Apparent Power
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = es_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_s[es_id][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                # State-of-Charge, [MWh]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = es_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'SoC, [MWh]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_soc[es_id][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = decimal_style
                row_idx = row_idx + 1

                # State-of-Charge, [%]
                sheet.cell(row=row_idx, column=1).value = operator_type
                sheet.cell(row=row_idx, column=2).value = tn_node_id
                sheet.cell(row=row_idx, column=3).value = es_id
                sheet.cell(row=row_idx, column=4).value = node_id
                sheet.cell(row=row_idx, column=5).value = int(year)
                sheet.cell(row=row_idx, column=6).value = day
                sheet.cell(row=row_idx, column=7).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=8).value = 'Expected'
                sheet.cell(row=row_idx, column=9).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 10).value = expected_soc_percent[es_id][p]
                    sheet.cell(row=row_idx, column=p + 10).number_format = percent_style
                row_idx = row_idx + 1

    return row_idx


def _write_relaxation_slacks_results_to_excel(planning_problem, workbook, results):
    _write_relaxation_slacks_results_network_operators_to_excel(planning_problem, workbook, results)


def _write_relaxation_slacks_results_no_coordination_to_excel(planning_problem, workbook, results):
    _write_relaxation_slacks_results_network_operators_to_excel(planning_problem, workbook, results)


def _write_relaxation_slacks_results_network_operators_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Relaxation Slacks TSO, DSOs')

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Operator'
    sheet.cell(row=row_idx, column=2).value = 'ADN Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Resource ID'
    sheet.cell(row=row_idx, column=4).value = 'Year'
    sheet.cell(row=row_idx, column=5).value = 'Day'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=8).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 9).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    tso_results = results['tso']['results']
    transmission_network = planning_problem.transmission_network.network
    tn_params = planning_problem.transmission_network.params
    if tn_params.slacks:
        row_idx = _write_relaxation_slacks_results_per_operator(transmission_network, sheet, 'TSO', row_idx, tso_results, tn_params)

    # Write results -- DSOs
    for tn_node_id in results['dso']:
        dso_results = results['dso'][tn_node_id]['results']
        distribution_network = planning_problem.distribution_networks[tn_node_id].network
        dn_params = planning_problem.distribution_networks[tn_node_id].params
        if dn_params.slacks:
            row_idx = _write_relaxation_slacks_results_per_operator(distribution_network, sheet, 'DSO', row_idx, dso_results, dn_params, tn_node_id=tn_node_id)


def _write_relaxation_slacks_results_per_operator(network, sheet, operator_type, row_idx, results, params, tn_node_id='-'):

    decimal_style = '0.000000'
    voltage_decimal_style = '0.000000'
    voltage_quantities = (
        ('squared_down', 'Voltage squared slack, down [p.u.^2]'),
        ('squared_up', 'Voltage squared slack, up [p.u.^2]'),
        ('physical_down', 'Voltage permitted relaxation, down [p.u.]'),
        ('physical_up', 'Voltage permitted relaxation, up [p.u.]'),
        ('violation_down', 'Voltage realized violation, down [p.u.]'),
        ('violation_up', 'Voltage realized violation, up [p.u.]'),
    )

    for year in results:
        for day in results[year]:
            for s_m in results[year][day]['scenarios']:
                for s_o in results[year][day]['scenarios'][s_m]:

                    # Voltage slacks
                    if params.slacks.grid_operation.voltage:
                        for node in network[year][day].nodes:

                            node_id = node.bus_i
                            voltage_results = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['voltage']
                            for quantity, label in voltage_quantities:
                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = node_id
                                sheet.cell(row=row_idx, column=4).value = int(year)
                                sheet.cell(row=row_idx, column=5).value = day
                                sheet.cell(row=row_idx, column=6).value = label
                                sheet.cell(row=row_idx, column=7).value = s_m
                                sheet.cell(row=row_idx, column=8).value = s_o
                                for p in range(network[year][day].num_instants):
                                    sheet.cell(row=row_idx, column=p + 9).value = voltage_results[quantity][node_id][p]
                                    sheet.cell(row=row_idx, column=p + 9).number_format = voltage_decimal_style
                                row_idx = row_idx + 1

                    # Branch flow slacks
                    if params.slacks.grid_operation.branch_flow:
                        for branch in network[year][day].branches:

                            branch_id = branch.branch_id
                            branch_slacks = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['branch_flow']
                            for quantity, label in (
                                ('flow_ij_sqr', 'Flow_ij_sqr'),
                                ('flow_ji_sqr', 'Flow_ji_sqr'),
                            ):
                                if branch_id not in branch_slacks[quantity]:
                                    continue
                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = branch_id
                                sheet.cell(row=row_idx, column=4).value = int(year)
                                sheet.cell(row=row_idx, column=5).value = day
                                sheet.cell(row=row_idx, column=6).value = label
                                sheet.cell(row=row_idx, column=7).value = s_m
                                sheet.cell(row=row_idx, column=8).value = s_o
                                for p in range(network[year][day].num_instants):
                                    slack_sqr = branch_slacks[quantity][branch_id][p]
                                    sheet.cell(row=row_idx, column=p + 9).value = slack_sqr
                                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                row_idx = row_idx + 1

                    # Node balance
                    for node in network[year][day].nodes:

                        node_id = node.bus_i

                        # - p
                        if params.slacks.node_balance.active_power:
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Node balance, p'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                slack_node_balance_p = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['node_balance']['p'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = slack_node_balance_p
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            row_idx = row_idx + 1

                        # - q
                        if params.slacks.node_balance.reactive_power:
                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Node balance, q'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                slack_node_balance_q = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['node_balance']['q'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = slack_node_balance_q
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            row_idx = row_idx + 1

                    # SharedESS
                    for shared_energy_storage in network[year][day].shared_energy_storages:

                        node_id = shared_energy_storage.bus

                        # - Day balance
                        if params.slacks.shared_ess.day_balance:

                            sheet.cell(row=row_idx, column=1).value = operator_type
                            sheet.cell(row=row_idx, column=2).value = tn_node_id
                            sheet.cell(row=row_idx, column=3).value = node_id
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Shared Energy Storage, soc_final'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                soc_final = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['shared_energy_storages']['soc_final'][node_id]
                                sheet.cell(row=row_idx, column=p + 9).value = soc_final
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            row_idx = row_idx + 1

                    # Load flexibility
                    if params.fl_reg:
                        for load in network[year][day].loads:

                            load_id = load.load_id

                            # - Day balance
                            if params.slacks.flexibility.day_balance:

                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = load_id
                                sheet.cell(row=row_idx, column=4).value = int(year)
                                sheet.cell(row=row_idx, column=5).value = day
                                sheet.cell(row=row_idx, column=6).value = 'Flex. balance, p'
                                sheet.cell(row=row_idx, column=7).value = s_m
                                sheet.cell(row=row_idx, column=8).value = s_o
                                for p in range(network[year][day].num_instants):
                                    day_balance_p = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['flexibility']['day_balance'][load_id]['p']
                                    sheet.cell(row=row_idx, column=p + 9).value = day_balance_p
                                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                row_idx = row_idx + 1

                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = load_id
                                sheet.cell(row=row_idx, column=4).value = int(year)
                                sheet.cell(row=row_idx, column=5).value = day
                                sheet.cell(row=row_idx, column=6).value = 'Flex. balance, q'
                                sheet.cell(row=row_idx, column=7).value = s_m
                                sheet.cell(row=row_idx, column=8).value = s_o
                                for p in range(network[year][day].num_instants):
                                    day_balance_q = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['flexibility']['day_balance'][load_id]['q']
                                    sheet.cell(row=row_idx, column=p + 9).value = day_balance_q
                                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                row_idx = row_idx + 1

                    # ESS
                    if params.es_reg:

                        for energy_storage in network[year][day].energy_storages:

                            es_id = energy_storage.es_id

                            # - Day balance
                            if params.slacks.ess.day_balance:

                                sheet.cell(row=row_idx, column=1).value = operator_type
                                sheet.cell(row=row_idx, column=2).value = tn_node_id
                                sheet.cell(row=row_idx, column=3).value = es_id
                                sheet.cell(row=row_idx, column=4).value = int(year)
                                sheet.cell(row=row_idx, column=5).value = day
                                sheet.cell(row=row_idx, column=6).value = 'Energy Storage, soc_final'
                                sheet.cell(row=row_idx, column=7).value = s_m
                                sheet.cell(row=row_idx, column=8).value = s_o
                                for p in range(network[year][day].num_instants):
                                    soc_final = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['soc_final'][es_id]
                                    sheet.cell(row=row_idx, column=p + 9).value = soc_final
                                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                row_idx = row_idx + 1

    return row_idx


# ======================================================================================================================
#   NETWORK diagram functions (plot)
# ======================================================================================================================
def _plot_networkx_diagram(planning_problem):

    for year in planning_problem.years:
        for day in planning_problem.days:

            transmission_network = planning_problem.transmission_network.network[year][day]

            node_labels = {}
            ref_nodes, pv_nodes, pq_nodes = [], [], []
            res_pv_nodes = [gen.bus for gen in transmission_network.generators if gen.gen_type == GEN_RES_SOLAR]
            res_wind_nodes = [gen.bus for gen in transmission_network.generators if gen.gen_type == GEN_RES_WIND]
            adn_nodes = planning_problem.active_distribution_network_nodes

            branches = []
            line_list, open_line_list = [], []
            transf_list, open_transf_list = [], []
            for branch in transmission_network.branches:
                if branch.is_transformer:
                    branches.append({'type': 'transformer', 'data': branch})
                else:
                    branches.append({'type': 'line', 'data': branch})

            # Build graph
            graph = nx.Graph()
            for i in range(len(transmission_network.nodes)):
                node = transmission_network.nodes[i]
                graph.add_node(node.bus_i)
                node_labels[node.bus_i] = '{}'.format(node.bus_i)
                if node.type == BUS_REF:
                    ref_nodes.append(node.bus_i)
                elif node.type == BUS_PV:
                    pv_nodes.append(node.bus_i)
                elif node.type == BUS_PQ:
                    if node.bus_i not in (res_pv_nodes + res_wind_nodes + adn_nodes):
                        pq_nodes.append(node.bus_i)
            for i in range(len(branches)):
                branch = branches[i]
                if branch['type'] == 'line':
                    graph.add_edge(branch['data'].fbus, branch['data'].tbus)
                    if branch['data'].status == 1:
                        line_list.append((branch['data'].fbus, branch['data'].tbus))
                    else:
                        open_line_list.append((branch['data'].fbus, branch['data'].tbus))
                if branch['type'] == 'transformer':
                    graph.add_edge(branch['data'].fbus, branch['data'].tbus)
                    if branch['data'].status == 1:
                        transf_list.append((branch['data'].fbus, branch['data'].tbus))
                    else:
                        open_transf_list.append((branch['data'].fbus, branch['data'].tbus))

            # Plot diagram
            pos = nx.spring_layout(graph, k=0.50, iterations=1000)
            fig, ax = plt.subplots(figsize=(12, 8))
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=ref_nodes, node_color='red', node_size=250, label='Reference bus')
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=pv_nodes, node_color='lightgreen', node_size=250, label='Conventional generator')
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=pq_nodes, node_color='lightblue', node_size=250, label='PQ buses')
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=res_pv_nodes, node_color='yellow', node_size=250, label='RES, PV')
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=res_wind_nodes, node_color='blue', node_size=250, label='RES, Wind')
            nx.draw_networkx_nodes(graph, ax=ax, pos=pos, nodelist=adn_nodes, node_color='orange', node_size=250, label='ADN buses')
            nx.draw_networkx_labels(graph, ax=ax, pos=pos, labels=node_labels, font_size=12)
            nx.draw_networkx_edges(graph, ax=ax, pos=pos, edgelist=line_list, width=1.50, edge_color='black')
            nx.draw_networkx_edges(graph, ax=ax, pos=pos, edgelist=transf_list, width=2.00, edge_color='blue', label='Transformer')
            nx.draw_networkx_edges(graph, ax=ax, pos=pos, edgelist=open_line_list, style='dashed', width=1.50, edge_color='red')
            nx.draw_networkx_edges(graph, ax=ax, pos=pos, edgelist=open_transf_list, style='dashed', width=2.00, edge_color='red')
            plt.legend(scatterpoints=1, frameon=False, prop={'size': 12})
            plt.axis('off')

            filename = os.path.join(planning_problem.diagrams_dir, f'{planning_problem.name}_{year}_{day}')
            plt.savefig(f'{filename}.pdf', bbox_inches='tight')
            plt.savefig(f'{filename}.png', bbox_inches='tight')
            plt.close(fig)


# ======================================================================================================================
#   Aux functions
# ======================================================================================================================
def _get_initial_candidate_solution(planning_problem):
    candidate_solution = {'investment': {}, 'total_capacity': {}}
    for e in range(len(planning_problem.active_distribution_network_nodes)):
        node_id = planning_problem.active_distribution_network_nodes[e]
        candidate_solution['investment'][node_id] = dict()
        candidate_solution['total_capacity'][node_id] = dict()
        for year in planning_problem.years:
            candidate_solution['investment'][node_id][year] = dict()
            candidate_solution['investment'][node_id][year]['s'] = 0.00
            candidate_solution['investment'][node_id][year]['e'] = 0.00
            candidate_solution['total_capacity'][node_id][year] = dict()
            candidate_solution['total_capacity'][node_id][year]['s'] = 0.00
            candidate_solution['total_capacity'][node_id][year]['e'] = 0.00
    return candidate_solution


def _get_test_candidate_solution(planning_problem, node_id, investment_year, s_inv=1.00, e_inv=2.00):

    candidate_solution = _get_initial_candidate_solution(planning_problem)

    if node_id not in candidate_solution['investment']:
        raise ValueError(f'Unknown shared ESS node {node_id}.')

    if investment_year not in planning_problem.years:
        raise ValueError(f'Unknown investment year {investment_year}.')

    candidate_solution['investment'][node_id][investment_year]['s'] = s_inv
    candidate_solution['investment'][node_id][investment_year]['e'] = e_inv

    _rebuild_candidate_total_capacities(planning_problem, candidate_solution)

    return candidate_solution


def _check_interface_nodes_base_voltage_consistency(planning_problem):
    for year in planning_problem.years:
        for day in planning_problem.days:
            for node_id in planning_problem.distribution_networks:
                tn_node_base_kv = planning_problem.transmission_network.network[year][day].get_node_base_kv(node_id)
                dn_ref_node_id = planning_problem.distribution_networks[node_id].network[year][day].get_reference_node_id()
                dn_node_base_kv = planning_problem.distribution_networks[node_id].network[year][day].get_node_base_kv(dn_ref_node_id)
                if not isclose(tn_node_base_kv, dn_node_base_kv, rel_tol=5e-2):
                    print(f'[ERROR] Distribution Network {planning_problem.distribution_networks[node_id].name}, TN node {node_id}. Inconsistent TN-DN base voltage, year {year}, day {day}! Check network(s). Exiting')
                    exit(ERROR_SPECIFICATION_FILE)


def _add_adn_node_to_transmission_network(planning_problem):
    for year in planning_problem.years:
        for day in planning_problem.days:
            for node_id in planning_problem.distribution_networks:
                if planning_problem.transmission_network.network[year][day].adn_load_exists(node_id):
                    adn_load_idx = planning_problem.transmission_network.network[year][day].get_adn_load_idx(node_id)
                    adn_load = planning_problem.transmission_network.network[year][day].loads[adn_load_idx]
                    adn_load.load_id = f'ADN_{node_id}'
                    adn_load.pd = np.zeros(adn_load.pd.shape)
                    adn_load.qd = np.zeros(adn_load.qd.shape)
                    adn_load.fl_reg = True
                    adn_load.status = 1
                else:
                    adn_load = Load()
                    adn_load.bus = node_id
                    adn_load.load_id = f'ADN_{node_id}'
                    adn_load.pd = np.zeros((planning_problem.transmission_network.num_oper_scenarios, planning_problem.num_instants))
                    adn_load.qd = np.zeros((planning_problem.transmission_network.num_oper_scenarios, planning_problem.num_instants))
                    adn_load.fl_reg = True
                    adn_load.status = 1
                    planning_problem.transmission_network.network[year][day].loads.append(adn_load)


def _add_shared_energy_storage_to_transmission_network(planning_problem):
    for year in planning_problem.years:
        for day in planning_problem.days:
            s_base = planning_problem.transmission_network.network[year][day].baseMVA
            for node_id in planning_problem.distribution_networks:
                shared_energy_storage = SharedEnergyStorage()
                shared_energy_storage.bus = node_id
                shared_energy_storage.dn_name = planning_problem.distribution_networks[node_id].name
                shared_energy_storage.s = shared_energy_storage.s / s_base
                shared_energy_storage.e = shared_energy_storage.e / s_base
                planning_problem.transmission_network.network[year][day].shared_energy_storages.append(shared_energy_storage)


def _add_shared_energy_storage_to_distribution_network(planning_problem):
    for year in planning_problem.years:
        for day in planning_problem.days:
            for node_id in planning_problem.distribution_networks:
                s_base = planning_problem.distribution_networks[node_id].network[year][day].baseMVA
                shared_energy_storage = SharedEnergyStorage()
                shared_energy_storage.bus = planning_problem.distribution_networks[node_id].network[year][day].get_reference_node_id()
                shared_energy_storage.dn_name = planning_problem.distribution_networks[node_id].network[year][day].name
                shared_energy_storage.s = shared_energy_storage.s / s_base
                shared_energy_storage.e = shared_energy_storage.e / s_base
                planning_problem.distribution_networks[node_id].network[year][day].shared_energy_storages.append(shared_energy_storage)


def _print_candidate_solution(candidate_solution):

    print('[INFO] Candidate solution:')

    # Header
    print('\t\t{:3}\t{:10}\t'.format('', 'Capacity'), end='')
    for node_id in candidate_solution['total_capacity']:
        for year in candidate_solution['total_capacity'][node_id]:
            print(f'{year}\t', end='')
        print()
        break

    # Values
    for node_id in candidate_solution['total_capacity']:
        print('\t\t{:3}\t{:10}\t'.format(node_id, 'S, [MVA]'), end='')
        for year in candidate_solution['total_capacity'][node_id]:
            print("{:.3f}\t".format(candidate_solution['total_capacity'][node_id][year]['s']), end='')
        print()
        print('\t\t{:3}\t{:10}\t'.format(node_id, 'E, [MVAh]'), end='')
        for year in candidate_solution['total_capacity'][node_id]:
            print("{:.3f}\t".format(candidate_solution['total_capacity'][node_id][year]['e']), end='')
        print()
