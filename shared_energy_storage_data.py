import os
from math import isclose
import pandas as pd
import pyomo.opt as po
import pyomo.environ as pe
from openpyxl import Workbook
from shared_energy_storage import SharedEnergyStorage
from shared_energy_storage_parameters import SharedEnergyStorageParameters
from helper_functions import *


# ======================================================================================================================
#  SHARED ENERGY STORAGE Information
# ======================================================================================================================
class SharedEnergyStorageData:

    def __init__(self):
        self.name = str()
        self.data_dir = str()
        self.results_dir = str()
        self.plots_dir = str()
        self.data_file = str()
        self.params_file = str()
        self.years = list()
        self.days = list()
        self.num_instants = 0
        self.discount_factor = 5e-2
        self.shared_energy_storages = dict()
        self.prob_market_scenarios = dict()         # Probability of market (price) scenarios
        self.cost_investment = dict()
        self.params = SharedEnergyStorageParameters()
        self.active_distribution_network_nodes = list()
        self.solver_recovery_diagnostics = list()

    def build_master_problem(self):
        return _build_master_problem(self)

    def build_subproblem(self):
        subproblem = dict()
        for node_id in self.active_distribution_network_nodes:
            subproblem[node_id] = _build_subproblem(self, node_id)
        return subproblem

    def optimize_master_problem(self, model, from_warm_start=False):
        print('[INFO] \t\t - Running Shared ESS optimization (master problem)...')
        return _optimize(model, self.params.lp_solver_params, from_warm_start=from_warm_start)

    def optimize(self, models, from_warm_start=False):
        print('[INFO] \t\t - Running Shared ESS optimization (subproblem)...')
        results = dict()
        for node_id in self.active_distribution_network_nodes:
            print(f'[INFO] \t\t\t - Node {node_id}...')
            results[node_id] = _optimize(
                models[node_id],
                self.params.solver_params,
                from_warm_start=from_warm_start,
                node_id=node_id,
                diagnostic_sink=self.solver_recovery_diagnostics,
            )
        return results

    def get_primal_value(self, models):
        objective = 0.00
        for node_id in self.active_distribution_network_nodes:
            objective += pe.value(models[node_id].objective)
        return objective

    def get_feasibility_penalty(self, models):
        penalty = 0.00
        for node_id in self.active_distribution_network_nodes:
            penalty += pe.value(models[node_id].feasibility_penalty)
        return penalty

    def get_feasibility_violation(self, models):
        return self.get_feasibility_penalty(models) / PENALTY_ESSO_SLACK

    def get_salvage_value(self, models):
        salvage_value = 0.00
        for node_id in self.active_distribution_network_nodes:
            salvage_value += pe.value(models[node_id].salvage_value)
        return salvage_value

    def get_salvage_value_sensitivities(self, models):
        return _get_salvage_value_sensitivities(self, models)

    def get_salvage_value_results(self, models):
        return _get_salvage_value_results(self, models)

    def update_model_with_candidate_solution(self, models, candidate_solution):
        _update_model_with_candidate_solution(self, models, candidate_solution)

    def get_candidate_solution(self, model):
        return _get_candidate_solution(self, model)

    def load_candidate_solution_into_master_model(self, model, candidate_solution):
        _load_candidate_solution_into_master_model(self, model, candidate_solution)

    def map_available_capacity_sensitivities_to_investments(self, models, sensitivities):
        return _map_available_capacity_sensitivities_to_investments(self, models, sensitivities)

    def read_shared_energy_storage_data_from_file(self):
        filename = os.path.join(self.data_dir, 'SharedESS', self.data_file)
        _read_shared_energy_storage_data_from_file(self, filename)

    def read_parameters_from_file(self):
        filename = os.path.join(self.data_dir, 'SharedESS', self.params_file)
        self.params.read_parameters_from_file(filename)

    def create_shared_energy_storages(self, planning_problem):
        for year in planning_problem.years:
            self.shared_energy_storages[year] = list()
            for node_id in planning_problem.transmission_network.active_distribution_network_nodes:
                shared_energy_storage = SharedEnergyStorage()
                shared_energy_storage.bus = node_id
                shared_energy_storage.dn_name = planning_problem.distribution_networks[node_id].name
                self.shared_energy_storages[year].append(shared_energy_storage)

    def get_shared_energy_storage_idx(self, node_id):
        repr_years = [year for year in self.years]
        for i in range(len(self.shared_energy_storages[repr_years[0]])):
            shared_energy_storage = self.shared_energy_storages[repr_years[0]][i]
            if shared_energy_storage.bus == node_id:
                return i
        print(f'[ERROR] Network {self.name}. Node {node_id} does not have a shared energy storage system! Check network.')
        exit(ERROR_NETWORK_FILE)

    def process_results(self, models):
        results = dict()
        results['capacity'] = self.get_available_capacity(models)
        results['operation'] = dict()
        results['operation']['aggregated'] = self.process_results_aggregated(models)
        results['operation']['detailed'] = self.process_results_detailed(models)
        results['soh'] = dict()
        results['soh']['aggregated'] = self.process_soh_results_aggregated(models)
        results['soh']['detailed'] = self.process_soh_results_detailed(models)
        results['salvage_value'] = self.get_salvage_value_results(models)
        results['relaxation_variables'] = dict()
        results['relaxation_variables']['investment'] = self.process_relaxation_variables_investment(models)
        if self.params.slacks:
            results['relaxation_variables']['degradation'] = dict()
            results['relaxation_variables']['degradation']['aggregated'] = dict()
            results['relaxation_variables']['degradation']['detailed'] = self.process_relaxation_variables_degradation_detailed(
                models)
            results['relaxation_variables']['operation'] = dict()
            results['relaxation_variables']['operation']['aggregated'] = self.process_relaxation_variables_operation_aggregated(
                models)
            results['relaxation_variables']['operation']['detailed'] = self.process_relaxation_variables_operation_detailed(
                models)
        return results

    def process_results_aggregated(self, models):
        return _process_results_aggregated(self, models)

    def process_results_detailed(self, models):
        return _process_results_detailed(self, models)

    def process_soh_results_aggregated(self, models):
        return _process_soh_results_aggregated(self, models)

    def process_soh_results_detailed(self, models):
        return _process_soh_results_detailed(self, models)

    def process_relaxation_variables_investment(self, models):
        return _process_relaxation_variables_investment(self, models)

    def process_relaxation_variables_degradation_detailed(self, models):
        return _process_relaxation_variables_degradation_detailed(self, models)

    def process_relaxation_variables_operation_aggregated(self, models):
        return _process_relaxation_variables_operation_aggregated(self, models)

    def process_relaxation_variables_operation_detailed(self, models):
        return _process_relaxation_variables_operation_detailed(self, models)

    def write_optimization_results_to_excel(self, models):
        results = self.process_results(models)
        _write_optimization_results_to_excel(self, self.results_dir, results)

    def update_data_with_candidate_solution(self, candidate_solution):
        for year in self.years:
            for shared_ess in self.shared_energy_storages[year]:
                shared_ess.s = candidate_solution[shared_ess.bus][year]['s']
                shared_ess.e = candidate_solution[shared_ess.bus][year]['e']
                shared_ess.e_init = candidate_solution[shared_ess.bus][year]['e'] * ENERGY_STORAGE_RELATIVE_INIT_SOC
                shared_ess.e_min = candidate_solution[shared_ess.bus][year]['e'] * ENERGY_STORAGE_MIN_ENERGY_STORED
                shared_ess.e_max = candidate_solution[shared_ess.bus][year]['e'] * ENERGY_STORAGE_MAX_ENERGY_STORED

    def get_updated_capacities(self, model):
        available_capacities = dict()
        for node_id in self.active_distribution_network_nodes:
            available_capacities[node_id] = dict()
            years = list(self.years)
            for y in range(len(years)):
                s_available, e_available = self.get_available_capacities(model[node_id], y)
                available_capacities[node_id][years[y]] = {'s_available': s_available, 'e_available': e_available}
        return available_capacities

    def get_available_capacities(self, model, year_idx):
        s_available = 0.00
        e_available = 0.00
        for y_inv in model.years:
            s_available += pe.value(model.es_s_available_per_unit[y_inv, year_idx])
            e_available += pe.value(model.es_e_available_per_unit[y_inv, year_idx])
        return s_available, e_available

    def get_available_capacity(self, models):
        return _get_available_capacity(self, models)

    def get_investment_cost_and_rated_capacity(self, model):
        return _get_investment_cost_and_rated_capacity(self, model)

    def write_ess_costs_to_excel(self, workbook, shared_ess_cost):
        _write_ess_costs_to_excel(self, workbook, shared_ess_cost)

    def write_ess_capacity_results_to_excel(self, workbook, shared_ess_capacity, write_investment=True):
        if write_investment:
            _write_ess_capacity_investment_to_excel(self, workbook, shared_ess_capacity['investment'], initial_sheet=False)
        _write_ess_capacity_rated_available_to_excel(self, workbook, shared_ess_capacity)

    def write_salvage_value_results_to_excel(self, workbook, salvage_value_results):
        _write_terminal_salvage_value_to_excel(workbook, salvage_value_results)

    def write_relaxation_slacks_results_to_excel(self, workbook, results):
        _write_investment_relaxation_slacks_results_to_excel(self, workbook, results['relaxation_variables']['investment'])
        if self.params.slacks:
            _write_detailed_degradation_relaxation_slacks_results_to_excel(self, workbook, results['relaxation_variables']['degradation']['detailed'])
            _write_aggregated_operation_relaxation_slacks_results_to_excel(self, workbook, results['relaxation_variables']['operation']['aggregated'])
            _write_detailed_operation_relaxation_slacks_results_to_excel(self, workbook, results['relaxation_variables']['operation']['detailed'])


# ======================================================================================================================
#  MASTER PROBLEM  functions
# ======================================================================================================================
def _build_master_problem(shared_ess_data):

    years = [year for year in shared_ess_data.years]

    model = pe.ConcreteModel()
    model.name = "ESS Optimization -- Benders' Master Problem"

    # ------------------------------------------------------------------------------------------------------------------
    # Sets
    model.years = range(len(shared_ess_data.years))
    model.scenarios_market = range(len(shared_ess_data.prob_market_scenarios))
    model.energy_storages = range(len(shared_ess_data.active_distribution_network_nodes))

    # ------------------------------------------------------------------------------------------------------------------
    # Decision variables
    model.es_s_investment = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)  # Investment in power capacity in year y
    model.es_e_investment = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)  # Investment in energy capacity in year y
    model.es_s_rated = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)       # Rated power capacity (considering calendar life)
    model.es_e_rated = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals)       # Rated energy capacity (considering calendar life, not degradation)
    model.alpha = pe.Var(domain=pe.Reals)                                                          # Local approximation of operational recourse
    model.alpha.setlb(-shared_ess_data.params.budget * 1e3)

    # ------------------------------------------------------------------------------------------------------------------
    # Constraints
    # - Yearly Power and Energy ratings as a function of yearly investments
    model.rated_s_capacity = pe.ConstraintList()
    model.rated_e_capacity = pe.ConstraintList()
    for e in model.energy_storages:
        total_s_capacity_per_year = [0.0 for _ in model.years]
        total_e_capacity_per_year = [0.0 for _ in model.years]
        for y in model.years:
            year = years[y]
            num_years = shared_ess_data.years[year]
            shared_energy_storage = shared_ess_data.shared_energy_storages[year][e]
            tcal_norm = round(shared_energy_storage.t_cal / num_years)
            max_tcal_norm = min(y + tcal_norm, len(shared_ess_data.years))
            for x in range(y, max_tcal_norm):
                total_s_capacity_per_year[x] += model.es_s_investment[e, y]
                total_e_capacity_per_year[x] += model.es_e_investment[e, y]
        for y in model.years:
            model.rated_s_capacity.add(model.es_s_rated[e, y] == total_s_capacity_per_year[y])
            model.rated_e_capacity.add(model.es_e_rated[e, y] == total_e_capacity_per_year[y])

    # - Maximum Energy Capacity (related to space constraints)
    model.energy_storage_maximum_capacity = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:
            model.energy_storage_maximum_capacity.add(model.es_e_rated[e, y] <= shared_ess_data.params.max_capacity)

    # - Energy-to-Power Ratio (related to ESS technology)
    model.energy_storage_power_to_energy_factor = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:
            model.energy_storage_power_to_energy_factor.add(model.es_e_investment[e, y] >= model.es_s_investment[e, y] * shared_ess_data.params.min_energy_to_power_ratio)
            model.energy_storage_power_to_energy_factor.add(model.es_e_investment[e, y] <= model.es_s_investment[e, y] * shared_ess_data.params.max_energy_to_power_ratio)

    # - Maximum Investment
    investment_cost_total = 0.0
    model.energy_storage_investment = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:
            year = years[y]
            for s_m in model.scenarios_market:
                omega_m = shared_ess_data.prob_market_scenarios[s_m]
                c_inv_s = shared_ess_data.cost_investment['power'][s_m][year]
                c_inv_e = shared_ess_data.cost_investment['energy'][s_m][year]
                annualization = 1 / ((1 + shared_ess_data.discount_factor) ** (int(year) - int(years[0])))
                investment_cost_total += annualization * omega_m * model.es_s_investment[e, y] * c_inv_s
                investment_cost_total += annualization * omega_m * model.es_e_investment[e, y] * c_inv_e
    model.energy_storage_investment.add(investment_cost_total <= shared_ess_data.params.budget)

    # Benders-type local sensitivity cuts
    model.benders_cuts = pe.ConstraintList()

    # Objective function
    investment_cost = 0.0
    for e in model.energy_storages:
        for y in model.years:
            year = years[y]
            for s_m in model.scenarios_market:

                omega_m = shared_ess_data.prob_market_scenarios[s_m]
                c_inv_s = shared_ess_data.cost_investment['power'][s_m][year]
                c_inv_e = shared_ess_data.cost_investment['energy'][s_m][year]
                annualization = 1 / ((1 + shared_ess_data.discount_factor) ** (int(year) - int(years[0])))

                # Investment Cost
                investment_cost += annualization * omega_m * model.es_s_investment[e, y] * c_inv_s
                investment_cost += annualization * omega_m * model.es_e_investment[e, y] * c_inv_e

    model.investment_cost = pe.Expression(expr=investment_cost)
    model.objective = pe.Objective(sense=pe.minimize, expr=model.investment_cost + model.alpha)

    # Define that we want the duals
    model.ipopt_zL_out = pe.Suffix(direction=pe.Suffix.IMPORT)  # Ipopt bound multipliers (obtained from solution)
    model.ipopt_zU_out = pe.Suffix(direction=pe.Suffix.IMPORT)
    model.ipopt_zL_in = pe.Suffix(direction=pe.Suffix.EXPORT)  # Ipopt bound multipliers (sent to solver)
    model.ipopt_zU_in = pe.Suffix(direction=pe.Suffix.EXPORT)
    model.dual = pe.Suffix(direction=pe.Suffix.IMPORT_EXPORT)

    return model


# ======================================================================================================================
#  OPERATIONAL PLANNING functions
# ======================================================================================================================
def _build_subproblem(shared_ess_data, node_id):

    model = pe.ConcreteModel()
    model.name = 'ESSO, Operational Planning'
    repr_days = [day for day in shared_ess_data.days]
    repr_years = [year for year in shared_ess_data.years]
    shared_ess_idx = shared_ess_data.get_shared_energy_storage_idx(node_id)

    # ------------------------------------------------------------------------------------------------------------------
    # Sets
    model.years = range(len(shared_ess_data.years))
    model.days = range(len(shared_ess_data.days))
    model.periods = range(shared_ess_data.num_instants)

    # Track ESSO constraints that belong to each investment cohort.
    # These constraints can be deactivated when the corresponding investment capacity is zero.
    model._esso_cohort_constraints = {
        y_inv: [] for y_inv in model.years
    }

    model._esso_cohort_inactive = {
        y_inv: False for y_inv in model.years
    }

    # ------------------------------------------------------------------------------------------------------------------
    # Variables
    model.es_s_investment_fixed = pe.Param(model.years, mutable=True, initialize=0.00)
    model.es_e_investment_fixed = pe.Param(model.years, mutable=True, initialize=0.00)
    model.es_s_investment = pe.Var(model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.es_e_investment = pe.Var(model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.slack_es_s_investment_up = pe.Var(model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.slack_es_s_investment_down = pe.Var(model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.slack_es_e_investment_up = pe.Var(model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.slack_es_e_investment_down = pe.Var(model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.es_s_rated = pe.Var(model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.es_e_rated = pe.Var(model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.es_snet = pe.Var(model.years, model.days, model.periods, domain=pe.Reals, initialize=0.0)
    model.es_pnet = pe.Var(model.years, model.days, model.periods, domain=pe.Reals, initialize=0.0)
    model.es_qnet = pe.Var(model.years, model.days, model.periods, domain=pe.Reals, initialize=0.0)
    if shared_ess_data.params.slacks:
        model.slack_es_snet_up = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.0)
        model.slack_es_snet_down = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.0)
        model.slack_es_snet_def_up = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.0)
        model.slack_es_snet_def_down = pe.Var(model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.0)

    model.es_s_rated_per_unit = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.es_e_rated_per_unit = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.es_s_available_per_unit = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.es_e_available_per_unit = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=0.0)
    model.es_s_rated_per_unit.fix(0.00)
    model.es_e_rated_per_unit.fix(0.00)

    model.es_sch_per_unit = pe.Var(model.years, model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
    model.es_sdch_per_unit = pe.Var(model.years, model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
    if shared_ess_data.params.slacks:
        model.slack_es_ch_comp_per_unit = pe.Var(model.years, model.years, model.days, model.periods, domain=pe.NonNegativeReals, initialize=0.00)
    model.es_avg_ch_dch_per_unit = pe.Var(model.years, model.years, domain=pe.Reals, initialize=0.00)
    model.es_soh_per_unit = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=1.00, bounds=(0.00, 1.00))
    model.es_degradation_per_unit = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=0.00, bounds=(0.00, 1.00))
    model.es_soh_per_unit_cumul = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=1.00, bounds=(0.00, 1.00))
    model.es_degradation_per_unit_cumul = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=0.00, bounds=(0.00, 1.00))
    if shared_ess_data.params.slacks:
        model.slack_es_soh_per_unit_up = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=0.00)
        model.slack_es_soh_per_unit_down = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=0.00)
        model.slack_es_soh_per_unit_cumul_up = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=0.00)
        model.slack_es_soh_per_unit_cumul_down = pe.Var(model.years, model.years, domain=pe.NonNegativeReals, initialize=0.00)
    model.es_soh_per_unit.fix(1.00)
    model.es_soh_per_unit_cumul.fix(1.00)

    # ------------------------------------------------------------------------------------------------------------------
    # Constraints
    # - Sinv and Einv fixing constraints
    model.energy_storage_capacity_fixing = pe.ConstraintList()
    for y in model.years:
        model.energy_storage_capacity_fixing.add(model.es_s_investment[y] == model.es_s_investment_fixed[y] + model.slack_es_s_investment_up[y] - model.slack_es_s_investment_down[y])
        model.energy_storage_capacity_fixing.add(model.es_e_investment[y] == model.es_e_investment_fixed[y] + model.slack_es_e_investment_up[y] - model.slack_es_e_investment_down[y])

    # - Rated capacities of each investment
    model.rated_s_capacity_unit = pe.ConstraintList()
    model.rated_e_capacity_unit = pe.ConstraintList()
    for y_inv in model.years:
        shared_energy_storage = shared_ess_data.shared_energy_storages[repr_years[y_inv]][shared_ess_idx]
        tcal_norm = round(shared_energy_storage.t_cal / (shared_ess_data.years[repr_years[y_inv]]))
        max_tcal_norm = min(y_inv + tcal_norm, len(shared_ess_data.years))
        for y in range(y_inv, max_tcal_norm):
            model.es_s_rated_per_unit[y_inv, y].fixed = False
            model.es_e_rated_per_unit[y_inv, y].fixed = False
            model.rated_s_capacity_unit.add(model.es_s_rated_per_unit[y_inv, y] == model.es_s_investment[y_inv])
            model.rated_e_capacity_unit.add(model.es_e_rated_per_unit[y_inv, y] == model.es_e_investment[y_inv])

    # - Rated yearly capacities as a function of yearly investments
    model.rated_s_capacity = pe.ConstraintList()
    model.rated_e_capacity = pe.ConstraintList()
    for y in model.years:
        total_s_capacity = 0.00
        total_e_capacity = 0.00
        for y_inv in model.years:
            total_s_capacity += model.es_s_rated_per_unit[y_inv, y]
            total_e_capacity += model.es_e_rated_per_unit[y_inv, y]
        model.rated_s_capacity.add(model.es_s_rated[y] == total_s_capacity)
        model.rated_e_capacity.add(model.es_e_rated[y] == total_e_capacity)

    # - Available capacities of each investment
    model.available_s_capacity_unit = pe.ConstraintList()
    model.available_e_capacity_unit = pe.ConstraintList()
    for y_inv in model.years:
        for y in model.years:
            model.available_s_capacity_unit.add(model.es_s_available_per_unit[y_inv, y] == model.es_s_rated_per_unit[y_inv, y])
            model.available_e_capacity_unit.add(model.es_e_available_per_unit[y_inv, y] == model.es_e_rated_per_unit[y_inv, y] * model.es_soh_per_unit_cumul[y_inv, y])

    # - Sum of charging and discharging power for the yearly average day (aux, used to estimate degradation of ESSs)
    model.energy_storage_charging_discharging = pe.ConstraintList()
    for y_inv in model.years:
        for y in model.years:
            avg_ch_dch = 0.0
            for d in model.days:
                day = repr_days[d]
                num_days = shared_ess_data.days[day]
                for p in model.periods:
                    sch = model.es_sch_per_unit[y_inv, y, d, p]
                    sdch = model.es_sdch_per_unit[y_inv, y, d, p]
                    avg_ch_dch += (num_days / 365.00) * (sch + sdch)
            _add_esso_cohort_constraint(model, 'energy_storage_charging_discharging', y_inv, y, model.es_avg_ch_dch_per_unit[y_inv, y] == avg_ch_dch)

    # - Capacity degradation
    model.energy_storage_capacity_degradation = pe.ConstraintList()
    for y_inv in model.years:

        num_years = shared_ess_data.years[repr_years[y_inv]]
        shared_energy_storage = shared_ess_data.shared_energy_storages[repr_years[y_inv]][shared_ess_idx]
        tcal_norm = round(shared_energy_storage.t_cal / (shared_ess_data.years[repr_years[y_inv]]))
        max_tcal_norm = min(y_inv + tcal_norm, len(shared_ess_data.years))

        for y in range(y_inv, max_tcal_norm):

            model.es_soh_per_unit[y_inv, y].fixed = False
            model.es_soh_per_unit_cumul[y_inv, y].fixed = False

            # Daily degradation
            _add_esso_cohort_constraint(model, 'energy_storage_capacity_degradation', y_inv, y, model.es_degradation_per_unit[y_inv, y] * (2 * shared_energy_storage.cl_nom * model.es_e_rated_per_unit[y_inv, y]) == model.es_avg_ch_dch_per_unit[y_inv, y])

            # Annual SoH
            if shared_ess_data.params.slacks:
                _add_esso_cohort_constraint(model, 'energy_storage_capacity_degradation', y_inv, y, model.es_soh_per_unit[y_inv, y] == 1.00 - model.es_degradation_per_unit[y_inv, y] + model.slack_es_soh_per_unit_up[y_inv, y] - model.slack_es_soh_per_unit_down[y_inv, y])
            else:
                _add_esso_cohort_constraint(model, 'energy_storage_capacity_degradation', y_inv, y, model.es_soh_per_unit[y_inv, y] == 1.00 - model.es_degradation_per_unit[y_inv, y])

            # Previous cumulative SoH
            prev_soh = 1.00
            if y > y_inv:
                prev_soh = model.es_soh_per_unit_cumul[y_inv, y - 1]

            # Cumulative SoH
            if shared_ess_data.params.slacks:
                _add_esso_cohort_constraint(model, 'energy_storage_capacity_degradation', y_inv, y, model.es_soh_per_unit_cumul[y_inv, y] == prev_soh * (model.es_soh_per_unit[y_inv, y] ** (365.00 * num_years)) + model.slack_es_soh_per_unit_cumul_up[y_inv, y] - model.slack_es_soh_per_unit_cumul_down[y_inv, y])
            else:
                _add_esso_cohort_constraint(model, 'energy_storage_capacity_degradation', y_inv, y, model.es_soh_per_unit_cumul[y_inv, y] == prev_soh * (model.es_soh_per_unit[y_inv, y] ** (365.00 * num_years)))

            # Minimum admissible capacity / SoH
            _add_esso_cohort_constraint(model, 'energy_storage_capacity_degradation', y_inv, y, model.es_e_available_per_unit[y_inv, y] >= shared_energy_storage.soh_min * model.es_e_rated_per_unit[y_inv, y])

            # Cumulative degradation
            _add_esso_cohort_constraint(model, 'energy_storage_capacity_degradation', y_inv, y, model.es_degradation_per_unit_cumul[y_inv, y] == 1.00 - model.es_soh_per_unit_cumul[y_inv, y])

    # - P, Q, S, SoC, per unit as a function of available capacities
    model.energy_storage_limits = pe.ConstraintList()
    model.energy_storage_complementarity = pe.ConstraintList()
    for y_inv in model.years:
        for y in model.years:
            s_max = model.es_s_rated_per_unit[y_inv, y]
            for d in model.days:
                for p in model.periods:

                    sch = model.es_sch_per_unit[y_inv, y, d, p]
                    sdch = model.es_sdch_per_unit[y_inv, y, d, p]

                    _add_esso_cohort_constraint(model, 'energy_storage_limits', y_inv, y, sch <= s_max)
                    _add_esso_cohort_constraint(model, 'energy_storage_limits', y_inv, y, sdch <= s_max)

                    if shared_ess_data.params.slacks:
                        _add_esso_cohort_constraint(model, 'energy_storage_complementarity', y_inv, y, sch * sdch <= model.slack_es_ch_comp_per_unit[y_inv, y, d, p] + ESS_COMPLEMENTARITY_TOLERANCE)
                    else:
                        _add_esso_cohort_constraint(model, 'energy_storage_complementarity', y_inv, y, sch * sdch <= ESS_COMPLEMENTARITY_TOLERANCE)

    # - Shared ESS operation, aggregated
    model.energy_storage_operation_agg = pe.ConstraintList()
    for y in model.years:
        for d in model.days:
            for p in model.periods:

                agg_snet = 0.00
                for y_inv in model.years:
                    agg_snet += (model.es_sch_per_unit[y_inv, y, d, p] - model.es_sdch_per_unit[y_inv, y, d, p])

                if shared_ess_data.params.slacks:
                    model.energy_storage_operation_agg.add(model.es_snet[y, d, p] == agg_snet + model.slack_es_snet_up[y, d, p] - model.slack_es_snet_down[y, d, p])
                else:
                    model.energy_storage_operation_agg.add(model.es_snet[y, d, p] == agg_snet)

                if shared_ess_data.params.slacks:
                    model.energy_storage_operation_agg.add(model.es_snet[y, d, p] ** 2 == model.es_pnet[y, d, p] ** 2 + model.es_qnet[y, d, p] ** 2 + model.slack_es_snet_def_up[y, d, p] - model.slack_es_snet_def_down[y, d, p])
                else:
                    model.energy_storage_operation_agg.add(model.es_snet[y, d, p] ** 2 == model.es_pnet[y, d, p] ** 2 + model.es_qnet[y, d, p] ** 2)

    # ------------------------------------------------------------------------------------------------------------------
    # Objective function
    slack_penalty = 0.0
    for y_inv in model.years:

        # Slacks for investment fixing
        slack_penalty += PENALTY_ESSO_SLACK * (model.slack_es_s_investment_up[y_inv] + model.slack_es_s_investment_down[y_inv])
        slack_penalty += PENALTY_ESSO_SLACK * (model.slack_es_e_investment_up[y_inv] + model.slack_es_e_investment_down[y_inv])

        if shared_ess_data.params.slacks:

            # Degradation
            for y in model.years:
                slack_penalty += PENALTY_ESSO_SLACK * (model.slack_es_soh_per_unit_up[y_inv, y] + model.slack_es_soh_per_unit_down[y_inv, y])
                slack_penalty += PENALTY_ESSO_SLACK * (model.slack_es_soh_per_unit_cumul_up[y_inv, y] + model.slack_es_soh_per_unit_cumul_down[y_inv, y])

            # Complementarity
            for y in model.years:
                for d in model.days:
                    for p in model.periods:
                        slack_penalty += PENALTY_ESSO_SLACK * (model.slack_es_ch_comp_per_unit[y_inv, y, d, p])

            # Expected power slacks
            for d in model.days:
                for p in model.periods:
                    slack_penalty += PENALTY_ESSO_SLACK * (model.slack_es_snet_up[y_inv, d, p] + model.slack_es_snet_down[y_inv, d, p])
                    slack_penalty += PENALTY_ESSO_SLACK * (model.slack_es_snet_def_up[y_inv, d, p] + model.slack_es_snet_def_down[y_inv, d, p])

    salvage_value = _build_terminal_salvage_value_expression(shared_ess_data, model, shared_ess_idx)

    model.feasibility_penalty = pe.Expression(expr=slack_penalty)
    model.salvage_value = pe.Expression(expr=salvage_value)
    model.salvage_credit = pe.Expression(expr=-model.salvage_value)
    model.objective = pe.Objective(
        sense=pe.minimize,
        expr=model.feasibility_penalty,
    )

    # Define that we want the duals
    model.ipopt_zL_out = pe.Suffix(direction=pe.Suffix.IMPORT)  # Ipopt bound multipliers (obtained from solution)
    model.ipopt_zU_out = pe.Suffix(direction=pe.Suffix.IMPORT)
    model.ipopt_zL_in = pe.Suffix(direction=pe.Suffix.EXPORT)  # Ipopt bound multipliers (sent to solver)
    model.ipopt_zU_in = pe.Suffix(direction=pe.Suffix.EXPORT)
    model.dual = pe.Suffix(direction=pe.Suffix.IMPORT_EXPORT)

    return model


def _build_terminal_salvage_value_expression(shared_ess_data, model, shared_ess_idx):
    params = shared_ess_data.params.salvage_value
    if not params.enabled:
        return 0.00

    terminal_year_idx = len(shared_ess_data.years) - 1
    terminal_discount = _get_terminal_discount_factor(shared_ess_data)
    salvage_value = 0.00

    for y_inv in model.years:
        year_inv = list(shared_ess_data.years)[y_inv]
        shared_energy_storage = shared_ess_data.shared_energy_storages[year_inv][shared_ess_idx]
        min_soh = shared_energy_storage.soh_min
        if not 0.00 <= min_soh < 1.00:
            raise ValueError(f'Invalid minimum SoH {min_soh} for salvage valuation.')

        e_rated = model.es_e_rated_per_unit[y_inv, terminal_year_idx]
        e_available = model.es_e_available_per_unit[y_inv, terminal_year_idx]
        usable_energy_above_eol = (e_available - min_soh * e_rated) / (1.00 - min_soh)
        residual_energy = (
            params.recycling_floor_fraction * e_rated
            + (1.00 - params.recycling_floor_fraction) * usable_energy_above_eol
        )
        expected_unit_cost = _get_expected_energy_investment_cost(shared_ess_data, year_inv)
        _, _, remaining_life_fraction = _get_remaining_calendar_life(
            shared_ess_data, y_inv, shared_energy_storage
        )
        salvage_value += (
            terminal_discount
            * params.energy_recovery_fraction
            * expected_unit_cost
            * remaining_life_fraction
            * residual_energy
        )

    return salvage_value


def _get_terminal_discount_factor(shared_ess_data):
    terminal_years = sum(float(num_years) for num_years in shared_ess_data.years.values())
    return 1.00 / ((1.00 + shared_ess_data.discount_factor) ** terminal_years)


def _get_remaining_calendar_life(shared_ess_data, investment_year_idx, shared_energy_storage):
    if shared_energy_storage.t_cal <= 0.00:
        raise ValueError('Shared ESS calendar life must be positive for salvage valuation.')
    represented_years = list(shared_ess_data.years.values())
    age_at_terminal = sum(float(value) for value in represented_years[investment_year_idx:])
    remaining_life = max(float(shared_energy_storage.t_cal) - age_at_terminal, 0.00)
    remaining_life_fraction = min(remaining_life / float(shared_energy_storage.t_cal), 1.00)
    return age_at_terminal, remaining_life, remaining_life_fraction


def _get_expected_energy_investment_cost(shared_ess_data, year):
    expected_cost = 0.00
    for scenario, probability in enumerate(shared_ess_data.prob_market_scenarios):
        expected_cost += probability * shared_ess_data.cost_investment['energy'][scenario][year]
    return expected_cost


def _get_salvage_value_sensitivities(shared_ess_data, models):
    years = list(shared_ess_data.years)
    sensitivities = {
        's': {year: {} for year in years},
        'e': {year: {} for year in years},
    }
    params = shared_ess_data.params.salvage_value
    terminal_year_idx = len(years) - 1
    terminal_discount = _get_terminal_discount_factor(shared_ess_data)

    for node_id in shared_ess_data.active_distribution_network_nodes:
        model = models[node_id]
        shared_ess_idx = shared_ess_data.get_shared_energy_storage_idx(node_id)
        for y_inv, year_inv in enumerate(years):
            sensitivities['s'][year_inv][node_id] = 0.00
            sensitivities['e'][year_inv][node_id] = 0.00
            if not params.enabled or model.es_e_rated_per_unit[y_inv, terminal_year_idx].fixed:
                continue

            shared_energy_storage = shared_ess_data.shared_energy_storages[year_inv][shared_ess_idx]
            min_soh = shared_energy_storage.soh_min
            terminal_soh = pe.value(model.es_soh_per_unit_cumul[y_inv, terminal_year_idx])
            normalized_health = (terminal_soh - min_soh) / (1.00 - min_soh)
            residual_fraction = (
                params.recycling_floor_fraction
                + (1.00 - params.recycling_floor_fraction) * normalized_health
            )
            _, _, remaining_life_fraction = _get_remaining_calendar_life(
                shared_ess_data, y_inv, shared_energy_storage
            )
            expected_unit_cost = _get_expected_energy_investment_cost(shared_ess_data, year_inv)
            sensitivities['e'][year_inv][node_id] = -(
                terminal_discount
                * params.energy_recovery_fraction
                * expected_unit_cost
                * remaining_life_fraction
                * residual_fraction
            )

    return sensitivities


def _get_salvage_value_results(shared_ess_data, models):
    years = list(shared_ess_data.years)
    terminal_year_idx = len(years) - 1
    terminal_representative_year = years[terminal_year_idx]
    terminal_horizon_years = sum(float(value) for value in shared_ess_data.years.values())
    terminal_date = int(years[0]) + terminal_horizon_years
    if terminal_date.is_integer():
        terminal_date = int(terminal_date)
    terminal_discount = _get_terminal_discount_factor(shared_ess_data)
    params = shared_ess_data.params.salvage_value
    cohorts = list()

    for node_id in shared_ess_data.active_distribution_network_nodes:
        model = models[node_id]
        shared_ess_idx = shared_ess_data.get_shared_energy_storage_idx(node_id)
        for y_inv, year_inv in enumerate(years):
            shared_energy_storage = shared_ess_data.shared_energy_storages[year_inv][shared_ess_idx]
            min_soh = shared_energy_storage.soh_min
            age_at_terminal, remaining_life, remaining_life_fraction = _get_remaining_calendar_life(
                shared_ess_data, y_inv, shared_energy_storage
            )
            active_in_terminal_block = not model.es_e_rated_per_unit[y_inv, terminal_year_idx].fixed
            salvage_eligible_at_terminal = (
                active_in_terminal_block and remaining_life > SMALL_TOLERANCE
            )
            e_rated = pe.value(model.es_e_rated_per_unit[y_inv, terminal_year_idx])
            e_available = pe.value(model.es_e_available_per_unit[y_inv, terminal_year_idx])
            terminal_soh = None
            normalized_health = None
            residual_fraction = None
            salvage_value = 0.00

            if active_in_terminal_block:
                terminal_soh = pe.value(model.es_soh_per_unit_cumul[y_inv, terminal_year_idx])
                normalized_health = (terminal_soh - min_soh) / (1.00 - min_soh)
                residual_fraction = (
                    params.recycling_floor_fraction
                    + (1.00 - params.recycling_floor_fraction) * normalized_health
                )
                if params.enabled and salvage_eligible_at_terminal:
                    salvage_value = (
                        terminal_discount
                        * params.energy_recovery_fraction
                        * _get_expected_energy_investment_cost(shared_ess_data, year_inv)
                        * remaining_life_fraction
                        * e_rated
                        * residual_fraction
                    )

            cohorts.append({
                'node_id': node_id,
                'investment_year': year_inv,
                'terminal_representative_year': terminal_representative_year,
                'terminal_date': terminal_date,
                'active_in_terminal_block': active_in_terminal_block,
                'salvage_eligible_at_terminal': salvage_eligible_at_terminal,
                'calendar_life': shared_energy_storage.t_cal,
                'age_at_terminal': age_at_terminal,
                'remaining_calendar_life': remaining_life,
                'remaining_calendar_life_fraction': remaining_life_fraction,
                'terminal_rated_energy': e_rated,
                'terminal_available_energy': e_available,
                'minimum_soh': min_soh,
                'terminal_soh': terminal_soh,
                'normalized_health': normalized_health,
                'residual_fraction': residual_fraction,
                'expected_unit_energy_cost': _get_expected_energy_investment_cost(
                    shared_ess_data, year_inv
                ),
                'terminal_discount_factor': terminal_discount,
                'salvage_value': salvage_value,
            })

    return {
        'enabled': params.enabled,
        'energy_recovery_fraction': params.energy_recovery_fraction,
        'recycling_floor_fraction': params.recycling_floor_fraction,
        'cost_basis': params.cost_basis,
        'health_basis': params.health_basis,
        'calendar_life_basis': params.calendar_life_basis,
        'terminal_horizon_years': terminal_horizon_years,
        'terminal_discount_factor': terminal_discount,
        'cohorts': cohorts,
        'total_salvage_value': sum(cohort['salvage_value'] for cohort in cohorts),
    }


def _create_solver(model, params, from_warm_start=False, node_id=None, option_overrides=None, log_suffix=None):

    solver = po.SolverFactory(params.solver, executable=params.solver_path)
    options = dict()
    if params.verbose and params.solver.lower() == 'ipopt':
        options['print_level'] = 6
    if params.options:
        options.update(params.options)
    if option_overrides:
        options.update(option_overrides)

    solver_log_path = None
    if params.solver.lower() == 'ipopt':
        if 'output_file' not in options:
            options['output_file'] = (
                f'optim_log_node_{node_id}.txt' if node_id is not None else 'optim_log.txt'
            )
        if log_suffix:
            path_stem, path_extension = os.path.splitext(options['output_file'])
            options['output_file'] = f'{path_stem}_{log_suffix}{path_extension}'
        options['file_append'] = 'yes'
        solver_log_path = os.path.abspath(options['output_file'])

    for key, value in options.items():
        solver.options[key] = value

    if from_warm_start and params.solver.lower() == 'ipopt':
        model.ipopt_zL_in.update(model.ipopt_zL_out)
        model.ipopt_zU_in.update(model.ipopt_zU_out)
        solver.options['warm_start_init_point'] = 'yes'
        solver.options['warm_start_bound_push'] = 1e-9
        solver.options['warm_start_bound_frac'] = 1e-9
        solver.options['warm_start_slack_bound_frac'] = 1e-9
        solver.options['warm_start_slack_bound_push'] = 1e-9
        solver.options['warm_start_mult_bound_push'] = 1e-9

    return solver, solver_log_path


def _run_solver_attempt(model, params, solve_context, from_warm_start=False, node_id=None, option_overrides=None, log_suffix=None):

    solver, solver_log_path = _create_solver(
        model,
        params,
        from_warm_start=from_warm_start,
        node_id=node_id,
        option_overrides=option_overrides,
        log_suffix=log_suffix,
    )
    result = None
    try:
        result = solver.solve(model, tee=params.verbose, load_solutions=False)
    except (ValueError, RuntimeError) as error:
        print(f'[WARNING] Shared ESS solver execution failed for {solve_context}: {error}')
    return result, solver_log_path


def _is_recoverable_shared_ess_failure(result, params, node_id):
    if node_id is None or params.solver.lower() != 'ipopt' or result is None:
        return False
    recovery_options = getattr(params, 'recovery_options', None)
    if not recovery_options or not hasattr(result, 'solver'):
        return False
    return result.solver.termination_condition == po.TerminationCondition.internalSolverError


def _format_solver_options(options):
    return ', '.join(f'{key}={value}' for key, value in sorted(options.items()))


def _optimize(model, params, from_warm_start=False, node_id=None, diagnostic_sink=None):

    solve_context = f'ESS node={node_id}' if node_id is not None else 'master problem'
    primary_result, primary_log_path = _run_solver_attempt(
        model,
        params,
        solve_context,
        from_warm_start=from_warm_start,
        node_id=node_id,
    )
    result = primary_result
    recovery_result = None
    recovery_log_path = None
    recovery_attempted = _is_recoverable_shared_ess_failure(primary_result, params, node_id)

    if recovery_attempted:
        recovery_options = params.recovery_options
        print(
            f'[WARNING] Shared ESS primary solve did not converge for {solve_context}: '
            f'{solver_result_summary(primary_result)} | warm_start={from_warm_start}'
        )
        if primary_log_path:
            print(f'[WARNING] IPOPT primary log for {solve_context}: {primary_log_path}')
        print(
            f'[INFO] Retrying Shared ESS solve once for {solve_context} with '
            f'{_format_solver_options(recovery_options)}.'
        )
        recovery_result, recovery_log_path = _run_solver_attempt(
            model,
            params,
            solve_context,
            from_warm_start=from_warm_start,
            node_id=node_id,
            option_overrides=recovery_options,
            log_suffix='recovery',
        )
        result = recovery_result if recovery_result is not None else primary_result

    if solver_result_succeeded(result):
        try:
            model.solutions.load_from(result)
        except ValueError as error:
            print(f'[WARNING] Shared ESS solution could not be loaded for {solve_context}: {error}')
            result = None
        if recovery_attempted and result is not None:
            print(f'[INFO] Shared ESS recovery solve succeeded for {solve_context}.')
    else:
        attempt_label = 'recovery' if recovery_attempted else 'solver'
        failed_attempt_result = recovery_result if recovery_attempted else result
        print(
            f'[WARNING] Shared ESS {attempt_label} did not converge for {solve_context}: '
            f'{solver_result_summary(failed_attempt_result)} | warm_start={from_warm_start}'
        )
        final_log_path = recovery_log_path if recovery_attempted else primary_log_path
        if final_log_path:
            print(f'[WARNING] IPOPT {attempt_label} log for {solve_context}: {final_log_path}')

    if recovery_attempted and diagnostic_sink is not None:
        diagnostic_sink.append({
            'subsystem': 'esso',
            'node_id': node_id,
            'warm_start': from_warm_start,
            'primary_result': solver_result_summary(primary_result),
            'recovery_result': solver_result_summary(recovery_result),
            'recovery_options': _format_solver_options(params.recovery_options),
            'recovery_succeeded': solver_result_succeeded(result),
            'primary_log': primary_log_path,
            'recovery_log': recovery_log_path,
        })

    return result


def _update_model_with_candidate_solution(shared_ess_data, models, candidate_solution):
    repr_years = [year for year in shared_ess_data.years]
    for node_id in models:
        model = models[node_id]
        for y_inv in model.years:
            year = repr_years[y_inv]
            s_candidate = candidate_solution[node_id][year]['s']
            e_candidate = candidate_solution[node_id][year]['e']
            model.es_s_investment_fixed[y_inv].set_value(s_candidate)
            model.es_e_investment_fixed[y_inv].set_value(e_candidate)
            model.es_s_investment[y_inv].fix(s_candidate)
            model.es_e_investment[y_inv].fix(e_candidate)
            _configure_esso_cohort_state(model, y_inv, s_candidate, e_candidate, shared_ess_data.params.slacks)


def _get_candidate_solution(self, model):
    years = [year for year in self.years]
    candidate_solution = {'investment': {}, 'total_capacity': {}}
    for e in model.energy_storages:
        node_id = self.shared_energy_storages[years[0]][e].bus
        candidate_solution['investment'][node_id] = dict()
        candidate_solution['total_capacity'][node_id] = dict()
        for y in model.years:
            year = years[y]
            candidate_solution['investment'][node_id][year] = dict()
            candidate_solution['investment'][node_id][year]['s'] = abs(pe.value(model.es_s_investment[e, y]))
            candidate_solution['investment'][node_id][year]['e'] = abs(pe.value(model.es_e_investment[e, y]))
            candidate_solution['total_capacity'][node_id][year] = dict()
            candidate_solution['total_capacity'][node_id][year]['s'] = abs(pe.value(model.es_s_rated[e, y]))
            candidate_solution['total_capacity'][node_id][year]['e'] = abs(pe.value(model.es_e_rated[e, y]))
    return candidate_solution


def _load_candidate_solution_into_master_model(shared_ess_data, model, candidate_solution):
    years = list(shared_ess_data.years)
    for e in model.energy_storages:
        node_id = shared_ess_data.shared_energy_storages[years[0]][e].bus
        for y, year in enumerate(years):
            investment = candidate_solution['investment'][node_id][year]
            total_capacity = candidate_solution['total_capacity'][node_id][year]
            model.es_s_investment[e, y].set_value(investment['s'])
            model.es_e_investment[e, y].set_value(investment['e'])
            model.es_s_rated[e, y].set_value(total_capacity['s'])
            model.es_e_rated[e, y].set_value(total_capacity['e'])


def _map_available_capacity_sensitivities_to_investments(shared_ess_data, models, sensitivities):

    years = list(shared_ess_data.years)
    investment_sensitivities = {'s': dict(), 'e': dict()}
    for year_inv in years:
        investment_sensitivities['s'][year_inv] = dict()
        investment_sensitivities['e'][year_inv] = dict()

    for node_id in shared_ess_data.active_distribution_network_nodes:

        model = models[node_id]

        for y_inv, year_inv in enumerate(years):

            if model._esso_cohort_inactive.get(y_inv, False):
                investment_sensitivities['s'][year_inv][node_id] = None
                investment_sensitivities['e'][year_inv][node_id] = None
                continue

            sensitivity_s = 0.00
            sensitivity_e = 0.00
            sensitivity_s_available = True
            sensitivity_e_available = True

            for y, year in enumerate(years):

                if not model.es_s_rated_per_unit[y_inv, y].fixed:
                    value_s = sensitivities['s'][year][node_id]
                    if value_s is None:
                        sensitivity_s_available = False
                    else:
                        sensitivity_s += value_s

                if not model.es_e_rated_per_unit[y_inv, y].fixed:
                    value_e = sensitivities['e'][year][node_id]
                    if value_e is None:
                        sensitivity_e_available = False
                    else:
                        # Local chain rule: hold the converged SoH trajectory fixed.
                        soh = pe.value(model.es_soh_per_unit_cumul[y_inv, y])
                        sensitivity_e += value_e * soh

            investment_sensitivities['s'][year_inv][node_id] = sensitivity_s if sensitivity_s_available else None
            investment_sensitivities['e'][year_inv][node_id] = sensitivity_e if sensitivity_e_available else None

    return investment_sensitivities


def _add_esso_cohort_constraint(model, constraint_name, y_inv, y, expr):
    constraint_list = getattr(model, constraint_name)
    constraint_list.add(expr)
    constraint_idx = len(constraint_list)
    model._esso_cohort_constraints[y_inv].append((constraint_name, constraint_idx, y))


def _set_esso_variable_state(variable, active, inactive_value):
    if active:
        if variable.fixed:
            variable.unfix()
    else:
        variable.fix(inactive_value)


def _esso_cohort_pair_is_within_lifetime(model, y_inv, y):
    # These variables are already fixed to zero outside the calendar lifetime when the model is constructed.
    return (not model.es_s_rated_per_unit[y_inv, y].fixed and not model.es_e_rated_per_unit[y_inv, y].fixed)


def _configure_esso_cohort_state(model, y_inv, s_capacity, e_capacity, slacks_enabled):

    tolerance = SHARED_ESS_ZERO_CAPACITY_TOLERANCE
    s_is_zero = abs(s_capacity) <= tolerance
    e_is_zero = abs(e_capacity) <= tolerance
    if s_is_zero != e_is_zero:
        raise ValueError(f'Inconsistent shared ESS cohort capacity: S={s_capacity}, E={e_capacity}.')
    inactive = s_is_zero and e_is_zero

    model._esso_cohort_inactive[y_inv] = inactive

    # ------------------------------------------------------------------
    # Variables
    # ------------------------------------------------------------------
    for y in model.years:

        within_lifetime = _esso_cohort_pair_is_within_lifetime(model, y_inv, y)
        active_pair = (not inactive and within_lifetime)

        # Annual degradation quantities
        _set_esso_variable_state(model.es_avg_ch_dch_per_unit[y_inv, y], active_pair, 0.0)
        _set_esso_variable_state(model.es_degradation_per_unit[y_inv, y], active_pair, 0.0)
        _set_esso_variable_state(model.es_degradation_per_unit_cumul[y_inv, y], active_pair, 0.0)
        _set_esso_variable_state(model.es_soh_per_unit[y_inv, y], active_pair, 1.0)
        _set_esso_variable_state(model.es_soh_per_unit_cumul[y_inv, y], active_pair, 1.0)
        if slacks_enabled:
            _set_esso_variable_state(model.slack_es_soh_per_unit_up[y_inv, y], active_pair, 0.0)
            _set_esso_variable_state(model.slack_es_soh_per_unit_down[y_inv, y], active_pair, 0.0)
            _set_esso_variable_state(model.slack_es_soh_per_unit_cumul_up[y_inv, y], active_pair, 0.0)
            _set_esso_variable_state(model.slack_es_soh_per_unit_cumul_down[y_inv, y], active_pair, 0.0)

        # --------------------------------------------------------------
        # Time-dependent operation
        # --------------------------------------------------------------
        for d in model.days:
            for p in model.periods:
                _set_esso_variable_state(model.es_sch_per_unit[y_inv, y, d, p], active_pair, 0.0)
                _set_esso_variable_state(model.es_sdch_per_unit[y_inv, y, d, p], active_pair, 0.0)
                if slacks_enabled:
                    _set_esso_variable_state(model.slack_es_ch_comp_per_unit[y_inv, y, d, p], active_pair, 0.0)

    # ------------------------------------------------------------------
    # Constraints
    # ------------------------------------------------------------------
    for constraint_name, constraint_idx, y in (model._esso_cohort_constraints[y_inv]):
        constraint = getattr(model, constraint_name)[constraint_idx]
        active_constraint = (not inactive and _esso_cohort_pair_is_within_lifetime(model, y_inv, y))
        if active_constraint:
            constraint.activate()
        else:
            constraint.deactivate()

    return inactive



# ======================================================================================================================
#  NETWORK PLANNING read functions
# ======================================================================================================================
def _read_shared_energy_storage_data_from_file(shared_ess_data, filename):

    try:
        num_scenarios, shared_ess_data.prob_market_scenarios = _get_operational_scenarios_info_from_excel_file(filename, 'Scenarios')
        investment_costs = dict()
        investment_costs['power'] = _get_investment_costs_from_excel_file(filename, 'Investment Cost, Power', num_scenarios, shared_ess_data.years)
        investment_costs['energy'] = _get_investment_costs_from_excel_file(filename, 'Investment Cost, Energy', num_scenarios, shared_ess_data.years)
        shared_ess_data.cost_investment = investment_costs
    except:
        print(f'[ERROR] File {filename}. Exiting...')
        exit(ERROR_OPERATIONAL_DATA_FILE)


def _get_operational_scenarios_info_from_excel_file(filename, sheet_name):

    num_scenarios = 0
    prob_scenarios = list()

    try:
        df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
        if is_int(df.iloc[0, 1]):
            num_scenarios = int(df.iloc[0, 1])
        for i in range(num_scenarios):
            if is_number(df.iloc[0, i + 2]):
                prob_scenarios.append(float(df.iloc[0, i + 2]))
    except:
        print('[ERROR] Workbook {}. Sheet {} does not exist.'.format(filename, sheet_name))
        exit(1)

    if num_scenarios != len(prob_scenarios):
        print('[WARNING] EnergyStorage file. Number of scenarios different from the probability vector!')

    if round(sum(prob_scenarios), 2) != 1.00:
        print('[ERROR] Probability of scenarios does not add up to 100%. Check file {}. Exiting.'.format(filename))
        exit(ERROR_OPERATIONAL_DATA_FILE)

    return num_scenarios, prob_scenarios


def _get_investment_costs_from_excel_file(filename, sheet_name, num_scenarios, years):

    try:
        df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
        data = dict()
        for i in range(num_scenarios):
            data[i] = dict()
            for year in years:
                year_found = False
                for j in range(len(df.columns) - 1):
                    year_excel = int(df.iloc[0, j + 1])
                    if year == year_excel:
                        year_found = True
                        if is_number(df.iloc[i + 1, j + 1]):
                            data[i][year] = float(df.iloc[i + 1, j + 1])
                if not year_found:
                    print('[ERROR] Workbook {}. Year {} not found!'.format(filename, year))
                    exit(ERROR_MARKET_DATA_FILE)
        return data
    except:
        print('[ERROR] Workbook {}. Sheet {} does not exist.'.format(filename, sheet_name))
        exit(ERROR_MARKET_DATA_FILE)


# ======================================================================================================================
#   Shared ESS -- Process results
# ======================================================================================================================
def _process_results_aggregated(shared_ess_data, models):

    processed_results = dict()

    repr_days = [day for day in shared_ess_data.days]
    repr_years = [year for year in shared_ess_data.years]
    for year in repr_years:
        processed_results[year] = dict()
        for day in repr_days:
            processed_results[year][day] = dict()
            for node_id in shared_ess_data.active_distribution_network_nodes:
                processed_results[year][day][node_id] = dict()
                processed_results[year][day][node_id]['p'] = list()
                processed_results[year][day][node_id]['q'] = list()

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for y in models[node_id].years:
            year = repr_years[y]
            for d in models[node_id].days:
                day = repr_days[d]
                for p in models[node_id].periods:
                    p_net = pe.value(models[node_id].es_pnet[y, d, p])
                    q_net = pe.value(models[node_id].es_qnet[y, d, p])
                    processed_results[year][day][node_id]['p'].append(p_net)
                    processed_results[year][day][node_id]['q'].append(q_net)

    return processed_results


def _process_results_detailed(shared_ess_data, models):

    repr_days = [day for day in shared_ess_data.days]
    repr_years = [year for year in shared_ess_data.years]

    processed_results = dict()
    for year_inv in repr_years:
        processed_results[year_inv] = dict()
        for year_curr in repr_years:
            processed_results[year_inv][year_curr] = dict()
            for day in repr_days:
                processed_results[year_inv][year_curr][day] = dict()
                for node_id in shared_ess_data.active_distribution_network_nodes:
                    processed_results[year_inv][year_curr][day][node_id] = dict()
                    processed_results[year_inv][year_curr][day][node_id]['s'] = list()

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for y_inv in models[node_id].years:
            year_inv = repr_years[y_inv]
            for y_curr in models[node_id].years:
                year_curr = repr_years[y_curr]
                for d in models[node_id].days:
                    day = repr_days[d]
                    for p in models[node_id].periods:
                        s_net = pe.value(models[node_id].es_sch_per_unit[y_inv, y_curr, d, p] - models[node_id].es_sdch_per_unit[y_inv, y_curr, d, p])
                        processed_results[year_inv][year_curr][day][node_id]['s'].append(s_net)

    return processed_results


def _process_soh_results_aggregated(shared_ess_data, models):

    repr_years = [year for year in shared_ess_data.years]

    processed_results = dict()
    for year in repr_years:
        processed_results[year] = {
            's_rated': dict(), 'e_rated': dict(),
            's_available': dict(), 'e_available': dict(),
            'soh': dict(), 'degradation': dict()
        }

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for y in models[node_id].years:
            year = repr_years[y]
            s_rated = pe.value(models[node_id].es_s_rated[y])
            e_rated = pe.value(models[node_id].es_e_rated[y])
            s_available, e_available = shared_ess_data.get_available_capacities(models[node_id], y)
            soh = 1.00
            if not isclose(e_available, 0.00, abs_tol=SMALL_TOLERANCE):
                soh = e_available / e_rated
            degradation = 1 - soh
            processed_results[year]['s_rated'][node_id] = s_rated
            processed_results[year]['e_rated'][node_id] = e_rated
            processed_results[year]['s_available'][node_id] = s_available
            processed_results[year]['e_available'][node_id] = e_available
            processed_results[year]['soh'][node_id] = soh
            processed_results[year]['degradation'][node_id] = degradation

    return processed_results


def _process_soh_results_detailed(shared_ess_data, models):

    repr_years = [year for year in shared_ess_data.years]

    processed_results = dict()
    for year_inv in repr_years:
        processed_results[year_inv] = dict()
        for year_curr in repr_years:
            processed_results[year_inv][year_curr] = {
                's_rated': dict(), 'e_rated': dict(),
                's_available': dict(), 'e_available': dict(),
                'soh_unit': dict(), 'degradation_unit': dict(),
                'soh_cumul': dict(), 'degradation_cumul': dict()
            }

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for y_inv in models[node_id].years:
            year_inv = repr_years[y_inv]
            for y_curr in models[node_id].years:
                year_curr = repr_years[y_curr]
                s_rated = pe.value(models[node_id].es_s_rated_per_unit[y_inv, y_curr])
                e_rated = pe.value(models[node_id].es_e_rated_per_unit[y_inv, y_curr])
                s_available = pe.value(models[node_id].es_s_available_per_unit[y_inv, y_curr])
                e_available = pe.value(models[node_id].es_e_available_per_unit[y_inv, y_curr])
                soh_unit = pe.value(models[node_id].es_soh_per_unit[y_inv, y_curr])
                degradation_unit = pe.value(models[node_id].es_degradation_per_unit[y_inv, y_curr])
                soh_cumul = pe.value(models[node_id].es_soh_per_unit_cumul[y_inv, y_curr])
                degradation_cumul = pe.value(models[node_id].es_degradation_per_unit_cumul[y_inv, y_curr])
                processed_results[year_inv][year_curr]['s_rated'][node_id] = s_rated
                processed_results[year_inv][year_curr]['e_rated'][node_id] = e_rated
                processed_results[year_inv][year_curr]['s_available'][node_id] = s_available
                processed_results[year_inv][year_curr]['e_available'][node_id] = e_available
                processed_results[year_inv][year_curr]['soh_unit'][node_id] = soh_unit
                processed_results[year_inv][year_curr]['degradation_unit'][node_id] = degradation_unit
                processed_results[year_inv][year_curr]['soh_cumul'][node_id] = soh_cumul
                processed_results[year_inv][year_curr]['degradation_cumul'][node_id] = degradation_cumul

    return processed_results


def _process_relaxation_variables_investment(shared_ess_data, models):

    repr_years = [year for year in shared_ess_data.years]

    processed_results = dict()
    for year_inv in repr_years:
        processed_results[year_inv] = dict()
        for node_id in shared_ess_data.active_distribution_network_nodes:
            processed_results[year_inv][node_id] = dict()

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for y_inv in models[node_id].years:
            year_inv = repr_years[y_inv]
            processed_results[year_inv][node_id] = dict()
            processed_results[year_inv][node_id]['s_up'] = pe.value(models[node_id].slack_es_s_investment_up[y_inv])
            processed_results[year_inv][node_id]['s_down'] = pe.value(models[node_id].slack_es_s_investment_down[y_inv])
            processed_results[year_inv][node_id]['e_up'] = pe.value(models[node_id].slack_es_e_investment_up[y_inv])
            processed_results[year_inv][node_id]['e_down'] = pe.value(models[node_id].slack_es_e_investment_down[y_inv])

    return processed_results


def _process_relaxation_variables_degradation_detailed(shared_ess_data, models):

    repr_years = [year for year in shared_ess_data.years]

    processed_results = dict()
    for year_inv in repr_years:
        processed_results[year_inv] = dict()
        for year_curr in repr_years:
            processed_results[year_inv][year_curr] = dict()
            for node_id in shared_ess_data.active_distribution_network_nodes:
                processed_results[year_inv][year_curr][node_id] = dict()

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for y_inv in models[node_id].years:
            year_inv = repr_years[y_inv]
            for y_curr in models[node_id].years:
                year_curr = repr_years[y_curr]
                if shared_ess_data.params.slacks:
                    # - Degradation per unit
                    soh_per_unit_up = pe.value(models[node_id].slack_es_soh_per_unit_up[y_inv, y_curr])
                    soh_per_unit_down = pe.value(models[node_id].slack_es_soh_per_unit_down[y_inv, y_curr])
                    soh_per_unit_cumul_up = pe.value(models[node_id].slack_es_soh_per_unit_cumul_up[y_inv, y_curr])
                    soh_per_unit_cumul_down = pe.value(models[node_id].slack_es_soh_per_unit_cumul_down[y_inv, y_curr])
                    processed_results[year_inv][year_curr][node_id]['soh_per_unit_up'] = soh_per_unit_up
                    processed_results[year_inv][year_curr][node_id]['soh_per_unit_down'] = soh_per_unit_down
                    processed_results[year_inv][year_curr][node_id]['soh_per_unit_cumul_up'] = soh_per_unit_cumul_up
                    processed_results[year_inv][year_curr][node_id]['soh_per_unit_cumul_down'] = soh_per_unit_cumul_down

    return processed_results


def _process_relaxation_variables_operation_aggregated(shared_ess_data, models):

    repr_days = [day for day in shared_ess_data.days]
    repr_years = [year for year in shared_ess_data.years]

    processed_results = dict()
    for year in repr_years:
        processed_results[year] = dict()
        for day in repr_days:
            processed_results[year][day] = dict()
            for node_id in shared_ess_data.active_distribution_network_nodes:
                processed_results[year][day][node_id] = dict()
                if shared_ess_data.params.slacks:
                    processed_results[year][day][node_id]['snet_up'] = list()
                    processed_results[year][day][node_id]['snet_down'] = list()
                    processed_results[year][day][node_id]['snet_def_up'] = list()
                    processed_results[year][day][node_id]['snet_def_down'] = list()

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for y in models[node_id].years:
            year = repr_years[y]
            for d in models[node_id].days:
                day = repr_days[d]
                for p in models[node_id].periods:
                    slack_es_snet_up = pe.value(models[node_id].slack_es_snet_up[y, d, p])
                    slack_es_snet_down = pe.value(models[node_id].slack_es_snet_down[y, d, p])
                    slack_es_snet_def_up = pe.value(models[node_id].slack_es_snet_def_up[y, d, p])
                    slack_es_snet_def_down = pe.value(models[node_id].slack_es_snet_def_down[y, d, p])
                    processed_results[year][day][node_id]['snet_up'].append(slack_es_snet_up)
                    processed_results[year][day][node_id]['snet_down'].append(slack_es_snet_down)
                    processed_results[year][day][node_id]['snet_def_up'].append(slack_es_snet_def_up)
                    processed_results[year][day][node_id]['snet_def_down'].append(slack_es_snet_def_down)

    return processed_results


def _process_relaxation_variables_operation_detailed(shared_ess_data, models):

    repr_days = [day for day in shared_ess_data.days]
    repr_years = [year for year in shared_ess_data.years]

    processed_results = dict()
    for year_inv in repr_years:
        processed_results[year_inv] = dict()
        for year_curr in repr_years:
            processed_results[year_inv][year_curr] = dict()
            for day in repr_days:
                processed_results[year_inv][year_curr][day] = dict()
                for node_id in shared_ess_data.active_distribution_network_nodes:
                    processed_results[year_inv][year_curr][day][node_id] = dict()
                    processed_results[year_inv][year_curr][day][node_id]['comp'] = list()

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for y_inv in models[node_id].years:
            year_inv = repr_years[y_inv]
            for y_curr in models[node_id].years:
                year_curr = repr_years[y_curr]
                for d in models[node_id].days:
                    day = repr_days[d]
                    if shared_ess_data.params.slacks:
                        # - Complementarity
                        for p in models[node_id].periods:
                            comp = pe.value(models[node_id].slack_es_ch_comp_per_unit[y_inv, y_curr, d, p])
                            processed_results[year_inv][year_curr][day][node_id]['comp'].append(comp)

    return processed_results


def _get_available_capacity(shared_ess_data, models):

    years = [year for year in shared_ess_data.years]
    ess_capacity = {'investment': dict(), 'rated': dict(), 'available': dict()}

    # - Investment in Power and Energy Capacity (per year)
    # - Power and Energy capacities available (per representative day)
    for node_id in shared_ess_data.active_distribution_network_nodes:

        ess_capacity['investment'][node_id] = dict()
        ess_capacity['rated'][node_id] = dict()
        ess_capacity['available'][node_id] = dict()

        for y in models[node_id].years:

            year = years[y]

            ess_capacity['investment'][node_id][year] = dict()
            ess_capacity['investment'][node_id][year]['power'] = pe.value(models[node_id].es_s_investment[y])
            ess_capacity['investment'][node_id][year]['energy'] = pe.value(models[node_id].es_e_investment[y])

            ess_capacity['rated'][node_id][year] = dict()
            ess_capacity['rated'][node_id][year]['power'] = pe.value(models[node_id].es_s_rated[y])
            ess_capacity['rated'][node_id][year]['energy'] = pe.value(models[node_id].es_e_rated[y])

            s_available, e_available = shared_ess_data.get_available_capacities(models[node_id], y)
            soh = 0.00
            if not isclose(e_available, 0.00, abs_tol=SMALL_TOLERANCE):
                soh = e_available / pe.value(models[node_id].es_e_rated[y])
            ess_capacity['available'][node_id][year] = dict()
            ess_capacity['available'][node_id][year]['power'] = s_available
            ess_capacity['available'][node_id][year]['energy'] = e_available
            ess_capacity['available'][node_id][year]['soh'] = soh
            ess_capacity['available'][node_id][year]['degradation_factor'] = 1 - soh

    return ess_capacity


def _get_investment_cost_and_rated_capacity(shared_ess_data, model):

    years = [year for year in shared_ess_data.years]
    ess_investment = {'capacity': dict(), 'cost': dict()}

    # - Investment in Power and Energy Capacity (per year)
    # - Power and Energy capacities available (per representative day)
    for e in model.energy_storages:

        node_id = shared_ess_data.shared_energy_storages[years[0]][e].bus
        ess_investment['capacity'][node_id] = dict()
        ess_investment['cost'][node_id] = dict()

        for y in model.years:

            year = years[y]

            ess_investment['capacity'][node_id][year] = {
                'power': pe.value(model.es_s_investment[e, y]),
                'energy': pe.value(model.es_e_investment[e, y])
            }

            ess_investment['cost'][node_id][year] = {'power': {}, 'energy': {}}
            expected_cost_power = 0.00
            expected_cost_energy = 0.00
            for s_m in model.scenarios_market:
                omega_market = shared_ess_data.prob_market_scenarios[s_m]
                ess_investment['cost'][node_id][year]['power'][s_m] = shared_ess_data.cost_investment['power'][s_m][year] * pe.value(model.es_s_investment[e, y])
                ess_investment['cost'][node_id][year]['energy'][s_m] = shared_ess_data.cost_investment['energy'][s_m][year] * pe.value(model.es_e_investment[e, y])
                expected_cost_power += omega_market * ess_investment['cost'][node_id][year]['power'][s_m]
                expected_cost_energy += omega_market * ess_investment['cost'][node_id][year]['energy'][s_m]
            ess_investment['cost'][node_id][year]['power']['expected'] = expected_cost_power
            ess_investment['cost'][node_id][year]['energy']['expected'] = expected_cost_energy

    return ess_investment


# ======================================================================================================================
#   Shared ESS -- Write Results
# ======================================================================================================================
def _write_optimization_results_to_excel(shared_ess_data, data_dir, results):

    wb = Workbook()

    _write_ess_capacity_investment_to_excel(shared_ess_data, wb, results['capacity']['investment'])
    _write_ess_capacity_rated_available_to_excel(shared_ess_data, wb, results['capacity'])
    _write_aggregated_shared_energy_storage_operation_results_to_excel(shared_ess_data, wb, results['operation']['aggregated'])
    _write_detailed_shared_energy_storage_operation_results_to_excel(shared_ess_data, wb, results['operation']['detailed'])
    _write_aggregated_shared_energy_storage_soh_results_to_excel(shared_ess_data, wb, results['soh']['aggregated'])
    _write_detailed_shared_energy_storage_soh_results_to_excel(shared_ess_data, wb, results['soh']['detailed'])
    _write_terminal_salvage_value_to_excel(wb, results['salvage_value'])
    shared_ess_data.write_relaxation_slacks_results_to_excel(wb, results)

    results_filename = os.path.join(data_dir, f'{shared_ess_data.name}_shared_ess_results.xlsx')
    try:
        wb.save(results_filename)
        print('[INFO] S-MPOPF Results written to {}.'.format(results_filename))
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = os.path.join(data_dir, f'{shared_ess_data.name}_shared_ess_results_{current_time}.xlsx')
        print('[INFO] S-MPOPF Results written to {}.'.format(backup_filename))
        wb.save(backup_filename)


def _write_terminal_salvage_value_to_excel(workbook, salvage_results):
    sheet = workbook.create_sheet('Terminal Salvage')
    columns = [
        ('node_id', 'Node ID', '0'),
        ('investment_year', 'Investment Year', '0'),
        ('terminal_representative_year', 'Terminal Representative Year', '0'),
        ('terminal_date', 'Terminal Valuation Date', '0'),
        ('active_in_terminal_block', 'Active in Terminal Representative Block', 'General'),
        ('salvage_eligible_at_terminal', 'Eligible at Terminal Date', 'General'),
        ('calendar_life', 'Calendar Life, [years]', '0.000000'),
        ('age_at_terminal', 'Age at Terminal Date, [years]', '0.000000'),
        ('remaining_calendar_life', 'Remaining Calendar Life, [years]', '0.000000'),
        ('remaining_calendar_life_fraction', 'Remaining Calendar Life Fraction', '0.000000'),
        ('terminal_rated_energy', 'Terminal Rated Energy, [MVAh]', '0.000000'),
        ('terminal_available_energy', 'Terminal Available Energy, [MVAh]', '0.000000'),
        ('minimum_soh', 'Minimum SoH, [p.u.]', '0.000000'),
        ('terminal_soh', 'Terminal SoH, [p.u.]', '0.000000'),
        ('normalized_health', 'Normalized Health Above Minimum, [p.u.]', '0.000000'),
        ('residual_fraction', 'Recoverable Residual Fraction, [p.u.]', '0.000000'),
        ('expected_unit_energy_cost', 'Expected Installation Energy Cost, [m.u./MVAh]', '0.000000'),
        ('terminal_discount_factor', 'Terminal Discount Factor', '0.000000'),
        ('salvage_value', 'Terminal Salvage Value, [NPV m.u.]', '0.000000'),
    ]

    metadata = [
        ('Enabled', salvage_results.get('enabled')),
        ('Energy Recovery Fraction', salvage_results.get('energy_recovery_fraction')),
        ('Recycling Floor Fraction', salvage_results.get('recycling_floor_fraction')),
        ('Cost Basis', salvage_results.get('cost_basis')),
        ('Health Basis', salvage_results.get('health_basis')),
        ('Calendar-Life Basis', salvage_results.get('calendar_life_basis')),
        ('Terminal Horizon, [years]', salvage_results.get('terminal_horizon_years')),
        ('Terminal Discount Factor', salvage_results.get('terminal_discount_factor')),
        ('Total Terminal Salvage Value, [NPV m.u.]', salvage_results.get('total_salvage_value')),
    ]
    for row_idx, (label, value) in enumerate(metadata, start=1):
        sheet.cell(row=row_idx, column=1).value = label
        sheet.cell(row=row_idx, column=2).value = value

    header_row = len(metadata) + 2
    for column_idx, (_, label, _) in enumerate(columns, start=1):
        sheet.cell(row=header_row, column=column_idx).value = label
    for row_idx, cohort in enumerate(salvage_results.get('cohorts', []), start=header_row + 1):
        for column_idx, (key, _, number_format) in enumerate(columns, start=1):
            value = cohort.get(key)
            if value is None:
                continue
            sheet.cell(row=row_idx, column=column_idx).value = value
            sheet.cell(row=row_idx, column=column_idx).number_format = number_format


def _write_ess_costs_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Shared ESS Cost')
    num_style = '0.00'

    years = [year for year in shared_ess_data.years]
    total_cost_power = dict()
    total_cost_energy = dict()
    for year in years:
        total_cost_power[year] = 0.00
        total_cost_energy[year] = 0.00

    # Write Header
    line_idx = 1
    sheet.cell(row=line_idx, column=1).value = 'Node'
    sheet.cell(row=line_idx, column=2).value = 'Quantity'
    sheet.cell(row=line_idx, column=3).value = 'Scenario'
    for y in range(len(years)):
        year = years[y]
        sheet.cell(row=line_idx, column=y + 4).value = int(year)

    # Write power and energy capacity investment
    for node_id in shared_ess_data.active_distribution_network_nodes:

        # Power
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'S, [MVA]'
        sheet.cell(row=line_idx, column=3).value = 'N/A'
        for y in range(len(years)):
            year = years[y]
            sheet.cell(row=line_idx, column=y + 4).value = results['capacity'][node_id][year]['power']
            sheet.cell(row=line_idx, column=y + 4).number_format = num_style

        # Energy
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'E, [MVAh]'
        sheet.cell(row=line_idx, column=3).value = 'N/A'
        for y in range(len(years)):
            year = years[y]
            sheet.cell(row=line_idx, column=y + 4).value = results['capacity'][node_id][year]['energy']
            sheet.cell(row=line_idx, column=y + 4).number_format = num_style

    # Write investment costs
    for node_id in shared_ess_data.active_distribution_network_nodes:
        for s_m in range(len(shared_ess_data.prob_market_scenarios)):

            # Cost Power
            line_idx = line_idx + 1
            sheet.cell(row=line_idx, column=1).value = node_id
            sheet.cell(row=line_idx, column=2).value = 'Cost S, [€]'
            sheet.cell(row=line_idx, column=3).value = s_m
            for y in range(len(years)):
                year = years[y]
                sheet.cell(row=line_idx, column=y + 4).value = results['cost'][node_id][year]['power'][s_m]
                sheet.cell(row=line_idx, column=y + 4).number_format = num_style

            # Cost Energy
            line_idx = line_idx + 1
            sheet.cell(row=line_idx, column=1).value = node_id
            sheet.cell(row=line_idx, column=2).value = 'Cost E, [€]'
            sheet.cell(row=line_idx, column=3).value = s_m
            for y in range(len(years)):
                year = years[y]
                sheet.cell(row=line_idx, column=y + 4).value = results['cost'][node_id][year]['energy'][s_m]
                sheet.cell(row=line_idx, column=y + 4).number_format = num_style

        # Expected Cost Power
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'Cost S, [€]'
        sheet.cell(row=line_idx, column=3).value = 'Expected'
        for y in range(len(years)):
            year = years[y]
            sheet.cell(row=line_idx, column=y + 4).value = results['cost'][node_id][year]['power']['expected']
            sheet.cell(row=line_idx, column=y + 4).number_format = num_style
            total_cost_power[year] += results['cost'][node_id][year]['power']['expected']

        # Expected Cost Energy
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'Cost E, [€]'
        sheet.cell(row=line_idx, column=3).value = 'Expected'
        for y in range(len(years)):
            year = years[y]
            sheet.cell(row=line_idx, column=y + 4).value = results['cost'][node_id][year]['energy']['expected']
            sheet.cell(row=line_idx, column=y + 4).number_format = num_style
            total_cost_energy[year] += results['cost'][node_id][year]['energy']['expected']

    # - Total
    line_idx = line_idx + 1
    sheet.cell(row=line_idx, column=1).value = 'Total'
    sheet.cell(row=line_idx, column=2).value = 'Cost, [€]'
    sheet.cell(row=line_idx, column=3).value = 'Expected'
    for y in range(len(years)):
        year = years[y]
        sheet.cell(row=line_idx, column=y + 4).value = total_cost_power[year] + total_cost_energy[year]
        sheet.cell(row=line_idx, column=y + 4).number_format = num_style


def _write_ess_capacity_investment_to_excel(shared_ess_data, workbook, results, initial_sheet=True):

    if initial_sheet:
        sheet = workbook.worksheets[0]
        sheet.title = 'Capacity Investment'
    else:
        sheet = workbook.create_sheet('Capacity Investment')

    years = [year for year in shared_ess_data.years]

    num_style = '0.00'

    # Write Header
    line_idx = 1
    sheet.cell(row=line_idx, column=1).value = 'Node'
    sheet.cell(row=line_idx, column=2).value = 'Quantity'
    for y in range(len(years)):
        year = years[y]
        sheet.cell(row=line_idx, column=y + 3).value = int(year)

    # Write investment values, power and energy
    for node_id in results:

        # Power capacity
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'S, [MVA]'
        for y in range(len(years)):
            year = years[y]
            sheet.cell(row=line_idx, column=y + 3).value = results[node_id][year]['power']
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style

        # Energy capacity
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'E, [MVAh]'
        for y in range(len(years)):
            year = years[y]
            sheet.cell(row=line_idx, column=y + 3).value = results[node_id][year]['energy']
            sheet.cell(row=line_idx, column=y + 3).number_format = num_style


def _write_ess_capacity_rated_available_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Capacity Available')

    num_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    row_idx, col_idx = 1, 1
    sheet.cell(row=row_idx, column=col_idx).value = 'Node'
    col_idx = col_idx + 1
    sheet.cell(row=row_idx, column=col_idx).value = 'Quantity'
    col_idx = col_idx + 1
    for year in shared_ess_data.years:
        sheet.cell(row=row_idx, column=col_idx).value = int(year)
        col_idx = col_idx + 1

    # Write investment values, power and energy
    for node_id in results['rated']:

        # Power, rated
        col_idx = 1
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = node_id
        col_idx = col_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = 'Srated, [MVA]'
        col_idx = col_idx + 1
        for year in shared_ess_data.years:
            sheet.cell(row=row_idx, column=col_idx).value = results['rated'][node_id][year]['power']
            sheet.cell(row=row_idx, column=col_idx).number_format = num_style
            col_idx = col_idx + 1

        # Capacity, rated
        col_idx = 1
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = node_id
        col_idx = col_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = 'Erated, [MVA]'
        col_idx = col_idx + 1
        for year in shared_ess_data.years:
            sheet.cell(row=row_idx, column=col_idx).value = results['rated'][node_id][year]['energy']
            sheet.cell(row=row_idx, column=col_idx).number_format = num_style
            col_idx = col_idx + 1

        # Power, available
        col_idx = 1
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = node_id
        col_idx = col_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = 'Savailable, [MVA]'
        col_idx = col_idx + 1
        for year in shared_ess_data.years:
            sheet.cell(row=row_idx, column=col_idx).value = results['available'][node_id][year]['power']
            sheet.cell(row=row_idx, column=col_idx).number_format = num_style
            col_idx = col_idx + 1

        # Capacity, available
        col_idx = 1
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = node_id
        col_idx = col_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = 'Eavailable, [MVAh]'
        col_idx = col_idx + 1
        for year in shared_ess_data.years:
            sheet.cell(row=row_idx, column=col_idx).value = results['available'][node_id][year]['energy']
            sheet.cell(row=row_idx, column=col_idx).number_format = num_style
            col_idx = col_idx + 1

        # SoH
        col_idx = 1
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = node_id
        col_idx = col_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = 'SoH, [%]'
        col_idx = col_idx + 1
        for year in shared_ess_data.years:
            sheet.cell(row=row_idx, column=col_idx).value = results['available'][node_id][year]['soh']
            sheet.cell(row=row_idx, column=col_idx).number_format = perc_style
            col_idx = col_idx + 1

        # Degradation factor
        col_idx = 1
        row_idx = row_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = node_id
        col_idx = col_idx + 1
        sheet.cell(row=row_idx, column=col_idx).value = 'Degradation factor, [%]'
        col_idx = col_idx + 1
        for year in shared_ess_data.years:
            sheet.cell(row=row_idx, column=col_idx).value = results['available'][node_id][year]['degradation_factor']
            sheet.cell(row=row_idx, column=col_idx).number_format = perc_style
            col_idx = col_idx + 1


def _write_aggregated_shared_energy_storage_operation_results_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Operation, aggregated')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    for p in range(shared_ess_data.num_instants):
        sheet.cell(row=row_idx, column=p + 5).value = p
    row_idx = row_idx + 1

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for year in results:
            for day in results[year]:

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                for p in range(shared_ess_data.num_instants):
                    pnet = results[year][day][node_id]['p'][p]
                    sheet.cell(row=row_idx, column=p + 5).value = pnet
                    sheet.cell(row=row_idx, column=p + 5).number_format = decimal_style
                row_idx = row_idx + 1

                # - Reactive Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
                for p in range(shared_ess_data.num_instants):
                    qnet = results[year][day][node_id]['q'][p]
                    sheet.cell(row=row_idx, column=p + 5).value = qnet
                    sheet.cell(row=row_idx, column=p + 5).number_format = decimal_style
                row_idx = row_idx + 1


def _write_detailed_shared_energy_storage_operation_results_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Operation, Detailed')

    row_idx = 1
    perc_style = '0.00%'
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year Investment'
    sheet.cell(row=row_idx, column=3).value = 'Year Current'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    for p in range(shared_ess_data.num_instants):
        sheet.cell(row=row_idx, column=p + 6).value = p
    row_idx = row_idx + 1

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for year_inv in results:
            for year_curr in results[year_inv]:
                for day in results[year_inv][year_curr]:

                    # - Apparent Power
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year_inv)
                    sheet.cell(row=row_idx, column=3).value = int(year_curr)
                    sheet.cell(row=row_idx, column=4).value = day
                    sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                    for p in range(shared_ess_data.num_instants):
                        snet = results[year_inv][year_curr][day][node_id]['s'][p]
                        sheet.cell(row=row_idx, column=p + 6).value = snet
                        sheet.cell(row=row_idx, column=p + 6).number_format = decimal_style
                    row_idx = row_idx + 1


def _write_aggregated_shared_energy_storage_soh_results_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Degradation, aggregated')

    row_idx = 1
    perc_style = '0.00%'
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Quantity'
    sheet.cell(row=row_idx, column=4).value = 'Value'
    row_idx = row_idx + 1

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for year in results:

            s_rated = results[year]['s_rated'][node_id]
            e_rated = results[year]['e_rated'][node_id]
            s_available = results[year]['s_available'][node_id]
            e_available = results[year]['e_available'][node_id]
            soh = results[year]['soh'][node_id]
            degradation = results[year]['degradation'][node_id]

            # - Srated
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = int(year)
            sheet.cell(row=row_idx, column=3).value = 'Srated, [MVA]'
            sheet.cell(row=row_idx, column=4).value = s_rated
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            row_idx = row_idx + 1

            # - Erated
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = int(year)
            sheet.cell(row=row_idx, column=3).value = 'Erated, [MVA]'
            sheet.cell(row=row_idx, column=4).value = e_rated
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            row_idx = row_idx + 1

            # - Savailable
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = int(year)
            sheet.cell(row=row_idx, column=3).value = 'Savailable, [MVA]'
            sheet.cell(row=row_idx, column=4).value = s_available
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            row_idx = row_idx + 1

            # - Eavailable
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = int(year)
            sheet.cell(row=row_idx, column=3).value = 'Eavailable, [MVA]'
            sheet.cell(row=row_idx, column=4).value = e_available
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            row_idx = row_idx + 1

            # - SoH
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = int(year)
            sheet.cell(row=row_idx, column=3).value = 'SoH, [%]'
            sheet.cell(row=row_idx, column=4).value = soh
            sheet.cell(row=row_idx, column=4).number_format = perc_style
            row_idx = row_idx + 1

            # - Degradation
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = int(year)
            sheet.cell(row=row_idx, column=3).value = 'Degradation, [%]'
            sheet.cell(row=row_idx, column=4).value = degradation
            sheet.cell(row=row_idx, column=4).number_format = perc_style
            row_idx = row_idx + 1


def _write_detailed_shared_energy_storage_soh_results_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Degradation, detailed')

    row_idx = 1
    perc_style = '0.00%'
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year Investment'
    sheet.cell(row=row_idx, column=3).value = 'Year Current'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Value'
    row_idx = row_idx + 1

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for year_inv in results:
            for year_curr in results[year_inv]:

                s_rated = results[year_inv][year_curr]['s_rated'][node_id]
                e_rated = results[year_inv][year_curr]['e_rated'][node_id]
                s_available = results[year_inv][year_curr]['s_available'][node_id]
                e_available = results[year_inv][year_curr]['e_available'][node_id]
                soh_unit = results[year_inv][year_curr]['soh_unit'][node_id]
                degradation_unit = results[year_inv][year_curr]['degradation_unit'][node_id]
                soh_cumul = results[year_inv][year_curr]['soh_cumul'][node_id]
                degradation_cumul = results[year_inv][year_curr]['degradation_cumul'][node_id]

                # - Srated, average day
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'Srated, [MVA]'
                sheet.cell(row=row_idx, column=5).value = s_rated
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                row_idx = row_idx + 1

                # - Erated, average day
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'Erated, [MVAh]'
                sheet.cell(row=row_idx, column=5).value = e_rated
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                row_idx = row_idx + 1

                # - Savailable, average day
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'Savailable, [MVA]'
                sheet.cell(row=row_idx, column=5).value = s_available
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                row_idx = row_idx + 1

                # - Eavailable, average day
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'Eavailable, [MVAh]'
                sheet.cell(row=row_idx, column=5).value = e_available
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoH, average day
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'SoH unit, [%]'
                sheet.cell(row=row_idx, column=5).value = soh_unit
                sheet.cell(row=row_idx, column=5).number_format = perc_style
                row_idx = row_idx + 1

                # - Degradation, average day
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'Degradation unit, [%]'
                sheet.cell(row=row_idx, column=5).value = degradation_unit
                sheet.cell(row=row_idx, column=5).number_format = perc_style
                row_idx = row_idx + 1

                # - SoH, cumulative
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'SoH cumul., [%]'
                sheet.cell(row=row_idx, column=5).value = soh_cumul
                sheet.cell(row=row_idx, column=5).number_format = perc_style
                row_idx = row_idx + 1

                # - Degradation, cumulative
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'Degradation cumul., [%]'
                sheet.cell(row=row_idx, column=5).value = degradation_cumul
                sheet.cell(row=row_idx, column=5).number_format = perc_style
                row_idx = row_idx + 1


def _write_investment_relaxation_slacks_results_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Slacks investment')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year Investment'
    sheet.cell(row=row_idx, column=3).value = 'Quantity'
    sheet.cell(row=row_idx, column=4).value = 'Value'
    row_idx = row_idx + 1

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for year_inv in results:

            # - Sup
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = int(year_inv)
            sheet.cell(row=row_idx, column=3).value = 'S, up'
            sheet.cell(row=row_idx, column=4).value = results[year_inv][node_id]['s_up']
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            row_idx = row_idx + 1

            # - Sdown
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = int(year_inv)
            sheet.cell(row=row_idx, column=3).value = 'S, down'
            sheet.cell(row=row_idx, column=4).value = results[year_inv][node_id]['s_down']
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            row_idx = row_idx + 1

            # - Eup
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = int(year_inv)
            sheet.cell(row=row_idx, column=3).value = 'E, up'
            sheet.cell(row=row_idx, column=4).value = results[year_inv][node_id]['e_up']
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            row_idx = row_idx + 1

            # - Edown
            sheet.cell(row=row_idx, column=1).value = node_id
            sheet.cell(row=row_idx, column=2).value = int(year_inv)
            sheet.cell(row=row_idx, column=3).value = 'E, down'
            sheet.cell(row=row_idx, column=4).value = results[year_inv][node_id]['e_down']
            sheet.cell(row=row_idx, column=4).number_format = decimal_style
            row_idx = row_idx + 1


def _write_detailed_degradation_relaxation_slacks_results_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Slacks degradation, detailed')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year Investment'
    sheet.cell(row=row_idx, column=3).value = 'Year Current'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Value'
    row_idx = row_idx + 1

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for year_inv in results:
            for year_curr in results[year_inv]:

                # - SoH per unit, up
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'SoH unit, up'
                sheet.cell(row=row_idx, column=5).value = results[year_inv][year_curr][node_id]['soh_per_unit_up']
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoH per unit, down
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'SoH unit, down'
                sheet.cell(row=row_idx, column=5).value = results[year_inv][year_curr][node_id]['soh_per_unit_down']
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoH per unit (cumulative), up
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'SoH cumul., up'
                sheet.cell(row=row_idx, column=5).value = results[year_inv][year_curr][node_id]['soh_per_unit_cumul_up']
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoH per unit (cumulative), down
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year_inv)
                sheet.cell(row=row_idx, column=3).value = int(year_curr)
                sheet.cell(row=row_idx, column=4).value = 'SoH cumul., down'
                sheet.cell(row=row_idx, column=5).value = results[year_inv][year_curr][node_id]['soh_per_unit_cumul_down']
                sheet.cell(row=row_idx, column=5).number_format = decimal_style
                row_idx = row_idx + 1


def _write_aggregated_operation_relaxation_slacks_results_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Slacks operation, aggregated')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    for p in range(shared_ess_data.num_instants):
        sheet.cell(row=row_idx, column=p + 5).value = p
    row_idx = row_idx + 1

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for year in results:
            for day in results[year]:

                # - Snet, up
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Snet, up'
                for p in range(shared_ess_data.num_instants):
                    snet_up = results[year][day][node_id]['snet_up'][p]
                    sheet.cell(row=row_idx, column=p + 5).value = snet_up
                    sheet.cell(row=row_idx, column=p + 5).number_format = decimal_style
                row_idx = row_idx + 1

                # - Snet, down
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Snet, down'
                for p in range(shared_ess_data.num_instants):
                    snet_down = results[year][day][node_id]['snet_down'][p]
                    sheet.cell(row=row_idx, column=p + 5).value = snet_down
                    sheet.cell(row=row_idx, column=p + 5).number_format = decimal_style
                row_idx = row_idx + 1

                # - Snet definition, up
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Snet definition, up'
                for p in range(shared_ess_data.num_instants):
                    snet_def_up = results[year][day][node_id]['snet_def_up'][p]
                    sheet.cell(row=row_idx, column=p + 5).value = snet_def_up
                    sheet.cell(row=row_idx, column=p + 5).number_format = decimal_style
                row_idx = row_idx + 1

                # - Snet definition, down
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Snet definition, down'
                for p in range(shared_ess_data.num_instants):
                    snet_def_down = results[year][day][node_id]['snet_def_down'][p]
                    sheet.cell(row=row_idx, column=p + 5).value = snet_def_down
                    sheet.cell(row=row_idx, column=p + 5).number_format = decimal_style
                row_idx = row_idx + 1


def _write_detailed_operation_relaxation_slacks_results_to_excel(shared_ess_data, workbook, results):

    sheet = workbook.create_sheet('Slacks operation, detailed')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year Investment'
    sheet.cell(row=row_idx, column=3).value = 'Year Current'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    for p in range(shared_ess_data.num_instants):
        sheet.cell(row=row_idx, column=p + 6).value = p
    row_idx = row_idx + 1

    for node_id in shared_ess_data.active_distribution_network_nodes:
        for year_inv in results:
            for year_curr in results[year_inv]:
                for day in results[year_inv][year_curr]:

                    if shared_ess_data.params.slacks:

                        # - Complementarity
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year_inv)
                        sheet.cell(row=row_idx, column=3).value = int(year_curr)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Complementary'
                        for p in range(shared_ess_data.num_instants):
                            comp = results[year_inv][year_curr][day][node_id]['comp'][p]
                            sheet.cell(row=row_idx, column=p + 6).value = comp
                            sheet.cell(row=row_idx, column=p + 6).number_format = decimal_style
                        row_idx = row_idx + 1

    return results
