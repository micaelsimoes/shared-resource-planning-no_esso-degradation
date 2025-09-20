import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from sklearn.preprocessing import MinMaxScaler
from copulas.multivariate import GaussianMultivariate
from math import isclose
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from network import Network
from network_parameters import NetworkParameters
from helper_functions import *


# ======================================================================================================================
#   Class NETWORK DATA -- Contains information of the Network over the planning period (years, days)
# ======================================================================================================================
class NetworkData:

    def __init__(self):
        self.name = str()
        self.data_dir = str()
        self.results_dir = str()
        self.diagrams_dir = str()
        self.years = dict()
        self.days = dict()
        self.num_instants = int()
        self.num_oper_scenarios = int()
        self.plot_operational_data = bool()
        self.discount_factor = float()
        self.network = dict()
        self.operational_data_file = str()
        self.params_file = str()
        self.params = NetworkParameters()
        self.cost_energy_p = dict()
        self.cost_flex = dict()
        self.is_transmission = False
        self.active_distribution_network_nodes = list()

    def build_model(self):
        network_models = dict()
        for year in self.years:
            network_models[year] = dict()
            for day in self.days:
                network_models[year][day] = self.network[year][day].build_model(self.params)
        return network_models

    def optimize(self, model, from_warm_start=False):
        results = dict()
        for year in self.years:
            results[year] = dict()
            for day in self.days:
                results[year][day] = self.network[year][day].run_smopf(model[year][day], self.params, from_warm_start=from_warm_start)
        return results

    def get_pq_map(self, t=None, num_steps=8, print_pq_map=False):
        pq_map = dict()
        for year in self.years:
            pq_map[year] = dict()
            for day in self.days:
                pq_map[year][day] = self.network[year][day].get_pq_map(t=t, params=self.params, num_steps=num_steps, print_pq_map=print_pq_map)
        return pq_map

    def build_pq_map_model(self, t=None, num_steps=8):
        network_models = dict()
        for year in self.years:
            network_models[year] = dict()
            for day in self.days:
                network_models[year][day] = dict()
                if t:
                    network_models[year][day] = self.network[year][day].build_pq_map_model(t, self.params)
                else:
                    for t in range(self.num_instants):
                        network_models[year][day][t] = self.network[year][day].network.build_pq_map_model(t, self.params)
        return network_models

    def update_of_to_settlement(self, model):
        for year in self.years:
            for day in self.days:
                self.network[year][day].update_of_to_settlement(model[year][day])

    def get_primal_value(self, model):
        obj = 0.0
        years = [year for year in self.years]
        for year in self.years:
            annualization = 1 / ((1 + self.discount_factor) ** (int(year) - int(years[0])))
            for day in self.days:
                obj += self.network[year][day].get_primal_value(model[year][day]) * self.years[year] * self.days[day] * annualization
        return obj

    def get_sensitivities(self, model):
        return _get_sensitivities(self, model)

    def update_model_with_candidate_solution(self, model, candidate_solution):
        _update_model_with_candidate_solution(self, model, candidate_solution)

    def read_network_data(self):
        _read_network_data(self)

    def read_network_parameters(self):
        filename = os.path.join(self.data_dir, self.name, self.params_file)
        self.params.read_parameters_from_file(filename)

    def update_data_with_candidate_solution(self, candidate_solution):
        _update_data_with_candidate_solution(self, candidate_solution)

    def process_results(self, model, results=dict()):
        return _process_results(self, model, results)

    def process_results_interface(self, model):
        results = dict()
        for year in self.years:
            results[year] = dict()
            for day in self.days:
                results[year][day] = self.network[year][day].process_results_interface(model[year][day])
        return results

    def process_results_summary_detail(self, model):
        results = dict()
        for year in self.years:
            results[year] = dict()
            for day in self.days:
                results[year][day] = self.network[year][day].process_results_summary_detail(model[year][day], self.params)
        return results

    def write_optimization_results_to_excel(self, results, filename=str()):
        if not filename:
            filename = self.name
        _write_optimization_results_to_excel(self, self.results_dir, results, filename=filename)


    def plot_operational_data_scenarios(self):
        years_to_plot = list(self.years)[0]
        _plot_load_data_scenarios(self, years_to_plot=[years_to_plot], save_dir=self.diagrams_dir)
        _plot_res_data_scenarios(self, years_to_plot=[years_to_plot], save_dir=self.diagrams_dir)


# ======================================================================================================================
#  NETWORK DATA read function
# ======================================================================================================================
def _read_network_data(network_planning):

    filename = os.path.join(network_planning.data_dir, network_planning.name, network_planning.operational_data_file)

    try:
        base_data = _read_network_base_profiles(filename)
        base_data['initial_year'] = list(network_planning.years)[0]
    except:
        print(f'[ERROR] Reading operational data, network {network_planning.name} from file. Exiting...')
        exit(ERROR_SPECIFICATION_FILE)

    synthetic_profiles = _generate_operational_scenarios(base_data)

    for year in network_planning.years:

        network_planning.network[year] = dict()

        for day in network_planning.days:

            # Create Network object
            network_planning.network[year][day] = Network()
            network_planning.network[year][day].name = network_planning.name
            network_planning.network[year][day].data_dir = network_planning.data_dir
            network_planning.network[year][day].results_dir = network_planning.results_dir
            network_planning.network[year][day].diagrams_dir = network_planning.diagrams_dir
            network_planning.network[year][day].year = int(year)
            network_planning.network[year][day].day = day
            network_planning.network[year][day].num_instants = network_planning.num_instants
            network_planning.network[year][day].num_oper_scenarios = network_planning.num_oper_scenarios
            network_planning.network[year][day].prob_operation_scenarios = [(1 / network_planning.num_oper_scenarios)] * network_planning.num_oper_scenarios
            network_planning.network[year][day].is_transmission = network_planning.is_transmission

            # Read info from file(s)
            network_planning.network[year][day].read_network_from_json_file()
            network_planning.network[year][day].update_network_operational_data(base_data, synthetic_profiles)

            if network_planning.params.print_to_screen:
                network_planning.network[year][day].print_network_to_screen()
            if network_planning.params.plot_diagram:
                network_planning.network[year][day].plot_diagram()


def _read_network_base_profiles(filename):

    base_operational_data = {
        'characterization': pd.read_excel(filename, sheet_name='Characterization'),
        'growth_factors': pd.read_excel(filename, sheet_name='Growth Factors'),
        'pc': pd.read_excel(filename, sheet_name='Pc'),
        'qc': pd.read_excel(filename, sheet_name='Qc'),
        'pg': pd.read_excel(filename, sheet_name='Pg'),
        'qg': pd.read_excel(filename, sheet_name='Qg'),
        'flex': pd.read_excel(filename, sheet_name='Flex')
    }

    return base_operational_data


def _generate_operational_scenarios(base_profiles, n_samples=100, bandwidth=0.10):

    synthetic_profiles = {
        'consumption': generate_consumption_profiles(base_profiles, n_samples=n_samples, bandwidth=bandwidth),
        'generation': generate_res_generation_profiles(base_profiles, n_samples=n_samples, bandwidth=bandwidth),
        'flexibility': generate_flexibility_profiles(base_profiles, n_samples=n_samples, bandwidth=bandwidth)
    }

    return synthetic_profiles


def generate_consumption_profiles(base_operational_data, n_samples=100, bandwidth=0.10):

    print('[INFO]\t - Generating load stochastic scenarios...')

    pc_df = base_operational_data['pc']
    qc_df = base_operational_data['qc']
    load_ids = pc_df['LoadID'].unique()
    seasons = pc_df['Season'].unique()
    synthetic_profiles = {}

    for season in seasons:

        synthetic_profiles[season] = {}

        for load_id in load_ids:

            # Filter
            pc_subset = pc_df[(pc_df['Season'] == season) & (pc_df['LoadID'] == load_id)]
            qc_subset = qc_df[(qc_df['Season'] == season) & (qc_df['LoadID'] == load_id)]
            if pc_subset.empty or qc_subset.empty:
                print(f'[ERROR] No data provided for load {load_id}, season {season}')
                exit(ERROR_NETWORK_FILE)

            # Prepare data
            pc_hours = pc_subset.iloc[:, 3:].copy()
            qc_hours = qc_subset.iloc[:, 3:].copy()
            if pc_hours.shape != qc_hours.shape:
                print(f"[ERROR] Shape mismatch between Pc and Qc, load {load_id}")
                exit(ERROR_NETWORK_FILE)

            # - Rename columns to distinguish them
            pc_hours.columns = [f'Pc_{i}' for i in range(24)]
            qc_hours.columns = [f'Qc_{i}' for i in range(24)]
            combined = pd.concat([pc_hours, qc_hours], axis=1)

            # Normalize
            scaler = MinMaxScaler()
            combined_scaled = scaler.fit_transform(combined)

            # Fit model
            model = GaussianMultivariate(distribution=CustomGaussianKDE(bandwidth=bandwidth))
            model.fit(pd.DataFrame(combined_scaled, columns=combined.columns))

            # Sample
            samples = model.sample(n_samples)
            samples = scaler.inverse_transform(samples)

            # Save
            synthetic_profiles[season][load_id] = {
                'pc': pd.DataFrame(samples[:,:24]),
                'qc': pd.DataFrame(samples[:,24:])
            }

    return synthetic_profiles


def generate_res_generation_profiles(base_operational_data, n_samples=100, bandwidth=0.10):

    print('[INFO]\t - Generating RES generation stochastic scenarios...')

    pg_df = base_operational_data['pg']
    seasons = pg_df['Season'].unique()
    synthetic_profiles = {}

    for season in seasons:

        synthetic_profiles[season] = {}
        season_data = pg_df[pg_df['Season'] == season]

        for gen_type in ['PV', 'Wind']:

            subset = season_data[season_data['GenType'] == gen_type]
            if subset.empty:
                print(f"[WARNING] No {gen_type} data for season {season}")
                continue

            gen_hours = subset.iloc[:, 3:].copy().dropna()

            if gen_hours.empty:
                print(f"[WARNING] No valid {gen_type} hourly data in season {season}")
                continue

            # Normalize and fit copula
            scaler = MinMaxScaler()
            scaled = scaler.fit_transform(gen_hours)

            model = GaussianMultivariate(distribution=CustomGaussianKDE(bandwidth=bandwidth))
            model.fit(pd.DataFrame(scaled))

            # Sample
            samples = model.sample(n_samples)
            samples = np.abs(scaler.inverse_transform(samples))

            synthetic_profiles[season][gen_type] = {
                'pg': pd.DataFrame(samples),
                'qg': pd.DataFrame(np.zeros(samples.shape))
            }

    return synthetic_profiles


def generate_flexibility_profiles(base_operational_data, n_samples=100, bandwidth=0.10):

    print('[INFO]\t - Generating flexibility stochastic scenarios...')

    flex_df = base_operational_data['flex']
    seasons = flex_df['Season'].unique()
    synthetic_profiles = {}

    for season in seasons:

        synthetic_profiles[season] = {}

        # Prepare data
        pc_df = flex_df[flex_df['Type']=='Flexible Load, [MW]'].iloc[:, 3:].copy()
        flex_up_df = flex_df[flex_df['Type']=='Maximum Flexible Load, [MW]'].iloc[:, 3:].copy()
        flex_down_df = flex_df[flex_df['Type']=='Minimum Flexible Load, [MW]'].iloc[:, 3:].copy()
        if pc_df.shape != flex_up_df.shape or pc_df.shape != flex_down_df.shape:
            print(" [ERROR] Shape mismatch, flexibility data")
            exit(ERROR_NETWORK_FILE)

        # - Rename columns to distinguish them
        pc_df.columns = [f'FL_{i}' for i in range(24)]
        flex_up_df.columns = [f'FL_up_{i}' for i in range(24)]
        flex_down_df.columns = [f'FL_down_{i}' for i in range(24)]
        combined = pd.concat([pc_df, flex_up_df, flex_down_df], axis=1).ffill().bfill()

        # Normalize
        scaler = MinMaxScaler()
        combined_scaled = scaler.fit_transform(combined)

        # Fit model
        model = GaussianMultivariate(distribution=CustomGaussianKDE(bandwidth=bandwidth))
        model.fit(pd.DataFrame(combined_scaled, columns=combined.columns))

        # Sample
        samples = model.sample(n_samples)
        samples = scaler.inverse_transform(samples)

        # Save
        synthetic_profiles[season] = {
            'pc': pd.DataFrame(samples[:,:24]),
            'pc_flex_up': pd.DataFrame(samples[:,24:48]),
            'pc_flex_down': pd.DataFrame(samples[:,48:])
        }

    return synthetic_profiles


# ======================================================================================================================
#  NETWORK PLANNING results functions
# ======================================================================================================================
def _process_results(network_planning, models, optimization_results):
    processed_results = dict()
    processed_results['results'] = dict()
    processed_results['of_value'] = _get_objective_function_value(network_planning, models)
    for year in network_planning.years:
        processed_results['results'][year] = dict()
        for day in network_planning.days:
            model = models[year][day]
            result = optimization_results[year][day]
            network = network_planning.network[year][day]
            processed_results['results'][year][day] = network.process_results(model, network_planning.params, result)
    return processed_results


def _get_objective_function_value(network_planning, models):

    years = [year for year in network_planning.years]

    of_value = 0.0
    initial_year = years[0]
    if network_planning.is_transmission:
        for y in range(len(network_planning.years)):
            year = years[y]
            num_years = network_planning.years[year]
            annualization = 1 / ((1 + network_planning.discount_factor) ** (int(year) - int(initial_year)))
            for day in network_planning.days:
                num_days = network_planning.days[day]
                network = network_planning.network[year][day]
                model = models[year][day]
                of_value += annualization * num_days * num_years * network.compute_objective_function_value(model, network_planning.params)
    return of_value


def _write_optimization_results_to_excel(network_planning, data_dir, processed_results, filename):

    wb = Workbook()

    _write_main_info_to_excel(network_planning, wb, processed_results)
    if network_planning.params.obj_type == OBJ_MIN_COST:
        _write_market_cost_values_to_excel(network_planning, wb)
    _write_shared_network_energy_storage_results_to_excel(network_planning, wb, processed_results['results'])
    _write_network_voltage_results_to_excel(network_planning, wb, processed_results['results'])
    _write_network_consumption_results_to_excel(network_planning, wb, processed_results['results'])
    _write_network_generation_results_to_excel(network_planning, wb, processed_results['results'])
    _write_network_branch_results_to_excel(network_planning, wb, processed_results['results'], 'losses')
    _write_network_branch_results_to_excel(network_planning, wb, processed_results['results'], 'ratio')
    _write_network_branch_loading_results_to_excel(network_planning, wb, processed_results['results'])
    _write_network_branch_power_flow_results_to_excel(network_planning, wb, processed_results['results'])
    if network_planning.params.es_reg:
        _write_network_energy_storage_results_to_excel(network_planning, wb, processed_results['results'])
    _write_relaxation_slacks_scenarios_results_to_excel(network_planning, wb, processed_results['results'])

    results_filename = os.path.join(data_dir, f'{filename}.xlsx')
    try:
        wb.save(results_filename)
        print('[INFO] S-MPOPF Results written to {}.'.format(results_filename))
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = os.path.join(data_dir, f'{network_planning.name}_results_{current_time}.xlsx')
        print('[INFO] S-MPOPF Results written to {}.'.format(backup_filename))
        wb.save(backup_filename)


def _write_main_info_to_excel(network_planning, workbook, results):

    sheet = workbook.worksheets[0]
    sheet.title = 'Main Info'

    decimal_style = '0.00'
    line_idx = 1

    # Write Header
    col_idx = 2
    for year in network_planning.years:
        for _ in network_planning.days:
            write_value(sheet, line_idx, col_idx, year)
            col_idx += 1
    col_idx = 2
    line_idx += 1
    for _ in network_planning.years:
        for day in network_planning.days:
            write_value(sheet, line_idx, col_idx, day)
            col_idx += 1

    # Objective function value
    col_idx = 2
    line_idx += 1
    obj_string = 'Objective'
    if network_planning.params.obj_type == OBJ_MIN_COST:
        obj_string += ' (cost), [€]'
    elif network_planning.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
        obj_string += ' (congestion management)'
    sheet.cell(row=line_idx, column=1).value = obj_string
    for year in network_planning.years:
        for day in network_planning.days:
            if results['results'][year][day]:
                value = results['results'][year][day]['obj']
                write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
            else:
                sheet.cell(row=line_idx, column=col_idx).value = 'N/A'
            col_idx += 1

    # Total Load
    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Load, [MWh]'
    for year in network_planning.years:
        for day in network_planning.days:
            value = results['results'][year][day]['total_load']['p']
            write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
            col_idx += 1

    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Load, [MVArh]'
    for year in network_planning.years:
        for day in network_planning.days:
            value = results['results'][year][day]['total_load']['q']
            write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
            col_idx += 1

    # Flexibility used
    if network_planning.params.fl_reg:
        col_idx = 2
        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = 'Flexibility used, [MWh]'
        for year in network_planning.years:
            for day in network_planning.days:
                value = results['results'][year][day]['flex_used']['p']
                write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
                col_idx += 1

        col_idx = 2
        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = 'Flexibility used, [MVArh]'
        for year in network_planning.years:
            for day in network_planning.days:
                value = results['results'][year][day]['flex_used']['q']
                write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
                col_idx += 1

    # Total Load curtailed
    if network_planning.params.l_curt:

        col_idx = 2
        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = 'Load curtailed, [MWh]'
        for year in network_planning.years:
            for day in network_planning.days:
                value = results['results'][year][day]['load_curt']['p']
                write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
                col_idx += 1

        col_idx = 2
        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = 'Load curtailed, [MVArh]'
        for year in network_planning.years:
            for day in network_planning.days:
                value = results['results'][year][day]['load_curt']['q']
                write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
                col_idx += 1

    # Total Generation
    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Generation, [MWh]'
    for year in network_planning.years:
        for day in network_planning.days:
            value = results['results'][year][day]['total_gen']['p']
            write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
            col_idx += 1

    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Generation, [MVArh]'
    for year in network_planning.years:
        for day in network_planning.days:
            value = results['results'][year][day]['total_gen']['q']
            write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
            col_idx += 1

    # Total Renewable Generation
    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Renewable generation, [MWh]'
    for year in network_planning.years:
        for day in network_planning.days:
            value = results['results'][year][day]['total_renewable_gen']['p']
            write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
            col_idx += 1

    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Renewable generation, [MVArh]'
    for year in network_planning.years:
        for day in network_planning.days:
            value = results['results'][year][day]['total_renewable_gen']['q']
            write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
            col_idx += 1

    # Renewable Generation Curtailed
    if network_planning.params.rg_curt:

        col_idx = 2
        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = 'Renewable generation curtailed, [MWh]'
        for year in network_planning.years:
            for day in network_planning.days:
                value = results['results'][year][day]['gen_curt']['p']
                write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
                col_idx += 1

        col_idx = 2
        line_idx += 1
        sheet.cell(row=line_idx, column=1).value = 'Renewable generation curtailed, [MVArh]'
        for year in network_planning.years:
            for day in network_planning.days:
                value = results['results'][year][day]['gen_curt']['q']
                write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
                col_idx += 1

    # Losses
    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Losses, [MWh]'
    for year in network_planning.years:
        for day in network_planning.days:
            value = results['results'][year][day]['losses']
            write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
            col_idx += 1

    # Execution time
    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Execution time, [s]'
    for year in network_planning.years:
        for day in network_planning.days:
            value = results['results'][year][day]['runtime']
            write_value(sheet, line_idx, col_idx, value, number_format=decimal_style)
            col_idx += 1

    # Number of price (market) scenarios
    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Number of market scenarios'
    for year in network_planning.years:
        for day in network_planning.days:
            value = len(network_planning.network[year][day].prob_market_scenarios)
            write_value(sheet, line_idx, col_idx, value)
            col_idx += 1

    # Number of operation (generation and consumption) scenarios
    col_idx = 2
    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Number of operation scenarios'
    for year in network_planning.years:
        for day in network_planning.days:
            value = len(network_planning.network[year][day].prob_operation_scenarios)
            write_value(sheet, line_idx, col_idx, value)
            col_idx += 1


def _write_market_cost_values_to_excel(network_planning, workbook):

    decimal_style = '0.00'
    perc_style = '0.00%'

    line_idx = 1
    sheet = workbook.create_sheet('Market Cost Info')

    # Write Header
    sheet.cell(row=line_idx, column=1).value = 'Cost'
    sheet.cell(row=line_idx, column=2).value = 'Year'
    sheet.cell(row=line_idx, column=3).value = 'Day'
    sheet.cell(row=line_idx, column=4).value = 'Scenario'
    sheet.cell(row=line_idx, column=5).value = 'Probability, [%]'
    for p in range(network_planning.num_instants):
        sheet.cell(row=line_idx, column=p + 6).value = p

    # Write active and reactive power costs per scenario
    for year in network_planning.years:
        for day in network_planning.days:
            network = network_planning.network[year][day]
            for s_o in range(len(network.prob_market_scenarios)):
                line_idx += 1
                sheet.cell(row=line_idx, column=1).value= 'Active power, [€/MW]'
                sheet.cell(row=line_idx, column=2).value= year
                sheet.cell(row=line_idx, column=3).value= day
                sheet.cell(row=line_idx, column=4).value= s_o
                sheet.cell(row=line_idx, column=5).value= network.prob_market_scenarios[s_o]
                sheet.cell(row=line_idx, column=5).number_format = perc_style
                for p in range(network.num_instants):
                    sheet.cell(row=line_idx, column=p + 6).value= network.cost_energy_p[s_o][p]
                    sheet.cell(row=line_idx, column=p + 6).number_format = decimal_style


def _write_network_voltage_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Voltage')

    row_idx = 1
    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_vmag = dict()
            expected_vang = dict()

            for node in network.nodes:
                expected_vmag[node.bus_i] = [0.0 for _ in range(network.num_instants)]
                expected_vang[node.bus_i] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['voltage']['vmag']:

                        v_min, v_max = network.get_node_voltage_limits(node_id)

                        # Voltage magnitude
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Vmag, [p.u.]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            v_mag = results[year][day]['scenarios'][s_m][s_o]['voltage']['vmag'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = v_mag
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if v_mag > v_max + VIOLATION_TOLERANCE or v_mag < v_min - VIOLATION_TOLERANCE:
                                sheet.cell(row=row_idx, column=p + 7).fill = violation_fill
                            expected_vmag[node_id][p] += v_mag * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Voltage angle
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Vang, [º]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            v_ang = results[year][day]['scenarios'][s_m][s_o]['voltage']['vang'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = v_ang
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            expected_vang[node_id][p] += v_ang * omega_m * omega_s
                        row_idx = row_idx + 1

            for node in network.nodes:

                node_id = node.bus_i
                v_min, v_max = network.get_node_voltage_limits(node_id)

                # Expected voltage magnitude
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Vmag, [p.u.]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_vmag[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    if expected_vmag[node_id][p] > v_max + VIOLATION_TOLERANCE or expected_vmag[node_id][p] < v_min - VIOLATION_TOLERANCE:
                        sheet.cell(row=row_idx, column=p + 7).fill = violation_fill
                row_idx = row_idx + 1

                # Expected voltage angle
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Vang, [º]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_vang[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1


def _write_network_consumption_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Consumption')

    row_idx = 1
    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Load ID'
    sheet.cell(row=row_idx, column=2).value = 'Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Year'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 8).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_pc = dict()
            expected_qc = dict()
            expected_pc_flex = dict()
            expected_qc_flex = dict()
            expected_pc_curt = dict()
            expected_qc_curt = dict()
            expected_pnet = dict()
            expected_qnet = dict()
            for load in network.loads:
                expected_pc[load.load_id] = [0.0 for _ in range(network.num_instants)]
                expected_qc[load.load_id] = [0.0 for _ in range(network.num_instants)]
                expected_pc_flex[load.load_id] = [0.0 for _ in range(network.num_instants)]
                expected_qc_flex[load.load_id] = [0.0 for _ in range(network.num_instants)]
                expected_pc_curt[load.load_id] = [0.0 for _ in range(network.num_instants)]
                expected_qc_curt[load.load_id] = [0.0 for _ in range(network.num_instants)]
                expected_pnet[load.load_id] = [0.0 for _ in range(network.num_instants)]
                expected_qnet[load.load_id] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]

                    for load in network.loads:

                        load_id = load.load_id
                        node_id = load.bus

                        # - Active Power
                        sheet.cell(row=row_idx, column=1).value = load_id
                        sheet.cell(row=row_idx, column=2).value = node_id
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Pc, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            pc = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc'][load_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = pc
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_pc[load_id][p] += pc * omega_m * omega_s
                        row_idx = row_idx + 1

                        if network_planning.params.fl_reg:

                            # - Flexibility, Pc
                            sheet.cell(row=row_idx, column=1).value = load_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'Pc_flex, [MW]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(network.num_instants):
                                pc_flex = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc_flex'][load_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = pc_flex
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                expected_pc_flex[load_id][p] += pc_flex * omega_m * omega_s
                            row_idx = row_idx + 1

                        if network_planning.params.l_curt:

                            # - Active power curtailment
                            sheet.cell(row=row_idx, column=1).value = load_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'Pc_curt, [MW]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(network.num_instants):
                                pc_curt = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc_curt'][load_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = pc_curt
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                if not isclose(pc_curt, 0.00, abs_tol= VIOLATION_TOLERANCE):
                                    sheet.cell(row=row_idx, column=p + 8).fill = violation_fill
                                expected_pc_curt[load_id][p] += pc_curt * omega_m * omega_s
                            row_idx = row_idx + 1

                        if network_planning.params.fl_reg or network_planning.params.l_curt:

                            # - Active power net consumption
                            sheet.cell(row=row_idx, column=1).value = load_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'Pc_net, [MW]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(network.num_instants):
                                p_net = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc_net'][load_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = p_net
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                expected_pnet[load_id][p] += p_net * omega_m * omega_s
                            row_idx = row_idx + 1

                        # - Reactive power
                        sheet.cell(row=row_idx, column=1).value = load_id
                        sheet.cell(row=row_idx, column=2).value = node_id
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Qc, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            qc = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc'][load_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = qc
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_qc[load_id][p] += qc * omega_m * omega_s
                        row_idx = row_idx + 1

                        if network_planning.params.fl_reg:

                            sheet.cell(row=row_idx, column=1).value = load_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'Qc_flex, [MVAr]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(network.num_instants):
                                qc_flex = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc_flex'][load_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = qc_flex
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                expected_qc_flex[load_id][p] += qc_flex * omega_m * omega_s
                            row_idx = row_idx + 1

                        if network_planning.params.l_curt:

                            # - Reactive power curtailment
                            sheet.cell(row=row_idx, column=1).value = load_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'Qc_curt, [MW]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(network.num_instants):
                                qc_curt = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc_curt'][load_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = qc_curt
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                if not isclose(qc_curt, 0.00, abs_tol=VIOLATION_TOLERANCE):
                                    sheet.cell(row=row_idx, column=p + 8).fill = violation_fill
                                expected_qc_curt[load_id][p] += qc_curt * omega_m * omega_s
                            row_idx = row_idx + 1

                            # - Reactive power net consumption
                            sheet.cell(row=row_idx, column=1).value = load_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'Qc_net, [MW]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(network.num_instants):
                                q_net = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc_net'][load_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = q_net
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                expected_qnet[load_id][p] += q_net * omega_m * omega_s
                            row_idx = row_idx + 1

                        if network_planning.params.fl_reg or network_planning.params.l_curt:

                            # - Reactive power net consumption
                            sheet.cell(row=row_idx, column=1).value = load_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = 'Qc_net, [MVAr]'
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(network.num_instants):
                                q_net = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc_net'][load_id][p]
                                sheet.cell(row=row_idx, column=p + 8).value = q_net
                                sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                expected_qnet[load_id][p] += q_net * omega_m * omega_s
                            row_idx = row_idx + 1

            for load in network.loads:

                load_id = load.load_id
                node_id = load.bus

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = load_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Pc, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_pc[load_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                if network_planning.params.fl_reg:

                    # - Flexibility, Pc
                    sheet.cell(row=row_idx, column=1).value = load_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = int(year)
                    sheet.cell(row=row_idx, column=4).value = day
                    sheet.cell(row=row_idx, column=5).value = 'Pc_flex, [MW]'
                    sheet.cell(row=row_idx, column=6).value = 'Expected'
                    sheet.cell(row=row_idx, column=7).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 8).value = expected_pc_flex[load_id][p]
                        sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                    row_idx = row_idx + 1

                if network_planning.params.l_curt:

                    # - Load curtailment (active power)
                    sheet.cell(row=row_idx, column=1).value = load_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = int(year)
                    sheet.cell(row=row_idx, column=4).value = day
                    sheet.cell(row=row_idx, column=5).value = 'Pc_curt, [MW]'
                    sheet.cell(row=row_idx, column=6).value = 'Expected'
                    sheet.cell(row=row_idx, column=7).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 8).value = expected_pc_curt[load_id][p]
                        sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                        if not isclose(expected_pc_curt[load_id][p], 0.00, abs_tol=VIOLATION_TOLERANCE):
                            sheet.cell(row=row_idx, column=p + 8).fill = violation_fill
                    row_idx = row_idx + 1

                if network_planning.params.fl_reg or network_planning.params.l_curt:

                    # - Active power net consumption
                    sheet.cell(row=row_idx, column=1).value = load_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = int(year)
                    sheet.cell(row=row_idx, column=4).value = day
                    sheet.cell(row=row_idx, column=5).value = 'Pc_net, [MW]'
                    sheet.cell(row=row_idx, column=6).value = 'Expected'
                    sheet.cell(row=row_idx, column=7).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 8).value = expected_pnet[load_id][p]
                        sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                    row_idx = row_idx + 1

                # - Reactive power
                sheet.cell(row=row_idx, column=1).value = load_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Qc, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_qc[load_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                if network_planning.params.fl_reg:

                    # - Flexibility, Qc
                    sheet.cell(row=row_idx, column=1).value = load_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = int(year)
                    sheet.cell(row=row_idx, column=4).value = day
                    sheet.cell(row=row_idx, column=5).value = 'Qc_flex, [MVAr]'
                    sheet.cell(row=row_idx, column=6).value = 'Expected'
                    sheet.cell(row=row_idx, column=7).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 8).value = expected_qc_flex[load_id][p]
                        sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                    row_idx = row_idx + 1

                if network_planning.params.l_curt:

                    # - Load curtailment (reactive power)
                    sheet.cell(row=row_idx, column=1).value = load_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = int(year)
                    sheet.cell(row=row_idx, column=4).value = day
                    sheet.cell(row=row_idx, column=5).value = 'Qc_curt, [MW]'
                    sheet.cell(row=row_idx, column=6).value = 'Expected'
                    sheet.cell(row=row_idx, column=7).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 8).value = expected_qc_curt[load_id][p]
                        sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                        if not isclose(expected_qc_curt[load_id][p], 0.00, abs_tol=VIOLATION_TOLERANCE):
                            sheet.cell(row=row_idx, column=p + 8).fill = violation_fill
                    row_idx = row_idx + 1

                    # - Reactive power net consumption
                    sheet.cell(row=row_idx, column=1).value = load_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = int(year)
                    sheet.cell(row=row_idx, column=4).value = day
                    sheet.cell(row=row_idx, column=5).value = 'Qc_net, [MW]'
                    sheet.cell(row=row_idx, column=6).value = 'Expected'
                    sheet.cell(row=row_idx, column=7).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 8).value = expected_qnet[load_id][p]
                        sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                    row_idx = row_idx + 1

                if network_planning.params.fl_reg or network_planning.params.l_curt:

                    # - Reactive power net consumption
                    sheet.cell(row=row_idx, column=1).value = load_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = int(year)
                    sheet.cell(row=row_idx, column=4).value = day
                    sheet.cell(row=row_idx, column=5).value = 'Qc_net, [MVAr]'
                    sheet.cell(row=row_idx, column=6).value = 'Expected'
                    sheet.cell(row=row_idx, column=7).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 8).value = expected_qnet[load_id][p]
                        sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                    row_idx = row_idx + 1


def _write_network_generation_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Generation')

    row_idx = 1
    decimal_style = '0.00'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Generator ID'
    sheet.cell(row=row_idx, column=2).value = 'Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Type'
    sheet.cell(row=row_idx, column=4).value = 'Year'
    sheet.cell(row=row_idx, column=5).value = 'Day'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=8).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 9).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_pg = dict()
            expected_pg_net = dict()
            expected_qg = dict()
            expected_qg_net = dict()
            expected_sg = dict()
            expected_sg_curt = dict()
            expected_sg_net = dict()
            for generator in network.generators:
                expected_pg[generator.gen_id] = [0.0 for _ in range(network.num_instants)]
                expected_pg_net[generator.gen_id] = [0.0 for _ in range(network.num_instants)]
                expected_qg[generator.gen_id] = [0.0 for _ in range(network.num_instants)]
                expected_qg_net[generator.gen_id] = [0.0 for _ in range(network.num_instants)]
                expected_sg[generator.gen_id] = [0.0 for _ in range(network.num_instants)]
                expected_sg_curt[generator.gen_id] = [0.0 for _ in range(network.num_instants)]
                expected_sg_net[generator.gen_id] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for generator in network.generators:

                        gen_id = generator.gen_id
                        node_id = generator.bus
                        gen_type = network.get_gen_type(gen_id)

                        # Active Power
                        sheet.cell(row=row_idx, column=1).value = gen_id
                        sheet.cell(row=row_idx, column=2).value = node_id
                        sheet.cell(row=row_idx, column=3).value = gen_type
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Pg, [MW]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            pg = results[year][day]['scenarios'][s_m][s_o]['generation']['pg'][gen_id][p]
                            sheet.cell(row=row_idx, column=p + 9).value = pg
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_pg[gen_id][p] += pg * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Active Power net
                        if generator.is_curtaillable() and network_planning.params.rg_curt:

                            sheet.cell(row=row_idx, column=1).value = gen_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = gen_type
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Pg_net, [MW]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network.num_instants):
                                pg_net = results[year][day]['scenarios'][s_m][s_o]['generation']['pg_net'][gen_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = pg_net
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                expected_pg_net[gen_id][p] += pg_net * omega_m * omega_s
                            row_idx = row_idx + 1

                        # Reactive Power
                        sheet.cell(row=row_idx, column=1).value = gen_id
                        sheet.cell(row=row_idx, column=2).value = node_id
                        sheet.cell(row=row_idx, column=3).value = gen_type
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Qg, [MVAr]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            qg = results[year][day]['scenarios'][s_m][s_o]['generation']['qg'][gen_id][p]
                            sheet.cell(row=row_idx, column=p + 9).value = qg
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_qg[gen_id][p] += qg * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Reactive Power net
                        if generator.is_curtaillable() and network_planning.params.rg_curt:

                            sheet.cell(row=row_idx, column=1).value = gen_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = gen_type
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Qg_net, [MVAr]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network.num_instants):
                                qg_net = results[year][day]['scenarios'][s_m][s_o]['generation']['qg_net'][gen_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = qg_net
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                expected_qg_net[gen_id][p] += qg_net * omega_m * omega_s
                            row_idx = row_idx + 1

                        # Apparent Power
                        if generator.is_curtaillable() and network_planning.params.rg_curt:

                            sheet.cell(row=row_idx, column=1).value = gen_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = gen_type
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Sg, [MVA]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network.num_instants):
                                sg = results[year][day]['scenarios'][s_m][s_o]['generation']['sg'][gen_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = sg
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                expected_sg[gen_id][p] += sg * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Apparent Power curtailed
                            sheet.cell(row=row_idx, column=1).value = gen_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = gen_type
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Sg_curt, [MVA]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network.num_instants):
                                sg_curt = results[year][day]['scenarios'][s_m][s_o]['generation']['sg_curt'][gen_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = sg_curt
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                if not isclose(sg_curt, 0.00, abs_tol= VIOLATION_TOLERANCE):
                                    sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                                expected_sg_curt[gen_id][p] += sg_curt * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Apparent Power net
                            sheet.cell(row=row_idx, column=1).value = gen_id
                            sheet.cell(row=row_idx, column=2).value = node_id
                            sheet.cell(row=row_idx, column=3).value = gen_type
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Sg_net, [MVA]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network.num_instants):
                                sg_net = results[year][day]['scenarios'][s_m][s_o]['generation']['sg_net'][gen_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = sg_net
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                expected_sg_net[gen_id][p] += sg_net * omega_m * omega_s
                            row_idx = row_idx + 1

            for generator in network.generators:

                node_id = generator.bus
                gen_id = generator.gen_id
                gen_type = network.get_gen_type(gen_id)

                # Active Power
                sheet.cell(row=row_idx, column=1).value = gen_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = gen_type
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Pg, [MW]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_pg[gen_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                # Active Power net
                if generator.is_curtaillable() and network_planning.params.rg_curt:
                    sheet.cell(row=row_idx, column=1).value = gen_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = gen_type
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Pg_net, [MW]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_pg_net[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    row_idx = row_idx + 1

                # Reactive Power
                sheet.cell(row=row_idx, column=1).value = gen_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = gen_type
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Qg, [MVAr]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_qg[gen_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                # Reactive Power net
                if generator.is_curtaillable() and network_planning.params.rg_curt:

                    sheet.cell(row=row_idx, column=1).value = gen_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = gen_type
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Qg_net, [MVAr]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_qg_net[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    row_idx = row_idx + 1

                # Apparent Power
                if generator.is_curtaillable() and network_planning.params.rg_curt:

                    sheet.cell(row=row_idx, column=1).value = gen_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = gen_type
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Sg, [MVA]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_sg[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    row_idx = row_idx + 1

                    sheet.cell(row=row_idx, column=1).value = gen_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = gen_type
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Sg_curt, [MVA]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_sg_curt[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                        if not isclose(expected_sg_curt[gen_id][p], 0.00, abs_tol=VIOLATION_TOLERANCE):
                            sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                    row_idx = row_idx + 1

                    sheet.cell(row=row_idx, column=1).value = gen_id
                    sheet.cell(row=row_idx, column=2).value = node_id
                    sheet.cell(row=row_idx, column=3).value = gen_type
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Sg_net, [MVA]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_sg_net[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    row_idx = row_idx + 1


def _write_network_branch_results_to_excel(network_planning, workbook, results, result_type):

    sheet_name = str()
    aux_string = str()
    if result_type == 'losses':
        sheet_name = 'Branch Losses'
        aux_string = 'P, [MW]'
    elif result_type == 'ratio':
        sheet_name = 'Transformer Ratio'
        aux_string = 'Ratio'

    row_idx = 1
    decimal_style = '0.00'

    sheet = workbook.create_sheet(sheet_name)

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Branch ID'
    sheet.cell(row=row_idx, column=2).value = 'From Node ID'
    sheet.cell(row=row_idx, column=3).value = 'To Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Year'
    sheet.cell(row=row_idx, column=5).value = 'Day'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=8).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 9).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_values = dict()
            for branch in network.branches:
                expected_values[branch.branch_id] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for branch in network.branches:
                        branch_id = branch.branch_id
                        if not(result_type == 'ratio' and not branch.is_transformer):

                            sheet.cell(row=row_idx, column=1).value = branch_id
                            sheet.cell(row=row_idx, column=2).value = branch.fbus
                            sheet.cell(row=row_idx, column=3).value = branch.tbus
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = aux_string
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network.num_instants):
                                value = results[year][day]['scenarios'][s_m][s_o]['branches'][result_type][branch_id][p]
                                sheet.cell(row=row_idx, column=p + 9).value = value
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                expected_values[branch_id][p] += value * omega_m * omega_s
                            row_idx = row_idx + 1

            for branch in network.branches:
                branch_id = branch.branch_id
                if not (result_type == 'ratio' and not branch.is_transformer):

                    sheet.cell(row=row_idx, column=1).value = branch_id
                    sheet.cell(row=row_idx, column=2).value = branch.fbus
                    sheet.cell(row=row_idx, column=3).value = branch.tbus
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = aux_string
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network.num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_values[branch_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    row_idx = row_idx + 1


def _write_network_branch_loading_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Branch Loading')

    row_idx = 1
    perc_style = '0.00%'
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Branch ID'
    sheet.cell(row=row_idx, column=2).value = 'From Node ID'
    sheet.cell(row=row_idx, column=3).value = 'To Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Year'
    sheet.cell(row=row_idx, column=5).value = 'Day'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=8).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 9).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_values = {'flow_ij': {}}
            for branch in network.branches:
                expected_values['flow_ij'][branch.branch_id] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for branch in network.branches:

                        # flow ij, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = branch.tbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Flow_ij, [%]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['branch_flow']['flow_ij_perc'][branch.branch_id][p]
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                            if value > 1.00 + VIOLATION_TOLERANCE:
                                sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                            expected_values['flow_ij'][branch.branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

            for branch in network.branches:

                # flow ij, [%]
                sheet.cell(row=row_idx, column=1).value = branch.branch_id
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Flow_ij, [%]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    value = expected_values['flow_ij'][branch.branch_id][p]
                    sheet.cell(row=row_idx, column=p + 9).value = value
                    sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                    if value > 1.00 + VIOLATION_TOLERANCE:
                        sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                row_idx = row_idx + 1


def _write_network_branch_power_flow_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Power Flows')

    row_idx = 1
    decimal_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'BranchID'
    sheet.cell(row=row_idx, column=2).value = 'From Node ID'
    sheet.cell(row=row_idx, column=3).value = 'To Node ID'
    sheet.cell(row=row_idx, column=4).value = 'Year'
    sheet.cell(row=row_idx, column=5).value = 'Day'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=8).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 9).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_values = {'pij': {}, 'pji': {}, 'qij': {}, 'qji': {}, 'sij': {}, 'sji': {}}
            for branch in network.branches:
                branch_id = branch.branch_id
                expected_values['pij'][branch_id] = [0.0 for _ in range(network.num_instants)]
                expected_values['pji'][branch_id] = [0.0 for _ in range(network.num_instants)]
                expected_values['qij'][branch_id] = [0.0 for _ in range(network.num_instants)]
                expected_values['qji'][branch_id] = [0.0 for _ in range(network.num_instants)]
                expected_values['sij'][branch_id] = [0.0 for _ in range(network.num_instants)]
                expected_values['sji'][branch_id] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for branch in network.branches:

                        branch_id = branch.branch_id
                        rating = branch.rate
                        if rating == 0.0:
                            rating = BRANCH_UNKNOWN_RATING

                        # Pij, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = branch.tbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pij'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_values['pij'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Pij, [%]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = branch.tbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'P, [%]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pij'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                        row_idx = row_idx + 1

                        # Pji, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = branch.fbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pji'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_values['pji'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Pji, [%]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = branch.fbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'P, [%]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pji'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                        row_idx = row_idx + 1

                        # Qij, [MVAr]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = branch.tbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qij'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_values['qij'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Qij, [%]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = branch.tbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Q, [%]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qij'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                        row_idx = row_idx + 1

                        # Qji, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = branch.fbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qji'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_values['qji'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Qji, [%]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = branch.fbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Q, [%]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qji'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                        row_idx = row_idx + 1

                        # Sij, [MVA]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = branch.tbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sij'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_values['sij'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Sij, [%]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = branch.tbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'S, [%]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sij'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                        row_idx = row_idx + 1

                        # Sji, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = branch.fbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sji'][branch_id][p]
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_values['sji'][branch_id][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Sji, [%]
                        sheet.cell(row=row_idx, column=1).value = branch_id
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = branch.fbus
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'S, [%]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network.num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sji'][branch_id][p] / rating)
                            sheet.cell(row=row_idx, column=p + 9).value = value
                            sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                        row_idx = row_idx + 1

            for branch in network.branches:

                branch_id = branch.branch_id
                rating = branch.rate
                if rating == 0.0:
                    rating = BRANCH_UNKNOWN_RATING

                # Pij, [MW]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_values['pij'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                # Pij, [%]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'P, [%]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = abs(expected_values['pij'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                row_idx = row_idx + 1

                # Pji, [MW]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_values['pji'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                # Pji, [%]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'P, [%]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = abs(expected_values['pji'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                row_idx = row_idx + 1

                # Qij, [MVAr]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_values['qij'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                # Qij, [%]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Q, [%]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = abs(expected_values['qij'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                row_idx = row_idx + 1

                # Qji, [MVAr]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_values['qji'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                # Qji, [%]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Q, [%]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = abs(expected_values['qji'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                row_idx = row_idx + 1

                # Sij, [MVA]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_values['sij'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                # Sij, [%]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = branch.tbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'S, [%]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = abs(expected_values['sij'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                row_idx = row_idx + 1

                # Sji, [MVA]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_values['sji'][branch_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                # Sji, [%]
                sheet.cell(row=row_idx, column=1).value = branch_id
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = branch.fbus
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'S, [%]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = abs(expected_values['sji'][branch_id][p]) / rating
                    sheet.cell(row=row_idx, column=p + 9).number_format = perc_style
                row_idx = row_idx + 1


def _write_network_energy_storage_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Energy Storage')

    row_idx = 1
    decimal_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'ESS ID'
    sheet.cell(row=row_idx, column=2).value = 'Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Year'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 8).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_p = dict()
            expected_q = dict()
            expected_s = dict()
            expected_soc = dict()
            expected_soc_perc = dict()

            for energy_storage in network.energy_storages:
                es_id = energy_storage.es_id
                expected_p[es_id] = [0.0 for _ in range(network.num_instants)]
                expected_q[es_id] = [0.0 for _ in range(network.num_instants)]
                expected_s[es_id] = [0.0 for _ in range(network.num_instants)]
                expected_soc[es_id] = [0.0 for _ in range(network.num_instants)]
                expected_soc_perc[es_id] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for energy_storage in network.energy_storages:

                        es_id = energy_storage.es_id
                        node_id = energy_storage.bus

                        # - Active Power
                        sheet.cell(row=row_idx, column=1).value = es_id
                        sheet.cell(row=row_idx, column=2).value = node_id
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            pc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['p'][es_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = pc
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if pc != 'N/A':
                                expected_p[es_id][p] += pc * omega_m * omega_s
                            else:
                                expected_p[es_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - Reactive Power
                        sheet.cell(row=row_idx, column=1).value = es_id
                        sheet.cell(row=row_idx, column=2).value = node_id
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            qc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['q'][es_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = qc
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if qc != 'N/A':
                                expected_q[es_id][p] += qc * omega_m * omega_s
                            else:
                                expected_q[es_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - Apparent Power
                        sheet.cell(row=row_idx, column=1).value = es_id
                        sheet.cell(row=row_idx, column=2).value = node_id
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            sc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['s'][es_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = sc
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if sc != 'N/A':
                                expected_s[es_id][p] += sc * omega_m * omega_s
                            else:
                                expected_s[es_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - SoC, [MWh]
                        sheet.cell(row=row_idx, column=1).value = es_id
                        sheet.cell(row=row_idx, column=2).value = node_id
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'SoC, [MWh]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            soc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['soc'][es_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = soc
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            if soc != 'N/A':
                                expected_soc[es_id][p] += soc * omega_m * omega_s
                            else:
                                expected_soc[es_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - SoC, [%]
                        sheet.cell(row=row_idx, column=1).value = es_id
                        sheet.cell(row=row_idx, column=2).value = node_id
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network.num_instants):
                            soc_perc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['soc_percent'][es_id][p]
                            sheet.cell(row=row_idx, column=p + 8).value = soc_perc
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                            if soc_perc != 'N/A':
                                expected_soc_perc[es_id][p] += soc_perc * omega_m * omega_s
                            else:
                                expected_soc_perc[es_id][p] = 'N/A'
                        row_idx = row_idx + 1

            for energy_storage in network.energy_storages:

                es_id = energy_storage.es_id
                node_id = energy_storage.bus

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = es_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_p[es_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # - Reactive Power
                sheet.cell(row=row_idx, column=1).value = es_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_q[es_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # - Apparent Power
                sheet.cell(row=row_idx, column=1).value = es_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_s[es_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoC, [MWh]
                sheet.cell(row=row_idx, column=1).value = es_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [MWh]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc[es_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoC, [%]
                sheet.cell(row=row_idx, column=1).value = es_id
                sheet.cell(row=row_idx, column=2).value = node_id
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_soc_perc[es_id][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1


def _write_shared_network_energy_storage_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Shared Energy Storage')

    row_idx = 1
    decimal_style = '0.00'
    perc_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:

            network = network_planning.network[year][day]

            expected_p = dict()
            expected_q = dict()
            expected_s = dict()
            expected_soc = dict()
            expected_soc_perc = dict()

            for energy_storage in network.shared_energy_storages:
                expected_p[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]
                expected_q[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]
                expected_s[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]
                expected_soc[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]
                expected_soc_perc[energy_storage.bus] = [0.0 for _ in range(network.num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network.prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network.prob_operation_scenarios[s_o]
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['p']:

                        # - Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            pc = results[year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['p'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = pc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if pc != 'N/A':
                                expected_p[node_id][p] += pc * omega_m * omega_s
                            else:
                                expected_p[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - Reactive Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            qc = results[year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['q'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = qc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if qc != 'N/A':
                                expected_q[node_id][p] += qc * omega_m * omega_s
                            else:
                                expected_q[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - Apparent Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            sc = results[year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['s'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = sc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if sc != 'N/A':
                                expected_s[node_id][p] += sc * omega_m * omega_s
                            else:
                                expected_s[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - SoC, [MWh]
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [MWh]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            soc = results[year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['soc'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = soc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if soc != 'N/A':
                                expected_soc[node_id][p] += soc * omega_m * omega_s
                            else:
                                expected_soc[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

                        # - SoC, [%]
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network.num_instants):
                            soc_perc = results[year][day]['scenarios'][s_m][s_o]['shared_energy_storages']['soc_percent'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = soc_perc
                            sheet.cell(row=row_idx, column=p + 7).number_format = perc_style
                            if soc_perc != 'N/A':
                                expected_soc_perc[node_id][p] += soc_perc * omega_m * omega_s
                            else:
                                expected_soc_perc[node_id][p] = 'N/A'
                        row_idx = row_idx + 1

            for energy_storage in network.shared_energy_storages:

                node_id = energy_storage.bus

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_p[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - Reactive Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_q[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - Apparent Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_s[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoC, [MWh]
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [MWh]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - SoC, [%]
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc_perc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = perc_style
                row_idx = row_idx + 1


def _write_relaxation_slacks_scenarios_results_to_excel(network_planning, workbook, results):

    sheet = workbook.create_sheet('Relaxation Slacks, Operation')

    row_idx = 1
    decimal_style = '0.00'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Resource ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(network_planning.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    for year in results:
        for day in results[year]:
            network = network_planning.network[year][day]
            for s_m in results[year][day]['scenarios']:
                for s_o in results[year][day]['scenarios'][s_m]:

                    # Voltage slacks
                    if network_planning.params.slacks.grid_operation.voltage:
                        for node in network.nodes:

                            node_id = node.bus_i

                            # - slack_e
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Voltage, e'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                slack_e = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['voltage']['e'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = slack_e
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                            # - slack_f
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Voltage, f'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                slack_f = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['voltage']['f'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = slack_f
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                    # Branch flow slacks
                    if network_planning.params.slacks.grid_operation.branch_flow:
                        for branch in network.branches:

                            branch_id = branch.branch_id

                            sheet.cell(row=row_idx, column=1).value = branch_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Flow_ij_sqr'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                iij_sqr = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['branch_flow']['flow_ij_sqr'][branch_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = iij_sqr
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                    # Node balance
                    for node in network.nodes:

                        node_id = node.bus_i

                        # - slack_p
                        if network_planning.params.slacks.node_balance.active_power:
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Node balance, p'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                slack_p = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['node_balance']['p'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = slack_p
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                        # - slack_q
                        if network_planning.params.slacks.node_balance.reactive_power:
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Node balance, q'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                slack_q = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['node_balance']['q'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = slack_q
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                    # Shared ESS
                    for shared_energy_storage in network.shared_energy_storages:

                        node_id = shared_energy_storage.bus

                        # - Complementarity
                        if network_planning.params.slacks.shared_ess.complementarity:
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Shared Energy Storage, comp'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                comp = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['shared_energy_storages']['comp'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = comp
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                        # - Day balance
                        if network_planning.params.slacks.shared_ess.day_balance:

                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Shared Energy Storage, soc_final'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network_planning.num_instants):
                                soc_final_up = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['shared_energy_storages']['soc_final'][node_id]
                                sheet.cell(row=row_idx, column=p + 7).value = soc_final_up
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            row_idx = row_idx + 1

                    # Load flexibility
                    if network_planning.params.fl_reg:
                        for load in network.loads:

                            load_id = load.load_id

                            # - Day balance
                            if network_planning.params.slacks.flexibility.day_balance:

                                sheet.cell(row=row_idx, column=1).value = load_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'Flexibility, balance, p'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    day_balance_p = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['flexibility']['day_balance'][load_id]['p']
                                    sheet.cell(row=row_idx, column=p + 7).value = day_balance_p
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                                sheet.cell(row=row_idx, column=1).value = load_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'Flexibility, balance, q'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    day_balance_q = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['flexibility']['day_balance'][load_id]['q']
                                    sheet.cell(row=row_idx, column=p + 7).value = day_balance_q
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                    # ESS
                    if network_planning.params.es_reg:
                        for energy_storage in network.energy_storages:

                            es_id = energy_storage.es_id

                            # - Complementarity
                            if network_planning.params.slacks.ess.complementarity:
                                sheet.cell(row=row_idx, column=1).value = es_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'Energy Storage, comp'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    comp = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['comp'][es_id][p]
                                    sheet.cell(row=row_idx, column=p + 7).value = comp
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1

                            # - Day balance
                            if network_planning.params.slacks.ess.day_balance:

                                sheet.cell(row=row_idx, column=1).value = es_id
                                sheet.cell(row=row_idx, column=2).value = int(year)
                                sheet.cell(row=row_idx, column=3).value = day
                                sheet.cell(row=row_idx, column=4).value = 'Energy Storage, soc_final'
                                sheet.cell(row=row_idx, column=5).value = s_m
                                sheet.cell(row=row_idx, column=6).value = s_o
                                for p in range(network_planning.num_instants):
                                    soc_final = results[year][day]['scenarios'][s_m][s_o]['relaxation_slacks']['energy_storages']['soc_final'][es_id]
                                    sheet.cell(row=row_idx, column=p + 7).value = soc_final
                                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                row_idx = row_idx + 1


# ======================================================================================================================
#  Plot functions
# ======================================================================================================================
def _plot_load_data_scenarios(network_planning, years_to_plot, save_dir, save_format='pdf'):

    print('[INFO]\t - Plotting load scenarios...')

    hours = np.arange(network_planning.num_instants)
    xticks = np.arange(0, network_planning.num_instants, 4)
    xtick_labels = [f"{h:02d}:00" for h in xticks]

    for year in years_to_plot:
        for season in network_planning.days:

            network = network_planning.network[year][season]
            fig, axs = plt.subplots(2, 1, figsize=(12, 10), sharex=True)

            num_colors = len(network.loads)
            color_map = plt.cm.get_cmap('plasma', num_colors)
            colors = [color_map(i / num_colors) for i in range(num_colors)]

            for l in range(len(network.loads)):

                load = network.loads[l]
                pc = load.pd * network.baseMVA
                qc = load.qd * network.baseMVA

                pc_mean = pc.mean(axis=0)
                pc_std = pc.std(axis=0)
                qc_mean = qc.mean(axis=0)
                qc_std = qc.std(axis=0)

                # Plot
                color = colors[l]
                axs[0].plot(hours, pc_mean, label=f'Load {load.load_id}', color=color)
                axs[0].fill_between(hours, pc_mean - pc_std, pc_mean + pc_std, alpha=0.2, color=color)
                axs[0].set_ylabel("Active Power, [MW]", fontsize=14)

                axs[1].plot(hours, qc_mean, label=f'Load {load.load_id}', color=color)
                axs[1].fill_between(hours, qc_mean - qc_std, qc_mean + qc_std, alpha=0.2, color=color)
                axs[1].set_ylabel("Reactive Power, [MVAr]", fontsize=14)

            for ax in axs:
                ax.set_xticks(xticks)
                ax.set_xticklabels(xtick_labels)
                ax.set_xlim(0, 23)
                ax.set_xlabel("Hour", loc='center', fontsize=12)
                ax.grid(True, axis='x', which='both')
                ax.tick_params(axis='both', labelsize=12)
                ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.2f'))

            handles, labels = axs[0].get_legend_handles_labels()
            fig.legend(handles, labels, loc='right', fontsize='small', frameon=True)

            # plt.tight_layout()
            filename = os.path.join(save_dir, f"{network.name}_load_scenarios_{year}_{season}.{save_format}")
            plt.savefig(filename)
            plt.close(fig)


def _plot_res_data_scenarios(network_planning, years_to_plot, save_dir, save_format='pdf'):

    print('[INFO]\t - Plotting RES generation scenarios...')

    hours = np.arange(network_planning.num_instants)
    xticks = np.arange(0, network_planning.num_instants, 4)
    xtick_labels = [f"{h:02d}:00" for h in xticks]
    for year in years_to_plot:
        for season in network_planning.days:

            network = network_planning.network[year][season]
            fig, axs = plt.subplots(1, 1, figsize=(8, 6), sharex=True)

            num_colors = len(network.generators)
            color_map = plt.cm.get_cmap('viridis', num_colors)
            colors = [color_map(i / num_colors) for i in range(num_colors)]

            for g in range(len(network.generators)):

                generator = network.generators[g]

                if generator.is_curtaillable():
                    pg = generator.pg * network.baseMVA
                    pg_mean = pg.mean(axis=0)
                    pg_std = pg.std(axis=0)

                    # Plot
                    color = colors[g]
                    axs.plot(hours, pg_mean, label=f'Generator {generator.gen_id}', color=color)
                    axs.fill_between(hours, pg_mean - pg_std, pg_mean + pg_std, alpha=0.2, color=color)
                    axs.set_ylabel("Active Power, [MW]", fontsize=14)
                    axs.set_xticks(xticks)
                    axs.set_xticklabels(xtick_labels)
                    axs.set_xlim(0, 23)
                    axs.set_xlabel("Hour", loc='center', fontsize=12)
                    axs.grid(True, which='both')
                    axs.tick_params(axis='both', labelsize=12)
                    axs.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.2f'))
                    axs.legend(loc='best', fontsize='small', frameon=True)

            plt.tight_layout()
            filename = os.path.join(save_dir, f"{network.name}_RES_generation_scenarios_{year}_{season}.{save_format}")
            plt.savefig(filename)
            plt.close(fig)


# ======================================================================================================================
#  OTHER (auxiliary) functions
# ======================================================================================================================
def _get_sensitivities(network_planning, model):

    sensitivities = {'s': dict(), 'e': dict()}
    for year in network_planning.years:
        sensitivities['s'][year] = dict()
        sensitivities['e'][year] = dict()
        for node_id in network_planning.active_distribution_network_nodes:
            sensitivities['s'][year][node_id] = 0.00
            sensitivities['e'][year][node_id] = 0.00

    for year in network_planning.years:

        num_years = network_planning.years[year]

        for day in network_planning.days:

            num_days = network_planning.days[day]
            model_repr_day = model[year][day]

            for c in model_repr_day.shared_energy_storage_s_sensitivities:
                node_id = network_planning.active_distribution_network_nodes[c - 1]  # Note: the sensitivity constraints start at "1"
                sensitivity_s = model_repr_day.dual[model_repr_day.shared_energy_storage_s_sensitivities[c]] * network_planning.network[year][day].baseMVA
                sensitivities['s'][year][node_id] += (num_days / 365.00) * sensitivity_s

            for c in model_repr_day.shared_energy_storage_e_sensitivities:
                node_id = network_planning.active_distribution_network_nodes[c - 1]
                sensitivity_e = model_repr_day.dual[model_repr_day.shared_energy_storage_e_sensitivities[c]] * network_planning.network[year][day].baseMVA
                sensitivities['e'][year][node_id] += (num_days / 365.00) * sensitivity_e

        # Note: annualization is already considered in the master problem's OF
        for node_id in network_planning.active_distribution_network_nodes:
            sensitivities['s'][year][node_id] *= 365.00 * num_years
            sensitivities['e'][year][node_id] *= 365.00 * num_years

    return sensitivities


def _update_data_with_candidate_solution(network_planning, candidate_solution):
    if network_planning.is_transmission:
        for node_id in network_planning.active_distribution_network_nodes:
            for year in network_planning.years:
                for day in network_planning.days:
                    shared_ess_idx = network_planning.network[year][day].get_shared_energy_storage_idx(node_id)
                    network_planning.network[year][day].shared_energy_storages[shared_ess_idx].s = candidate_solution[node_id][year]['s'] / network_planning.network[year][day].baseMVA
                    network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e = candidate_solution[node_id][year]['e'] / network_planning.network[year][day].baseMVA
                    network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_init = candidate_solution[node_id][year]['e'] * ENERGY_STORAGE_RELATIVE_INIT_SOC / network_planning.network[year][day].baseMVA
                    network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_min = candidate_solution[node_id][year]['e'] * ENERGY_STORAGE_MIN_ENERGY_STORED / network_planning.network[year][day].baseMVA
                    network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_max = candidate_solution[node_id][year]['e'] * ENERGY_STORAGE_MAX_ENERGY_STORED / network_planning.network[year][day].baseMVA
    else:
        tn_node_id = network_planning.tn_connection_nodeid
        for year in network_planning.years:
            for day in network_planning.days:
                ref_node_id = network_planning.network[year][day].get_reference_node_id()
                shared_ess_idx = network_planning.network[year][day].get_shared_energy_storage_idx(ref_node_id)
                network_planning.network[year][day].shared_energy_storages[shared_ess_idx].s = candidate_solution[tn_node_id][year]['s'] / network_planning.network[year][day].baseMVA
                network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e = candidate_solution[tn_node_id][year]['e'] / network_planning.network[year][day].baseMVA
                network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_init = candidate_solution[tn_node_id][year]['e'] * ENERGY_STORAGE_RELATIVE_INIT_SOC / network_planning.network[year][day].baseMVA
                network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_min = candidate_solution[tn_node_id][year]['e'] * ENERGY_STORAGE_MIN_ENERGY_STORED / network_planning.network[year][day].baseMVA
                network_planning.network[year][day].shared_energy_storages[shared_ess_idx].e_max = candidate_solution[tn_node_id][year]['e'] * ENERGY_STORAGE_MAX_ENERGY_STORED / network_planning.network[year][day].baseMVA


def _update_model_with_candidate_solution(network, model, candidate_solution):
    if network.is_transmission:
        for year in network.years:
            for day in network.days:
                s_base = network.network[year][day].baseMVA
                for node_id in network.active_distribution_network_nodes:
                    shared_ess_idx = network.network[year][day].get_shared_energy_storage_idx(node_id)
                    model[year][day].shared_es_s_rated_fixed[shared_ess_idx].set_value(abs(candidate_solution[node_id][year]['s']) / s_base)
                    model[year][day].shared_es_e_rated_fixed[shared_ess_idx].set_value(abs(candidate_solution[node_id][year]['e']) / s_base)
    else:
        tn_node_id = network.tn_connection_nodeid
        for year in network.years:
            for day in network.days:
                s_base = network.network[year][day].baseMVA
                ref_node_id = network.network[year][day].get_reference_node_id()
                shared_ess_idx = network.network[year][day].get_shared_energy_storage_idx(ref_node_id)
                model[year][day].shared_es_s_rated_fixed[shared_ess_idx].set_value(abs(candidate_solution[tn_node_id][year]['s']) / s_base)
                model[year][day].shared_es_e_rated_fixed[shared_ess_idx].set_value(abs(candidate_solution[tn_node_id][year]['e']) / s_base)
